import sys
import io
from io import BytesIO
import csv
import subprocess
import os

# Kiểm tra và cài đặt dependencies tự động (chỉ khi chạy trực tiếp)
def check_and_install_dependencies():
    """Kiểm tra và cài đặt dependencies nếu thiếu"""
    # Tránh vòng lặp vô hạn
    if os.environ.get('APP_SKIP_CHECK') == '1':
        return
    
    required_packages = {
        'flask': 'flask',
        'openpyxl': 'openpyxl', 
        'flask_login': 'flask-login',
        'flask_wtf': 'flask-wtf',
        'werkzeug': 'werkzeug',
        'sqlalchemy': 'sqlalchemy',
        'flask_migrate': 'flask-migrate',
        'jinja2': 'jinja2',
        'reportlab': 'reportlab',
        'selenium': 'selenium',
        'webdriver_manager': 'webdriver-manager',
        'PIL': 'Pillow',
        'numpy': 'numpy'
    }
    
    missing_packages = []
    for module_name, package_name in required_packages.items():
        try:
            if module_name == 'PIL':
                __import__('PIL')
            else:
                __import__(module_name)
        except ImportError:
            missing_packages.append(package_name)
    
    if missing_packages:
        print("=" * 70)
        print("⚠️  PHÁT HIỆN THIẾU CÁC THƯ VIỆN CẦN THIẾT")
        print("=" * 70)
        print(f"\nCác thư viện thiếu: {', '.join(missing_packages)}")
        print("\nĐang tự động cài đặt...")
        print("-" * 70)
        
        try:
            # Cài đặt từ requirements.txt nếu có
            if os.path.exists('requirements.txt'):
                print("📦 Đang cài đặt từ requirements.txt...")
                result = subprocess.run(
                    [sys.executable, '-m', 'pip', 'install', '-r', 'requirements.txt'],
                    check=False
                )
                if result.returncode == 0:
                    print("✅ Cài đặt thành công!")
                    print("\n🔄 Đang khởi động lại ứng dụng...")
                    os.environ['APP_SKIP_CHECK'] = '1'
                    os.execv(sys.executable, [sys.executable] + sys.argv)
                else:
                    print("⚠️  Có lỗi khi cài đặt từ requirements.txt")
                    print("Đang thử cài đặt từng package...")
                    for pkg in missing_packages:
                        print(f"   Đang cài đặt {pkg}...")
                        subprocess.run([sys.executable, '-m', 'pip', 'install', pkg], 
                                     check=False)
            else:
                # Cài đặt từng package
                print("📦 Đang cài đặt các package thiếu...")
                for pkg in missing_packages:
                    print(f"   Đang cài đặt {pkg}...")
                    subprocess.run([sys.executable, '-m', 'pip', 'install', pkg], 
                                 check=False)
            
            print("\n✅ Đã cài đặt xong! Đang khởi động lại...")
            os.environ['APP_SKIP_CHECK'] = '1'
            os.execv(sys.executable, [sys.executable] + sys.argv)
        except Exception as e:
            print(f"\n❌ Lỗi khi cài đặt: {e}")
            print("\n" + "=" * 70)
            print("HƯỚNG DẪN CÀI ĐẶT THỦ CÔNG:")
            print("=" * 70)
            print("\n1. Cài đặt dependencies:")
            if os.path.exists('requirements.txt'):
                print(f"   {sys.executable} -m pip install -r requirements.txt")
            else:
                print(f"   {sys.executable} -m pip install {' '.join(missing_packages)}")
            print("\n2. Chạy lại ứng dụng:")
            print("   python app.py")
            print("=" * 70)
            sys.exit(1)
# ============================================================================
# GOOGLE SHEET BACKGROUND UPDATE
# ============================================================================

def update_google_sheet_background_safe(attendance_id, employee_team, employee_id, attendance_data):
    """
    Background task an toàn để cập nhật Google Sheet
    Không làm crash app nếu có lỗi
    """
    import sys
    from datetime import datetime
    
    # Log helper - chỉ log vào stderr để tránh I/O operation on closed file
    def _log(msg):
        try:
            print(msg, flush=True, file=sys.stderr)
        except Exception:
            pass
    
    try:
        _log(f"🔵 [SHEET_UPDATE] HÀM ĐƯỢC GỌI - ID: {attendance_id}")
        _log(f"   Team: {employee_team}, Employee ID: {employee_id}")
        
        _log(f"🔵 [SHEET_UPDATE] Tạo app context...")
        with app.app_context():
            _log(f"🔵 [SHEET_UPDATE] Đã vào app context")
            
            # Khởi tạo Google API
            _log(f"🔵 [SHEET_UPDATE] Khởi tạo GoogleDriveAPI...")
            try:
                google_api = GoogleDriveAPI()
                _log(f"🔵 [SHEET_UPDATE] Đã khởi tạo GoogleDriveAPI")
            except Exception as api_init_err:
                _log(f"❌ [SHEET_UPDATE] Lỗi khởi tạo GoogleDriveAPI: {api_init_err}")
                import traceback
                _log(f"   Traceback: {traceback.format_exc()}")
                return
            
            # Kiểm tra token
            _log(f"🔵 [SHEET_UPDATE] Kiểm tra token...")
            try:
                token_valid = google_api.ensure_valid_token()
                if not token_valid:
                    _log(f"❌ [SHEET_UPDATE] Token không hợp lệ")
                    return
                _log(f"✅ [SHEET_UPDATE] Token hợp lệ")
            except Exception as token_err:
                _log(f"❌ [SHEET_UPDATE] Lỗi kiểm tra token: {token_err}")
                import traceback
                _log(f"   Traceback: {traceback.format_exc()}")
                return
            
            if not google_api.sheets_service:
                _log(f"❌ [SHEET_UPDATE] Google Sheets service không khả dụng")
                return
            
            # Tìm tháng của ngày nghỉ
            attendance_date_str = attendance_data.get('date')
            attendance_month = None
            if attendance_date_str:
                try:
                    attendance_dt = datetime.strptime(attendance_date_str, "%Y-%m-%d")
                    attendance_month = attendance_dt.strftime("%Y%m")
                    _log(f"📅 [SHEET_UPDATE] Ngày: {attendance_date_str} -> Tháng: {attendance_month}")
                except ValueError as date_err:
                    _log(f"⚠️ [SHEET_UPDATE] Lỗi parse ngày: {date_err}")
            
            current_month = attendance_month or datetime.now().strftime("%Y%m")
            _log(f"🔍 [SHEET_UPDATE] Tìm file - Team: {employee_team}, Month: {current_month}, Employee ID: {employee_id}")
            
            # Tìm file
            try:
                target_file = google_api.find_team_timesheet(
                    folder_id=GOOGLE_DRIVE_FOLDER_ID,
                    team_name=employee_team,
                    month_year=current_month
                )
            except Exception as find_err:
                _log(f"❌ [SHEET_UPDATE] Lỗi khi tìm file: {find_err}")
                import traceback
                _log(f"   Traceback: {traceback.format_exc()}")
                return
            
            if not target_file:
                _log(f"❌ [SHEET_UPDATE] KHÔNG TÌM THẤY FILE - Team: {employee_team}, Month: {current_month}")
                return
            
            _log(f"✅ [SHEET_UPDATE] Tìm thấy file: {target_file.get('name', 'N/A')} (ID: {target_file.get('id', 'N/A')})")
            
            # Cập nhật sheet
            _log(f"🚀 [SHEET_UPDATE] Gọi update_timesheet_for_attendance...")
            _log(f"   Spreadsheet ID: {target_file['id']}")
            _log(f"   Sheet Name: {employee_id}")
            _log(f"   Data Keys: {list(attendance_data.keys())}")
            if 'leave_summary' in attendance_data:
                _log(f"   Leave Summary: {attendance_data['leave_summary']}")
            
            try:
                success = google_api.update_timesheet_for_attendance(
                    spreadsheet_id=target_file['id'],
                    sheet_name=str(employee_id),
                    attendance_data=attendance_data
                )
                _log(f"📊 [SHEET_UPDATE] Kết quả: {success}")
            
                if success:
                    _log(f"✅ [SHEET_UPDATE] Cập nhật thành công - File: {target_file['name']}, Sheet: {employee_id}")
                    try:
                        create_backup()
                        _log(f"🛡️ [SHEET_UPDATE] Đã tạo backup")
                    except Exception as backup_error:
                        _log(f"⚠️ [SHEET_UPDATE] Lỗi tạo backup: {backup_error}")
                else:
                    _log(f"❌ [SHEET_UPDATE] Cập nhật thất bại")
            except Exception as update_err:
                _log(f"❌ [SHEET_UPDATE] Lỗi khi gọi update_timesheet_for_attendance: {update_err}")
                import traceback
                _log(f"   Traceback: {traceback.format_exc()}")
            
    except Exception as e:
        import traceback
        try:
            _log(f"❌ [SHEET_UPDATE] LỖI TỔNG QUÁT: {str(e)}")
            _log(f"   Type: {type(e).__name__}")
            _log(f"   Traceback:\n{traceback.format_exc()}")
        except Exception:
            # Nếu ngay cả logging cũng lỗi, thử print trực tiếp
            try:
                print(f"❌ [SHEET_UPDATE] CRITICAL ERROR: {str(e)}", flush=True, file=sys.stderr)
            except Exception:
                pass
# Chỉ kiểm tra khi chạy trực tiếp file này
if __name__ == '__main__':
    check_and_install_dependencies()

# Bọc stdout/stderr chỉ khi có thuộc tính buffer (tránh lỗi trong IDLE, một số IDE)
try:
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    if hasattr(sys.stderr, "buffer"):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
except Exception:
    # Nếu có lỗi thì giữ nguyên stdout/stderr mặc định
    pass

# Một số môi trường (ví dụ Werkzeug reloader) có thể đóng stdout/stderr tạm thời.
# Các lệnh print trong app có thể ném ValueError: I/O operation on closed file.
# Để tránh crash 500, bọc print bằng hàm an toàn.
import builtins as _builtins

def _safe_print(*args, **kwargs):
    try:
        _builtins.print(*args, **kwargs)
    except ValueError:
        # stdout có thể đã bị đóng tạm; bỏ qua để không làm hỏng request
        pass

print = _safe_print

# Import các thư viện cần thiết
try:
    from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, get_flashed_messages, abort, send_file, make_response
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from flask_login import LoginManager, login_user, login_required, logout_user, current_user
    from flask_wtf import CSRFProtect
    from flask_wtf.csrf import generate_csrf
    from werkzeug.security import generate_password_hash, check_password_hash
except ImportError as e:
    if __name__ == '__main__':
        print(f"\n❌ Lỗi: Không thể import thư viện. {e}")
        print("\nVui lòng cài đặt dependencies:")
        if os.path.exists('requirements.txt'):
            print(f"   {sys.executable} -m pip install -r requirements.txt")
        else:
            print(f"   {sys.executable} -m pip install flask openpyxl flask-login flask-wtf werkzeug sqlalchemy flask-migrate")
        print("\nHoặc chạy lại app.py để tự động cài đặt.")
        sys.exit(1)
    else:
        raise
from datetime import datetime, timedelta, time, date
import os
import json
import uuid
from functools import wraps
from config import config
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy import func, text
import re
import pickle
import time as time_module

# Import Google API libraries
try:
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request as GoogleRequest
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
    GOOGLE_API_AVAILABLE = True
except ImportError:
    GOOGLE_API_AVAILABLE = False
    print("Google API libraries not available. Running in demo mode.")

# Phạm vi quyền truy cập Google API
GOOGLE_SCOPES = [
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets'
]

# ID folder Google Drive
GOOGLE_DRIVE_FOLDER_ID = '1dHF_x6fCJEs9krtmaZPabBIWiTr5xpB3'

# Chính sách: mẹ có con < 12 tháng được tính đủ công khi làm 7h/ngày (chỉ áp dụng ngày thường, không áp dụng cuối tuần/lễ)
def is_user_on_maternity_flex(user, target_date=None):
    """Check if user is in maternity flex window (7h required instead of 8h)."""
    if not user:
        return False
    if not getattr(user, 'is_maternity_flex', False):
        return False
    if target_date is None:
        target_date = datetime.utcnow().date()
    # Kiểm tra ngày bắt đầu (nếu có)
    flex_from = getattr(user, 'maternity_flex_from', None)
    if flex_from:
        try:
            from_date = flex_from if hasattr(flex_from, 'isoformat') and not isinstance(flex_from, str) else None
            if not from_date and isinstance(flex_from, date):
                from_date = flex_from
            if from_date and target_date < from_date:
                return False
        except Exception:
            pass  # nếu lỗi parse, bỏ qua from_date để không chặn quyền lợi
    until = getattr(user, 'maternity_flex_until', None)
    if until:
        try:
            until_date = until if hasattr(until, 'isoformat') and not isinstance(until, str) else None
            if not until_date and isinstance(until, date):
                until_date = until
            if until_date:
                return target_date <= until_date
        except Exception:
            return True  # Nếu có lỗi parse thì vẫn ưu tiên cho user
    return True


def get_required_daily_hours(user, target_date=None, holiday_type=None, shift_code=None):
    """
    Return required working hours for a given user/date.
    - Default: 8h
    - Maternity flex: hiển thị đủ 8h (chỉ áp dụng ca 1-4, ngày thường)
    - Holidays giữ nguyên 8h
    """
    base_hours = 8.0
    # Không giảm giờ nếu không phải ngày thường
    if holiday_type in ('vietnamese_holiday', 'weekend', 'japanese_holiday'):
        return base_hours
    # Với chính sách mẹ <12 tháng: chỉ áp dụng cho ca 1-4 ngày thường
    if shift_code in ('1', '2', '3', '4') and is_user_on_maternity_flex(user, target_date):
        return 8.0
    return base_hours

from collections import defaultdict
import secrets
from flask_migrate import Migrate
from jinja2 import Template
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
import base64
import traceback
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import registerFont
import zipfile
import webbrowser
import subprocess
import platform
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager


# License / activation
APP_LICENSE_KEY = os.environ.get('APP_LICENSE_KEY', 'LIC-W8B61JUL-F7OD')
# Endpoint verify license online (License Manager Pro)
# Có thể override bằng biến môi trường LICENSE_VERIFY_ENDPOINT nếu cần
LICENSE_VERIFY_ENDPOINT = os.environ.get(
    'LICENSE_VERIFY_ENDPOINT',
    'https://management-license.vercel.app/'
)

def _infer_license_data_from_text(text: str, license_key: str = "") -> dict:
    """
    Suy luận trạng thái license từ nội dung text/HTML khi server license không trả JSON.
    Tránh false-positive kiểu: trang HTML có chữ "đã hết hạn" ở phần thống kê/label nhưng key thực tế vẫn còn hạn.
    Trả về dict giống format JSON mong đợi (có thể rỗng nếu không suy ra được).
    """
    import re
    from datetime import datetime as _dt_mod

    raw_text = text or ""
    lower_text = raw_text.lower()

    key = (license_key or "").strip()
    key_lower = key.lower()

    # Ưu tiên parse trong vùng gần license key (nếu xuất hiện trong response)
    candidate_raw = raw_text
    candidate_lower = lower_text
    key_pos = None
    if key_lower and key_lower in lower_text:
        idx = lower_text.find(key_lower)
        key_pos = idx
        start = max(0, idx - 1500)
        end = min(len(lower_text), idx + 1500)
        candidate_raw = raw_text[start:end]
        candidate_lower = lower_text[start:end]
        key_pos = idx - start

    data: dict = {}

    # 1) Cố gắng bắt đúng dòng "Trạng thái: ..."
    active_pat = re.compile(
        r"trạng\s*thái\s*[:\-]\s*(đang\s*hoạt\s*động|còn\s*hạn|active|valid|hợp\s*lệ)",
        re.IGNORECASE
    )
    expired_pat = re.compile(
        r"trạng\s*thái\s*[:\-]\s*(đã\s*hết\s*hạn|hết\s*hạn|expired|invalid|không\s*hợp\s*lệ)",
        re.IGNORECASE
    )

    matches: list[tuple[str, int]] = []
    for m in active_pat.finditer(candidate_raw):
        matches.append(("active", m.start()))
    for m in expired_pat.finditer(candidate_raw):
        matches.append(("expired", m.start()))

    chosen: str | None = None
    if matches:
        if key_pos is not None:
            # Chọn match gần key nhất
            matches.sort(key=lambda t: abs(t[1] - key_pos))
            chosen = matches[0][0]
        else:
            # Nếu không thấy key trong response mà lại có nhiều trạng thái -> không đoán bừa
            unique = {m[0] for m in matches}
            if len(unique) == 1:
                chosen = next(iter(unique))

    if chosen == "expired":
        data["valid"] = False
        data["status"] = "expired"
        data["message"] = "License đã hết hạn (theo nội dung server license)."
    elif chosen == "active":
        data["valid"] = True
        data["status"] = "active"

    # 2) Nếu chưa xác định được bằng "Trạng thái", fallback theo cụm từ khoá trong vùng candidate (đặc biệt khi response là HTML hiển thị chi tiết license)
    if not data:
        if "license hợp lệ" in candidate_lower:
            data["valid"] = True
            data["status"] = "active"
        elif "không hợp lệ" in candidate_lower:
            data["valid"] = False
            data["status"] = "expired"
            data["message"] = "License không hợp lệ (theo nội dung server license)."
        elif "đã hết hạn" in candidate_lower:
            data["valid"] = False
            data["status"] = "expired"
            data["message"] = "License đã hết hạn (theo nội dung server license)."

    # Nếu vẫn chưa đoán được -> trả rỗng để caller quyết định
    if not data:
        return {}

    # Bóc tách days_remaining / expiry nếu có
    try:
        m = re.search(r"(\d+)\s*ngày", candidate_raw, re.IGNORECASE)
        if m:
            data["days_remaining"] = int(m.group(1))
    except Exception:
        pass

    try:
        m2 = re.search(r"Hết\s*hạn[: ]+([0-9: ]+\d{2}/\d{2}/\d{4})", candidate_raw, re.IGNORECASE)
        if m2:
            raw_expiry = m2.group(1).strip()
            try:
                expiry_dt = _dt_mod.strptime(raw_expiry, "%H:%M:%S %d/%m/%Y")
                data["expiry"] = expiry_dt.isoformat()
            except Exception:
                data["expiry"] = raw_expiry
    except Exception:
        pass

    return data


# Import database models
from database.models import (
    db,
    User,
    Attendance,
    Request,
    Department,
    AuditLog,
    PasswordResetToken,
    LeaveRequest,
    Holiday,
    Activation,
)
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
from email.header import Header
from utils.email_utils import send_leave_request_email, send_leave_request_email_async
from state.email_state import email_status
from queue import Queue
from sqlalchemy.exc import SQLAlchemyError

# Import utility functions
from utils.validators import (
    validate_input_sanitize,
    validate_employee_id,
    validate_str,
    validate_date,
    validate_time,
    validate_float,
    validate_int,
    validate_note,
    validate_reason,
    validate_holiday_type,
    validate_role_value,
    ValidationError,
)
from utils.session import check_session_timeout, update_session_activity, log_audit_action
from utils.signature_manager import signature_manager
from utils.logger import logger, security_logger, audit_logger, database_logger, api_logger
from utils.security_enhanced import security_manager, require_security_check
from utils.database_utils import safe_db_commit, safe_db_rollback, retry_db_operation

def has_role(user_id, required_role):
    """Check if user has a specific role"""
    user = db.session.get(User, user_id)
    if not user:
        return False
    return required_role in user.roles.split(',')

def check_approval_permission(user_id, attendance_id, current_role):
    """Check if user has permission to approve specific attendance"""
    user = db.session.get(User, user_id)
    if not user:
        return False, "❌ KHÔNG TÌM THẤY NGƯỜI DÙNG"
    
    attendance = db.session.get(Attendance, attendance_id)
    if not attendance:
        return False, "❌ KHÔNG TÌM THẤY BẢN GHI CHẤM CÔNG"
    
    # ADMIN và MANAGER có thể duyệt tất cả
    if current_role in ['ADMIN', 'MANAGER']:
        return True, ""
    
    # TEAM_LEADER có thể duyệt nhân viên cùng phòng ban (bao gồm cả bản thân)
    if current_role == 'TEAM_LEADER':
        if not attendance.user or attendance.user.department != user.department:
            return False, "❌ KHÔNG CÙNG PHÒNG BAN: Chỉ duyệt được nhân viên cùng phòng"
        return True, ""
    
    return False, "❌ KHÔNG CÓ QUYỀN DUYỆT CHẤM CÔNG"

def check_leave_approval_permission(user_id, request_id, current_role):
    """Check if user has permission to approve specific leave request"""
    user = db.session.get(User, user_id)
    if not user:
        return False, "❌ KHÔNG TÌM THẤY NGƯỜI DÙNG"
    
    leave_request = db.session.get(LeaveRequest, request_id)
    if not leave_request:
        return False, "❌ KHÔNG TÌM THẤY ĐƠN NGHỈ PHÉP"
    
    # ADMIN và MANAGER có thể duyệt tất cả
    if current_role in ['ADMIN', 'MANAGER']:
        return True, ""
    
    # TEAM_LEADER có thể duyệt nhân viên cùng phòng ban (bao gồm cả bản thân)
    if current_role == 'TEAM_LEADER':
        if not leave_request.user or leave_request.user.department != user.department:
            return False, "❌ KHÔNG CÙNG PHÒNG BAN: Chỉ duyệt được nhân viên cùng phòng"
        return True, ""
    
    return False, "❌ KHÔNG CÓ QUYỀN DUYỆT ĐƠN NGHỈ PHÉP"

def validate_overtime_comp_time(check_in, check_out, shift_start, shift_end, break_time, comp_time_regular, comp_time_overtime, comp_time_ot_before_22, comp_time_ot_after_22, date, next_day_checkout=False, holiday_type='normal', shift_code=None):
    """Simple validation for overtime compensation time"""
    # Basic validation - allow all if basic conditions are met
    return True, None

def convert_overtime_to_hhmm():
    """Convert overtime values to HH:MM format using optimized bulk processing"""
    from utils.query_optimizer import bulk_convert_overtime_optimized
    total_converted = bulk_convert_overtime_optimized()
    # print(f"Đã làm sạch lại overtime về dạng H:MM cho {total_converted} bản ghi.")

app = Flask(__name__)

# ----- Jinja filter: convert UTC datetime to Vietnam time (Asia/Ho_Chi_Minh) -----
from datetime import datetime
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
    _VN_TZ = ZoneInfo('Asia/Ho_Chi_Minh')
    _UTC_TZ = ZoneInfo('UTC')
except Exception:  # Fallback if zoneinfo not available
    _VN_TZ = None
    _UTC_TZ = None

def _to_vn_datetime(dt: datetime) -> datetime | None:
    if not dt:
        return None
    # If tz-aware, convert; if naive, assume UTC and convert
    try:
        if _VN_TZ is None:
            return dt  # best-effort fallback
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=_UTC_TZ)
        return dt.astimezone(_VN_TZ)
    except Exception:
        return dt

def _vn_datetime_format(dt: datetime, fmt: str = '%d/%m/%Y %H:%M') -> str:
    local_dt = _to_vn_datetime(dt)
    return local_dt.strftime(fmt) if local_dt else ''

app.jinja_env.filters['vn_datetime'] = _vn_datetime_format

# Filter để lọc notes - chỉ hiển thị phần text ghi chú, ẩn use_lunch_break
def _filter_leave_notes(notes_str):
    """
    Lọc notes để chỉ hiển thị phần text ghi chú thực sự, ẩn use_lunch_break
    Nếu notes là JSON chứa use_lunch_break, chỉ trả về _original_notes (nếu có)
    Nếu notes là text thường, trả về như cũ
    """
    if not notes_str:
        return ''
    
    try:
        import json
        # Thử parse JSON
        notes_data = json.loads(notes_str)
        if isinstance(notes_data, dict):
            # Nếu có _original_notes, trả về phần text ghi chú thực sự
            if '_original_notes' in notes_data:
                return notes_data['_original_notes']
            # Nếu chỉ có use_lunch_break và không có text ghi chú, trả về rỗng
            if 'use_lunch_break' in notes_data and len(notes_data) == 1:
                return ''
            # Nếu có các key khác, có thể là JSON hợp lệ nhưng không phải format của chúng ta
            # Trả về rỗng để tránh hiển thị JSON
            return ''
    except (json.JSONDecodeError, TypeError, ValueError):
        # Không phải JSON, trả về như cũ (text thường)
        return notes_str
    
    return notes_str

app.jinja_env.filters['filter_leave_notes'] = _filter_leave_notes

# Dictionary để lưu trạng thái email gửi
# in-memory state moved to state/email_state.py for a single source of import

# --- Persistent email status model ---
class EmailStatusRecord(db.Model):
    __tablename__ = 'email_status_records'
    __table_args__ = {
        'extend_existing': True
    }
    id = db.Column(db.Integer, primary_key=True)
    request_id = db.Column(db.Integer, index=True, nullable=False, unique=True)
    status = db.Column(db.String(50), nullable=False)
    message = db.Column(db.Text, nullable=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

def _ensure_email_status_table():
    try:
        # Create table if it does not exist without affecting others
        EmailStatusRecord.__table__.create(bind=db.engine, checkfirst=True)
    except Exception as e:
        pass

def upsert_email_status(request_id: int, status: str, message: str):
    try:
        _ensure_email_status_table()
        record = EmailStatusRecord.query.filter_by(request_id=request_id).first()
        if record is None:
            record = EmailStatusRecord(request_id=request_id, status=status, message=message)
            db.session.add(record)
        else:
            record.status = status
            record.message = message
        db.session.commit()
        # print(f"[EmailStatus] upsert request_id={request_id} -> {status}")
    except Exception as e:
        db.session.rollback()
        # print(f"[EmailStatus] upsert error: {e}")

def get_email_status_record(request_id: int):
    try:
        _ensure_email_status_table()
        return EmailStatusRecord.query.filter_by(request_id=request_id).first()
    except Exception as e:
        # print(f"[EmailStatus] get error: {e}")
        return None

# --- Google Drive API Integration ---
class GoogleDriveAPI:
    """Class để quản lý Google Drive API"""
    
    def __init__(self, auto_authenticate=True):
        """Khởi tạo Google Drive API client
        
        Args:
            auto_authenticate: Nếu True, tự động authenticate khi khởi tạo. 
                              Nếu False, chỉ load token từ file mà không authenticate.
        """
        self.creds = None
        self.drive_service = None
        self.sheets_service = None
        self.token_file = 'token.pickle'
        self.last_refresh_file = 'last_token_refresh.txt'
        
        if not GOOGLE_API_AVAILABLE:
            print("Lỗi: Google API libraries không có sẵn!")
            print("Hãy cài đặt: pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client")
            return
        
        if not os.path.exists('credentials.json'):
            print("Lỗi: Không tìm thấy file credentials.json!")
            print("Hãy tạo file credentials.json trong thư mục hiện tại.")
            return
        
        # Chỉ load token từ file nếu không auto-authenticate
        if not auto_authenticate:
            if os.path.exists('token.pickle'):
                try:
                    with open('token.pickle', 'rb') as token:
                        self.creds = pickle.load(token)
                except Exception as e:
                    print(f"Lỗi khi load token: {e}")
            return
        
        # Chỉ authenticate nếu được phép và cần thiết
        if auto_authenticate:
            print("Phát hiện file credentials.json - Bắt đầu xác thực...")
            self.authenticate(allow_browser_auth=False)  # Không tự động mở browser
            
            # Tự động gia hạn token nếu cần
            self.auto_refresh_token_if_needed()
    
    def authenticate(self, allow_browser_auth=False):
        """Xác thực với Google API
        
        Args:
            allow_browser_auth: Nếu True, cho phép mở browser để authenticate.
                               Nếu False, chỉ thử refresh token, không mở browser.
        """
        if not GOOGLE_API_AVAILABLE:
            print("Google API libraries not available")
            return
            
        try:
            # Kiểm tra token đã lưu
            if os.path.exists('token.pickle'):
                with open('token.pickle', 'rb') as token:
                    self.creds = pickle.load(token)
            
            # Nếu không có credentials hợp lệ
            if not self.creds or not self.creds.valid:
                # Thử refresh nếu có refresh_token
                if self.creds and self.creds.expired and self.creds.refresh_token:
                    try:
                        self.creds.refresh(GoogleRequest())
                        # Lưu credentials sau khi refresh
                        with open('token.pickle', 'wb') as token:
                            pickle.dump(self.creds, token)
                    except Exception as refresh_err:
                        # Refresh thất bại
                        if allow_browser_auth:
                            # Chỉ mở browser nếu được phép
                            flow = InstalledAppFlow.from_client_secrets_file(
                                'credentials.json', GOOGLE_SCOPES)
                            self.creds = flow.run_local_server(port=0)
                            # Lưu credentials
                            with open('token.pickle', 'wb') as token:
                                pickle.dump(self.creds, token)
                        else:
                            # Không mở browser, chỉ báo lỗi
                            print(f"⚠️ Token hết hạn và không thể refresh tự động. Cần authenticate thủ công.")
                            return  # Không raise exception để không làm crash app
                elif allow_browser_auth:
                    # Chỉ mở browser nếu được phép và không có token
                    flow = InstalledAppFlow.from_client_secrets_file(
                        'credentials.json', GOOGLE_SCOPES)
                    self.creds = flow.run_local_server(port=0)
                    # Lưu credentials
                    with open('token.pickle', 'wb') as token:
                        pickle.dump(self.creds, token)
                else:
                    # Không có token và không được phép mở browser
                    print("⚠️ Không có token và không được phép mở browser để authenticate.")
                    return  # Không raise exception để không làm crash app
            
            # Khởi tạo services nếu có credentials
            if self.creds and self.creds.valid:
                self.drive_service = build('drive', 'v3', credentials=self.creds)
                self.sheets_service = build('sheets', 'v4', credentials=self.creds)
                print("✅ Xác thực thành công!")
        except Exception as e:
            print(f"⚠️ Lỗi xác thực: {e}")
            # Không crash app, chỉ log lỗi
    
    def auto_refresh_token_if_needed(self):
        """Tự động gia hạn token nếu cần thiết"""
        try:
            # Kiểm tra xem có cần gia hạn không
            if self.should_refresh_token():
                print("🔄 Tự động gia hạn token...")
                if self.refresh_token():
                    self.save_last_refresh_time()
                    print("✅ Token đã được gia hạn thành công!")
                else:
                    print("⚠️ Không thể gia hạn token, cần xác thực lại")
            else:
                print("✅ Token vẫn còn hiệu lực, không cần gia hạn")
        except Exception as e:
            print(f"⚠️  Lỗi khi gia hạn token: {e}")
    
    def ensure_valid_token(self):
        """Đảm bảo token luôn hợp lệ trước khi sử dụng API"""
        try:
            # Kiểm tra xem có credentials không
            if not self.creds:
                print("❌ Không có credentials, cần xác thực lại")
                return False
            
            # Kiểm tra xem token có hết hạn không
            if self.creds.expired:
                print("🔄 Token đã hết hạn, đang gia hạn...")
                if self.creds.refresh_token:
                    try:
                        self.creds.refresh(GoogleRequest())
                        # Lưu token mới
                        with open(self.token_file, 'wb') as token:
                            pickle.dump(self.creds, token)
                        # Cập nhật services
                        self.drive_service = build('drive', 'v3', credentials=self.creds)
                        self.sheets_service = build('sheets', 'v4', credentials=self.creds)
                        print("✅ Token đã được gia hạn thành công!")
                        return True
                    except Exception as refresh_error:
                        error_str = str(refresh_error)
                        if 'invalid_grant' in error_str.lower():
                            print("❌ Lỗi invalid_grant: Refresh token không hợp lệ")
                            print("💡 Cần tạo token mới bằng cách chạy: python refresh_token.py")
                            # Xóa token cũ để tránh lỗi lặp lại
                            try:
                                if os.path.exists(self.token_file):
                                    backup_name = f"token_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pickle"
                                    shutil.copy2(self.token_file, backup_name)
                                    os.remove(self.token_file)
                                    print(f"💾 Đã backup và xóa token cũ: {backup_name}")
                            except Exception:
                                pass
                        else:
                            print(f"❌ Lỗi khi refresh token: {refresh_error}")
                        return False
                else:
                    print("❌ Không có refresh_token, cần xác thực lại")
                    return False
            
            # Kiểm tra xem services có hoạt động không
            if not self.drive_service or not self.sheets_service:
                print("🔄 Khởi tạo lại services...")
                self.drive_service = build('drive', 'v3', credentials=self.creds)
                self.sheets_service = build('sheets', 'v4', credentials=self.creds)
                print("✅ Services đã được khởi tạo lại!")
            
            return True
        except Exception as e:
            print(f"❌ Lỗi khi đảm bảo token hợp lệ: {e}")
            return False
    
    def should_refresh_token(self):
        """Kiểm tra xem có cần gia hạn token không"""
        try:
            # Kiểm tra xem có credentials không
            if not self.creds:
                return True  # Không có credentials, cần xác thực
            
            # Kiểm tra xem token có hết hạn không
            if self.creds.expired:
                print("⚠️ Token đã hết hạn!")
                return True  # Token đã hết hạn, cần gia hạn
            
            # Kiểm tra file lần gia hạn cuối
            if not os.path.exists(self.last_refresh_file):
                print("⚠️ Chưa có file lưu thời gian gia hạn, cần gia hạn")
                return True  # Chưa có file, cần gia hạn
            
            # Đọc thời gian gia hạn cuối
            with open(self.last_refresh_file, 'r') as f:
                last_refresh_str = f.read().strip()
            
            last_refresh = datetime.fromisoformat(last_refresh_str)
            now = datetime.now()
            time_diff = (now - last_refresh).total_seconds()
            
            # Nếu đã qua 45 phút thì cần gia hạn (Google tokens thường hết hạn sau 1 giờ)
            # Gia hạn sớm 15 phút để đảm bảo không bao giờ hết hạn
            if time_diff >= 45 * 60:  # 45 phút
                print(f"⚠️ Token sắp hết hạn (đã {time_diff/60:.1f} phút), cần gia hạn")
                return True
            
            # Cảnh báo khi còn 10 phút nữa sẽ gia hạn
            if time_diff >= 35 * 60:  # 35 phút
                print(f"ℹ️ Token sẽ được gia hạn trong {45 - time_diff/60:.1f} phút nữa")
            
            return False
        except Exception as e:
            print(f"❌ Lỗi khi kiểm tra thời gian gia hạn: {e}")
            return True  # Nếu có lỗi thì gia hạn để an toàn
    
    def refresh_token(self):
        """Gia hạn token"""
        try:
            if not self.creds:
                print("Không có credentials để gia hạn")
                return False
            
            # Luôn thử gia hạn token nếu có refresh_token
            if self.creds.refresh_token:
                print("Đang gia hạn token...")
                self.creds.refresh(GoogleRequest())
                
                # Lưu token mới
                with open(self.token_file, 'wb') as token:
                    pickle.dump(self.creds, token)
                
                # Cập nhật services
                self.drive_service = build('drive', 'v3', credentials=self.creds)
                self.sheets_service = build('sheets', 'v4', credentials=self.creds)
                
                print("✅ Token đã được gia hạn thành công!")
                return True
            else:
                print("Không có refresh_token, cần xác thực lại")
                return False
        except Exception as e:
            print(f"Lỗi khi gia hạn token: {e}")
            return False
    
    def save_last_refresh_time(self):
        """Lưu thời gian gia hạn cuối"""
        try:
            with open(self.last_refresh_file, 'w') as f:
                f.write(datetime.now().isoformat())
        except Exception as e:
            print(f"Lỗi khi lưu thời gian gia hạn: {e}")
    
    def update_sheet_value(self, spreadsheet_id, sheet_name, row, column, new_value):
        """
        Cập nhật giá trị trong Google Sheet với các tham số cụ thể
        
        Args:
            spreadsheet_id (str): ID của spreadsheet
            sheet_name (str): Tên sheet
            row (int): Số dòng (bắt đầu từ 1)
            column (str): Tên cột (A, B, C, ..., M, N, O, ...)
            new_value (str): Giá trị mới
        
        Returns:
            bool: True nếu thành công, False nếu thất bại
        """
        try:
            # Đảm bảo token luôn hợp lệ trước khi sử dụng API
            if not self.ensure_valid_token():
                print("❌ Không thể đảm bảo token hợp lệ")
                return False
            # Tạo range từ các tham số
            range_name = f"{sheet_name}!{column}{row}"
            
            print(f"\n🔧 CẬP NHẬT GIÁ TRỊ TRONG SHEET:")
            print(f"   Sheet: {sheet_name}")
            print(f"   Ô: {column}{row}")
            print(f"   Giá trị mới: {new_value}")
            print(f"   Range: {range_name}")
            
            # Cập nhật giá trị
            body = {
                'values': [[new_value]]
            }
            
            result = self.sheets_service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=range_name,
                valueInputOption='USER_ENTERED',
                body=body
            ).execute()
            
            print(f"✅ Cập nhật thành công!")
            print(f"   Số ô đã cập nhật: {result.get('updatedCells', 0)}")
            
            # Căn giữa cell sau khi cập nhật
            try:
                self.center_align_cells(spreadsheet_id, sheet_name, [range_name])
            except Exception as e:
                print(f"⚠️ Không thể căn giữa cell: {e}")
            
            return True
            
        except Exception as e:
            print(f"❌ Lỗi khi cập nhật sheet: {e}")
            return False
    
    def batch_update_values(self, spreadsheet_id, data_ranges):
        """Cập nhật nhiều ô theo lô bằng A1 notation.

        Args:
            spreadsheet_id (str): ID của spreadsheet
            data_ranges (list[dict]): Mỗi phần tử có dạng {'range': 'Sheet!A1', 'values': [[value]]}

        Returns:
            bool: True nếu thành công, False nếu thất bại
        """
        try:
            if not self.ensure_valid_token():
                print("❌ Không thể đảm bảo token hợp lệ")
                return False
            body = {
                'valueInputOption': 'USER_ENTERED',
                'data': data_ranges
            }
            result = self.sheets_service.spreadsheets().values().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=body
            ).execute()
            updated = result.get('totalUpdatedCells', 0)
            print(f"✅ Batch update thành công, số ô cập nhật: {updated}")
            return True
        except Exception as e:
            print(f"❌ Batch update thất bại: {e}")
            return False

    def center_align_cells(self, spreadsheet_id, sheet_name, ranges):
        """Căn giữa các cells trong Google Sheet.
        
        Args:
            spreadsheet_id (str): ID của spreadsheet
            sheet_name (str): Tên sheet
            ranges (list[str]): Danh sách các range cần căn giữa (ví dụ: ['Sheet!G5', 'Sheet!K5'])
        
        Returns:
            bool: True nếu thành công, False nếu thất bại
        """
        try:
            if not self.ensure_valid_token():
                print("❌ Không thể đảm bảo token hợp lệ")
                return False
            
            if not ranges:
                return True
            
            # Lấy sheet ID
            sheet_id = self._get_sheet_id(spreadsheet_id, sheet_name)
            if sheet_id is None:
                print("⚠️ Không thể lấy sheet ID")
                return False
            
            # Chuyển đổi A1 notation sang GridRange
            def a1_to_grid_range(a1_range, sheet_id):
                """Chuyển đổi A1 notation (Sheet!G5) sang GridRange format."""
                try:
                    # Tách phần range (bỏ phần sheet name)
                    if '!' in a1_range:
                        range_part = a1_range.split('!')[1]
                    else:
                        range_part = a1_range
                    
                    # Parse cột và dòng (ví dụ: G5 -> column=6, row=4 (0-based))
                    import re
                    match = re.match(r'([A-Z]+)(\d+)', range_part)
                    if not match:
                        return None
                    
                    col_str = match.group(1)
                    row_str = match.group(2)
                    
                    # Chuyển cột sang index (A=0, B=1, ..., G=6, ...)
                    col_index = 0
                    for char in col_str:
                        col_index = col_index * 26 + (ord(char) - ord('A') + 1)
                    col_index -= 1  # 0-based
                    
                    # Chuyển dòng sang index (1-based -> 0-based)
                    row_index = int(row_str) - 1
                    
                    return {
                        'sheetId': sheet_id,
                        'startRowIndex': row_index,
                        'endRowIndex': row_index + 1,
                        'startColumnIndex': col_index,
                        'endColumnIndex': col_index + 1
                    }
                except Exception as e:
                    print(f"⚠️ Lỗi parse range {a1_range}: {e}")
                    return None
            
            # Tạo requests để căn giữa các cells
            requests = []
            for range_str in ranges:
                grid_range = a1_to_grid_range(range_str, sheet_id)
                if grid_range:
                    requests.append({
                        'repeatCell': {
                            'range': grid_range,
                            'cell': {
                                'userEnteredFormat': {
                                    'horizontalAlignment': 'CENTER',
                                    'verticalAlignment': 'MIDDLE',
                                    'textFormat': {
                                        'fontFamily': 'Google Sans',
                                        'fontSize': 9
                                    }
                                }
                            },
                            'fields': 'userEnteredFormat.horizontalAlignment,userEnteredFormat.verticalAlignment,userEnteredFormat.textFormat.fontFamily,userEnteredFormat.textFormat.fontSize'
                        }
                    })
            
            if not requests:
                print("⚠️ Không có requests hợp lệ để căn giữa")
                return False
            
            body = {
                'requests': requests
            }
            
            result = self.sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=body
            ).execute()
            
            print(f"✅ Căn giữa thành công cho {len(requests)} cells")
            return True
        except Exception as e:
            print(f"⚠️ Không thể căn giữa cells: {e}")
            import traceback
            print(traceback.format_exc())
            return False

    def _get_sheet_id(self, spreadsheet_id, sheet_name):
        """Lấy sheet ID từ tên sheet."""
        try:
            if not self.ensure_valid_token():
                return None
            spreadsheet = self.sheets_service.spreadsheets().get(
                spreadsheetId=spreadsheet_id
            ).execute()
            sheets = spreadsheet.get('sheets', [])
            for sheet in sheets:
                if sheet['properties']['title'] == sheet_name:
                    return sheet['properties']['sheetId']
            return None
        except Exception as e:
            print(f"⚠️ Không thể lấy sheet ID: {e}")
            return None

    def _column_index_to_letter(self, index_zero_based):
        # 0 -> A, 1 -> B, ...
        letter = ''
        index = index_zero_based
        while True:
            index, remainder = divmod(index, 26)
            letter = chr(65 + remainder) + letter
            if index == 0:
                break
            index -= 1
        return letter

    def _normalize_cell(self, s):
        try:
            if s is None:
                return ''
            return str(s).strip().lower()
        except Exception:
            return ''

    def _date_variants(self, date_str_iso):
        # Trả về các biến thể để dò khớp ngày trong sheet
        # Dựa trên cấu trúc thực tế: 2025/12/1 (YYYY/MM/D) - không có số 0 ở đầu
        try:
            # Thử parse nhiều format đầu vào
            dt = None
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                try:
                    dt = datetime.strptime(date_str_iso, fmt)
                    break
                except ValueError:
                    continue
            
            if dt is None:
                print(f"⚠️ Không thể parse ngày: {date_str_iso}")
                return [date_str_iso]
            
            # Lấy các giá trị ngày, tháng, năm không có số 0 ở đầu
            day_no_zero = str(dt.day)  # 1 thay vì 01
            month_no_zero = str(dt.month)  # 12 thay vì 12 (giữ nguyên nếu >= 10)
            year = str(dt.year)
            
            # Tạo danh sách biến thể bao gồm cả có và không có số 0
            variants = [
                # Format có số 0 ở đầu (chuẩn)
                dt.strftime('%Y/%m/%d'),  # 2025/12/01
                dt.strftime('%Y-%m-%d'),  # 2025-12-01
                dt.strftime('%d/%m/%Y'),  # 01/12/2025
                dt.strftime('%m/%d/%Y'),  # 12/01/2025
                dt.strftime('%d-%m-%Y'),  # 01-12-2025
                # Format KHÔNG có số 0 ở đầu (như trong Google Sheet thực tế)
                f"{year}/{month_no_zero}/{day_no_zero}",  # 2025/12/1
                f"{year}-{month_no_zero}-{day_no_zero}",  # 2025-12-1
                f"{day_no_zero}/{month_no_zero}/{year}",  # 1/12/2025
                f"{month_no_zero}/{day_no_zero}/{year}",  # 12/1/2025
                f"{day_no_zero}-{month_no_zero}-{year}",  # 1-12-2025
                # Format có số 0 ở tháng nhưng không có ở ngày
                f"{year}/{dt.strftime('%m')}/{day_no_zero}",  # 2025/12/1
                f"{year}-{dt.strftime('%m')}-{day_no_zero}",  # 2025-12-1
            ]
            
            # Loại bỏ trùng lặp và giữ thứ tự
            seen = set()
            unique_variants = []
            for v in variants:
                if v not in seen:
                    seen.add(v)
                    unique_variants.append(v)
            
            return unique_variants
        except Exception as e:
            print(f"⚠️ Lỗi parse ngày {date_str_iso}: {e}")
            return [date_str_iso]

    def _read_sheet_values(self, spreadsheet_id, sheet_name, a1_range='A1:ZZ1000'):
        try:
            if not self.ensure_valid_token():
                print("❌ Không thể đảm bảo token hợp lệ")
                return []
            resp = self.sheets_service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range=f"{sheet_name}!{a1_range}"
            ).execute()
            return resp.get('values', [])
        except Exception as e:
            print(f"❌ Lỗi khi đọc sheet: {e}")
            return []

    def _build_header_map(self, header_row):
        """Xây dựng map từ header row"""
        header_map = {}
        for idx, cell in enumerate(header_row):
            normalized = self._normalize_cell(cell)
            if normalized:
                header_map[normalized] = idx
        return header_map

    def _find_best_header(self, header_map, candidates):
        """Tìm header tốt nhất từ danh sách candidates"""
        for cand in candidates:
            normalized = self._normalize_cell(cand)
            if normalized in header_map:
                return header_map[normalized]
        return None

    def _find_row_by_date(self, all_rows, date_str_iso, date_header_index=None):
        """Tìm dòng theo ngày trong sheet"""
        variants = self._date_variants(date_str_iso)
        date_col_index = date_header_index if date_header_index is not None else 0
        
        for i, row in enumerate(all_rows):
            if len(row) <= date_col_index:
                continue
            cell_value = str(row[date_col_index]).strip()
            for variant in variants:
                if variant in cell_value:
                    return i + 1  # Trả về số dòng (1-based)
        print(f"❌ Không tìm thấy dòng phù hợp")
        return None

    def update_timesheet_for_attendance(self, spreadsheet_id, sheet_name, attendance_data):
        """Cập nhật nhiều trường theo ngày dựa trên tiêu đề cột; fallback về cập nhật 1 ô.

        attendance_data tối thiểu nên có: date (YYYY-MM-DD), status, note, total_hours,
        regular_work_hours, break_time, overtime_before_22, overtime_after_22,
        doi_ung, approved_by, approved_at
        """
        import sys
        from datetime import datetime as dt
        timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        try:
            print(f"\n{'='*80}")
            print(f"🚀 [UPDATE_TIMESHEET_START] {timestamp} - Bắt đầu cập nhật timesheet")
            print(f"   📊 Spreadsheet ID: {spreadsheet_id}")
            print(f"   📋 Sheet Name: {sheet_name}")
            print(f"   📅 Date: {attendance_data.get('date', 'Unknown')}")
            print(f"   👤 Employee: {attendance_data.get('user_name', 'Unknown')}")
            print(f"{'='*80}")
            try:
                sys.stdout.flush()
            except Exception:
                pass
            # Đảm bảo bộ sao lưu định kỳ đã chạy nền
            try:
                ensure_backup_scheduler_started()
            except Exception as _e:
                print(f"⚠️ Không thể khởi động backup scheduler: {_e}")
            
            timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            print(f"🔍 [TOKEN_CHECK] {timestamp} - Kiểm tra token...")
            try:
                sys.stdout.flush()
            except Exception:
                pass
            
            if not self.ensure_valid_token():
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                print(f"❌ [TOKEN_INVALID] {timestamp} - Không thể đảm bảo token hợp lệ")
                try:
                    sys.stdout.flush()
                except Exception:
                    pass
                return False

            timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            print(f"✅ [TOKEN_VALID] {timestamp} - Token hợp lệ")
            print(f"🔍 [READ_SHEET] {timestamp} - Đang đọc dữ liệu từ sheet...")
            try:
                sys.stdout.flush()
            except Exception:
                pass
            
            rows = self._read_sheet_values(spreadsheet_id, sheet_name)
            
            timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            print(f"📊 [READ_SHEET_DONE] {timestamp} - Số dòng đọc được: {len(rows)}")
            try:
                sys.stdout.flush()
            except Exception:
                pass
            
            if not rows:
                print("⚠️ Không đọc được dữ liệu sheet, fallback về 1 ô")
                return self.update_sheet_value(
                    spreadsheet_id, sheet_name, 35, 'M',
                    f"Phê duyệt bởi {attendance_data.get('approved_by','')} - {attendance_data.get('approved_at','')}"
                )

            print(f"📋 Header row: {rows[0] if rows else 'Empty'}")
            header = rows[0]
            header_map = self._build_header_map(header)
            print(f"🗺️ Header map: {header_map}")

            # Xác định cột ngày và dòng tương ứng với ngày
            # Dựa trên cấu trúc thực tế: cột A (index 0) chứa ngày theo format 2025/10/13
            date_iso = attendance_data.get('date', '')
            # Cột A luôn là cột ngày (index 0) trong timesheet
            date_col_index = 0
            
            print(f"🔍 Tìm dòng theo ngày: {date_iso}")
            print(f"📅 Các biến thể ngày: {self._date_variants(date_iso)}")
            print(f"🎯 Tìm kiếm trong cột {date_col_index} (cột A)")
            
            target_row_index = self._find_row_by_date(rows, date_iso, date_col_index)
            
            if target_row_index:
                print(f"✅ Tìm thấy dòng {target_row_index} cho ngày {date_iso}")
            else:
                print(f"❌ Không tìm thấy dòng cho ngày {date_iso}")
                # Debug: in ra một vài dòng đầu để kiểm tra
                print("📋 Một vài dòng đầu trong sheet:")
                for i, row in enumerate(rows[:15]):
                    if len(row) > 0:
                        print(f"   Dòng {i+1}: '{row[0] if len(row) > 0 else 'Empty'}'")
                
                # Debug: tìm kiếm thủ công trong cột A
                print(f"🔍 Tìm kiếm thủ công ngày {date_iso} trong cột A:")
                for i, row in enumerate(rows):
                    if len(row) > 0 and row[0]:
                        cell_value = str(row[0]).strip()
                        if date_iso in cell_value or any(variant in cell_value for variant in self._date_variants(date_iso)):
                            print(f"   ✅ Tìm thấy khớp ở dòng {i+1}: '{cell_value}'")
                        elif '2025' in cell_value and '10' in cell_value and '13' in cell_value:
                            print(f"   🔍 Có thể khớp ở dòng {i+1}: '{cell_value}'")

            if not target_row_index:
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                print(f"\n{'='*80}")
                print(f"❌ [ROW_NOT_FOUND] {timestamp} - KHÔNG TÌM THẤY DÒNG THEO NGÀY")
                print(f"   Date: {date_iso}")
                print(f"   Sheet: {sheet_name}")
                print(f"   Spreadsheet ID: {spreadsheet_id}")
                print(f"   Số dòng trong sheet: {len(rows)}")
                print(f"   ⚠️ Fallback về cập nhật 1 ô (M35)")
                print(f"{'='*80}\n")
                try:
                    sys.stdout.flush()
                except Exception:
                    pass
                result = self.update_sheet_value(
                    spreadsheet_id, sheet_name, 35, 'M',
                    f"Phê duyệt bởi {attendance_data.get('approved_by','')} - {attendance_data.get('approved_at','')}"
                )
                print(f"📊 [FALLBACK_UPDATE] Kết quả fallback: {result}")
                try:
                    sys.stdout.flush()
                except Exception:
                    pass
                return result

            # Mapping cột cụ thể theo yêu cầu
            # Cột G: Giờ vào (Actual clock-in time)
            # Cột K: Giờ ra (Actual Clock-out time: Until 10PM)
            # Cột E: Tổng giờ nghỉ + đối ứng (Planned Working Hours)
            # Cột M: Giờ công (Actual working hours & JP_Hols Work)
            # Cột N: Tăng ca trước 22h (【Until 10 PM】WeekDay OT & WeekEnd & JP_hols OT & Viet_hols)
            # Cột O: Tăng ca sau 22h (【After 10 PM】WeekDay OT & WeekEnd & JP_hols OT & Viet_hols)
            # Cột P: Ghi chú số ngày nghỉ (ưu tiên dùng cho đơn nghỉ phép)
            
            # Sử dụng cột cố định thay vì tìm kiếm header
            column_mapping = {
                'check_in': 'G',      # Giờ vào
                'check_out': 'K',     # Giờ ra  
                'break_comp_total': 'E',  # Tổng giờ nghỉ + đối ứng
                'regular_work_hours': 'M',  # Giờ công
                'overtime_before_22': 'N',  # Tăng ca trước 22h
                'overtime_after_22': 'O',   # Tăng ca sau 22h
                'leave_summary': 'P'  # Ghi chú số ngày nghỉ
            }

            updates = []

            def _is_effective_zero(v):
                try:
                    # None hoặc chuỗi rỗng
                    if v is None:
                        return True
                    s = str(v).strip()
                    if s == '':
                        return True
                    # Các dạng số 0 phổ biến
                    zero_like = {'0', '0.0', '0.00', '00', '00.0', '00.00'}
                    if s in zero_like:
                        return True
                    # Các dạng thời gian 0
                    time_like_zero = {'0:00', '00:00', '0:0'}
                    if s in time_like_zero:
                        return True
                    # Trường hợp "0 h" hoặc "0 giờ"
                    lower = s.lower()
                    if lower in {'0 h', '0h', '0 giờ'}:
                        return True
                except Exception:
                    return False
                return False

            update_ranges = []  # Lưu các ranges để căn giữa sau
            
            def add_update(field_key, value):
                # Bỏ qua giá trị 0/0:00 theo yêu cầu
                if _is_effective_zero(value):
                    return
                col_letter = column_mapping.get(field_key)
                if col_letter:
                    a1 = f"{sheet_name}!{col_letter}{target_row_index}"
                    updates.append({'range': a1, 'values': [[str(value)]]})
                    update_ranges.append(a1)  # Lưu range để căn giữa

            # Tính tổng giờ nghỉ + đối ứng (cột E)
            print(f"\n🧮 TÍNH TOÁN DỮ LIỆU:")
            break_time = attendance_data.get('break_time', '0:00')
            doi_ung_total = attendance_data.get('doi_ung_total', '0:00')  # HH:MM tổng đối ứng
            print(f"   ⏰ Giờ nghỉ: {break_time}")
            print(f"   🔄 Tổng đối ứng: {doi_ung_total}")
            
            # Chuyển đổi giờ nghỉ và đối ứng sang phút để cộng
            def hhmm_to_minutes(hhmm_str):
                try:
                    if not hhmm_str or hhmm_str == '0:00':
                        return 0
                    if ':' in hhmm_str:
                        h, m = hhmm_str.split(':')
                        return int(h) * 60 + int(m)
                    return 0
                except (ValueError, AttributeError, TypeError):
                    return 0
            
            def minutes_to_hhmm(total_minutes):
                hours = total_minutes // 60
                minutes = total_minutes % 60
                return f"{hours}:{minutes:02d}"
            
            break_minutes = hhmm_to_minutes(break_time)
            doi_ung_minutes = hhmm_to_minutes(doi_ung_total)
            total_break_comp_minutes = break_minutes + doi_ung_minutes
            total_break_comp_hhmm = minutes_to_hhmm(total_break_comp_minutes)
            
            print(f"   📊 Giờ nghỉ (phút): {break_minutes}")
            print(f"   📊 Tổng đối ứng (phút): {doi_ung_minutes}")
            print(f"   📊 Tổng nghỉ + đối ứng: {total_break_comp_hhmm}")

            print(f"\n📝 CHUẨN BỊ CẬP NHẬT CÁC CỘT:")

            def _is_full_leave_day(raw_value):
                """Chỉ xem là nghỉ tròn ngày khi giá trị thực sự biểu thị 1.0 ngày."""
                if raw_value is None:
                    return False
                if isinstance(raw_value, bool):
                    return raw_value
                if isinstance(raw_value, (int, float)):
                    return abs(float(raw_value) - 1.0) < 1e-9
                if isinstance(raw_value, str):
                    normalized = raw_value.strip().lower()
                    if normalized in {'true', 'full', 'full_day'}:
                        return True
                    if normalized in {'false', '0', '0.0', 'half'}:
                        return False
                    try:
                        return abs(float(normalized) - 1.0) < 1e-9
                    except ValueError:
                        return False
                return False

            def _to_float_leave_value(raw_value):
                """Chuyển mọi kiểu dữ liệu thành float nếu có thể."""
                try:
                    if raw_value is None:
                        return None
                    if isinstance(raw_value, bool):
                        return 1.0 if raw_value else 0.0
                    if isinstance(raw_value, (int, float)):
                        return float(raw_value)
                    if isinstance(raw_value, str):
                        normalized = raw_value.strip().lower()
                        if normalized in {'', 'none'}:
                            return None
                        normalized = normalized.replace(',', '.')
                        return float(normalized)
                except ValueError:
                    return None
                return None

            def _detect_leave_day_value():
                """Lấy giá trị số ngày nghỉ từ attendance_data hoặc từ summary."""
                value = _to_float_leave_value(attendance_data.get('leave_fraction_days'))
                if value is not None:
                    return value
            
                summary_text = attendance_data.get('leave_summary')
                if summary_text:
                    try:
                        match = re.search(r'(\d+(?:[.,]\d+)?)\s*ngày', summary_text, re.IGNORECASE)
                        if match:
                            return _to_float_leave_value(match.group(1))
                    except Exception:
                        pass
                return None

            leave_day_value = _detect_leave_day_value()

            def _is_half_leave_day():
                return leave_day_value is not None and abs(leave_day_value - 0.5) < 1e-9

            memo_only = bool(attendance_data.get('memo_only'))

            # Nếu là ngày nghỉ tròn 1 ngày (full_leave_day) → xóa sạch các cột thời gian G,H,I,J,K,M
            full_leave_day = _is_full_leave_day(attendance_data.get('full_leave_day'))
            if full_leave_day and not memo_only:
                print("   🔸 Full leave day: clearing columns G,H,I,J,K,M")
                for col_letter in ['G', 'H', 'I', 'J', 'K', 'M']:
                    a1 = f"{sheet_name}!{col_letter}{target_row_index}"
                    updates.append({'range': a1, 'values': [['']]})
                    update_ranges.append(a1)
            else:
                # Không phải full-day
                is_half_day = _is_half_leave_day()

                if memo_only:
                    # Chế độ chỉ memo (ví dụ: nghỉ 30 phút) → không đụng vào E,G,K,M,N,O
                    print("   🔸 Chế độ memo_only: chỉ cập nhật cột P (Memo), bỏ qua các cột thời gian")
                else:
                    # Nếu là nghỉ 0.5 ngày: chỉ thao tác trên E, K, P (không động vào G, M, N, O)
                    if not is_half_day:
                        # Ngày làm bình thường / các loại khác: cập nhật đầy đủ
                        print(f"   🔸 Cột G (Giờ vào): {attendance_data.get('check_in')}")
                        add_update('check_in', attendance_data.get('check_in'))
                        
                        print(f"   🔸 Cột K (Giờ ra): {attendance_data.get('check_out')}")
                        add_update('check_out', attendance_data.get('check_out'))
                        
                        print(f"   🔸 Cột M (Giờ công): {attendance_data.get('regular_work_hours')}")
                        add_update('regular_work_hours', attendance_data.get('regular_work_hours'))
                    else:
                        print("   🔸 Ngày nghỉ 0.5: bỏ qua cập nhật G, M, N, O (chỉ xử lý E, K, P)")
                    
                    break_comp_cell_value = total_break_comp_hhmm
                    if is_half_day:
                        print("   🔸 Cột E (Tổng nghỉ + đối ứng): phát hiện nghỉ 0.5 ngày → set 0:00")
                        break_comp_cell_value = '0:00'
                    else:
                        print(f"   🔸 Cột E (Tổng nghỉ + đối ứng): {break_comp_cell_value}")
                    add_update('break_comp_total', break_comp_cell_value)
                    
                    # Chỉ cập nhật OT cho các ngày không phải nghỉ 0.5
                    if not is_half_day:
                        print(f"   🔸 Cột N (Tăng ca <22h): {attendance_data.get('overtime_before_22')}")
                        add_update('overtime_before_22', attendance_data.get('overtime_before_22'))
                        
                        print(f"   🔸 Cột O (Tăng ca >22h): {attendance_data.get('overtime_after_22')}")
                        add_update('overtime_after_22', attendance_data.get('overtime_after_22'))

            # Ghi chú số ngày nghỉ (nếu có) - đọc giá trị hiện tại và nối thêm nếu có
            leave_summary_value = attendance_data.get('leave_summary')
            if leave_summary_value:
                # Đọc giá trị hiện tại của cột P
                col_p_index = 15  # Cột P là index 15 (A=0, B=1, ..., P=15)
                current_p_value = None
                if target_row_index and len(rows) >= target_row_index:
                    row_data = rows[target_row_index - 1]  # target_row_index là 1-based
                    if len(row_data) > col_p_index:
                        current_p_value = str(row_data[col_p_index]).strip() if row_data[col_p_index] else ""
                
                # Nối thêm nội dung mới vào giá trị hiện tại (với dấu "; " phân cách)
                if current_p_value:
                    # Kiểm tra xem nội dung mới đã có trong cột P chưa (tránh trùng lặp)
                    if leave_summary_value not in current_p_value:
                        final_p_value = f"{current_p_value}; {leave_summary_value}"
                    else:
                        final_p_value = current_p_value  # Đã có rồi, không thêm nữa
                else:
                    final_p_value = leave_summary_value
                
                print(f"   🔸 Cột P (Ghi chú):")
                print(f"      - Giá trị hiện tại: {current_p_value or '(trống)'}")
                print(f"      - Thêm mới: {leave_summary_value}")
                print(f"      - Giá trị cuối: {final_p_value}")
                add_update('leave_summary', final_p_value)
            else:
                print(f"   🔸 Cột P (Ghi chú): Không có nội dung mới")
            
            # Xử lý trừ giờ nghỉ trưa nếu có nghỉ phép và không phải full leave day
            use_lunch_break = attendance_data.get('use_lunch_break')
            if leave_summary_value and not full_leave_day and use_lunch_break is not None and not memo_only:
                # Helper functions để chuyển đổi HH:MM <-> phút
                def hhmm_to_minutes(hhmm_str):
                    try:
                        if not hhmm_str or str(hhmm_str).strip() == '':
                            return 0
                        hhmm_str = str(hhmm_str).strip()
                        if hhmm_str == '0:00' or hhmm_str == '0' or hhmm_str == '00:00':
                            return 0
                        if ':' in hhmm_str:
                            parts = hhmm_str.split(':')
                            h = int(parts[0]) if parts[0] else 0
                            m = int(parts[1]) if len(parts) > 1 and parts[1] else 0
                            return h * 60 + m
                        return 0
                    except Exception:
                        return 0
                
                def minutes_to_hhmm(total_minutes):
                    if total_minutes < 0:
                        total_minutes = 0
                    hours = total_minutes // 60
                    minutes = total_minutes % 60
                    return f"{hours}:{minutes:02d}"
                
                # Tính số giờ cần trừ
                lunch_break_hours = 4 if use_lunch_break else 5
                lunch_break_minutes = lunch_break_hours * 60
                
                # Xử lý cột K (Giờ ra) - luôn xử lý khi nghỉ 0.5 ngày
                col_k_index = 10  # Cột K là index 10 (A=0, B=1, ..., K=10)
                current_k_value = None
                if target_row_index and len(rows) >= target_row_index:
                    row_data = rows[target_row_index - 1]  # target_row_index là 1-based
                    if len(row_data) > col_k_index:
                        current_k_value = row_data[col_k_index]
                
                current_k_minutes = hhmm_to_minutes(current_k_value) if current_k_value else 0
                
                # Xử lý cột K nếu có giá trị > 0
                if current_k_minutes > 0:
                    try:
                        new_k_minutes = current_k_minutes - lunch_break_minutes
                        
                        if new_k_minutes < 0:
                            new_k_minutes = 0
                        
                        new_k_value = minutes_to_hhmm(new_k_minutes)
                        
                        print(f"   🔸 Xử lý giờ nghỉ trưa cho cột K (Giờ ra):")
                        print(f"      - Giá trị cột K hiện tại: {current_k_value or '0:00'} ({current_k_minutes} phút)")
                        print(f"      - Có dùng giờ nghỉ trưa: {use_lunch_break}")
                        print(f"      - Trừ: {lunch_break_hours} giờ ({lunch_break_minutes} phút)")
                        print(f"      - Giá trị cột K mới: {new_k_value} ({new_k_minutes} phút)")
                        
                        # Cập nhật lại cột K
                        a1_k = f"{sheet_name}!K{target_row_index}"
                        updates.append({'range': a1_k, 'values': [[new_k_value]]})
                        update_ranges.append(a1_k)
                    except Exception as k_err:
                        print(f"   ⚠️ Lỗi khi xử lý giờ nghỉ trưa cho cột K: {k_err}")
                else:
                    print(f"   ⚠️ Cột K không có giá trị hoặc bằng 0, bỏ qua xử lý giờ nghỉ trưa cho cột K")
                
                # Nếu là nghỉ 0.5 ngày thì KHÔNG đụng vào cột M (giờ công),
                # để công thức trên Google Sheet tự tính lại từ E/K.
                if not _is_half_leave_day():
                    # Xử lý cột M (Giờ công) - chỉ xử lý nếu có giá trị
                    col_m_index = 12  # Cột M là index 12 (A=0, B=1, ..., M=12)
                    current_m_value = None
                    if target_row_index and len(rows) >= target_row_index:
                        row_data = rows[target_row_index - 1]  # target_row_index là 1-based
                        if len(row_data) > col_m_index:
                            current_m_value = row_data[col_m_index]
                    
                    current_m_minutes = hhmm_to_minutes(current_m_value) if current_m_value else 0
                    
                    # Chỉ xử lý nếu cột M có giá trị > 0
                    if current_m_minutes > 0:
                        try:
                            new_m_minutes = current_m_minutes - lunch_break_minutes
                            
                            if new_m_minutes < 0:
                                new_m_minutes = 0
                            
                            new_m_value = minutes_to_hhmm(new_m_minutes)
                            
                            print(f"   🔸 Xử lý giờ nghỉ trưa cho cột M (Giờ công):")
                            print(f"      - Giá trị cột M hiện tại: {current_m_value or '0:00'} ({current_m_minutes} phút)")
                            print(f"      - Có dùng giờ nghỉ trưa: {use_lunch_break}")
                            print(f"      - Trừ: {lunch_break_hours} giờ ({lunch_break_minutes} phút)")
                            print(f"      - Giá trị cột M mới: {new_m_value} ({new_m_minutes} phút)")
                            
                            # Cập nhật lại cột M
                            a1_m = f"{sheet_name}!M{target_row_index}"
                            updates.append({'range': a1_m, 'values': [[new_m_value]]})
                            update_ranges.append(a1_m)
                        except Exception as m_err:
                            print(f"   ⚠️ Lỗi khi xử lý giờ nghỉ trưa cho cột M: {m_err}")
                    else:
                        print(f"   ⚠️ Cột M không có giá trị hoặc bằng 0, bỏ qua xử lý giờ nghỉ trưa cho cột M")
                
                # Xử lý cột E (Tổng nghỉ + đối ứng) - set về 00:00 nếu không dùng giờ nghỉ trưa
                if not use_lunch_break:
                    try:
                        print(f"   🔸 Xử lý cột E (Tổng nghỉ + đối ứng):")
                        print(f"      - Không dùng giờ nghỉ trưa, set cột E về 00:00")
                        
                        # Cập nhật cột E về 00:00
                        a1_e = f"{sheet_name}!E{target_row_index}"
                        updates.append({'range': a1_e, 'values': [['00:00']]})
                        update_ranges.append(a1_e)
                    except Exception as e_err:
                        print(f"   ⚠️ Lỗi khi xử lý cột E: {e_err}")

            # Xử lý đơn đi trễ/về sớm: trừ giờ từ cột G (đi trễ) hoặc cột K (về sớm)
            late_early_type = attendance_data.get('late_early_type')
            late_early_minutes = attendance_data.get('late_early_minutes', 0)
            if late_early_type and late_early_minutes > 0:
                # Helper functions để chuyển đổi HH:MM <-> phút (nếu chưa định nghĩa)
                def hhmm_to_minutes_le(hhmm_str):
                    try:
                        if not hhmm_str or str(hhmm_str).strip() == '':
                            return 0
                        hhmm_str = str(hhmm_str).strip()
                        if hhmm_str == '0:00' or hhmm_str == '0' or hhmm_str == '00:00':
                            return 0
                        if ':' in hhmm_str:
                            parts = hhmm_str.split(':')
                            h = int(parts[0]) if parts[0] else 0
                            m = int(parts[1]) if len(parts) > 1 and parts[1] else 0
                            return h * 60 + m
                        return 0
                    except Exception:
                        return 0
                
                def minutes_to_hhmm_le(total_minutes):
                    if total_minutes < 0:
                        total_minutes = 0
                    hours = total_minutes // 60
                    minutes = total_minutes % 60
                    return f"{hours}:{minutes:02d}"

                print(f"\n   🔸 Xử lý đơn đi trễ/về sớm:")
                print(f"      - Loại: {late_early_type}")
                print(f"      - Số phút: {late_early_minutes}")

                if late_early_type == 'late':
                    # Đi trễ: CỘNG giờ vào cột G (Giờ vào) - vì đến muộn hơn
                    col_g_index = 6  # Cột G là index 6 (A=0, B=1, ..., G=6)
                    current_g_value = None
                    if target_row_index and len(rows) >= target_row_index:
                        row_data = rows[target_row_index - 1]  # target_row_index là 1-based
                        if len(row_data) > col_g_index:
                            current_g_value = row_data[col_g_index]
                    
                    current_g_minutes = hhmm_to_minutes_le(current_g_value) if current_g_value else 0
                    
                    if current_g_minutes > 0:
                        try:
                            new_g_minutes = current_g_minutes + late_early_minutes  # CỘNG thêm
                            
                            new_g_value = minutes_to_hhmm_le(new_g_minutes)
                            
                            print(f"      - Cột G (Giờ vào) hiện tại: {current_g_value or '0:00'} ({current_g_minutes} phút)")
                            print(f"      - Cộng: {late_early_minutes} phút (đi trễ)")
                            print(f"      - Cột G mới: {new_g_value} ({new_g_minutes} phút)")
                            
                            # Cập nhật lại cột G
                            a1_g = f"{sheet_name}!G{target_row_index}"
                            updates.append({'range': a1_g, 'values': [[new_g_value]]})
                            update_ranges.append(a1_g)
                        except Exception as g_err:
                            print(f"      ⚠️ Lỗi khi xử lý cột G: {g_err}")
                    else:
                        print(f"      ⚠️ Cột G không có giá trị hoặc bằng 0, bỏ qua xử lý đi trễ")

                elif late_early_type == 'early':
                    # Về sớm: trừ giờ từ cột K (Giờ ra)
                    col_k_index = 10  # Cột K là index 10 (A=0, B=1, ..., K=10)
                    current_k_value = None
                    if target_row_index and len(rows) >= target_row_index:
                        row_data = rows[target_row_index - 1]  # target_row_index là 1-based
                        if len(row_data) > col_k_index:
                            current_k_value = row_data[col_k_index]
                    
                    current_k_minutes = hhmm_to_minutes_le(current_k_value) if current_k_value else 0
                    
                    if current_k_minutes > 0:
                        try:
                            new_k_minutes = current_k_minutes - late_early_minutes
                            
                            if new_k_minutes < 0:
                                new_k_minutes = 0
                            
                            new_k_value = minutes_to_hhmm_le(new_k_minutes)
                            
                            print(f"      - Cột K (Giờ ra) hiện tại: {current_k_value or '0:00'} ({current_k_minutes} phút)")
                            print(f"      - Trừ: {late_early_minutes} phút")
                            print(f"      - Cột K mới: {new_k_value} ({new_k_minutes} phút)")
                            
                            # Cập nhật lại cột K
                            a1_k = f"{sheet_name}!K{target_row_index}"
                            updates.append({'range': a1_k, 'values': [[new_k_value]]})
                            update_ranges.append(a1_k)
                        except Exception as k_err:
                            print(f"      ⚠️ Lỗi khi xử lý cột K: {k_err}")
                    else:
                        print(f"      ⚠️ Cột K không có giá trị hoặc bằng 0, bỏ qua xử lý về sớm")

            print(f"\n📊 TỔNG KẾT CẬP NHẬT:")
            print(f"   📝 Số ô sẽ cập nhật: {len(updates)}")
            for i, update in enumerate(updates, 1):
                print(f"   {i}. {update['range']} = {update['values'][0][0]}")
            
            if updates:
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                print(f"\n🚀 [BATCH_UPDATE_START] {timestamp} - Bắt đầu cập nhật batch ({len(updates)} ô)...")
                try:
                    sys.stdout.flush()
                except Exception:
                    pass
                
                ok = self.batch_update_values(spreadsheet_id, updates)
                
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                if ok:
                    print(f"\n{'='*80}")
                    print(f"✅ [BATCH_UPDATE_SUCCESS] {timestamp} - Cập nhật batch thành công!")
                    print(f"   Số ô đã cập nhật: {len(updates)}")
                    print(f"{'='*80}\n")
                    try:
                        sys.stdout.flush()
                    except Exception:
                        pass
                    
                    # Căn giữa tất cả các cells đã cập nhật
                    if update_ranges:
                        timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                        print(f"🎯 [CENTER_ALIGN_START] {timestamp} - Bắt đầu căn giữa {len(update_ranges)} cells...")
                        try:
                            sys.stdout.flush()
                        except Exception:
                            pass
                        center_ok = self.center_align_cells(spreadsheet_id, sheet_name, update_ranges)
                        if center_ok:
                            timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                            print(f"✅ [CENTER_ALIGN_SUCCESS] {timestamp} - Căn giữa thành công!")
                            try:
                                sys.stdout.flush()
                            except Exception:
                                pass
                        else:
                            timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                            print(f"⚠️ [CENTER_ALIGN_WARNING] {timestamp} - Không thể căn giữa, nhưng dữ liệu đã được cập nhật")
                            try:
                                sys.stdout.flush()
                            except Exception:
                                pass
                    
                    # Backup ngay sau khi cập nhật thành công
                    try:
                        create_backup()
                        print("🛡️ Đã tạo backup sau cập nhật timesheet")
                    except Exception as e:
                        print(f"⚠️ Không thể tạo backup sau cập nhật: {e}")
                    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    print(f"✅ [UPDATE_COMPLETE] {timestamp} - Hoàn thành cập nhật timesheet")
                    try:
                        sys.stdout.flush()
                    except Exception:
                        pass
                    return True
                else:
                    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    print(f"\n{'='*80}")
                    print(f"❌ [BATCH_UPDATE_FAILED] {timestamp} - CẬP NHẬT BATCH THẤT BẠI!")
                    print(f"   Spreadsheet ID: {spreadsheet_id}")
                    print(f"   Sheet Name: {sheet_name}")
                    print(f"   Số ô cần cập nhật: {len(updates)}")
                    print(f"   Row Index: {target_row_index}")
                    print(f"{'='*80}\n")
                    try:
                        sys.stdout.flush()
                    except Exception:
                        pass
                    return False

            timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            print(f"\n{'='*80}")
            print(f"⚠️ [NO_UPDATES] {timestamp} - KHÔNG CÓ DỮ LIỆU ĐỂ CẬP NHẬT")
            print(f"   Spreadsheet ID: {spreadsheet_id}")
            print(f"   Sheet Name: {sheet_name}")
            print(f"   Row Index: {target_row_index}")
            print(f"   Attendance Data Keys: {list(attendance_data.keys())}")
            print(f"   ⚠️ Fallback về cập nhật 1 ô (M35)")
            print(f"{'='*80}\n")
            try:
                sys.stdout.flush()
            except Exception:
                pass
            result = self.update_sheet_value(
                spreadsheet_id, sheet_name, 35, 'M',
                f"Phê duyệt bởi {attendance_data.get('approved_by','')} - {attendance_data.get('approved_at','')}"
            )
            print(f"📊 [FALLBACK_UPDATE] Kết quả fallback: {result}")
            try:
                sys.stdout.flush()
            except Exception:
                pass
            return result
        except Exception as e:
            timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            print(f"\n{'='*80}")
            print(f"❌ [UPDATE_EXCEPTION] {timestamp} - Lỗi trong update_timesheet_for_attendance")
            print(f"   Error: {str(e)}")
            print(f"   Type: {type(e).__name__}")
            import traceback
            print(f"   Traceback:")
            print(traceback.format_exc())
            print(f"{'='*80}\n")
            try:
                sys.stdout.flush()
            except Exception:
                pass
            return False
    
    def _find_month_folder(self, folder_id, month_year):
        """Tìm folder tháng trong folder năm"""
        try:
            # Tìm folder năm (2025)
            year = month_year[:4]  # 2025
            month = month_year[4:]  # 10
            
            # Tìm folder năm
            year_query = f"'{folder_id}' in parents and trashed=false and name='{year}' and mimeType='application/vnd.google-apps.folder'"
            year_results = self.drive_service.files().list(
                q=year_query,
                pageSize=10,
                fields="files(id, name, mimeType)"
            ).execute()
            
            year_folders = year_results.get('files', [])
            if not year_folders:
                print(f"❌ Không tìm thấy folder năm {year}")
                return None
            
            year_folder = year_folders[0]
            print(f"✅ Tìm thấy folder năm: {year_folder['name']}")
            
            # Tìm folder tháng trong folder năm
            month_query = f"'{year_folder['id']}' in parents and trashed=false and name='{month}' and mimeType='application/vnd.google-apps.folder'"
            month_results = self.drive_service.files().list(
                q=month_query,
                pageSize=10,
                fields="files(id, name, mimeType)"
            ).execute()
            
            month_folders = month_results.get('files', [])
            if not month_folders:
                print(f"❌ Không tìm thấy folder tháng {month}")
                return None
            
            return month_folders[0]
            
        except Exception as e:
            print(f"❌ Lỗi khi tìm folder tháng: {e}")
            return None
    
    def _search_timesheet_in_folder(self, folder_id, target_name, file_name):
        """Tìm file timesheet trong folder cụ thể - hỗ trợ cả có và không có tiền tố DMI-"""
        try:
            # Tạo danh sách các biến thể tên file để tìm kiếm
            search_variants = []
            
            # Bước 1: Tìm file chính xác với tháng (có và không có DMI-)
            search_variants.append(target_name)  # Tên gốc
            if not target_name.startswith('DMI-'):
                search_variants.append(f"DMI-{target_name}")  # Thêm DMI- nếu chưa có
            else:
                # Nếu đã có DMI-, thử tìm không có DMI-
                search_variants.append(target_name.replace('DMI-', '', 1))
            
            # Bước 2: Tìm file gốc (không có tháng) - có và không có DMI-
            search_variants.append(file_name)  # Tên gốc
            if not file_name.startswith('DMI-'):
                search_variants.append(f"DMI-{file_name}")  # Thêm DMI- nếu chưa có
            else:
                # Nếu đã có DMI-, thử tìm không có DMI-
                search_variants.append(file_name.replace('DMI-', '', 1))
            
            # Loại bỏ trùng lặp
            search_variants = list(dict.fromkeys(search_variants))
            
            # Tìm kiếm với từng biến thể
            for variant in search_variants:
                query = f"'{folder_id}' in parents and trashed=false and name contains '{variant}'"
                results = self.drive_service.files().list(
                    q=query,
                    pageSize=100,
                    fields="files(id, name, mimeType, size, modifiedTime, webViewLink, capabilities)"
                ).execute()
                
                files = results.get('files', [])
                if files:
                    print(f"✅ Tìm thấy file với tên: {variant}")
                    return files
            
            # Bước 3: Tìm kiếm linh hoạt với từ khóa chính (loại bỏ DMI- để tìm kiếm)
            print(f"🔄 Thử tìm kiếm linh hoạt với từ khóa...")
            # Loại bỏ DMI- để tìm kiếm
            base_file_name = file_name.replace('DMI-', '').strip()
            keywords = self._extract_keywords(base_file_name)
            
            for keyword in keywords:
                print(f"   Tìm kiếm với từ khóa: {keyword}")
                keyword_query = f"'{folder_id}' in parents and trashed=false and name contains '{keyword}'"
                keyword_results = self.drive_service.files().list(
                    q=keyword_query,
                    pageSize=100,
                    fields="files(id, name, mimeType, size, modifiedTime, webViewLink, capabilities)"
                ).execute()
                
                keyword_files = keyword_results.get('files', [])
                if keyword_files:
                    # Lọc các file phù hợp nhất (có chứa từ khóa và có "timesheet" hoặc "sheet")
                    filtered_files = [
                        f for f in keyword_files 
                        if 'timesheet' in f['name'].lower() or 'sheet' in f['name'].lower()
                    ]
                    if filtered_files:
                        print(f"✅ Tìm thấy file với từ khóa: {keyword}")
                        return filtered_files
                    else:
                        print(f"✅ Tìm thấy file với từ khóa: {keyword} (không có timesheet trong tên)")
                        return keyword_files
            
            print(f"❌ Không tìm thấy file nào với tất cả từ khóa")
            return []
            
        except Exception as e:
            print(f"❌ Lỗi khi tìm kiếm trong folder: {e}")
            return []
    
    def find_team_timesheet(self, folder_id, team_name, month_year="202510"):
        """Tìm file timesheet của team cụ thể"""
        try:
            # Đảm bảo token luôn hợp lệ trước khi sử dụng API
            if not self.ensure_valid_token():
                print("❌ Không thể đảm bảo token hợp lệ")
                return None
            
            # Lấy tên file từ mapping
            file_name = self.get_department_file_mapping(team_name)
            # Loại bỏ DMI- nếu có để tạo target_name (sẽ tìm cả hai biến thể)
            base_file_name = file_name.replace('DMI-', '').strip() if file_name.startswith('DMI-') else file_name
            target_name = f"{base_file_name}-{month_year}"
            
            print(f"\n🔍 Đang tìm file timesheet của team: {team_name}")
            print(f"📁 Mapping từ database: {file_name}")
            print(f"📁 Tên file tìm kiếm: {target_name} (sẽ thử cả DMI-{target_name} nếu cần)")
            print(f"{'='*60}")
            
            # Bước 1: Tìm trong folder con 2025/10/ trước
            print(f"🔍 Tìm kiếm trong folder con 2025/10/...")
            target_folder = self._find_month_folder(folder_id, month_year)
            
            if target_folder:
                print(f"✅ Tìm thấy folder tháng: {target_folder['name']}")
                files = self._search_timesheet_in_folder(target_folder['id'], target_name, file_name)
                if files:
                    return files[0]  # Trả về file đầu tiên tìm thấy
            else:
                print(f"❌ Không tìm thấy folder tháng {month_year}")
            
            # Bước 2: Fallback - tìm trong folder gốc
            print(f"🔄 Fallback: Tìm trong folder gốc...")
            files = self._search_timesheet_in_folder(folder_id, target_name, file_name)
            
            if files:
                return files[0]  # Trả về file đầu tiên tìm thấy
            
            print(f"❌ Không tìm thấy file timesheet cho team {team_name}")
            return None
            
        except Exception as e:
            print(f"❌ Lỗi khi tìm file timesheet: {e}")
            return None
    
    def list_all_timesheets(self, folder_id):
        """Lấy danh sách tất cả file timesheet trong folder"""
        try:
            if not self.ensure_valid_token():
                print("❌ Không thể đảm bảo token hợp lệ")
                return []
            
            query = f"'{folder_id}' in parents and trashed=false and (mimeType='application/vnd.google-apps.spreadsheet' or mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')"
            results = self.drive_service.files().list(
                q=query,
                pageSize=100,
                fields="files(id, name, mimeType, size, modifiedTime, webViewLink)"
            ).execute()
            
            files = results.get('files', [])
            return files
            
        except Exception as e:
            print(f"❌ Lỗi khi lấy danh sách timesheet: {e}")
            return []
    
    def _get_file_type(self, mime_type):
        """Chuyển đổi mime type sang tên dễ hiểu"""
        types = {
            'application/vnd.google-apps.spreadsheet': 'Google Sheets',
            'application/vnd.google-apps.document': 'Google Docs',
            'application/vnd.google-apps.folder': 'Folder',
            'application/pdf': 'PDF',
            'text/csv': 'CSV',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'Excel'
        }
        return types.get(mime_type, mime_type)
    
    def get_department_file_mapping(self, department):
        """Mapping phòng ban với tên file timesheet - ưu tiên đọc từ database"""
        # Thử đọc từ database trước
        try:
            from database.models import Department
            dept = Department.query.filter_by(name=department, is_active=True).first()
            if dept and dept.timesheet_file:
                return dept.timesheet_file
        except Exception:
            pass
        
        # Fallback về hardcoded mapping cho backward compatibility
        mapping = {
            'BUD A': 'Bud_TimeSheet',
            'BUD B': 'Bud_TimeSheet',
            'BUD C': 'Bud_TimeSheet',
            'CREEK&RIVER': 'Creek&River_timesheet',
            'KIRI': 'KIRI TIME SHEET',
            'OFFICE': 'BACKOFFICE_TIMESHEET',
            'YORK': 'Chirashi_TimeSheet',
            'COMO': 'Chirashi_TimeSheet',
            'IT': 'IT_TimeSheet',
            'SCOPE': 'SCOPE_TimeSheet'
        }
        return mapping.get(department, f"{department}_TimeSheet")

    def get_all_department_mappings(self):
        """Lấy tất cả mapping phòng ban - kết hợp database và hardcoded"""
        # Bắt đầu với hardcoded mapping
        result = {
            'BUD A': 'Bud_TimeSheet',
            'BUD B': 'Bud_TimeSheet',
            'BUD C': 'Bud_TimeSheet',
            'CREEK&RIVER': 'Creek&River_timesheet',
            'KIRI': 'KIRI TIME SHEET',
            'OFFICE': 'BACKOFFICE_TIMESHEET',
            'YORK': 'Chirashi_TimeSheet',
            'COMO': 'Chirashi_TimeSheet',
            'IT': 'IT_TimeSheet',
            'SCOPE': 'SCOPE_TimeSheet'
        }
        
        # Thêm/override từ database
        try:
            from database.models import Department
            db_depts = Department.query.filter_by(is_active=True).all()
            for dept in db_depts:
                if dept.timesheet_file:
                    result[dept.name] = dept.timesheet_file
        except Exception:
            pass
        
        return result
    
    def _extract_keywords(self, file_name):
        """Trích xuất từ khóa chính từ tên file để tìm kiếm linh hoạt - loại bỏ DMI-"""
        keywords = []
        
        # Loại bỏ tiền tố DMI- nếu có
        clean_file_name = file_name.replace('DMI-', '').strip() if file_name.startswith('DMI-') else file_name
        
        # Chuẩn hóa tên file trước khi xử lý
        normalized_name = self._normalize_text(clean_file_name)
        
        # Chia thành từ
        words = normalized_name.split()
        
        # Thêm từ khóa theo độ ưu tiên
        for word in words:
            if len(word) > 2:  # Chỉ lấy từ có độ dài > 2
                keywords.append(word)
        
        # Thêm từ khóa đặc biệt cho các trường hợp cụ thể (so sánh với text đã chuẩn hóa)
        if 'bud' in normalized_name:
            keywords.extend(['bud', 'timesheet'])
        elif 'creek' in normalized_name or 'river' in normalized_name:
            keywords.extend(['creek', 'river', 'timesheet'])
        elif 'kiri' in normalized_name or 'kirinuki' in normalized_name:
            keywords.extend(['kiri', 'kirinuki', 'time', 'sheet'])
        elif 'backoffice' in normalized_name or 'back' in normalized_name:
            keywords.extend(['backoffice', 'back', 'timesheet'])
        elif 'chirashi' in normalized_name:
            keywords.extend(['chirashi', 'timesheet'])
        elif 'scope' in normalized_name:
            keywords.extend(['scope', 'timesheet'])
        elif 'it' in normalized_name:
            keywords.extend(['it', 'timesheet'])
        
        # Loại bỏ trùng lặp và sắp xếp theo độ ưu tiên
        unique_keywords = list(dict.fromkeys(keywords))
        return unique_keywords
    
    def _normalize_text(self, text):
        """Chuẩn hóa text để so sánh dễ dàng"""
        if not text:
            return ""
        
        # Chuyển về chữ thường và loại bỏ ký tự đặc biệt
        import re
        normalized = re.sub(r'[&_\-\.]', ' ', text.lower())
        # Loại bỏ khoảng trắng thừa
        normalized = ' '.join(normalized.split())
        return normalized

# ====== BACKUP SCHEDULER ======
import threading
import shutil
import requests
import time

_backup_scheduler_lock = threading.Lock()
_backup_scheduler_started = False

# ====== TOKEN KEEP-ALIVE SCHEDULER ======
_token_keepalive_lock = threading.Lock()
_token_keepalive_started = False

# ====== LICENSE ONLINE CHECK SCHEDULER ======
_license_check_lock = threading.Lock()
_license_check_started = False
_license_is_valid = True  # Global flag để track license status


def _force_shutdown_app(reason: str = ""):
    """
    Dừng toàn bộ chương trình NGAY LẬP TỨC (dùng cho trường hợp license không hợp lệ / hết hạn).
    Sử dụng os._exit để chắc chắn dừng cả Flask và mọi thread nền.
    """
    try:
        print(f"[LICENSE] Ứng dụng sẽ dừng ngay lập tức. Lý do: {reason}", flush=True)
    except Exception:
        pass
    # Thoát process ngay lập tức, không chạy clean-up handlers
    os._exit(1)


def _license_check_worker(interval_seconds: int = 300):
    """
    Thread nền kiểm tra license online liên tục qua License Manager Pro.
    Nếu license hết hạn/không hợp lệ -> chặn truy cập app (không thoát app).
    """
    global _license_is_valid, _license_warning_state
    from datetime import datetime

    # Đợi vài giây cho app & database khởi động xong
    print(f"[LICENSE] License checker worker đang khởi động, sẽ bắt đầu kiểm tra sau 5 giây...", flush=True)
    time.sleep(5)
    print(f"[LICENSE] Bắt đầu kiểm tra license định kỳ (mỗi {interval_seconds} giây)...", flush=True)

    while True:
        try:
            with app.app_context():
                activation = None
                try:
                    activation = get_activation_record()
                except Exception as e:
                    print(f"[LICENSE] Lỗi lấy activation record: {e}")

                # Ưu tiên APP_LICENSE_KEY (biến môi trường), fallback sang DB
                license_key = None
                # Ưu tiên APP_LICENSE_KEY trước (để override key trong DB nếu cần)
                license_key = (APP_LICENSE_KEY or '').strip()
                # Nếu không có APP_LICENSE_KEY, mới lấy từ DB
                if not license_key and activation is not None:
                    license_key = (getattr(activation, 'license_key', None) or '').strip()

                if not license_key:
                    print("[LICENSE] Không tìm thấy license key để verify. Đã chặn truy cập app.")
                    global _license_is_valid
                    _license_is_valid = False
                    time.sleep(interval_seconds)
                    continue

                # Ưu tiên dùng API endpoint, fallback về ?verify= nếu API không có
                verify_url = f"{LICENSE_VERIFY_ENDPOINT.rstrip('/')}/api/verify?verify={license_key}"
                try:
                    resp = requests.get(verify_url, timeout=10)
                except Exception as e:
                    # Lỗi mạng tạm thời: log lại nhưng KHÔNG dừng app ngay
                    try:
                        from datetime import datetime, timedelta
                        next_check_time = datetime.now() + timedelta(seconds=interval_seconds)
                        next_check_str = next_check_time.strftime("%H:%M:%S")
                        print(f"[LICENSE] Lỗi kết nối server license: {e}. Sẽ thử lại sau {interval_seconds} giây (lần tiếp theo: {next_check_str})", flush=True)
                    except Exception:
                        print(f"[LICENSE] Lỗi kết nối server license: {e}. Sẽ thử lại sau {interval_seconds} giây.", flush=True)
                    time.sleep(interval_seconds)
                    continue

                if resp.status_code != 200:
                    try:
                        from datetime import datetime, timedelta
                        next_check_time = datetime.now() + timedelta(seconds=interval_seconds)
                        next_check_str = next_check_time.strftime("%H:%M:%S")
                        print(f"[LICENSE] Server license trả về status {resp.status_code}. Sẽ thử lại sau {interval_seconds} giây (lần tiếp theo: {next_check_str})", flush=True)
                    except Exception:
                        print(f"[LICENSE] Server license trả về status {resp.status_code}. Sẽ thử lại sau {interval_seconds} giây.", flush=True)
                    time.sleep(interval_seconds)
                    continue

                try:
                    data = resp.json()
                except Exception as e:
                    # Server license không trả về JSON chuẩn (thường là HTML hiển thị thông tin license).
                    # Thay vì dừng app ngay, ta cố gắng phân tích nội dung text để suy ra trạng thái license.
                    print(f"[LICENSE] Không parse được JSON từ server license: {e}. Thử phân tích nội dung text...")
                    data = _infer_license_data_from_text(resp.text or "", license_key=license_key)
                    if not data:
                        # Không đoán được trạng thái từ nội dung -> log lại và tạm cho qua để không khóa app nhầm.
                        try:
                            from datetime import datetime, timedelta
                            next_check_time = datetime.now() + timedelta(seconds=interval_seconds)
                            next_check_str = next_check_time.strftime("%H:%M:%S")
                            print(f"[LICENSE] Không xác định được trạng thái license từ nội dung server. Sẽ thử lại sau {interval_seconds} giây (lần tiếp theo: {next_check_str})", flush=True)
                        except Exception:
                            print(f"[LICENSE] Không xác định được trạng thái license từ nội dung server. Sẽ thử lại sau {interval_seconds} giây.", flush=True)
                        time.sleep(interval_seconds)
                        continue

                is_valid = bool(data.get("valid", False))
                status = str(data.get("status", "")).lower()
                days_remaining = data.get("days_remaining", None)
                expiry_str = data.get("expiry", "")

                # Kiểm tra hết hạn dựa trên các field trả về
                expired = False

                # Tính toán days_remaining từ expiry để đảm bảo chính xác (override giá trị từ API nếu có)
                if expiry_str:
                    try:
                        # Parse expiry date (có thể là UTC với 'Z')
                        if 'Z' in expiry_str:
                            # Chuyển UTC về local time để khớp với UI (UI dùng new Date() = local time)
                            from datetime import timezone
                            expiry_dt_utc = datetime.fromisoformat(expiry_str.replace('Z', '+00:00'))
                            # Chuyển UTC sang local time
                            expiry_dt = expiry_dt_utc.astimezone().replace(tzinfo=None)
                        else:
                            expiry_dt = datetime.fromisoformat(expiry_str)
                            if expiry_dt.tzinfo:
                                expiry_dt = expiry_dt.astimezone().replace(tzinfo=None)
                        
                        # Dùng local time như UI (new Date() trong JavaScript)
                        now_dt = datetime.now()
                        
                        # Tính days_remaining chính xác (dùng floor như UI: Math.floor(diffMs / (1000 * 60 * 60 * 24)))
                        diff_time = expiry_dt - now_dt
                        # Tính số milliseconds (giống JavaScript) - có thể âm nếu đã hết hạn
                        diff_ms = diff_time.total_seconds() * 1000
                        # Math.floor(diffMs / (1000 * 60 * 60 * 24))
                        calculated_days_remaining = int(diff_ms / (1000 * 60 * 60 * 24))
                        # Override days_remaining từ API bằng giá trị tính toán chính xác
                        # Nếu âm nghĩa là đã hết hạn, nếu dương nghĩa là còn hạn
                        days_remaining = calculated_days_remaining
                        if calculated_days_remaining < 0:
                            expired = True
                            print(f"[LICENSE] Đã tính lại days_remaining từ expiry: {calculated_days_remaining} ngày (ĐÃ HẾT HẠN) (từ API: {data.get('days_remaining', 'N/A')})")
                        else:
                            print(f"[LICENSE] Đã tính lại days_remaining từ expiry: {calculated_days_remaining} ngày (từ API: {data.get('days_remaining', 'N/A')})")
                    except Exception as e:
                        print(f"[LICENSE] Lỗi parse expiry '{expiry_str}': {e}")

                # 1) Nếu có days_remaining
                if isinstance(days_remaining, (int, float)):
                    # Nếu còn ≤ 1 ngày nhưng CHƯA hết hạn -> gửi cảnh báo cho toàn hệ thống (giống cơ chế refresh token)
                    # Áp dụng cho cả trường hợp days_remaining == 1 và days_remaining == 0 nhưng vẫn còn giờ/phút/giây.
                    if 0 <= days_remaining <= 1 and not expired:
                        try:
                            warning_message = (
                                "⚠️ ỨNG DỤNG CHẤM CÔNG & NGHỈ PHÉP SẮP HẾT HẠN LICENSE.\n\n"
                                "- Toàn bộ NHÂN VIÊN cần NHANH CHÓNG hoàn thành việc nhập đầy đủ dữ liệu chấm công, đăng ký nghỉ phép, tăng ca... trước khi key hết hạn.\n"
                                "- Các TRƯỞNG NHÓM / QUẢN LÝ / QUẢN TRỊ VIÊN cần PHÊ DUYỆT TẤT CẢ các đơn chấm công, nghỉ phép, tăng ca của cấp dưới trong thời gian sớm nhất.\n\n"
                                "LƯU Ý QUAN TRỌNG:\n"
                                "- Sau khi license hết hạn, hệ thống chấm công và nghỉ phép sẽ TẠM DỪNG HOẠT ĐỘNG, không thể tiếp tục nhập liệu hay phê duyệt.\n\n"
                                "ĐỀ NGHỊ ADMIN SỚM LIÊN HỆ DEVELOPER ĐỂ GIA HẠN LICENSE:\n"
                                "Nguyễn Công Đạt - 0375097105."
                            )
                            # Dùng cơ chế publish_token_status để hiện banner cảnh báo (giống refresh token)
                            try:
                                publish_token_status(
                                    'expired',
                                    warning_message,
                                    needs_reauth=False
                                )
                            except Exception:
                                pass
                            # Gửi thêm qua Telegram nếu có cấu hình
                            try:
                                send_telegram_message(warning_message)
                            except Exception:
                                pass
                        except Exception:
                            # Không để lỗi cảnh báo làm hỏng luồng chính
                            pass
                    else:
                        # Nếu days_remaining > 1 (hoặc âm nhưng đã được xử lý expired ở dưới) -> clear cảnh báo LICENSE nếu đang bật
                        try:
                            from datetime import datetime
                            with _license_warning_lock:
                                _license_warning_state = {
                                    'active': False,
                                    'payload': None,
                                    'updated_at': datetime.now().isoformat(),
                                }
                        except Exception:
                            pass

                    if days_remaining < 0:
                        expired = True

                if (not is_valid) or expired or (status not in ("active", "đang hoạt động", "")):
                    # Khi license hết hạn hoặc không hợp lệ -> chặn truy cập app nhưng không thoát
                    _license_is_valid = False
                    base_msg = data.get("message", "License không hợp lệ hoặc đã hết hạn")
                    contact_msg = (
                        f"{base_msg}\n\n"
                        "Vui lòng liên hệ ADMIN để gia hạn:\n"
                        "Nguyễn Công Đạt - 0375097105."
                    )
                    # In rõ license key đang sử dụng để dễ debug khi hết hạn/không hợp lệ
                    try:
                        print(f"[LICENSE] License key đang dùng: {license_key}")
                    except Exception:
                        pass
                    print(f"[LICENSE] License KHÔNG HỢP LỆ / HẾT HẠN - Đã chặn truy cập app: {contact_msg}")
                    try:
                        # Thử gửi thông báo qua Telegram nếu đã cấu hình
                        send_telegram_message(f"[LICENSE EXPIRED]\n{contact_msg}")
                    except Exception:
                        pass
                    # KHÔNG gọi _force_shutdown_app nữa, chỉ set flag để chặn truy cập
                else:
                    # License hợp lệ -> cho phép truy cập
                    if not _license_is_valid:
                        print(f"[LICENSE] License đã được gia hạn - Cho phép truy cập lại app")
                    _license_is_valid = True

                    # Nếu trước đó có cảnh báo LICENSE, clear cache để UI ẩn banner ở lần load sau
                    try:
                        with _license_warning_lock:
                            _license_warning_state = {
                                'active': False,
                                'payload': None,
                                'updated_at': datetime.now().isoformat(),
                            }
                    except Exception:
                        pass

                # Nếu tới đây là license vẫn hợp lệ
                try:
                    exp_info = expiry_str or "N/A"
                    print(f"[LICENSE] License hợp lệ. Hết hạn: {exp_info}, days_remaining={days_remaining}")
                except Exception:
                    pass

        except Exception as e:
            print(f"[LICENSE] Lỗi không mong đợi trong license_check_worker: {e}")

        # Log thông tin về lần check tiếp theo
        try:
            from datetime import datetime, timedelta
            next_check_time = datetime.now() + timedelta(seconds=interval_seconds)
            next_check_str = next_check_time.strftime("%H:%M:%S")
            print(f"[LICENSE] Sẽ kiểm tra lại sau {interval_seconds} giây (lần tiếp theo: {next_check_str})", flush=True)
        except Exception:
            print(f"[LICENSE] Sẽ kiểm tra lại sau {interval_seconds} giây", flush=True)

        # Ngủ trước khi kiểm tra lại
        time.sleep(interval_seconds)


def _check_license_once() -> tuple[bool, bool, str, str]:
    """
    Chạy 1 lần logic kiểm tra license.
    Trả về: (is_valid, expired, status, message)
    KHÔNG tự tắt app trong hàm này, để caller quyết định.
    """
    from datetime import datetime as _dt_mod

    with app.app_context():
        activation = None
        try:
            activation = get_activation_record()
        except Exception as e:
            print(f"[LICENSE] Lỗi lấy activation record: {e}")

        # Ưu tiên APP_LICENSE_KEY (biến môi trường), fallback sang DB
        license_key = None
        # Ưu tiên APP_LICENSE_KEY trước (để override key trong DB nếu cần)
        license_key = (APP_LICENSE_KEY or '').strip()
        print(f"[LICENSE] License key from APP_LICENSE_KEY: {license_key}", flush=True)
        # Nếu không có APP_LICENSE_KEY, mới lấy từ DB
        if not license_key and activation is not None:
            license_key = (getattr(activation, 'license_key', None) or '').strip()
            print(f"[LICENSE] License key from DB: {license_key}", flush=True)

        if not license_key:
            msg = "Không có license key"
            print(f"[LICENSE] {msg}")
            return False, True, "missing", msg

        # Ưu tiên dùng API endpoint, fallback về ?verify= nếu API không có
        verify_url = f"{LICENSE_VERIFY_ENDPOINT.rstrip('/')}/api/verify?verify={license_key}"
        print(f"[LICENSE] Calling API: {verify_url}", flush=True)
        try:
            resp = requests.get(verify_url, timeout=10)
            print(f"[LICENSE] API Status Code: {resp.status_code}", flush=True)
        except Exception as e:
            msg = f"Lỗi kết nối server license: {e}"
            print(f"[LICENSE] {msg}")
            # Lỗi mạng: coi như chưa xác định, nhưng không đánh expired, để caller quyết định
            return True, False, "unknown", msg

        if resp.status_code != 200:
            msg = f"Server license trả về status {resp.status_code}"
            print(f"[LICENSE] {msg}")
            return True, False, "unknown", msg

        # Parse JSON hoặc text như trong worker
        try:
            data = resp.json()
            # Debug: log full response để kiểm tra
            print(f"[LICENSE] API Response Full: {resp.text[:500]}", flush=True)
            print(f"[LICENSE] API Response Parsed: valid={data.get('valid')}, status={data.get('status')}, message={data.get('message')}, key={data.get('key')}", flush=True)
        except Exception as e:
            print(f"[LICENSE] Không parse được JSON từ server license: {e}. Thử phân tích nội dung text...")
            print(f"[LICENSE] Response text: {resp.text[:200]}", flush=True)
            data = _infer_license_data_from_text(resp.text or "", license_key=license_key)
            if not data:
                msg = "Không xác định được trạng thái license từ nội dung server."
                print(f"[LICENSE] {msg}")
                return True, False, "unknown", msg

        is_valid = bool(data.get("valid", False))
        status = str(data.get("status", "")).lower()
        days_remaining = data.get("days_remaining", None)
        expiry_str = data.get("expiry", "")

        expired = False
        if isinstance(days_remaining, (int, float)) and days_remaining < 0:
            expired = True
        if expiry_str and not expired:
            try:
                normalized = expiry_str.replace('Z', '+00:00') if 'Z' in expiry_str else expiry_str
                expiry_dt = _dt_mod.fromisoformat(normalized)
                from datetime import timezone
                now_dt = _dt_mod.now() if expiry_dt.tzinfo is None else _dt_mod.now(timezone.utc)
                if expiry_dt < now_dt:
                    expired = True
            except Exception as e:
                print(f"[LICENSE] Lỗi parse expiry '{expiry_str}': {e}")

        msg = data.get("message", "")
        # Debug: log kết quả cuối cùng
        print(f"[LICENSE] Check result: is_valid={is_valid}, expired={expired}, status={status}, msg={msg}", flush=True)
        return is_valid, expired, status, msg

def ensure_license_check_started(interval_seconds: int = 300):
    """
    Đảm bảo thread kiểm tra license online chỉ khởi động một lần.
    """
    global _license_check_started
    if _license_check_started:
        return
    with _license_check_lock:
        if _license_check_started:
            return
        try:
            t = threading.Thread(
                target=_license_check_worker,
                args=(interval_seconds,),
                daemon=True
            )
            t.start()
            _license_check_started = True
            print(f"[LICENSE] License online checker started (mỗi {interval_seconds} giây).")
        except Exception as e:
            print(f"[LICENSE] Không thể khởi động license online checker: {e}")

# Telegram Bot Configuration
BOT_TOKEN = "7970479477:AAFzt-MNjgY57DOVLvWTNSBuoYxYlSxKZpo"
CHAT_ID = "6070177456"

def _ensure_dir(path):
    try:
        os.makedirs(path, exist_ok=True)
    except Exception as e:
        print(f"⚠️ Không thể tạo thư mục {path}: {e}")

def _list_backups(backup_dir, prefix="attendance", suffix=".db"):
    try:
        files = [
            os.path.join(backup_dir, f) for f in os.listdir(backup_dir)
            if f.startswith(prefix) and f.endswith(suffix)
        ]
        files.sort(key=lambda p: os.path.getmtime(p))
        return files
    except Exception:
        return []

def send_telegram_file(file_path, caption=None):
    """Gửi file lên Telegram"""
    try:
        if not BOT_TOKEN or not CHAT_ID:
            print("⚠️ Chưa cấu hình BOT_TOKEN hoặc CHAT_ID")
            return False
        
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"
        
        with open(file_path, 'rb') as file:
            files = {'document': file}
            data = {
                'chat_id': CHAT_ID,
                'caption': caption or f"🛡️ Backup database - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            }
            
            response = requests.post(url, files=files, data=data, timeout=30)
            
            if response.status_code == 200:
                print(f"📤 Đã gửi backup lên Telegram: {os.path.basename(file_path)}")
                return True
            else:
                print(f"❌ Lỗi gửi Telegram: {response.status_code} - {response.text}")
                return False
                
    except Exception as e:
        print(f"❌ Lỗi khi gửi file lên Telegram: {e}")
        return False

def create_backup(backup_dir="backups", retention=3, send_to_telegram=True):
    """
    Sao lưu database: ưu tiên instance/attendance.db; fallback attendance.db tại root.
    Giữ tối đa 'retention' bản backup gần nhất.
    Gửi file backup lên Telegram nếu send_to_telegram=True.
    """
    try:
        _ensure_dir(backup_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        src_candidates = [
            os.path.join("instance", "attendance.db"),
            os.path.join(os.getcwd(), "attendance.db"),
        ]
        src = None
        for c in src_candidates:
            if os.path.exists(c):
                src = c
                break
        if not src:
            print("⚠️ Không tìm thấy file database để backup")
            return False
        base_name = f"attendance_{timestamp}.db"
        dst = os.path.join(backup_dir, base_name)
        shutil.copy2(src, dst)
        print(f"✅ Đã backup database: {dst}")
        
        # Gửi lên Telegram nếu được yêu cầu
        if send_to_telegram:
            try:
                send_telegram_file(dst, f"🛡️ Backup database - {timestamp}")
            except Exception as e:
                print(f"⚠️ Không thể gửi backup lên Telegram: {e}")
        
        # Retention
        backups = _list_backups(backup_dir, prefix="attendance_", suffix=".db")
        if len(backups) > retention:
            to_delete = backups[:len(backups)-retention]
            for f in to_delete:
                try:
                    os.remove(f)
                    print(f"🧹 Xóa bản backup cũ: {f}")
                except Exception as e:
                    print(f"⚠️ Không xóa được backup cũ {f}: {e}")
        return True
    except Exception as e:
        print(f"❌ Lỗi khi backup database: {e}")
        return False

def _backup_worker(interval_minutes=180, backup_dir="backups", retention=3, send_to_telegram=True):
    """Worker chạy nền để backup định kỳ."""
    # Đợi interval đầu tiên trước khi chạy backup lần đầu (tránh tạo backup ngay khi khởi động)
    try:
        interval_seconds = max(60, int(interval_minutes) * 60)
    except Exception:
        interval_seconds = 3 * 60 * 60  # Fallback 3 giờ
    
    while True:
        # Ngủ trước, sau đó mới chạy backup (tránh tạo backup ngay khi khởi động)
        try:
            time_module.sleep(interval_seconds)
        except Exception:
            # Fallback ngủ 3 giờ nếu cấu hình lỗi
            time_module.sleep(3 * 60 * 60)
        
        # Sau khi ngủ xong, mới chạy backup
        try:
            create_backup(backup_dir=backup_dir, retention=retention, send_to_telegram=send_to_telegram)
        except Exception as e:
            print(f"⚠️ Lỗi trong backup worker: {e}")

def ensure_backup_scheduler_started(interval_minutes=180, backup_dir="backups", retention=3, send_to_telegram=True):
    """
    Khởi chạy thread backup định kỳ một lần duy nhất.
    """
    global _backup_scheduler_started
    if _backup_scheduler_started:
        return
    with _backup_scheduler_lock:
        if _backup_scheduler_started:
            return
        try:
            t = threading.Thread(
                target=_backup_worker,
                kwargs={
                    'interval_minutes': interval_minutes,
                    'backup_dir': backup_dir,
                    'retention': retention,
                    'send_to_telegram': send_to_telegram,
                },
                daemon=True
            )
            t.start()
            _backup_scheduler_started = True
            telegram_status = " + Telegram" if send_to_telegram else ""
            print(f"🛡️ Backup scheduler started: every {interval_minutes} minutes, dir='{backup_dir}', retention={retention}{telegram_status}")
        except Exception as e:
            print(f"⚠️ Không thể khởi chạy backup scheduler: {e}")

# ====== TOKEN KEEP-ALIVE FUNCTIONS ======

def _token_keepalive_worker(interval_minutes=30):
    """Worker chạy nền để giữ token sống - KHÔNG tự động authenticate"""
    while True:
        try:
            # Chỉ load token từ file, không tự động authenticate
            google_api = GoogleDriveAPI(auto_authenticate=False)
            
            # Nếu không có token hoặc token không hợp lệ, chỉ thông báo, không authenticate
            if not google_api.creds:
                print(f"⚠️ [Token Keep-Alive] Không có token. Cần admin bấm nút Refresh Token để ủy quyền.")
                try:
                    publish_token_status('expired', 'Không có token. Vui lòng bấm nút Refresh Token để ủy quyền.', needs_reauth=True)
                except Exception:
                    pass
                time_module.sleep(interval_minutes * 60)
                continue
            
            # Kiểm tra và refresh token nếu cần (chỉ refresh tự động, không authenticate)
            if google_api.creds.expired:
                if google_api.creds.refresh_token:
                    try:
                        print(f"🔄 [Token Keep-Alive] Đang refresh token tự động...")
                        google_api.creds.refresh(GoogleRequest())
                        # Lưu token mới
                        with open(google_api.token_file, 'wb') as token:
                            pickle.dump(google_api.creds, token)
                        google_api.save_last_refresh_time()
                        print(f"✅ [Token Keep-Alive] Token đã được refresh thành công!")
                        # Notify admins that token is valid
                        try:
                            publish_token_status('valid', 'Token đã được refresh tự động thành công!')
                        except Exception:
                            pass
                    except Exception as refresh_err:
                        error_str = str(refresh_err).lower()
                        print(f"⚠️ [Token Keep-Alive] Không thể refresh token tự động: {refresh_err}")
                        # Notify admins that token needs reauth
                        try:
                            if 'invalid_grant' in error_str:
                                publish_token_status('expired', 'Token không hợp lệ. Vui lòng bấm nút Refresh Token để ủy quyền lại.', needs_reauth=True)
                            else:
                                publish_token_status('expired', 'Token hết hạn. Vui lòng bấm nút Refresh Token để ủy quyền lại.', needs_reauth=True)
                        except Exception:
                            pass
                else:
                    print(f"⚠️ [Token Keep-Alive] Token hết hạn và không có refresh_token. Cần admin bấm nút Refresh Token.")
                    try:
                        publish_token_status('expired', 'Token hết hạn. Vui lòng bấm nút Refresh Token để ủy quyền lại.', needs_reauth=True)
                    except Exception:
                        pass
            else:
                print(f"ℹ️ [Token Keep-Alive] Token vẫn còn hiệu lực")
            
        except Exception as e:
            print(f"❌ [Token Keep-Alive] Lỗi: {e}")
        
        # Ngủ theo khoảng thời gian
        try:
            time_module.sleep(max(60, int(interval_minutes) * 60))
        except Exception:
            # Fallback ngủ 30 phút nếu cấu hình lỗi
            time_module.sleep(30 * 60)

def ensure_token_keepalive_started(interval_minutes=30):
    """
    Khởi chạy thread giữ token sống một lần duy nhất.
    """
    global _token_keepalive_started
    if _token_keepalive_started:
        return
    with _token_keepalive_lock:
        if _token_keepalive_started:
            return
        try:
            t = threading.Thread(
                target=_token_keepalive_worker,
                kwargs={
                    'interval_minutes': interval_minutes,
                },
                daemon=True
            )
            t.start()
            _token_keepalive_started = True
            print(f"🔑 Token Keep-Alive started: every {interval_minutes} minutes")
        except Exception as e:
            print(f"⚠️ Không thể khởi chạy Token Keep-Alive: {e}")

# ====== YEARLY SCHEDULE RESET SCHEDULER ======
_yearly_reset_scheduler_lock = threading.Lock()
_yearly_reset_scheduler_started = False
_last_reset_year = None
_last_notification_days = set()  # Lưu các ngày đã gửi thông báo để tránh gửi trùng

def send_telegram_message(message):
    """Gửi tin nhắn văn bản lên Telegram"""
    try:
        if not BOT_TOKEN or not CHAT_ID:
            print("⚠️ Chưa cấu hình BOT_TOKEN hoặc CHAT_ID")
            return False
        
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        data = {
            'chat_id': CHAT_ID,
            'text': message,
            'parse_mode': 'HTML'
        }
        
        response = requests.post(url, json=data, timeout=30)
        
        if response.status_code == 200:
            print(f"📤 Đã gửi thông báo lên Telegram")
            return True
        else:
            print(f"❌ Lỗi gửi Telegram: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        print(f"❌ Lỗi khi gửi tin nhắn lên Telegram: {e}")
        return False

def get_admin_users():
    """Lấy danh sách tất cả admin users"""
    try:
        with app.app_context():
            from database.models import User
            admins = User.query.filter(
                User.roles.like('%ADMIN%'),
                User.is_deleted == False,
                User.is_active == True
            ).all()
            return admins
    except Exception as e:
        print(f"❌ Lỗi khi lấy danh sách admin: {e}")
        return []

def check_december_data_complete(year):
    """Kiểm tra dữ liệu tháng 12 đã đầy đủ chưa"""
    try:
        with app.app_context():
            from database.models import Attendance, User
            from datetime import date
            
            # Lấy tất cả users đang hoạt động
            active_users = User.query.filter_by(is_deleted=False, is_active=True).all()
            
            # Kiểm tra từng user có dữ liệu đầy đủ trong tháng 12 không
            incomplete_users = []
            december_start = date(year, 12, 1)
            december_end = date(year, 12, 31)
            
            for user in active_users:
                # Đếm số ngày có dữ liệu chấm công trong tháng 12
                attendance_count = Attendance.query.filter(
                    Attendance.user_id == user.id,
                    Attendance.date >= december_start,
                    Attendance.date <= december_end
                ).count()
                
                # Kiểm tra nếu thiếu dữ liệu (ít hơn 20 ngày làm việc - có thể điều chỉnh)
                # Tháng 12 thường có khoảng 22-23 ngày làm việc (trừ cuối tuần và lễ)
                if attendance_count < 15:  # Ngưỡng tối thiểu 15 ngày
                    incomplete_users.append({
                        'name': user.name,
                        'employee_id': user.employee_id,
                        'attendance_count': attendance_count
                    })
            
            return {
                'complete': len(incomplete_users) == 0,
                'incomplete_users': incomplete_users,
                'total_users': len(active_users),
                'checked_users': len(active_users) - len(incomplete_users)
            }
    except Exception as e:
        print(f"❌ Lỗi khi kiểm tra dữ liệu tháng 12: {e}")
        return {
            'complete': False,
            'incomplete_users': [],
            'error': str(e)
        }

def reset_yearly_schedule():
    """Reset/xóa dữ liệu lịch cũ khi bắt đầu năm mới"""
    try:
        with app.app_context():
            from database.models import Attendance
            from datetime import date
            
            current_year = datetime.now().year
            previous_year = current_year - 1
            
            # Xóa tất cả dữ liệu chấm công của năm trước
            # Sử dụng strftime cho SQLite hoặc extract cho PostgreSQL/MySQL
            try:
                # Thử dùng extract trước (PostgreSQL/MySQL)
                deleted_count = Attendance.query.filter(
                    db.extract('year', Attendance.date) == previous_year
                ).delete()
            except Exception:
                # Fallback cho SQLite: dùng strftime
                from sqlalchemy import func
                deleted_count = Attendance.query.filter(
                    func.strftime('%Y', Attendance.date) == str(previous_year)
                ).delete()
            
            db.session.commit()
            
            print(f"✅ Đã xóa {deleted_count} bản ghi chấm công của năm {previous_year}")
            
            # Gửi thông báo cho admin
            message = f"🔄 <b>RESET LỊCH HÀNG NĂM</b>\n\n"
            message += f"✅ Đã hoàn tất reset lịch vào ngày 1/1/{current_year}\n"
            message += f"📊 Đã xóa {deleted_count} bản ghi chấm công của năm {previous_year}\n"
            message += f"📅 Hệ thống đã sẵn sàng cho năm mới {current_year}"
            
            send_telegram_message(message)
            
            return True, deleted_count
    except Exception as e:
        print(f"❌ Lỗi khi reset lịch hàng năm: {e}")
        error_msg = f"❌ <b>LỖI RESET LỊCH</b>\n\nĐã xảy ra lỗi khi reset lịch hàng năm: {str(e)}"
        send_telegram_message(error_msg)
        return False, 0

def send_yearly_reset_reminder(days_until_reset):
    """Gửi thông báo nhắc nhở admin trước ngày reset"""
    try:
        current_year = datetime.now().year
        next_year = current_year + 1
        
        # Kiểm tra dữ liệu tháng 12
        check_result = check_december_data_complete(current_year)
        
        message = f"⏰ <b>NHẮC NHỞ RESET LỊCH HÀNG NĂM</b>\n\n"
        
        if days_until_reset == 7:
            message += f"📅 Còn <b>7 ngày</b> nữa đến ngày reset lịch (1/1/{next_year})\n\n"
        elif days_until_reset == 3:
            message += f"📅 Còn <b>3 ngày</b> nữa đến ngày reset lịch (1/1/{next_year})\n\n"
        elif days_until_reset == 1:
            message += f"📅 Còn <b>1 ngày</b> nữa đến ngày reset lịch (1/1/{next_year})\n\n"
            message += f"⚠️ <b>LƯU Ý QUAN TRỌNG:</b>\n"
            message += f"• Hệ thống sẽ tự động xóa tất cả dữ liệu chấm công của năm {current_year}\n"
            message += f"• Vui lòng đảm bảo đã sao lưu dữ liệu quan trọng\n"
            message += f"• Nhắc nhở nhân viên hoàn tất nhập dữ liệu tháng 12\n\n"
        
        # Thông tin về dữ liệu tháng 12
        if check_result.get('complete', False):
            message += f"✅ <b>Dữ liệu tháng 12:</b> Đã đầy đủ ({check_result.get('checked_users', 0)}/{check_result.get('total_users', 0)} nhân viên)\n"
        else:
            incomplete = check_result.get('incomplete_users', [])
            if incomplete:
                message += f"⚠️ <b>Dữ liệu tháng 12:</b> Có {len(incomplete)} nhân viên chưa đầy đủ:\n"
                for user_info in incomplete[:5]:  # Chỉ hiển thị 5 người đầu
                    message += f"   • {user_info['name']} (ID: {user_info['employee_id']}) - {user_info['attendance_count']} ngày\n"
                if len(incomplete) > 5:
                    message += f"   ... và {len(incomplete) - 5} nhân viên khác\n"
            else:
                message += f"⚠️ <b>Dữ liệu tháng 12:</b> Không thể kiểm tra (có lỗi)\n"
        
        message += f"\n📋 <b>Hành động cần thực hiện:</b>\n"
        message += f"• Kiểm tra và đảm bảo nhân viên đã nhập đầy đủ dữ liệu tháng 12\n"
        message += f"• Sao lưu dữ liệu quan trọng trước ngày 1/1/{next_year}\n"
        message += f"• Hệ thống sẽ tự động reset vào 00:00 ngày 1/1/{next_year}"
        
        send_telegram_message(message)
        print(f"📤 Đã gửi thông báo nhắc nhở reset lịch ({days_until_reset} ngày trước)")
        
    except Exception as e:
        print(f"❌ Lỗi khi gửi thông báo nhắc nhở: {e}")

def _yearly_reset_worker():
    """Worker chạy nền để kiểm tra và reset lịch hàng năm"""
    global _last_reset_year, _last_notification_days
    
    while True:
        try:
            now = datetime.now()
            current_year = now.year
            current_month = now.month
            current_day = now.day
            
            # Kiểm tra nếu là ngày 1/1 và chưa reset năm này
            if current_month == 1 and current_day == 1:
                if _last_reset_year != current_year:
                    print(f"🔄 [YEARLY RESET] Phát hiện ngày 1/1/{current_year}, bắt đầu reset lịch...")
                    success, deleted_count = reset_yearly_schedule()
                    if success:
                        _last_reset_year = current_year
                        _last_notification_days.clear()  # Reset danh sách thông báo
                        print(f"✅ [YEARLY RESET] Đã hoàn tất reset lịch năm {current_year}")
                    else:
                        print(f"❌ [YEARLY RESET] Lỗi khi reset lịch năm {current_year}")
            
            # Kiểm tra và gửi thông báo nhắc nhở
            # Tính số ngày còn lại đến 1/1 năm sau
            next_year = current_year + 1
            next_jan_1 = datetime(next_year, 1, 1)
            days_until_reset = (next_jan_1 - now).days
            
            # Gửi thông báo vào các mốc: 7 ngày, 3 ngày, 1 ngày trước
            if days_until_reset in [7, 3, 1]:
                if days_until_reset not in _last_notification_days:
                    send_yearly_reset_reminder(days_until_reset)
                    _last_notification_days.add(days_until_reset)
            
            # Xóa các mốc đã qua khỏi danh sách để có thể gửi lại năm sau
            if days_until_reset < 0:
                _last_notification_days.clear()
            
            # Kiểm tra mỗi ngày một lần (vào lúc 00:00)
            # Tính thời gian đến 00:00 ngày hôm sau
            tomorrow = datetime(now.year, now.month, now.day) + timedelta(days=1)
            seconds_until_midnight = (tomorrow - now).total_seconds()
            
            # Ngủ đến 00:00 ngày hôm sau, nhưng tối thiểu 1 giờ để tránh lỗi
            sleep_seconds = max(3600, int(seconds_until_midnight))
            time.sleep(sleep_seconds)
            
        except Exception as e:
            print(f"⚠️ Lỗi trong yearly reset worker: {e}")
            # Ngủ 1 giờ trước khi thử lại
            time.sleep(3600)

def ensure_yearly_reset_scheduler_started():
    """Khởi chạy thread reset lịch hàng năm một lần duy nhất"""
    global _yearly_reset_scheduler_started
    if _yearly_reset_scheduler_started:
        return
    with _yearly_reset_scheduler_lock:
        if _yearly_reset_scheduler_started:
            return
        try:
            t = threading.Thread(
                target=_yearly_reset_worker,
                daemon=True
            )
            t.start()
            _yearly_reset_scheduler_started = True
            print(f"📅 Yearly reset scheduler started: tự động reset lịch vào ngày 1/1 hàng năm")
        except Exception as e:
            print(f"⚠️ Không thể khởi chạy yearly reset scheduler: {e}")

def start_all_background_services():
    """
    Khởi động tất cả các dịch vụ nền: backup + token keep-alive + yearly reset
    """
    print("🚀 Khởi động các dịch vụ nền...")
    
    # Khởi động backup scheduler
    try:
        ensure_backup_scheduler_started(interval_minutes=60, backup_dir="backups", retention=3, send_to_telegram=True)
    except Exception as e:
        print(f"⚠️ Lỗi khởi động backup scheduler: {e}")
    
    # Khởi động token keep-alive
    try:
        ensure_token_keepalive_started(interval_minutes=30)
    except Exception as e:
        print(f"⚠️ Lỗi khởi động token keep-alive: {e}")
    
    # Khởi động yearly reset scheduler
    try:
        ensure_yearly_reset_scheduler_started()
    except Exception as e:
        print(f"⚠️ Lỗi khởi động yearly reset scheduler: {e}")
    
    # Khởi động license online checker (mặc định 60 giây kiểm tra 1 lần)
    try:
        ensure_license_check_started(interval_seconds=60)
    except Exception as e:
        print(f"[LICENSE] Lỗi khởi động license online checker: {e}")
    
    print("✅ Tất cả dịch vụ nền đã được khởi động!")

# --- Helper function để xử lý định dạng thời gian SA/CH/AM/PM ---
def clean_time_format(time_str):
    """Xử lý định dạng thời gian có SA/CH/AM/PM"""
    if not time_str:
        return '00:00'
    
    # Xử lý đặc biệt cho các trường hợp 12:00
    if '12:00' in time_str:
        # 12:00 SA = 12:00 trưa (PM) - trong tiếng Việt
        if 'SA' in time_str:
            return '12:00'
        # 12:00 CH = 12:00 chiều (PM) - trong tiếng Việt  
        elif 'CH' in time_str:
            return '12:00'
        # 12:00 PM = 12:00 trưa (PM) - chuẩn quốc tế
        elif 'PM' in time_str:
            return '12:00'
        # 12:00 AM = 00:00 nửa đêm (AM) - chuẩn quốc tế
        elif 'AM' in time_str:
            return '00:00'
        else:
            return '12:00'
    else:
        # Xử lý định dạng thời gian có thể có SA/CH/AM/PM
        return time_str.replace('SA', '').replace('CH', '').replace('AM', '').replace('PM', '').strip()

# --- Google Drive API Routes ---
@app.route('/api/google-drive/update-sheet', methods=['POST'])
@login_required
def update_google_sheet():
    """API để cập nhật Google Sheets thay vì mở Chrome"""
    try:
        data = request.get_json()
        
        # Kiểm tra quyền admin hoặc manager
        user = db.session.get(User, session['user_id'])
        if not user or not any(role in ['ADMIN', 'MANAGER'] for role in user.roles.split(',')):
            return jsonify({'error': 'Không có quyền truy cập'}), 403
        
        # Lấy thông tin từ request
        spreadsheet_id = data.get('spreadsheet_id')
        sheet_name = data.get('sheet_name')
        row = data.get('row')
        column = data.get('column')
        new_value = data.get('new_value')
        
        if not all([spreadsheet_id, sheet_name, row, column, new_value is not None]):
            return jsonify({'error': 'Thiếu thông tin bắt buộc'}), 400
        
        # Khởi tạo Google Drive API
        google_api = GoogleDriveAPI()
        
        if not google_api.sheets_service:
            return jsonify({'error': 'Không thể kết nối Google API'}), 500
        
        # Cập nhật giá trị trong sheet
        success = google_api.update_sheet_value(
            spreadsheet_id=spreadsheet_id,
            sheet_name=sheet_name,
            row=int(row),
            column=column,
            new_value=str(new_value)
        )
        
        if success:
            return jsonify({
                'message': 'Cập nhật Google Sheets thành công',
                'spreadsheet_id': spreadsheet_id,
                'sheet_name': sheet_name,
                'cell': f"{column}{row}",
                'new_value': new_value
            }), 200
        else:
            return jsonify({'error': 'Không thể cập nhật Google Sheets'}), 500
            
    except Exception as e:
        print(f"Lỗi khi cập nhật Google Sheets: {e}")
        return jsonify({'error': 'Lỗi hệ thống'}), 500

@app.route('/api/google-drive/find-bud-timesheet', methods=['GET'])
@login_required
def find_bud_timesheet():
    """API để tìm file Bud_TimeSheet-202510"""
    try:
        # Kiểm tra quyền admin hoặc manager
        user = db.session.get(User, session['user_id'])
        if not user or not any(role in ['ADMIN', 'MANAGER'] for role in user.roles.split(',')):
            return jsonify({'error': 'Không có quyền truy cập'}), 403
        
        # Khởi tạo Google Drive API
        google_api = GoogleDriveAPI()
        
        if not google_api.drive_service:
            return jsonify({'error': 'Không thể kết nối Google API'}), 500
        
        # Tìm file Bud_TimeSheet-202510
        target_file = google_api.find_and_read_bud_timesheet(GOOGLE_DRIVE_FOLDER_ID)
        
        if target_file:
            return jsonify({
                'message': 'Tìm thấy file Bud_TimeSheet-202510',
                'file_id': target_file['id'],
                'file_name': target_file['name'],
                'file_type': google_api._get_file_type(target_file['mimeType']),
                'web_view_link': target_file.get('webViewLink', 'N/A')
            }), 200
        else:
            return jsonify({'error': 'Không tìm thấy file Bud_TimeSheet-202510'}), 404
            
    except Exception as e:
        print(f"Lỗi khi tìm file Bud_TimeSheet: {e}")
        return jsonify({'error': 'Lỗi hệ thống'}), 500

@app.route('/api/google-drive/list-timesheets', methods=['GET'])
@login_required
def list_all_timesheets():
    """API để lấy danh sách tất cả file timesheet"""
    try:
        # Kiểm tra quyền admin hoặc manager
        user = db.session.get(User, session['user_id'])
        if not user or not any(role in ['ADMIN', 'MANAGER'] for role in user.roles.split(',')):
            return jsonify({'error': 'Không có quyền truy cập'}), 403
        
        # Khởi tạo Google Drive API
        google_api = GoogleDriveAPI()
        
        if not google_api.drive_service:
            return jsonify({'error': 'Không thể kết nối Google API'}), 500
        
        # Lấy danh sách tất cả file timesheet
        timesheets = google_api.list_all_timesheets(GOOGLE_DRIVE_FOLDER_ID)
        
        if timesheets:
            return jsonify({
                'message': f'Tìm thấy {len(timesheets)} file timesheet',
                'count': len(timesheets),
                'files': [
                    {
                        'id': file['id'],
                        'name': file['name'],
                        'type': google_api._get_file_type(file['mimeType']),
                        'size': file.get('size', 'N/A'),
                        'modified_time': file['modifiedTime'],
                        'web_view_link': file.get('webViewLink', 'N/A')
                    }
                    for file in timesheets
                ]
            }), 200
        else:
            return jsonify({'error': 'Không tìm thấy file timesheet nào'}), 404
            
    except Exception as e:
        print(f"Lỗi khi lấy danh sách file timesheet: {e}")
        return jsonify({'error': 'Lỗi hệ thống'}), 500

@app.route('/api/google-drive/department-mapping', methods=['GET'])
@login_required
def get_department_mapping():
    """API để lấy mapping phòng ban với file timesheet"""
    try:
        # Kiểm tra quyền admin hoặc manager
        user = db.session.get(User, session['user_id'])
        if not user or not any(role in ['ADMIN', 'MANAGER'] for role in user.roles.split(',')):
            return jsonify({'error': 'Không có quyền truy cập'}), 403
        
        # Khởi tạo Google Drive API
        google_api = GoogleDriveAPI()
        
        # Lấy mapping phòng ban
        mappings = google_api.get_all_department_mappings()
        
        return jsonify({
            'message': 'Mapping phòng ban với file timesheet',
            'mappings': mappings,
            'note': 'York và Como đều sử dụng Chirashi_TimeSheet'
        }), 200
            
    except Exception as e:
        print(f"Lỗi khi lấy mapping phòng ban: {e}")
        return jsonify({'error': 'Lỗi hệ thống'}), 500

# --- Helper function để xử lý định dạng thời gian SA/CH/AM/PM ---
def clean_time_format(time_str):
    """Xử lý định dạng thời gian có SA/CH/AM/PM"""
    if not time_str:
        return '00:00'
    
    # Xử lý đặc biệt cho các trường hợp 12:00
    if '12:00' in time_str:
        # 12:00 SA = 12:00 trưa (PM) - trong tiếng Việt
        if 'SA' in time_str:
            return '12:00'
        # 12:00 CH = 12:00 chiều (PM) - trong tiếng Việt  
        elif 'CH' in time_str:
            return '12:00'
        # 12:00 PM = 12:00 trưa (PM) - chuẩn quốc tế
        elif 'PM' in time_str:
            return '12:00'
        # 12:00 AM = 00:00 nửa đêm (AM) - chuẩn quốc tế
        elif 'AM' in time_str:
            return '00:00'
        else:
            return '12:00'
    else:
        # Xử lý định dạng thời gian có thể có SA/CH/AM/PM
        return time_str.replace('SA', '').replace('CH', '').replace('AM', '').replace('PM', '').strip()

# --- Helper tính đơn vị nghỉ theo ca ---
def _compute_leave_units_generic(from_date_dt: datetime, from_time_str: str, to_date_dt: datetime, to_time_str: str) -> float:
    try:
        # Xử lý định dạng thời gian có SA/CH/AM/PM
        clean_from_time = clean_time_format(from_time_str)
        clean_to_time = clean_time_format(to_time_str)
        
        start_dt = datetime.combine(from_date_dt.date(), datetime.strptime(clean_from_time, '%H:%M').time())
        end_dt = datetime.combine(to_date_dt.date(), datetime.strptime(clean_to_time, '%H:%M').time())
    except Exception:
        # Fallback: tính theo số ngày lịch
        return max(0.0, (to_date_dt - from_date_dt).days + 1)
    if end_dt < start_dt:
        return 0.0
    workday_hours = 8.0
    half_hours = 4.0
    if start_dt.date() == end_dt.date():
        hours = (end_dt - start_dt).total_seconds() / 3600.0
        if hours <= 0:
            return 0.0
        # Logic tính theo thời gian làm việc thực tế (trừ giờ nghỉ)
        # 1 ngày = 8 tiếng làm việc, 0.5 ngày = 4 tiếng làm việc
        # Làm tròn đến 0.5
        days = round((hours / workday_hours) * 2) / 2.0
        return days
    # nhiều ngày
    end_of_first = datetime.combine(start_dt.date(), time(23,59,59))
    first_hours = (end_of_first - start_dt).total_seconds() / 3600.0
    first_unit = round((first_hours / workday_hours) * 2) / 2.0
    
    start_of_last = datetime.combine(end_dt.date(), time(0,0,0))
    last_hours = (end_dt - start_of_last).total_seconds() / 3600.0
    last_unit = round((last_hours / workday_hours) * 2) / 2.0
    
    middle_days = (to_date_dt.date() - from_date_dt.date()).days - 1
    middle_units = max(0, middle_days) * 1.0
    
    total_units = first_unit + middle_units + last_unit
    return round(total_units * 2) / 2.0

# Load configuration
config_name = os.environ.get('FLASK_CONFIG') or 'default'
app.config.from_object(config[config_name])

# Initialize CSRF protection
csrf = CSRFProtect(app)

# CSRF protection is enabled for all routes
# No need to disable in development

# Expose csrf_token() helper to Jinja templates
@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=generate_csrf)

# Time formatting helper (convert UTC -> local, e.g., UTC+7)
@app.context_processor
def inject_format_helpers():
    def format_local(dt, hours_offset=7):
        try:
            if not dt:
                return ''
            return (dt + timedelta(hours=hours_offset)).strftime('%d/%m/%Y %H:%M')
        except Exception:
            return dt.strftime('%d/%m/%Y %H:%M') if dt else ''
    return dict(format_local=format_local)

# Initialize database
db.init_app(app)
migrate = Migrate(app, db)

# Initialize signature manager
signature_manager.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Handler để API routes trả về JSON thay vì redirect HTML khi chưa đăng nhập
@login_manager.unauthorized_handler
def unauthorized():
    # Nếu request là API call (có header Accept: application/json hoặc path bắt đầu bằng /api/)
    if request.path.startswith('/api/') or request.headers.get('Accept', '').startswith('application/json'):
        return jsonify({'error': 'Chưa đăng nhập', 'day_type': 'normal', 'reason': 'Ngày thường (fallback)'}), 401
    # Ngược lại, redirect về login như bình thường
    return redirect(url_for('login'))

# Import rate limiting from utils
from utils.decorators import rate_limit

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ==========================
# License / Activation utils
# ==========================
def get_activation_record():
    """Lấy (hoặc tạo mặc định) bản ghi kích hoạt duy nhất."""
    activation = Activation.query.get(1)
    if not activation:
        activation = Activation(id=1, is_activated=False)
        db.session.add(activation)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
    return activation


def is_app_activated():
    """Kiểm tra ứng dụng đã được kích hoạt hay chưa."""
    activation = get_activation_record()
    return bool(activation and activation.is_activated)

# ====== LICENSE ACCESS CONTROL ======
@app.before_request
def check_license_before_request():
    """
    Chặn tất cả request khi license hết hạn.
    Cho phép truy cập static files và trang activate.
    """
    # Cho phép truy cập static files và trang activate
    if request.endpoint in ('static', 'activate') or request.path.startswith('/static/'):
        return None
    
    # Nếu license không hợp lệ, chặn tất cả request khác
    if not _license_is_valid:
        from flask import render_template_string
        contact_msg = (
            "License không hợp lệ hoặc đã hết hạn.\n\n"
            "Vui lòng liên hệ ADMIN để gia hạn:\n"
            "Nguyễn Công Đạt - 0375097105.\n\n"
            "Hệ thống sẽ tự động kiểm tra lại license sau 60 giây."
        )
        
        # Trả về trang HTML thông báo license hết hạn
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>License Hết Hạn</title>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    height: 100vh;
                    margin: 0;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                }
                .container {
                    background: white;
                    padding: 40px;
                    border-radius: 10px;
                    box-shadow: 0 10px 30px rgba(0,0,0,0.3);
                    max-width: 500px;
                    text-align: center;
                }
                h1 {
                    color: #e74c3c;
                    margin-bottom: 20px;
                }
                p {
                    color: #333;
                    line-height: 1.6;
                    white-space: pre-line;
                }
                .contact {
                    margin-top: 20px;
                    padding: 15px;
                    background: #f8f9fa;
                    border-radius: 5px;
                }
                .refresh-info {
                    margin-top: 20px;
                    color: #666;
                    font-size: 14px;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>⚠️ License Hết Hạn</h1>
                <p>{{ message }}</p>
                <div class="refresh-info">
                    Hệ thống đang tự động kiểm tra lại license mỗi 60 giây.<br>
                    Vui lòng đợi trong giây lát...
                </div>
            </div>
            <script>
                // Tự động reload sau 65 giây để kiểm tra lại license
                setTimeout(function() {
                    location.reload();
                }, 65000);
            </script>
        </body>
        </html>
        """
        return render_template_string(html_template, message=contact_msg), 403
    
    return None

# Routes
@app.route('/')
def index():
    # Nếu app chưa được kích hoạt thì bắt buộc vào trang kích hoạt
    if not is_app_activated():
        return redirect(url_for('activate'))
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
@rate_limit(max_requests=100, window_seconds=300)
def login():
    # Nếu app chưa kích hoạt, không cho login mà chuyển sang trang kích hoạt
    if not is_app_activated():
        return redirect(url_for('activate'))
    remembered_username = request.cookies.get('remembered_username', '')
    skip_auto_login = request.args.get('logout') == '1'

    if request.method == 'GET':
        remember_token = request.cookies.get('remember_token')
        if remember_token and not skip_auto_login:
            user = User.query.filter_by(remember_token=remember_token).first()
            if user and user.remember_token_expires and user.remember_token_expires > datetime.now():
                # Auto login with remember token
                session['user_id'] = user.id
                session['name'] = user.name
                session['employee_id'] = user.employee_id
                session['roles'] = user.roles.split(',')
                # Ưu tiên EMPLOYEE nếu user có vai trò này
                user_roles = user.roles.split(',')
                if 'EMPLOYEE' in user_roles:
                    session['current_role'] = 'EMPLOYEE'
                else:
                    session['current_role'] = user_roles[0]
                session['last_activity'] = datetime.now().isoformat()
                
                log_audit_action(
                    user_id=user.id,
                    action='AUTO_LOGIN',
                    table_name='users',
                    record_id=user.id,
                    new_values={'auto_login_time': datetime.now().isoformat()}
                )
                
                flash('Đăng nhập tự động thành công!', 'success')
                return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        employee_id_str = request.form.get('username', '').strip()
        password = request.form.get('password')
        remember = request.form.get('remember') == 'on'
        
        # Validate input
        if not employee_id_str or not password:
            flash('Vui lòng nhập đầy đủ mã nhân viên và mật khẩu!', 'error')
            return render_template(
                'login.html',
                messages=get_flashed_messages(with_categories=False),
                remembered_username=employee_id_str or remembered_username,
                remember_checked=remember or bool(remembered_username)
            )
        # Validate employee id format (digits only)
        try:
            employee_id = validate_employee_id(employee_id_str)
        except ValidationError as ve:
            flash(ve.message or 'Mã nhân viên không hợp lệ!', 'error')
            return render_template(
                'login.html',
                messages=get_flashed_messages(with_categories=False),
                remembered_username=employee_id_str or remembered_username,
                remember_checked=remember or bool(remembered_username)
            )
        
        if not validate_input_sanitize(password):
            flash('Mật khẩu không hợp lệ!', 'error')
            return render_template(
                'login.html',
                messages=get_flashed_messages(with_categories=False),
                remembered_username=employee_id_str or remembered_username,
                remember_checked=remember or bool(remembered_username)
            )
        
        try:
            user = User.query.filter_by(employee_id=employee_id).first()
            
            if user and user.check_password(password):
                # Check security before login
                is_allowed, message = security_manager.check_login_attempts(employee_id)
                if not is_allowed:
                    flash(message, 'error')
                    return render_template('login.html', messages=get_flashed_messages(with_categories=False))
                
                session['user_id'] = user.id
                session['name'] = user.name
                session['employee_id'] = user.employee_id
                session['roles'] = user.roles.split(',')
                # Ưu tiên EMPLOYEE nếu user có vai trò này
                user_roles = user.roles.split(',')
                logger.info(f"User login successful", user_id=user.id, employee_id=employee_id, roles=user_roles)
                if 'EMPLOYEE' in user_roles:
                    session['current_role'] = 'EMPLOYEE'
                else:
                    session['current_role'] = user_roles[0]
                session['last_activity'] = datetime.now().isoformat()
                
                # Clear failed attempts on successful login
                security_manager.clear_failed_attempts(employee_id)
                
                response = redirect(url_for('dashboard'))
                
                log_audit_action(
                    user_id=user.id,
                    action='LOGIN',
                    table_name='users',
                    record_id=user.id,
                    new_values={'login_time': datetime.now().isoformat()}
                )
                
                if remember:
                    # Generate secure remember token
                    remember_token = secrets.token_urlsafe(32)
                    user.remember_token = remember_token
                    user.remember_token_expires = datetime.now() + timedelta(days=30)
                    db.session.commit()
                    response.set_cookie('remember_token', remember_token, max_age=30*24*60*60, httponly=True, secure=app.config.get('SESSION_COOKIE_SECURE', False))
                    response.set_cookie('remembered_username', employee_id_str, max_age=30*24*60*60)
                else:
                    # Clear remember token if not checked
                    if user.remember_token:
                        user.remember_token = None
                        user.remember_token_expires = None
                        db.session.commit()
                    response.delete_cookie('remember_token')
                    response.delete_cookie('remembered_username')
                
                flash('Đăng nhập thành công!', 'success')
                return response
            
            # Record failed login attempt
            security_manager.record_failed_login(employee_id)
            flash('Mã nhân viên hoặc mật khẩu không đúng!', 'error')
        except Exception as e:
            security_logger.error("Login system error", 
                                error_type='LoginSystemError',
                                employee_id=employee_id,
                                error_details=str(e))
            flash('Đã xảy ra lỗi khi đăng nhập!', 'error')
    
    form_username = request.form.get('username', '').strip() if request.method == 'POST' else ''
    username_prefill = form_username or remembered_username
    remember_prefill = (request.form.get('remember') == 'on') if request.method == 'POST' else bool(username_prefill)

    return render_template(
        'login.html',
        messages=get_flashed_messages(with_categories=False),
        remembered_username=username_prefill,
        remember_checked=remember_prefill
    )

@app.route('/logout')
def logout():
    forget_device = request.args.get('forget') == '1'
    # Log logout if user was logged in
    if 'user_id' in session:
        log_audit_action(
            user_id=session['user_id'],
            action='LOGOUT',
            table_name='users',
            record_id=session['user_id'],
            new_values={'logout_time': datetime.now().isoformat()}
        )
        
        user = db.session.get(User, session['user_id'])
        if user and forget_device:
            user.remember_token = None
            user.remember_token_expires = None
            db.session.commit()
    
    session.clear()
    redirect_url = url_for('login') if forget_device else url_for('login', logout=1)
    response = redirect(redirect_url)
    if forget_device:
        response.delete_cookie('remember_token')
        response.delete_cookie('remembered_username')
    return response

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Process any pending DB updates from async email threads
    from utils.email_utils import process_db_updates
    process_db_updates()
    
    user = db.session.get(User, session['user_id'])
    if not user:
        session.clear()
        flash('Phiên đăng nhập không hợp lệ!', 'error')
        return redirect(url_for('login'))
    
    # Kiểm tra user có active không
    if not user.is_active:
        session.clear()
        flash('Tài khoản đã bị khóa!', 'error')
        return redirect(url_for('login'))
    
    # Kiểm tra session timeout
    if check_session_timeout():
        flash('Phiên đăng nhập đã hết hạn!', 'error')
        return redirect(url_for('login'))
    
    # Cập nhật thời gian hoạt động cuối
    update_session_activity()
    
    # Xử lý tham số role từ query string
    role_param = request.args.get('role')
    if role_param and role_param in user.roles.split(','):
        session['current_role'] = role_param
        # Set current_role from query param
    
    # Đảm bảo session có đầy đủ thông tin
    if 'roles' not in session:
        session['roles'] = user.roles.split(',')
    
    # Chỉ set current_role nếu chưa có hoặc không hợp lệ
    if 'current_role' not in session:
        # Ưu tiên EMPLOYEE nếu user có vai trò này (chỉ khi đăng nhập lần đầu)
        user_roles = user.roles.split(',')
        if 'EMPLOYEE' in user_roles:
            session['current_role'] = 'EMPLOYEE'
        else:
            session['current_role'] = user_roles[0]
        # print(f"DEBUG DASHBOARD: Set current_role to {session['current_role']} (no current_role in session)")
    elif session['current_role'] not in user.roles.split(','):
        # Ưu tiên EMPLOYEE nếu user có vai trò này (chỉ khi current_role không hợp lệ)
        user_roles = user.roles.split(',')
        if 'EMPLOYEE' in user_roles:
            session['current_role'] = 'EMPLOYEE'
        else:
            session['current_role'] = user_roles[0]
        # print(f"DEBUG DASHBOARD: Reset current_role to {session['current_role']} (not in user roles)")
    else:
        # print(f"DEBUG DASHBOARD: Keep current_role as {session['current_role']} (valid role)")

        if 'name' not in session:
            session['name'] = user.name
        if 'employee_id' not in session:
            session['employee_id'] = user.employee_id
    # Final current_role setup complete
    
    # Kiểm tra xem user đã có chữ ký cá nhân chưa
    has_signature = bool(user.personal_signature)

    # Nếu là ADMIN, kiểm tra cảnh báo ngày lễ để hiển thị trên dashboard
    if 'ADMIN' in (user.roles or '').split(','):
        try:
            from datetime import date as _date_mod
            current_year = datetime.now().year
            # Có ít nhất 1 ngày lễ trong năm hiện tại chưa?
            year_holidays = Holiday.query.filter(
                Holiday.date >= _date_mod(current_year, 1, 1),
                Holiday.date <= _date_mod(current_year, 12, 31)
            ).count()
            if year_holidays == 0:
                flash('⚠️ Hiện chưa cấu hình ngày lễ nào cho năm hiện tại. Vui lòng vào "Quản lý ngày lễ" để thêm.', 'warning')
        except Exception:
            # Không để lỗi phụ làm vỡ dashboard
            pass
    
    # Lấy danh sách đơn nghỉ phép của người dùng để hiển thị badge bị từ chối trên sidebar
    try:
        leave_requests = (LeaveRequest.query
                          .filter(LeaveRequest.user_id == user.id)
                          .order_by(LeaveRequest.created_at.desc())
                          .limit(100)
                          .all())
    except Exception:
        leave_requests = []

    # Lấy trạng thái cảnh báo LICENSE gần nhất để render sẵn vào HTML,
    # giúp banner hiển thị nhanh hơn mà không phải chờ request fetch() đầu tiên.
    try:
        with _license_warning_lock:
            license_warning_state = dict(_license_warning_state)
    except Exception:
        # Nếu có lỗi, fallback về state rỗng (không hiển thị cảnh báo)
        license_warning_state = {'active': False, 'payload': None, 'updated_at': None}

    return render_template(
        'dashboard.html',
        user=user,
        has_signature=has_signature,
        leave_requests=leave_requests,
        license_warning_state_json=json.dumps(license_warning_state, ensure_ascii=False)
    )

@app.route('/api/attendance', methods=['POST'])
@rate_limit(max_requests=500, window_seconds=60)
def record_attendance():
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    data = request.get_json()
    # print('DEBUG raw:', data)
    # print('DEBUG signature POST:', data.get('signature'))  # Thêm log signature
    # Validate input
    date = validate_date(data.get('date'))
    check_in = validate_time(data.get('check_in'))
    check_out = validate_time(data.get('check_out'))
    note = validate_note(data.get('note', ''))
    
    # Khai báo holiday_type trước khi sử dụng
    is_holiday = bool(data.get('is_holiday', False))
    holiday_type = validate_holiday_type(data.get('holiday_type'))
    
    # Chỉ chấp nhận HH:MM
    # Lễ Việt Nam không đi làm: break_time = 0:00, ngược lại = 1:00
    if holiday_type == 'vietnamese_holiday' and (not check_in or not check_out):
        raw_break_time = data.get('break_time', '00:00') or '00:00'
    else:
        raw_break_time = data.get('break_time', '01:00') or '01:00'
    if not (isinstance(raw_break_time, str) and re.match(r'^\d{1,2}:[0-5]\d$', raw_break_time)):
        return jsonify({'error': 'Thời gian nghỉ phải ở định dạng HH:MM'}), 400
    comp_time_regular_raw = data.get('comp_time_regular', '00:00') or '00:00'
    comp_time_overtime_raw = data.get('comp_time_overtime', '00:00') or '00:00'
    comp_time_ot_before_22_raw = data.get('comp_time_ot_before_22', '00:00') or '00:00'
    comp_time_ot_after_22_raw = data.get('comp_time_ot_after_22', '00:00') or '00:00'
    overtime_comp_time_raw = data.get('overtime_comp_time', '00:00') or '00:00'
    for fld, val in [('comp_time_regular', comp_time_regular_raw), ('comp_time_overtime', comp_time_overtime_raw), ('comp_time_ot_before_22', comp_time_ot_before_22_raw), ('comp_time_ot_after_22', comp_time_ot_after_22_raw), ('overtime_comp_time', overtime_comp_time_raw)]:
        if not (isinstance(val, str) and re.match(r'^\d{1,2}:[0-5]\d$', val)):
            return jsonify({'error': f'{fld} phải ở định dạng HH:MM'}), 400
    # Quy đổi HH:MM → giờ (float) tương thích trường hiện tại, nhưng mọi tính toán dùng seconds
    def hhmm_to_hours(hhmm):
        """Chuyển đổi an toàn HH:MM sang giờ thập phân"""
        if not hhmm or hhmm == "":
            return 0.0
        try:
            if isinstance(hhmm, (int, float)):
                return float(hhmm)
            if isinstance(hhmm, str) and ":" in hhmm:
                hh, mm = hhmm.split(':')
                return int(hh) + int(mm)/60
            else:
                # Thử chuyển đổi string số
                return float(hhmm)
        except (ValueError, TypeError) as e:
            # print(f"Warning: Failed to convert {repr(hhmm)} to hours: {e}")
            return 0.0
    break_time = hhmm_to_hours(raw_break_time)
    comp_time_regular = hhmm_to_hours(comp_time_regular_raw)
    comp_time_overtime = hhmm_to_hours(comp_time_overtime_raw)
    comp_time_ot_before_22 = hhmm_to_hours(comp_time_ot_before_22_raw)
    comp_time_ot_after_22 = hhmm_to_hours(comp_time_ot_after_22_raw)
    overtime_comp_time = hhmm_to_hours(overtime_comp_time_raw)
    shift_code = data.get('shift_code')
    shift_start = validate_time(data.get('shift_start'))
    shift_end = validate_time(data.get('shift_end'))
    next_day_checkout = bool(data.get('next_day_checkout', False))  # Flag cho tăng ca qua ngày mới
    # print('DEBUG validated:', 'shift_code:', shift_code, 'shift_start:', shift_start, 'shift_end:', shift_end)
    if not date:
        return jsonify({'error': 'Vui lòng chọn ngày chấm công hợp lệ'}), 400
    if not holiday_type:
        return jsonify({'error': 'Vui lòng chọn loại ngày hợp lệ'}), 400
    # Cho phép không nhập giờ vào/ra cho lễ Việt Nam (nhân viên được 8h mặc định)
    if holiday_type != 'vietnamese_holiday' and (not check_in or not check_out):
        return jsonify({'error': 'Vui lòng nhập đầy đủ giờ vào và giờ ra hợp lệ'}), 400
    if break_time is None:
        return jsonify({'error': 'Thời gian nghỉ không hợp lệ!'}), 400
    if comp_time_regular is None:
        return jsonify({'error': 'Giờ đối ứng trong ca không hợp lệ!'}), 400
    if comp_time_overtime is None:
        return jsonify({'error': 'Giờ đối ứng tăng ca không hợp lệ!'}), 400
    if comp_time_ot_before_22 is None or comp_time_ot_after_22 is None:
        return jsonify({'error': 'Giờ đối ứng tăng ca theo mốc (trước/sau 22h) không hợp lệ!'}), 400
    
    # Validation: Kiểm tra xem có tăng ca hay không trước khi cho phép đối ứng tăng ca
    is_valid, error_message = validate_overtime_comp_time(
        check_in, check_out, shift_start, shift_end, break_time, 
        comp_time_regular, comp_time_overtime, comp_time_ot_before_22, comp_time_ot_after_22, date, data.get('next_day_checkout', False), holiday_type, shift_code
    )
    if not is_valid:
        return jsonify({'error': error_message}), 400
    
    # Lễ Việt Nam không đi làm: không cần shift_code, shift_start, shift_end
    if holiday_type != 'vietnamese_holiday' and (not shift_code or not shift_start or not shift_end):
        return jsonify({'error': 'Vui lòng chọn ca làm việc hợp lệ!'}), 400
    # Tối ưu: Lấy user và existing_attendance trong 1 query
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    required_hours = get_required_daily_hours(user, date, holiday_type, shift_code)
    
    # Kiểm tra existing attendance với timeout để tránh deadlock
    try:
        existing_attendance = Attendance.query.filter_by(user_id=user.id, date=date).first()
    except Exception as e:
        # print(f"Database query error: {e}")
        return jsonify({'error': 'Lỗi truy vấn database, vui lòng thử lại'}), 500
    if existing_attendance:
        if existing_attendance.status != 'rejected':
            return jsonify({'error': 'Bạn đã chấm công cho ngày này rồi, không thể chấm công 2 lần trong 1 ngày.'}), 400
        else:
            db.session.delete(existing_attendance)
            db.session.commit()
    if date > datetime.now().date():
        return jsonify({'error': 'Không thể chấm công cho ngày trong tương lai!'}), 400
    # Tự động lấy chữ ký từ database thay vì yêu cầu user ký
    signature = data.get('signature', '')
    
    # Lấy chữ ký từ database theo thứ tự ưu tiên (với timeout)
    try:
        auto_signature = signature_manager.get_signature_from_database(user.id, 'EMPLOYEE')
    except Exception as e:
        # print(f"Signature query error: {e}")
        auto_signature = None  # Fallback nếu có lỗi
    signature_info = {
        'has_signature': False,
        'signature_type': 'none',
        'message': ''
    }
    
    if auto_signature:
        signature = auto_signature
        signature_info = {
            'has_signature': True,
            'signature_type': 'database',
            'message': f'Đã sử dụng chữ ký có sẵn từ database'
        }
        # print(f"✅ AUTO SIGNATURE: User {user.name} using signature from database")
    else:
        signature_info = {
            'has_signature': False,
            'signature_type': 'none',
            'message': 'Không có chữ ký trong database, sẽ sử dụng chữ ký mặc định'
        }
        # print(f"⚠️ NO AUTO SIGNATURE: User {user.name} has no signature in database")
    
    # Xử lý đặc biệt cho lễ Việt Nam không đi làm
    if holiday_type == 'vietnamese_holiday' and (not check_in or not check_out):
        # print(f"DEBUG: Creating Vietnamese holiday attendance without check-in/out")
        # Lễ Việt Nam không đi làm: set giá trị mặc định
        attendance = Attendance(
            user_id=user.id,
            date=date,
            break_time=0.0,  # Không có thời gian nghỉ khi không đi làm
            comp_time_regular_minutes=0,
            comp_time_overtime_minutes=0,
            comp_time_ot_before_22_minutes=0,
            comp_time_ot_after_22_minutes=0,
            overtime_comp_time_minutes=0,
            is_holiday=is_holiday,
            holiday_type=holiday_type,
            status='pending',
            overtime_before_22="0:00",
            overtime_after_22="0:00",
            shift_code='5',  # Ca 5 (Ca tự do) cho lễ Việt Nam
            signature=signature,
            check_in=None,  # Không có giờ vào
            check_out=None,  # Không có giờ ra
            shift_start=None,  # Không có giờ bắt đầu ca
            shift_end=None,  # Không có giờ kết thúc ca
            total_work_hours=required_hours,  # Tự động tính đủ công theo chính sách (mặc định 8h, ưu tiên 7h nếu áp dụng)
            required_hours=required_hours
        )
    else:
        # Logic bình thường cho các trường hợp khác
        # Chuyển đổi giờ sang phút cho các cột minutes mới
        def hours_to_minutes(hours):
            return int(round(hours * 60)) if hours else 0
        
        attendance = Attendance(
            user_id=user.id,
            date=date,
            break_time=break_time,
            comp_time_regular_minutes=hours_to_minutes(comp_time_regular),
            comp_time_overtime_minutes=hours_to_minutes(comp_time_overtime),
            comp_time_ot_before_22_minutes=hours_to_minutes(comp_time_ot_before_22),
            comp_time_ot_after_22_minutes=hours_to_minutes(comp_time_ot_after_22),
            overtime_comp_time_minutes=hours_to_minutes(overtime_comp_time),
            is_holiday=is_holiday,
            holiday_type=holiday_type,
            status='pending',
            overtime_before_22="0:00",
            overtime_after_22="0:00",
            shift_code=shift_code,
            signature=signature,
            required_hours=required_hours
        )
    
    # Nếu user có vai trò cao hơn, lưu chữ ký vào field tương ứng
    if 'TEAM_LEADER' in user.roles.split(','):
        attendance.team_leader_signature = signature
    if 'MANAGER' in user.roles.split(','):
        attendance.manager_signature = signature
    db.session.add(attendance)
    
    # Chỉ set check_in/check_out khi có giờ vào/ra (không áp dụng cho lễ Việt Nam không đi làm)
    if check_in and check_out:
        attendance.check_in = datetime.combine(date, check_in)
    
    # Xử lý giờ ra - nếu là tăng ca qua ngày mới thì cộng thêm 1 ngày
    if next_day_checkout:
        # Bật qua đêm: set check_out sang ngày hôm sau, cho phép cả trường hợp check_out_time > check_in_time
        # Kiểm tra thời gian làm việc có hợp lý không (tối thiểu 1 giờ)
        work_duration = (datetime.combine(date + timedelta(days=1), check_out) - datetime.combine(date, check_in)).total_seconds() / 3600
        if work_duration < 1.0:
            return jsonify({'error': 'Thời gian làm việc quá ngắn. Vui lòng kiểm tra lại giờ vào/ra.'}), 400
        attendance.check_out = datetime.combine(date + timedelta(days=1), check_out)
        # print(f"DEBUG: Tăng ca qua ngày mới - check_out: {attendance.check_out}")
    else:
        attendance.check_out = datetime.combine(date, check_out)
    
    attendance.shift_start = shift_start
    attendance.shift_end = shift_end
    
    attendance.note = note
    # Gán user object để update_work_hours() có thể check chính sách mẹ <12 tháng
    attendance.user = user
    # Chỉ gọi update_work_hours() khi có giờ vào/ra (trường hợp lễ Việt Nam không đi làm đã set total_work_hours=8.0)
    if check_in and check_out:
        attendance.update_work_hours()
    try:
        logger.info("Attempting to commit attendance record", 
                   user_id=user.id, date=date.isoformat(), holiday_type=holiday_type)
        
        db.session.commit()
        
        logger.info("Successfully committed attendance record", 
                   attendance_id=attendance.id, user_id=user.id)
        
        audit_logger.audit_action(
            action='CREATE_ATTENDANCE',
            table_name='attendances',
            record_id=attendance.id,
            new_values={
                'date': attendance.date.isoformat(),
                'check_in': attendance.check_in.isoformat() if attendance.check_in else None,
                'check_out': attendance.check_out.isoformat() if attendance.check_out else None,
                'status': attendance.status
            }
        )
        
        return jsonify({
            'message': 'Chấm công thành công',
            'work_hours': attendance.total_work_hours,
            'overtime_before_22': attendance.overtime_before_22,
            'overtime_after_22': attendance.overtime_after_22,
            'required_hours': required_hours,
            'signature_info': signature_info
        })
        
    except SQLAlchemyError as e:
        database_logger.error("Database error during attendance commit", 
                             error_type='SQLAlchemyError', 
                             user_id=user.id, 
                             error_details=str(e))
        db.session.rollback()
        return jsonify({'error': 'Lỗi cơ sở dữ liệu khi lưu chấm công'}), 500
        
    except ValidationError as e:
        logger.warning("Validation error during attendance creation", 
                      error_type='ValidationError',
                      user_id=user.id,
                      validation_error=e.message)
        return jsonify({'error': e.message}), 400
        
    except Exception as e:
        logger.critical("Unexpected error during attendance creation", 
                       error_type='UnexpectedError',
                       user_id=user.id,
                       error_details=str(e))
        db.session.rollback()
        return jsonify({'error': 'Lỗi hệ thống không xác định'}), 500

@app.route('/api/log-error', methods=['POST'])
def log_frontend_error():
    """Endpoint để log lỗi từ frontend"""
    try:
        data = request.get_json()
        api_logger.error(
            "Frontend error reported",
            error_type='FrontendError',
            frontend_error_type=data.get('type'),
            frontend_message=data.get('message'),
            url=data.get('url'),
            user_agent=data.get('userAgent'),
            timestamp=data.get('timestamp')
        )
        return jsonify({'status': 'logged'}), 200
    except Exception as e:
        logger.error("Failed to log frontend error", error_details=str(e))
        return jsonify({'status': 'error'}), 500

@app.route('/api/attendance/history')
def get_attendance_history():
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    try:
        user = db.session.get(User, session['user_id'])
        if not user:
            return jsonify({'error': 'Không tìm thấy người dùng'}), 404
        current_role = session.get('current_role', user.roles.split(',')[0])
        if request.args.get('all') == '1':
            if current_role != 'ADMIN':
                return jsonify({'error': 'Chỉ quản trị viên mới có thể xem lịch sử chấm công toàn bộ'}), 403
            if not has_role(session['user_id'], 'ADMIN'):
                return jsonify({'error': 'Bạn không có quyền truy cập dữ liệu toàn bộ'}), 403
            page = validate_int(request.args.get('page', 1), min_val=1)
            per_page = validate_int(request.args.get('per_page', 10), min_val=1, max_val=100)
            search = validate_input_sanitize(request.args.get('search', '').strip())
            department = validate_input_sanitize(request.args.get('department', '').strip())
            date_from = validate_date(request.args.get('date_from', '').strip()) if request.args.get('date_from') else None
            date_to = validate_date(request.args.get('date_to', '').strip()) if request.args.get('date_to') else None
            
            if page is None or per_page is None:
                return jsonify({'error': 'Tham số phân trang không hợp lệ'}), 400
                
            # Use optimized query for fastest performance
            from utils.query_optimizer import optimize_attendance_history_query
            attendances, total = optimize_attendance_history_query(
                search=search, department=department, date_from=date_from, date_to=date_to,
                user_id=user.id, page=page, per_page=per_page, is_admin=True
            )
            # Disable caching for admin history data
            history = []
            for att in attendances:
                att_dict = att.to_dict()
                att_dict['user_name'] = att.user.name if att.user else '-'
                att_dict['department'] = att.user.department if att.user else '-'
                att_dict['approver_name'] = att.approver.name if att.approver else '-'
                
                # Debug logging chỉ khi cần thiết
                if app.debug and att.id <= 5:  # Chỉ log 5 records đầu tiên trong debug mode
                    pass

                history.append(att_dict)

            return jsonify({
                'total': total,
                'page': page,
                'per_page': per_page,
                'data': history
            })
        else:
            # Lấy tham số lọc theo tháng
            month = validate_int(request.args.get('month', '').strip()) if request.args.get('month') else None
            year = validate_int(request.args.get('year', '').strip()) if request.args.get('year') else None
            
            # Nếu không có tham số tháng/năm, lấy tháng/năm hiện tại
            if not month or not year:
                from datetime import datetime
                now = datetime.now()
                month = month or now.month
                year = year or now.year
            
            # Tạo date_from và date_to cho tháng được chọn
            from datetime import datetime
            date_from = datetime(year, month, 1).date()
            if month == 12:
                date_to = datetime(year + 1, 1, 1).date()
            else:
                date_to = datetime(year, month + 1, 1).date()
            
            # Use optimized query for user's own records with month filter
            from utils.query_optimizer import optimize_attendance_history_query
            attendances, total = optimize_attendance_history_query(
                user_id=user.id, page=1, per_page=1000, is_admin=False,
                date_from=date_from, date_to=date_to
            )
            history = []
            for att in attendances:
                history.append(att.to_dict())
            resp = jsonify(history)
            resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            resp.headers['Pragma'] = 'no-cache'
            resp.headers['Expires'] = '0'
            return resp
    except Exception as e:
        # print(f"Error in get_attendance_history: {str(e)}")
        return jsonify({'error': 'Đã xảy ra lỗi khi lấy lịch sử chấm công'}), 500






def check_attendance_access_permission(user_id, attendance_id, action='read'):
    """Check if user has permission to access specific attendance record"""
    user = db.session.get(User, user_id)
    if not user:
        return False, "❌ KHÔNG TÌM THẤY NGƯỜI DÙNG"
    
    attendance = db.session.get(Attendance, attendance_id)
    if not attendance:
        return False, "❌ KHÔNG TÌM THẤY BẢN GHI CHẤM CÔNG"
    
    current_role = session.get('current_role', user.roles.split(',')[0])
    
    # ADMIN có thể truy cập tất cả
    if current_role == 'ADMIN':
        return True, ""
    
    # MANAGER có thể truy cập nhân viên cùng phòng ban
    if current_role == 'MANAGER':
        if not attendance.user or attendance.user.department != user.department:
            return False, "❌ KHÔNG CÙNG PHÒNG BAN: Chỉ xem được nhân viên cùng phòng"
        return True, ""
    
    # TEAM_LEADER có thể truy cập nhân viên cùng phòng ban
    if current_role == 'TEAM_LEADER':
        if not attendance.user or attendance.user.department != user.department:
            return False, "❌ KHÔNG CÙNG PHÒNG BAN: Chỉ xem được nhân viên cùng phòng"
        return True, ""
    
    # EMPLOYEE chỉ có thể truy cập bản ghi của chính mình
    if current_role == 'EMPLOYEE':
        if attendance.user_id != user_id:
            return False, "❌ CHỈ XEM ĐƯỢC BẢN GHI CỦA MÌNH"
        return True, ""
    
    return False, "❌ KHÔNG CÓ QUYỀN XEM BẢN GHI NÀY"

def check_request_access_permission(user_id, request_id, action='read'):
    """Check if user has permission to access specific request record"""
    user = db.session.get(User, user_id)
    if not user:
        return False, "❌ KHÔNG TÌM THẤY NGƯỜI DÙNG"
    
    req = Request.query.options(joinedload(Request.user)).get(request_id)
    if not req:
        return False, "❌ KHÔNG TÌM THẤY YÊU CẦU"
    
    current_role = session.get('current_role', user.roles.split(',')[0])
    
    # ADMIN có thể truy cập tất cả
    if current_role == 'ADMIN':
        return True, ""
    
    # MANAGER có thể truy cập yêu cầu của nhân viên cùng phòng ban
    if current_role == 'MANAGER':
        if not req.user or req.user.department != user.department:
            return False, "❌ KHÔNG CÙNG PHÒNG BAN: Chỉ xem được yêu cầu cùng phòng"
        return True, ""
    
    # TEAM_LEADER có thể truy cập yêu cầu của nhân viên cùng phòng ban
    if current_role == 'TEAM_LEADER':
        if not req.user or req.user.department != user.department:
            return False, "❌ KHÔNG CÙNG PHÒNG BAN: Chỉ xem được yêu cầu cùng phòng"
        return True, ""
    
    # EMPLOYEE chỉ có thể truy cập yêu cầu của chính mình
    if current_role == 'EMPLOYEE':
        if req.user_id != user_id:
            return False, "❌ CHỈ XEM ĐƯỢC YÊU CẦU CỦA MÌNH"
        return True, ""
    
    return False, "❌ KHÔNG CÓ QUYỀN XEM YÊU CẦU NÀY"

# Import session utilities from utils
from utils.session import check_session_timeout, update_session_activity, log_audit_action

def require_role(required_role):
    """Decorator to require specific role for route"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('user_id'):
                return redirect(url_for('login'))
            
            # Kiểm tra vai trò hiện tại trong session
            current_role = session.get('current_role')
            if current_role != required_role:
                flash(f'⚠️ CẦN CHUYỂN VAI TRÒ: Chuyển sang vai trò {required_role} để truy cập trang này', 'error')
                return redirect(url_for('dashboard'))
            
            # Kiểm tra user có role này trong database không
            if not has_role(session['user_id'], required_role):
                flash('❌ KHÔNG CÓ QUYỀN: Bạn không có quyền truy cập trang này', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Role-based route decorators
def require_admin(f):
    return require_role('ADMIN')(f)

def require_manager(f):
    return require_role('MANAGER')(f)

def require_team_lead(f):
    return require_role('TEAM_LEADER')(f)

def require_employee(f):
    return require_role('EMPLOYEE')(f)

@app.route('/admin/users')
@require_admin
def admin_users():
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Mặc định 10/trang, không cho chọn
    search = request.args.get('search', '', type=str).strip()
    department_filter = request.args.get('department', '', type=str).strip()
    maternity_filter = request.args.get('maternity', '', type=str).strip()  # 'active' = chỉ mẹ <12m đang áp dụng

    # Ngày hiện tại để tính toán chính sách mẹ <12 tháng
    today = date.today()

    query = User.query.filter_by(is_deleted=False)  # Chỉ hiển thị users chưa bị soft delete
    if search:
        # Cải thiện tìm kiếm: chuyển về lowercase và sử dụng func.lower() để đảm bảo không phân biệt hoa thường
        search_lower = search.lower().strip()
        # Tách từ khóa tìm kiếm thành các từ riêng lẻ
        search_words = search_lower.split()
        
        # Tạo điều kiện tìm kiếm đơn giản - tìm theo từng từ riêng lẻ
        name_conditions = []
        for word in search_words:
            name_conditions.append(func.lower(User.name).contains(word))
        
        # Thêm điều kiện tìm kiếm theo mã nhân viên
        name_conditions.append(func.lower(func.cast(User.employee_id, db.String)).contains(search_lower))
        
        # Kết hợp tất cả điều kiện với OR
        query = query.filter(db.or_(*name_conditions))
    if department_filter:
        query = query.filter(User.department == department_filter)

    # Lọc theo chính sách mẹ có con <12 tháng (đang hiệu lực)
    if maternity_filter == 'active':
        query = query.filter(
            User.is_maternity_flex.is_(True),
            db.or_(User.maternity_flex_from.is_(None), User.maternity_flex_from <= today),
            db.or_(User.maternity_flex_until.is_(None), today <= User.maternity_flex_until),
        )
    query = query.order_by(User.name.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    users = pagination.items

    # Bổ sung thông tin chính sách mẹ có con <12 tháng cho view
    for u in users:
        # Đang áp dụng: trong khoảng thời gian được hưởng chính sách
        flex_active = bool(
            u.is_maternity_flex
            and (u.maternity_flex_from is None or u.maternity_flex_from <= today)
            and (u.maternity_flex_until is None or today <= u.maternity_flex_until)
        )

        # Còn bao nhiêu ngày đến khi hết chính sách (nếu có ngày kết thúc)
        remaining_days = None
        if u.maternity_flex_until:
            remaining_days = (u.maternity_flex_until - today).days

        # Trạng thái "tạm ngưng" chỉ hiển thị khi đã cấu hình nhưng chưa tới ngày bắt đầu
        # Sau khi hết thời gian hưởng chính sách thì không coi là "tạm ngưng" nữa,
        # user trở về trạng thái nhân viên bình thường.
        flex_pending = bool(
            u.is_maternity_flex
            and not flex_active
            and u.maternity_flex_from is not None
            and u.maternity_flex_from > today
        )

        # gán tạm thuộc tính hiển thị
        u.flex_active = flex_active
        u.flex_pending = flex_pending
        u.flex_remaining_days = remaining_days

    # Lấy danh sách phòng ban từ database (unique, không null)
    db_departments = db.session.query(User.department).filter(
        User.is_deleted == False,
        User.department != None,
        User.department != ''
    ).distinct().order_by(User.department).all()
    departments = sorted(set([d[0] for d in db_departments if d[0]]))

    # Calculate statistics
    admin_count = sum(1 for user in users if 'ADMIN' in user.roles.split(','))
    active_count = sum(1 for user in users if user.is_active)
    department_count = len(set(user.department for user in users))
    # Tính toán phân trang đẹp (hiển thị 5 trang quanh trang hiện tại)
    start_page = max(1, pagination.page - 2)
    end_page = min(pagination.pages, pagination.page + 2)
    if end_page - start_page < 4:
        end_page = min(pagination.pages, start_page + 4)
        start_page = max(1, end_page - 4)
    page_range = range(start_page, end_page + 1)

    return render_template(
        'admin/users.html',
        users=users,
        admin_count=admin_count,
        active_count=active_count,
        department_count=department_count,
        pagination=pagination,
        search=search,
        departments=departments,
        department_filter=department_filter,
        maternity_filter=maternity_filter,
        per_page=per_page,
        page_range=page_range
    )

@app.route('/admin/departments', methods=['GET', 'POST'])
@require_admin
def admin_departments():
    """Quản lý phòng ban và mapping với Google Sheet"""
    from database.models import Department
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            name = request.form.get('name', '').strip().upper()
            timesheet_file = request.form.get('timesheet_file', '').strip()
            
            if not name or not timesheet_file:
                flash('Vui lòng điền đầy đủ thông tin!', 'error')
                return redirect(url_for('admin_departments'))
            
            # Kiểm tra trùng tên
            existing = Department.query.filter_by(name=name).first()
            if existing:
                flash(f'Phòng ban "{name}" đã tồn tại!', 'error')
                return redirect(url_for('admin_departments'))
            
            # Tạo mới
            new_dept = Department(name=name, timesheet_file=timesheet_file, is_active=True)
            db.session.add(new_dept)
            db.session.commit()
            
            flash(f'Đã thêm phòng ban "{name}" thành công!', 'success')
            
        elif action == 'edit':
            dept_id = request.form.get('dept_id', type=int)
            name = request.form.get('name', '').strip().upper()
            timesheet_file = request.form.get('timesheet_file', '').strip()
            
            dept = Department.query.get(dept_id)
            if dept:
                dept.name = name
                dept.timesheet_file = timesheet_file
                db.session.commit()
                
                flash(f'Đã cập nhật phòng ban "{name}" thành công!', 'success')
            else:
                flash('Không tìm thấy phòng ban!', 'error')
                
        elif action == 'delete':
            dept_id = request.form.get('dept_id', type=int)
            dept = Department.query.get(dept_id)
            if dept:
                name = dept.name
                db.session.delete(dept)
                db.session.commit()
                
                flash(f'Đã xóa phòng ban "{name}" thành công!', 'success')
            else:
                flash('Không tìm thấy phòng ban!', 'error')
        
        return redirect(url_for('admin_departments'))
    
    # GET: Hiển thị danh sách phòng ban
    departments = Department.query.order_by(Department.name).all()
    return render_template('admin/departments.html', departments=departments)


def _notify_missing_holidays():
    """Gửi thông báo cho admin khi chưa cấu hình ngày lễ (giống cơ chế refresh token)."""
    try:
        current_year = datetime.now().year
        message = (
            f"⚠️ CHƯA CẤU HÌNH NGÀY LỄ CHO NĂM {current_year}\n\n"
            "Vui lòng vào mục 'Quản lý ngày lễ' để thêm các ngày nghỉ lễ Việt Nam / Nhật Bản "
            "trước khi chạy bảng chấm công."
        )
        # Thông báo qua kênh SSE giống refresh token (nếu front-end đang lắng nghe)
        try:
            publish_token_status('expired', message, needs_reauth=False)
        except Exception:
            pass
        # Gửi thêm qua Telegram (nếu đã cấu hình BOT_TOKEN, CHAT_ID)
        try:
            send_telegram_message(message)
        except Exception:
            pass
    except Exception:
        # Không để lỗi thông báo làm hỏng request chính
        pass


@app.route('/admin/holidays', methods=['GET', 'POST'])
@require_admin
def admin_holidays():
    """Quản lý ngày lễ Việt Nam và Nhật Bản"""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            date_str = request.form.get('date', '').strip()
            holiday_type = request.form.get('holiday_type', '').strip()
            name = request.form.get('name', '').strip()
            description = request.form.get('description', '').strip()
            
            if not date_str or not holiday_type:
                flash('Vui lòng điền đầy đủ thông tin!', 'error')
                return redirect(url_for('admin_holidays'))
            
            if holiday_type not in ['vietnamese_holiday', 'japanese_holiday']:
                flash('Loại ngày lễ không hợp lệ!', 'error')
                return redirect(url_for('admin_holidays'))
            
            try:
                from datetime import datetime
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Ngày không hợp lệ!', 'error')
                return redirect(url_for('admin_holidays'))
            
            # Kiểm tra trùng ngày
            existing = Holiday.query.filter_by(date=date).first()
            if existing:
                flash(f'Ngày {date_str} đã được đánh dấu là {existing.holiday_type}!', 'error')
                return redirect(url_for('admin_holidays'))
            
            # Tạo mới
            # Lưu người tạo dựa trên session (hệ thống đang dùng session để quản lý đăng nhập)
            creator_id = session.get('user_id')
            new_holiday = Holiday(
                date=date,
                holiday_type=holiday_type,
                name=name if name else None,
                description=description if description else None,
                created_by=creator_id
            )
            db.session.add(new_holiday)
            db.session.commit()
            
            flash(f'Đã thêm ngày lễ thành công!', 'success')
            
        elif action == 'edit':
            holiday_id = request.form.get('holiday_id', type=int)
            date_str = request.form.get('date', '').strip()
            holiday_type = request.form.get('holiday_type', '').strip()
            name = request.form.get('name', '').strip()
            description = request.form.get('description', '').strip()
            
            holiday = Holiday.query.get(holiday_id)
            if not holiday:
                flash('Không tìm thấy ngày lễ!', 'error')
                return redirect(url_for('admin_holidays'))
            
            if not date_str or not holiday_type:
                flash('Vui lòng điền đầy đủ thông tin!', 'error')
                return redirect(url_for('admin_holidays'))
            
            if holiday_type not in ['vietnamese_holiday', 'japanese_holiday']:
                flash('Loại ngày lễ không hợp lệ!', 'error')
                return redirect(url_for('admin_holidays'))
            
            try:
                from datetime import datetime
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Ngày không hợp lệ!', 'error')
                return redirect(url_for('admin_holidays'))
            
            # Kiểm tra trùng ngày với ngày lễ khác
            existing = Holiday.query.filter_by(date=date).first()
            if existing and existing.id != holiday_id:
                flash(f'Ngày {date_str} đã được đánh dấu là {existing.holiday_type}!', 'error')
                return redirect(url_for('admin_holidays'))
            
            holiday.date = date
            holiday.holiday_type = holiday_type
            holiday.name = name if name else None
            holiday.description = description if description else None
            db.session.commit()
            
            flash(f'Đã cập nhật ngày lễ thành công!', 'success')
                
        elif action == 'delete':
            holiday_id = request.form.get('holiday_id', type=int)
            holiday = Holiday.query.get(holiday_id)
            if holiday:
                date_str = holiday.date.strftime('%d/%m/%Y')
                db.session.delete(holiday)
                db.session.commit()
                
                flash(f'Đã xóa ngày lễ {date_str} thành công!', 'success')
            else:
                flash('Không tìm thấy ngày lễ!', 'error')
        
        return redirect(url_for('admin_holidays'))
    
    # GET: Hiển thị danh sách ngày lễ
    holidays = Holiday.query.order_by(Holiday.date.desc()).all()
    return render_template('admin/holidays.html', holidays=holidays)


@app.route('/api/get-day-type', methods=['GET'])
def get_day_type():
    """API để lấy loại ngày dựa trên ngày được chọn"""
    # Kiểm tra session-based authentication (ứng dụng này dùng session thay vì Flask-Login)
    if 'user_id' not in session:
        return jsonify({'error': 'Chưa đăng nhập', 'day_type': 'normal', 'reason': 'Ngày thường (fallback)'}), 401
    
    try:
        
        date_str = request.args.get('date')
        print(f"DEBUG: date_str = {date_str}")
        if not date_str:
            return jsonify({'error': 'Vui lòng cung cấp ngày'}), 400
        
        try:
            from datetime import datetime
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Ngày không hợp lệ'}), 400
        
        # Kiểm tra thứ trong tuần (0 = Monday, 6 = Sunday)
        weekday = date.weekday()  # 0-6 (Monday-Sunday)
        is_weekend = weekday >= 5  # Saturday (5) or Sunday (6)
        
        # Kiểm tra ngày lễ trong database (với xử lý lỗi nếu bảng chưa tồn tại)
        holiday = None
        try:
            holiday = Holiday.query.filter_by(date=date).first()
        except Exception as e:
            # Nếu bảng holidays chưa tồn tại, bỏ qua và chỉ dựa vào thứ trong tuần
            print(f"Warning: Không thể query Holiday table: {e}")
            holiday = None
        
        # Độ ưu tiên: Lễ Việt Nam > Cuối tuần > Lễ Nhật
        if holiday and holiday.holiday_type == 'vietnamese_holiday':
            return jsonify({
                'day_type': 'vietnamese_holiday',
                'reason': 'Lễ Việt Nam',
                'holiday_name': holiday.name if holiday.name else None
            })
        elif is_weekend:
            return jsonify({
                'day_type': 'weekend',
                'reason': 'Cuối tuần'
            })
        elif holiday and holiday.holiday_type == 'japanese_holiday':
            return jsonify({
                'day_type': 'japanese_holiday',
                'reason': 'Lễ Nhật Bản',
                'holiday_name': holiday.name if holiday.name else None
            })
        else:
            return jsonify({
                'day_type': 'normal',
                'reason': 'Ngày thường'
            })
    except Exception as e:
        # Đảm bảo luôn trả về JSON ngay cả khi có lỗi
        import traceback
        print(f"Error in get_day_type: {e}")
        print(traceback.format_exc())
        return jsonify({
            'error': 'Lỗi khi xác định loại ngày',
            'day_type': 'normal',
            'reason': 'Ngày thường (fallback)'
        }), 500


@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@require_admin
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        try:
            name = validate_input_sanitize(request.form.get('name'))
            department = validate_input_sanitize(request.form.get('department'))
            email = request.form.get('email', '').strip()
            is_maternity_flex = request.form.get('is_maternity_flex') == 'on'
            flex_from_str = (request.form.get('maternity_flex_from') or '').strip()
            flex_until_str = (request.form.get('maternity_flex_until') or '').strip()
            
            if not name:
                flash('Tên người dùng không hợp lệ', 'error')
                return redirect(url_for('edit_user', user_id=user_id))
            if not department:
                flash('Phòng ban không hợp lệ', 'error')
                return redirect(url_for('edit_user', user_id=user_id))
            
            # Validate email nếu có
            if email:
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_pattern, email):
                    flash('Email không hợp lệ!', 'error')
                    return redirect(url_for('edit_user', user_id=user_id))
                
                # Kiểm tra email đã được sử dụng bởi user khác chưa
                existing_email_user = User.query.filter_by(email=email, is_deleted=False).first()
                if existing_email_user and existing_email_user.id != user_id:
                    flash('Email này đã được sử dụng bởi nhân viên khác!', 'error')
                    return redirect(url_for('edit_user', user_id=user_id))
            
            # Parse ngày linh hoạt (cho phép để trống)
            def _parse_date(val, label):
                if not val:
                    return None
                try:
                    return datetime.strptime(val, '%Y-%m-%d').date()
                except ValueError:
                    flash(f'Ngày {label} không hợp lệ (định dạng YYYY-MM-DD)', 'error')
                    raise
            
            flex_from_date = None
            flex_until_date = None
            try:
                flex_from_date = _parse_date(flex_from_str, 'bắt đầu áp dụng')
                flex_until_date = _parse_date(flex_until_str, 'hết hiệu lực')
            except ValueError:
                return redirect(url_for('edit_user', user_id=user_id))
            
            if flex_from_date and flex_until_date and flex_until_date < flex_from_date:
                flash('Ngày hết hiệu lực phải lớn hơn hoặc bằng ngày bắt đầu', 'error')
                return redirect(url_for('edit_user', user_id=user_id))
            
            # Get selected roles from checkboxes
            selected_roles = []
            role_mapping = {
                'EMPLOYEE': 'EMPLOYEE',
                'TEAM_LEADER': 'TEAM_LEADER', 
                'MANAGER': 'MANAGER',
                'ADMIN': 'ADMIN'
            }
            
            for role_key, role_value in role_mapping.items():
                if request.form.get(f'role_{role_key}') == 'on':
                    selected_roles.append(role_value)
            
            if not selected_roles:
                flash('Vui lòng chọn ít nhất một vai trò!', 'error')
                return redirect(url_for('edit_user', user_id=user_id))
            
            # Update user
            old_values = {
                'name': user.name,
                'department': user.department,
                'roles': user.roles,
                'email': user.email,
                'is_maternity_flex': user.is_maternity_flex,
                'maternity_flex_from': user.maternity_flex_from.isoformat() if user.maternity_flex_from else None,
                'maternity_flex_until': user.maternity_flex_until.isoformat() if user.maternity_flex_until else None
            }
            
            user.name = name
            user.roles = ','.join(selected_roles)
            user.department = department
            user.email = email if email else None
            user.is_maternity_flex = is_maternity_flex
            user.maternity_flex_from = flex_from_date
            user.maternity_flex_until = flex_until_date
            
            db.session.commit()
            
            # Log the action
            log_audit_action(
                user_id=session['user_id'],
                action='UPDATE_USER',
                table_name='users',
                record_id=user_id,
                old_values=old_values,
                new_values={
                    'name': name,
                    'department': department,
                    'roles': ','.join(selected_roles),
                    'email': email if email else None,
                    'is_maternity_flex': is_maternity_flex,
                    'maternity_flex_from': flex_from_date.isoformat() if flex_from_date else None,
                    'maternity_flex_until': flex_until_date.isoformat() if flex_until_date else None
                }
            )
            
            flash('Cập nhật người dùng thành công', 'success')
            return redirect(url_for('admin_users'))
        except Exception as e:
            # print(f"Error updating user: {str(e)}")
            flash('Đã xảy ra lỗi khi cập nhật người dùng!', 'error')
            return redirect(url_for('edit_user', user_id=user_id))
    # Lấy danh sách phòng ban từ bảng Department trong database
    from database.models import Department
    db_departments = Department.query.filter_by(is_active=True).order_by(Department.name).all()
    departments = [d.name for d in db_departments]
    
    return render_template('admin/edit_user.html', user=user, departments=departments)

@app.route('/admin/users/create', methods=['GET', 'POST'])
@require_admin
def create_user():
    if request.method == 'POST':
        try:
            # Validate input
            employee_id_str = request.form.get('employee_id')
            password = request.form.get('password')
            name = validate_input_sanitize(request.form.get('name'))
            department = validate_input_sanitize(request.form.get('department'))
            is_maternity_flex = request.form.get('is_maternity_flex') == 'on'
            flex_from_str = (request.form.get('maternity_flex_from') or '').strip()
            flex_until_str = (request.form.get('maternity_flex_until') or '').strip()
            
            # Validate employee_id
            employee_id = validate_employee_id(employee_id_str)
            if not employee_id:
                flash('Mã nhân viên không hợp lệ!', 'error')
                return render_template('admin/create_user.html')
            
            # Validate password
            if not validate_str(password, max_length=100):
                flash('Mật khẩu không hợp lệ!', 'error')
                return render_template('admin/create_user.html')
            
            # Validate name and department
            if not name:
                flash('Tên người dùng không hợp lệ', 'error')
                return render_template('admin/create_user.html')
            if not department:
                flash('Phòng ban không hợp lệ', 'error')
                return render_template('admin/create_user.html')
            
            def _parse_date(val, label):
                if not val:
                    return None
                try:
                    return datetime.strptime(val, '%Y-%m-%d').date()
                except ValueError:
                    flash(f'Ngày {label} không hợp lệ (định dạng YYYY-MM-DD)', 'error')
                    raise
            
            flex_from_date = None
            flex_until_date = None
            try:
                flex_from_date = _parse_date(flex_from_str, 'bắt đầu áp dụng')
                flex_until_date = _parse_date(flex_until_str, 'hết hiệu lực')
            except ValueError:
                return render_template('admin/create_user.html')
            
            if flex_from_date and flex_until_date and flex_until_date < flex_from_date:
                flash('Ngày hết hiệu lực phải lớn hơn hoặc bằng ngày bắt đầu', 'error')
                return render_template('admin/create_user.html')
            
            # Check if employee_id already exists (chỉ kiểm tra users chưa bị xóa)
            existing_user = User.query.filter_by(employee_id=employee_id, is_deleted=False).first()
            if existing_user:
                flash('Mã nhân viên đã tồn tại!', 'error')
                return render_template('admin/create_user.html')
            
            # Get selected roles from checkboxes
            selected_roles = []
            role_mapping = {
                'EMPLOYEE': 'EMPLOYEE',
                'TEAM_LEADER': 'TEAM_LEADER', 
                'MANAGER': 'MANAGER',
                'ADMIN': 'ADMIN'
            }
            
            for role_key, role_value in role_mapping.items():
                if request.form.get(f'role_{role_key}') == 'on':
                    selected_roles.append(role_value)
            
            if not selected_roles:
                flash('Vui lòng chọn ít nhất một vai trò!', 'error')
                return render_template('admin/create_user.html')
            
            # Create new user
            new_user = User(
                employee_id=employee_id,
                name=name,
                department=department,
                roles=','.join(selected_roles),
                is_active=True,
                is_maternity_flex=is_maternity_flex,
                maternity_flex_from=flex_from_date,
                maternity_flex_until=flex_until_date
            )
            new_user.set_password(password)
            
            db.session.add(new_user)
            db.session.commit()
            
            # Log the action
            log_audit_action(
                user_id=session['user_id'],
                action='CREATE_USER',
                table_name='users',
                record_id=new_user.id,
                new_values={
                    'employee_id': employee_id,
                    'name': name,
                    'department': department,
                    'roles': ','.join(selected_roles),
                    'is_maternity_flex': is_maternity_flex,
                    'maternity_flex_from': flex_from_date.isoformat() if flex_from_date else None,
                    'maternity_flex_until': flex_until_date.isoformat() if flex_until_date else None
                }
            )
            
            flash('Tạo người dùng thành công!', 'success')
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            # Lấy danh sách phòng ban từ bảng Department trong database
            from database.models import Department
            db_departments = Department.query.filter_by(is_active=True).order_by(Department.name).all()
            departments = [d.name for d in db_departments]
            return render_template('admin/create_user.html', departments=departments)
    
    # Lấy danh sách phòng ban từ bảng Department trong database
    from database.models import Department
    db_departments = Department.query.filter_by(is_active=True).order_by(Department.name).all()
    departments = [d.name for d in db_departments]
    return render_template('admin/create_user.html', departments=departments)

@app.route('/switch-role', methods=['POST'])
def switch_role():
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    data = request.get_json()
    role = validate_role_value(data.get('role'))
    if not role:
        return jsonify({'error': 'Vai trò không hợp lệ'}), 400
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    if role not in user.roles.split(','):
        return jsonify({'error': 'Vai trò không hợp lệ'}), 400
    old_role = session.get('current_role')
    session['current_role'] = role
    
    # Invalidate cache for role switch
    from utils.realtime_updates import invalidate_role_cache
    invalidate_role_cache(user.id, role)
    
    log_audit_action(
        user_id=user.id,
        action='SWITCH_ROLE',
        table_name='users',
        record_id=user.id,
        old_values={'current_role': old_role},
        new_values={'current_role': role}
    )
    
    response = jsonify({'message': 'Đã chuyển vai trò thành công'})
    # Ensure no caching of role switch response
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# API endpoint để submit request
@app.route('/api/request/submit', methods=['POST'])
def submit_request():
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    data = request.get_json()
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    # Validate input
    request_type = validate_input_sanitize(data.get('request_type'))
    start_date = validate_date(data.get('start_date'))
    end_date = validate_date(data.get('end_date'))
    reason = validate_reason(data.get('reason'))
    if not request_type:
        return jsonify({'error': 'Loại yêu cầu không hợp lệ'}), 400
    if not start_date:
        return jsonify({'error': 'Ngày bắt đầu không hợp lệ'}), 400
    if not end_date:
        return jsonify({'error': 'Ngày kết thúc không hợp lệ'}), 400
    if not reason:
        return jsonify({'error': 'Lý do không hợp lệ'}), 400
    if start_date > end_date:
        return jsonify({'error': 'Ngày bắt đầu phải trước ngày kết thúc'}), 400
    if start_date < datetime.now().date():
        return jsonify({'error': 'Không thể tạo yêu cầu cho ngày trong quá khứ'}), 400
    leader = User.query.filter_by(department=user.department, roles='TEAM_LEADER', is_deleted=False).first()
    if not leader:
        return jsonify({'error': 'Không tìm thấy trưởng nhóm cho phòng ban này'}), 400
    new_request = Request(
        user_id=user.id,
        request_type=request_type,
        start_date=start_date,
        end_date=end_date,
        reason=reason,
        current_approver_id=leader.id,
        step='leader',
        status='pending'
    )
    db.session.add(new_request)
    db.session.commit()
    return jsonify({'message': 'Gửi yêu cầu thành công'}), 201

# API endpoint để phê duyệt/từ chối request
@app.route('/api/request/<int:request_id>/approve', methods=['POST'])
def approve_request(request_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    data = request.get_json()
    action = data.get('action')  # 'approve' hoặc 'reject'
    # Chỉ yêu cầu có lý do (không bắt buộc 10 ký tự)
    reason_raw = data.get('reason', '') if data.get('action') == 'reject' else ''
    reason = validate_input_sanitize(reason_raw, max_length=500) if reason_raw else ''
    if action not in ['approve', 'reject']:
        return jsonify({'error': 'Hành động không hợp lệ'}), 400
    if action == 'reject' and not reason:
        return jsonify({'error': 'Vui lòng nhập lý do từ chối'}), 400
    has_permission, error_message = check_request_access_permission(session['user_id'], request_id, 'approve')
    if not has_permission:
        return jsonify({'error': error_message}), 403
    req = Request.query.options(joinedload(Request.user)).get_or_404(request_id)
    approver = db.session.get(User, session['user_id'])
    if req.current_approver_id != approver.id:
        return jsonify({'error': 'Bạn không có quyền phê duyệt yêu cầu này'}), 403
    if action == 'approve':
        if req.step == 'leader':
            manager = User.query.filter(
                User.department == req.user.department,
                User.roles.like('%MANAGER%'),
                User.is_deleted == False
            ).first()
            if not manager:
                return jsonify({'error': 'Không tìm thấy quản lý cho phòng ban này'}), 400
            req.current_approver_id = manager.id
            req.step = 'manager'
        elif req.step == 'manager':
            admin = User.query.filter(
                User.roles.like('%ADMIN%'),
                User.is_deleted == False
            ).first()
            if not admin:
                return jsonify({'error': 'Không tìm thấy quản trị viên'}), 400
            req.current_approver_id = admin.id
            req.step = 'admin'
        elif req.step == 'admin':
            req.status = 'approved'
            req.step = 'done'
    else:  # reject
        req.status = 'rejected'
        req.step = 'employee_edit'
        req.reject_reason = reason
        req.current_approver_id = req.user_id
    db.session.commit()
    return jsonify({'message': 'Cập nhật yêu cầu thành công'}), 200

@app.route('/api/attendance/<int:attendance_id>', methods=['DELETE'])
def delete_attendance(attendance_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    # Kiểm tra session timeout
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    
    # Cập nhật thời gian hoạt động cuối
    update_session_activity()
    
    # Kiểm tra quyền truy cập (chỉ EMPLOYEE có thể xóa bản ghi của chính mình)
    has_permission, error_message = check_attendance_access_permission(session['user_id'], attendance_id, 'delete')
    if not has_permission:
        return jsonify({'error': error_message}), 403
    
    att = db.session.get(Attendance, attendance_id)
    if not att:
        return jsonify({'error': 'Không tìm thấy bản ghi'}), 404
    if att.approved:
        return jsonify({'error': 'Bản ghi đã được phê duyệt, không thể xóa!'}), 400
    try:
        # Log attendance deletion
        log_audit_action(
            user_id=session['user_id'],
            action='DELETE_ATTENDANCE',
            table_name='attendances',
            record_id=attendance_id,
            old_values={
                'date': att.date.isoformat(),
                'check_in': att.check_in.isoformat() if att.check_in else None,
                'check_out': att.check_out.isoformat() if att.check_out else None,
                'status': att.status
            }
        )
        
        db.session.delete(att)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Lỗi khi xóa bản ghi!'}), 500

@app.route('/api/attendance/<int:attendance_id>', methods=['GET'])
def get_attendance(attendance_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    # Kiểm tra session timeout
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    
    # Cập nhật thời gian hoạt động cuối
    update_session_activity()
    
    # Kiểm tra quyền truy cập
    has_permission, error_message = check_attendance_access_permission(session['user_id'], attendance_id, 'read')
    if not has_permission:
        return jsonify({'error': error_message}), 403
    
    attendance = db.session.get(Attendance, attendance_id)
    if not attendance:
        return jsonify({'error': 'Không tìm thấy bản ghi'}), 404
    
    # Lấy thông tin người dùng
    user_info = {
        'name': attendance.user.name if attendance.user else 'Unknown',
        'employee_id': attendance.user.employee_id if attendance.user else 'Unknown',
        'department': attendance.user.department if attendance.user else 'Unknown'
    }
    
    # Lấy thông tin người phê duyệt nếu có
    approver_info = None
    if attendance.approved_by:
        approver = db.session.get(User, attendance.approved_by)
        if approver:
            approver_info = {
                'name': approver.name,
                'employee_id': approver.employee_id,
                'department': approver.department,
                'roles': approver.roles
            }
    
    return jsonify({
        'id': attendance.id,
        'date': attendance.date.strftime('%d/%m/%Y'),
        'check_in': attendance.check_in.strftime('%H:%M') if attendance.check_in else None,
        'check_out': attendance.check_out.strftime('%H:%M') if attendance.check_out else None,
        'break_time': attendance._format_hours_minutes(attendance.break_time),
        'comp_time_regular': attendance._format_minutes_to_hhmm(attendance.comp_time_regular_minutes),
        'comp_time_overtime': attendance._format_minutes_to_hhmm(attendance.comp_time_overtime_minutes),
        'comp_time_ot_before_22': attendance._format_minutes_to_hhmm(attendance.comp_time_ot_before_22_minutes),
        'comp_time_ot_after_22': attendance._format_minutes_to_hhmm(attendance.comp_time_ot_after_22_minutes),
        'overtime_comp_time': attendance._format_minutes_to_hhmm(attendance.overtime_comp_time_minutes),
        'is_holiday': attendance.is_holiday,
        'holiday_type': attendance.holiday_type,
        'note': attendance.note,
        'approved': attendance.approved,
        'status': attendance.status,
        'shift_code': attendance.shift_code,
        'shift_start': attendance.shift_start.strftime('%H:%M') if attendance.shift_start else None,
        'shift_end': attendance.shift_end.strftime('%H:%M') if attendance.shift_end else None,
        'signature': attendance.signature,
        'team_leader_signature': attendance.team_leader_signature,
        'manager_signature': attendance.manager_signature,
        'user_name': user_info['name'],
        'user_employee_id': user_info['employee_id'],
        'user_department': user_info['department'],
        'approver_info': approver_info,
        'approved_at': attendance.approved_at.isoformat() if attendance.approved_at else None
    })

@app.route('/api/attendance/<int:attendance_id>', methods=['PUT'])
def update_attendance(attendance_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    has_permission, error_message = check_attendance_access_permission(session['user_id'], attendance_id, 'update')
    if not has_permission:
        return jsonify({'error': error_message}), 403
    attendance = db.session.get(Attendance, attendance_id)
    if not attendance:
        return jsonify({'error': 'Không tìm thấy bản ghi'}), 404
    if attendance.approved:
        return jsonify({'error': 'Bản ghi đã được phê duyệt, không thể sửa!'}), 400
    data = request.get_json()
    # print('DEBUG signature PUT:', data.get('signature'))  # Thêm log signature
    # Validate input
    date = validate_date(data.get('date'))
    check_in = validate_time(data.get('check_in'))
    check_out = validate_time(data.get('check_out'))
    note = validate_note(data.get('note', ''))
    
    # Chuyển đổi HH:MM sang float cho các trường thời gian
    def hhmm_to_float(hhmm_str):
        """Chuyển đổi HH:MM sang float (giờ)"""
        if not hhmm_str or hhmm_str == "0:00":
            return 0.0
        try:
            if isinstance(hhmm_str, str) and ':' in hhmm_str:
                hours, minutes = hhmm_str.split(':')
                return float(hours) + float(minutes) / 60.0
            else:
                return float(hhmm_str)
        except (ValueError, TypeError):
            return 0.0
    
    break_time = hhmm_to_float(data.get('break_time', '1:00'))
    comp_time_regular = hhmm_to_float(data.get('comp_time_regular', '0:00'))
    comp_time_overtime = hhmm_to_float(data.get('comp_time_overtime', '0:00'))
    comp_time_ot_before_22 = hhmm_to_float(data.get('comp_time_ot_before_22', '0:00'))
    comp_time_ot_after_22 = hhmm_to_float(data.get('comp_time_ot_after_22', '0:00'))
    overtime_comp_time = hhmm_to_float(data.get('overtime_comp_time', '0:00'))
    is_holiday = bool(data.get('is_holiday', False))
    holiday_type = validate_holiday_type(data.get('holiday_type'))
    shift_code = data.get('shift_code')
    shift_start = validate_time(data.get('shift_start'))
    shift_end = validate_time(data.get('shift_end'))
    next_day_checkout = bool(data.get('next_day_checkout', False))  # Flag cho tăng ca qua ngày mới
    # Lấy thông tin user trước khi sử dụng
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    # Tự động lấy chữ ký từ database khi cập nhật
    signature = data.get('signature', '')
    
    # Nếu không có chữ ký hoặc chữ ký rỗng, lấy từ database
    if not signature:
        auto_signature = signature_manager.get_signature_from_database(user.id, 'EMPLOYEE')
        if auto_signature:
            signature = auto_signature
            attendance.signature = signature
            # print(f"✅ AUTO SIGNATURE UPDATE: User {user.name} using signature from database")
        else:
            pass  # Không còn log debug
    else:
        # Nếu có chữ ký mới, cập nhật
        attendance.signature = signature
        # Không còn log debug
    
    if not date:
        return jsonify({'error': 'Vui lòng chọn ngày chấm công hợp lệ'}), 400
    if not holiday_type:
        return jsonify({'error': 'Vui lòng chọn loại ngày hợp lệ'}), 400
    if not check_in or not check_out:
        return jsonify({'error': 'Vui lòng nhập đầy đủ giờ vào và giờ ra hợp lệ'}), 400
    if break_time is None:
        return jsonify({'error': 'Thời gian nghỉ không hợp lệ!'}), 400
    if comp_time_regular is None:
        return jsonify({'error': 'Giờ đối ứng trong ca không hợp lệ!'}), 400
    if comp_time_overtime is None:
        return jsonify({'error': 'Giờ đối ứng tăng ca không hợp lệ!'}), 400
    if comp_time_ot_before_22 is None or comp_time_ot_after_22 is None:
        return jsonify({'error': 'Giờ đối ứng tăng ca theo mốc (trước/sau 22h) không hợp lệ!'}), 400
    
    # Validation: Kiểm tra xem có tăng ca hay không trước khi cho phép đối ứng tăng ca
    is_valid, error_message = validate_overtime_comp_time(
        check_in, check_out, shift_start, shift_end, break_time, 
        comp_time_regular, comp_time_overtime, comp_time_ot_before_22, comp_time_ot_after_22, date, data.get('next_day_checkout', False), holiday_type, shift_code
    )
    if not is_valid:
        return jsonify({'error': error_message}), 400
    
    if not shift_code or not shift_start or not shift_end:
        return jsonify({'error': 'Vui lòng chọn ca làm việc hợp lệ!'}), 400
    
    # Kiểm tra xem có bản ghi khác cùng ngày không (trừ bản ghi hiện tại)
    
    existing_attendance = Attendance.query.filter(
        Attendance.user_id == user.id,
        Attendance.date == date,
        Attendance.id != attendance_id
    ).first()
    
    if existing_attendance:
        if existing_attendance.status != 'rejected':
            return jsonify({'error': 'Bạn đã chấm công cho ngày này rồi, không thể chấm công 2 lần trong 1 ngày.'}), 400
        else:
            db.session.delete(existing_attendance)
            db.session.commit()
    
    attendance.date = date
    attendance.check_in = datetime.combine(date, check_in)
    
    # Xử lý giờ ra - nếu là tăng ca qua ngày mới thì cộng thêm 1 ngày
    if next_day_checkout:
        attendance.check_out = datetime.combine(date + timedelta(days=1), check_out)
        # print(f"DEBUG UPDATE: Tăng ca qua ngày mới - check_out: {attendance.check_out}")
    else:
        attendance.check_out = datetime.combine(date, check_out)
    
    attendance.note = note
    attendance.break_time = break_time
    attendance.comp_time_regular_minutes = int(round(comp_time_regular * 60)) if comp_time_regular else 0
    attendance.comp_time_overtime_minutes = int(round(comp_time_overtime * 60)) if comp_time_overtime else 0
    attendance.comp_time_ot_before_22_minutes = int(round(comp_time_ot_before_22 * 60)) if comp_time_ot_before_22 else 0
    attendance.comp_time_ot_after_22_minutes = int(round(comp_time_ot_after_22 * 60)) if comp_time_ot_after_22 else 0
    attendance.overtime_comp_time_minutes = int(round(overtime_comp_time * 60)) if overtime_comp_time else 0
    attendance.is_holiday = is_holiday
    attendance.holiday_type = holiday_type
    attendance.shift_code = shift_code
    attendance.shift_start = shift_start
    attendance.shift_end = shift_end
    if attendance.status == 'rejected':
        attendance.status = 'pending'
    if date > datetime.now().date():
        return jsonify({'error': 'Không thể chấm công cho ngày trong tương lai!'}), 400
    attendance.update_work_hours()
    try:
        db.session.commit()
        log_audit_action(
            user_id=session['user_id'],
            action='UPDATE_ATTENDANCE',
            table_name='attendances',
            record_id=attendance_id,
            old_values={
                'date': attendance.date.isoformat(),
                'check_in': attendance.check_in.isoformat() if attendance.check_in else None,
                'check_out': attendance.check_out.isoformat() if attendance.check_out else None,
                'status': attendance.status
            },
            new_values={
                'date': date.isoformat(),
                'check_in': datetime.combine(date, check_in).isoformat(),
                'check_out': attendance.check_out.isoformat(),
                'status': attendance.status
            }
        )
        message = 'Cập nhật chấm công thành công'
        return jsonify({
            'message': message,
            'work_hours': attendance.total_work_hours,
            'overtime_before_22': attendance.overtime_before_22,
            'overtime_after_22': attendance.overtime_after_22
        })
    except Exception as e:
        # print(f"Database error: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Đã xảy ra lỗi khi cập nhật dữ liệu'}), 500

@app.route('/api/signature/check', methods=['POST'])
def check_signature_status():
    """API để kiểm tra trạng thái chữ ký cho phê duyệt"""
    # print(f"DEBUG: check_signature_status called with data: {request.get_json()}")
    
    if 'user_id' not in session:
        # print("DEBUG: No user_id in session")
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        # print("DEBUG: Session timeout")
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    
    data = request.get_json()
    attendance_id = data.get('attendance_id')
    request_id = data.get('request_id')
    check_session = data.get('check_session', False)
    current_role = session.get('current_role')
    user_id = session['user_id']
    
    # Nếu chỉ kiểm tra session signature
    if check_session:
        session_signature, session_meta = signature_manager.get_signature_from_session(user_id, current_role)
        return jsonify({
            'session_signature': session_signature if session_signature else None
        })
    
    # Sử dụng Signature Manager để kiểm tra trạng thái
    signature_status = signature_manager.check_signature_status(user_id, current_role, attendance_id)
    return jsonify(signature_status)

@app.route('/api/signature/validate-quality', methods=['POST'])
def validate_signature_quality():
    """验证签名质量"""
    if 'user_id' not in session:
        return jsonify({'error': 'Chưa đăng nhập'}), 401
    
    data = request.get_json()
    signature = data.get('signature')
    
    if not signature:
        return jsonify({
            'valid': False,
            'error': 'Không có dữ liệu chữ ký'
        })
    
    # 使用签名处理器验证质量
    quality_result = signature_manager.validate_signature_quality(signature)
    
    return jsonify(quality_result)

@app.route('/api/signature/fit-to-form', methods=['POST'])
def fit_signature_to_form():
    """Điều chỉnh chữ ký vừa khít với ô ký trong biểu mẫu"""
    if 'user_id' not in session:
        return jsonify({'error': 'Chưa đăng nhập'}), 401
    
    data = request.get_json()
    signature = data.get('signature')
    box_type = data.get('box_type', 'default')
    
    if not signature:
        return jsonify({
            'success': False,
            'error': 'Không có dữ liệu chữ ký'
        })
    
    # Điều chỉnh chữ ký vừa khít với ô
    fitted_signature = signature_manager.fit_signature_to_form_box(signature, box_type)
    
    # Kiểm tra xem có vừa không
    fit_result = signature_manager.validate_signature_fit(signature, box_type)
    
    return jsonify({
        'success': True,
        'fitted_signature': fitted_signature,
        'fit_result': fit_result
    })

@app.route('/api/signature/create-form-signatures', methods=['POST'])
def create_form_signatures():
    """Tạo chữ ký cho toàn bộ biểu mẫu"""
    if 'user_id' not in session:
        return jsonify({'error': 'Chưa đăng nhập'}), 401
    
    data = request.get_json()
    signatures = data.get('signatures', {})
    
    if not signatures:
        return jsonify({
            'success': False,
            'error': 'Không có dữ liệu chữ ký'
        })
    
    # Tạo chữ ký cho toàn bộ biểu mẫu
    form_signatures = signature_manager.create_form_signatures(signatures)
    
    return jsonify({
        'success': True,
        'form_signatures': form_signatures
    })

@app.route('/api/signature/save-session', methods=['POST'])
def save_signature_to_session():
    """API để lưu chữ ký vào session"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    
    data = request.get_json()
    signature = data.get('signature')
    signature_type = data.get('type', 'new')  # 'new', 'reused', 'session_reused', 'database_reused'
    dont_ask_again = data.get('dont_ask_again', False)
    
    if not signature:
        return jsonify({'error': 'Chữ ký không hợp lệ'}), 400
    
    current_role = session.get('current_role')
    user_id = session['user_id']
    
    # Lưu chữ ký vào session với flag don't ask again
    success = signature_manager.save_signature_to_session(
        user_id, current_role, signature, signature_type, dont_ask_again
    )
    
    if success:
        # Ghi log chi tiết
        signature_manager.log_signature_action(
            user_id=user_id,
            action='SAVE_SESSION',
            signature_type=signature_type,
            additional_data={'dont_ask_again': dont_ask_again}
        )
        
        message = 'Đã lưu chữ ký vào phiên'
        if dont_ask_again:
            message += ' và đặt không hỏi lại trong phiên này'
        
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'error': 'Không thể lưu chữ ký'}), 500

@app.route('/api/signature/clear-session', methods=['POST'])
def clear_session_signature():
    """API để xóa chữ ký khỏi session"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    
    current_role = session.get('current_role')
    user_id = session['user_id']
    
    success = signature_manager.clear_session_signature(user_id, current_role)
    
    if success:
        return jsonify({'success': True, 'message': 'Đã xóa chữ ký khỏi phiên'})
    else:
        return jsonify({'error': 'Không thể xóa chữ ký'}), 500

@app.route('/api/attendance/<int:attendance_id>/approve', methods=['POST'])
@rate_limit(max_requests=200, window_seconds=60)
def approve_attendance(attendance_id):
    """Phê duyệt chấm công - ĐÃ TỐI ƯU: Database commit trước, Google Sheet background"""
    import sys
    from datetime import datetime as dt
    
    # Log bắt đầu
    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    try:
        print(f"\n{'='*80}", flush=True, file=sys.stderr)
        print(f"🚀 [APPROVE_START] {timestamp} - Bắt đầu phê duyệt attendance ID: {attendance_id}", flush=True, file=sys.stderr)
        print(f"{'='*80}", flush=True, file=sys.stderr)
    except Exception:
        pass
    
    # Kiểm tra session
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    
    # Lấy user
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    current_role = session.get('current_role', user.roles.split(',')[0])
    if current_role not in ['TEAM_LEADER', 'MANAGER', 'ADMIN']:
        return jsonify({'error': 'Bạn không có quyền phê duyệt chấm công'}), 403
    
    # Kiểm tra quyền
    has_permission, error_message = check_approval_permission(user.id, attendance_id, current_role)
    if not has_permission:
        return jsonify({'error': error_message}), 403
    
    # Lấy dữ liệu
    data = request.get_json()
    action = data.get('action')
    reason = validate_reason(data.get('reason', '')) if data.get('action') == 'reject' else ''
    approver_signature = data.get('signature')
    
    if action not in ['approve', 'reject']:
        return jsonify({'error': 'Hành động không hợp lệ'}), 400
    
    # Lấy attendance
    attendance = db.session.get(Attendance, attendance_id)
    if not attendance:
        return jsonify({'error': 'Không tìm thấy bản ghi chấm công'}), 404
    
    try:
        # XỬ LÝ PHÊ DUYỆT
        if action == 'approve':
            # Xử lý chữ ký
            if current_role == 'ADMIN':
                approver_signature = None
                signature_type = 'admin_no_signature'
            else:
                if not user.has_personal_signature():
                    return jsonify({
                        'error': 'Bạn chưa có chữ ký cá nhân. Vui lòng thiết lập chữ ký trong phần Cài đặt trước khi phê duyệt.',
                        'redirect_to_settings': True
                    }), 400
                
                if user.has_personal_signature():
                    approver_signature = user.personal_signature
                    signature_type = 'personal_signature'
                else:
                    session_signature, session_meta = signature_manager.get_signature_from_session(user.id, current_role)
                    if session_signature and signature_manager.should_use_session_signature(user.id, current_role):
                        approver_signature = session_signature
                        signature_type = 'session_reused'
                    else:
                        db_signature = signature_manager.get_signature_from_database(user.id, current_role, attendance_id)
                        if db_signature:
                            approver_signature = db_signature
                            signature_type = 'database_reused'
                        elif approver_signature:
                            signature_type = 'new'
                        else:
                            return jsonify({'error': 'Chữ ký là bắt buộc khi phê duyệt. Vui lòng ký tên để xác nhận.'}), 400
            
            old_status = attendance.status
            
            # Cập nhật theo vai trò
            if current_role == 'TEAM_LEADER':
                if attendance.status != 'pending':
                    return jsonify({'error': 'Bản ghi không ở trạng thái chờ duyệt'}), 400
                attendance.status = 'pending_manager'
                attendance.approved_by = user.id
                attendance.approved_at = datetime.now()
                if approver_signature:
                    attendance.team_leader_signature = approver_signature
                attendance.team_leader_signer_id = user.id
                message = 'Đã chuyển lên Quản lý phê duyệt'
                
            elif current_role == 'MANAGER':
                if attendance.status != 'pending_manager':
                    return jsonify({'error': 'Bản ghi chưa được Trưởng nhóm phê duyệt'}), 400
                attendance.status = 'pending_admin'
                attendance.approved_by = user.id
                attendance.approved_at = datetime.now()
                if approver_signature:
                    attendance.manager_signature = approver_signature
                attendance.manager_signer_id = user.id
                message = 'Đã phê duyệt thành công'
                
            elif current_role == 'ADMIN':
                if attendance.status not in ['pending_manager', 'pending_admin']:
                    return jsonify({'error': 'Bản ghi chưa được cấp dưới phê duyệt'}), 400
                
                # Check Google API token trước khi ADMIN approve
                token_status = check_google_token_status()
                if not token_status.get('can_approve', False):
                    # Publish notification to all admins
                    publish_token_status('expired', token_status.get('message', 'Token hết hạn'), needs_reauth=True)
                    return jsonify({
                        'error': f"⚠️ Token Google API hết hạn. {token_status.get('message', 'Vui lòng refresh token trước khi phê duyệt.')}",
                        'error_code': 'token_expired',
                        'needs_reauth': True
                    }), 503
                
                attendance.status = 'approved'
                attendance.approved = True
                attendance.approved_by = user.id
                attendance.approved_at = datetime.now()
                message = 'Phê duyệt hoàn tất'
                
                # ===== DATABASE COMMIT TRƯỚC KHI XỬ LÝ GOOGLE SHEET =====
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                try:
                    print(f"💾 [DB_COMMIT] {timestamp} - Đang commit database...", flush=True, file=sys.stderr)
                except Exception:
                    pass
                
                try:
                    db.session.commit()
                    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    try:
                        print(f"✅ [DB_COMMIT_SUCCESS] {timestamp} - Database đã được commit", flush=True, file=sys.stderr)
                    except Exception:
                        pass
                except Exception as commit_error:
                    db.session.rollback()
                    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    try:
                        print(f"❌ [DB_COMMIT_ERROR] {timestamp} - Lỗi commit: {str(commit_error)}", flush=True, file=sys.stderr)
                    except Exception:
                        pass
                    return jsonify({'error': 'Lỗi lưu database'}), 500
                
                # ===== CHUẨN BỊ DỮ LIỆU CHO GOOGLE SHEET =====
                employee_team_for_thread = attendance.user.department if attendance.user else "Unknown"
                employee_id_for_thread = attendance.user.employee_id if attendance.user else None
                
                if not employee_id_for_thread:
                    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    try:
                        print(f"⚠️ [SHEET_SKIP] {timestamp} - Không có employee_id, bỏ qua Google Sheet", flush=True, file=sys.stderr)
                    except Exception:
                        pass
                    return jsonify({'message': message})
                
                # Chuẩn bị dữ liệu
                break_time_value = attendance._format_hours_minutes(attendance.break_time) if attendance.break_time else '0:00'
                note_value = attendance.note if attendance.note else ''
                
                comp_time_regular_value = attendance._format_minutes_to_hhmm(attendance.comp_time_regular_minutes)
                comp_time_overtime_value = attendance._format_minutes_to_hhmm(attendance.comp_time_overtime_minutes)
                comp_time_ot_before_22_value = attendance._format_minutes_to_hhmm(attendance.comp_time_ot_before_22_minutes)
                comp_time_ot_after_22_value = attendance._format_minutes_to_hhmm(attendance.comp_time_ot_after_22_minutes)
                overtime_comp_time_value = attendance._format_minutes_to_hhmm(attendance.overtime_comp_time_minutes)
                
                overtime_before_22_val = attendance.overtime_before_22 or '0:00'
                overtime_after_22_val = attendance.overtime_after_22 or '0:00'
                
                # Tính tổng đối ứng
                def hhmm_to_minutes_safe(v):
                    try:
                        if not v or v in ['0', '0:00']:
                            return 0
                        if isinstance(v, str) and ':' in v:
                            h, m = v.split(':', 1)
                            return int(h or '0') * 60 + int(m or '0')
                    except Exception:
                        pass
                    return 0
                
                total_comp_minutes = (
                    hhmm_to_minutes_safe(comp_time_regular_value) +
                    hhmm_to_minutes_safe(comp_time_ot_before_22_value) +
                    hhmm_to_minutes_safe(comp_time_ot_after_22_value) +
                    hhmm_to_minutes_safe(comp_time_overtime_value) +
                    hhmm_to_minutes_safe(overtime_comp_time_value)
                )
                total_comp_display = f"{total_comp_minutes // 60}:{total_comp_minutes % 60:02d}"
                
                doi_ung_parts = []
                if comp_time_regular_value not in [None, '', 0, '0', '0:00']:
                    doi_ung_parts.append(f"Bù giờ thường: {comp_time_regular_value}")
                if comp_time_overtime_value not in [None, '', 0, '0', '0:00']:
                    doi_ung_parts.append(f"Bù giờ tăng ca: {comp_time_overtime_value}")
                if comp_time_ot_before_22_value not in [None, '', 0, '0', '0:00']:
                    doi_ung_parts.append(f"Bù OT <22h: {comp_time_ot_before_22_value}")
                if comp_time_ot_after_22_value not in [None, '', 0, '0', '0:00']:
                    doi_ung_parts.append(f"Bù OT >22h: {comp_time_ot_after_22_value}")
                if overtime_comp_time_value not in [None, '', 0, '0', '0:00']:
                    doi_ung_parts.append(f"Đối ứng OT: {overtime_comp_time_value}")
                
                doi_ung_summary = f"{total_comp_display} [ " + ' | '.join(doi_ung_parts) + " ]" if doi_ung_parts else total_comp_display
                
                regular_work_display = attendance._format_hours_minutes(attendance.calculate_regular_work_hours())
                total_hours_value = getattr(attendance, 'total_hours', None) or getattr(attendance, 'total_work_hours', '')
                
                def to_hhmm_from_decimal(hours_val):
                    try:
                        if hours_val is None or hours_val == '':
                            return ''
                        if isinstance(hours_val, str):
                            if ':' in hours_val:
                                return hours_val
                            hours_float = float(hours_val)
                        else:
                            hours_float = float(hours_val)
                        total_minutes = int(round(hours_float * 60))
                        return f"{total_minutes // 60}:{total_minutes % 60:02d}"
                    except Exception:
                        return str(hours_val)
                
                total_hours_display = to_hhmm_from_decimal(total_hours_value)
                
                attendance_data = {
                    'id': attendance.id,
                    'user_name': attendance.user.name if attendance.user else 'Unknown',
                    'date': attendance.date.strftime('%Y-%m-%d') if attendance.date else '',
                    'check_in': attendance.check_in.strftime('%H:%M') if attendance.check_in else '',
                    'check_out': attendance.check_out.strftime('%H:%M') if attendance.check_out else '',
                    'total_hours': total_hours_display,
                    'regular_work_hours': regular_work_display,
                    'break_time': break_time_value,
                    'overtime_before_22': overtime_before_22_val,
                    'overtime_after_22': overtime_after_22_val,
                    'comp_time_regular': comp_time_regular_value,
                    'comp_time_overtime': comp_time_overtime_value,
                    'comp_time_ot_before_22': comp_time_ot_before_22_value,
                    'comp_time_ot_after_22': comp_time_ot_after_22_value,
                    'overtime_comp_time': overtime_comp_time_value,
                    'note': note_value,
                    'doi_ung': doi_ung_summary,
                    'doi_ung_total': total_comp_display,
                    'status': attendance.status,
                    'approved_by': user.name,
                    'approved_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # ===== CHẠY GOOGLE SHEET UPDATE TRONG BACKGROUND =====
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                print(f"🚀 [BACKGROUND_START] {timestamp} - Khởi động background thread cho Google Sheet")
                sys.stdout.flush()
                
                thread = threading.Thread(
                    target=update_google_sheet_background_safe,
                    args=(attendance_id, employee_team_for_thread, employee_id_for_thread, attendance_data),
                    daemon=True
                )
                thread.start()
                
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                print(f"✅ [BACKGROUND_STARTED] {timestamp} - Background thread đã được khởi động")
                sys.stdout.flush()
            
            # Log audit
            log_audit_action(
                user_id=user.id,
                action='APPROVE_ATTENDANCE',
                table_name='attendances',
                record_id=attendance_id,
                old_values={'status': old_status},
                new_values={'status': attendance.status, 'approved_by': user.id, 'approved_at': attendance.approved_at.isoformat()}
            )
            
            # Log chữ ký nếu có
            if approver_signature and current_role != 'ADMIN':
                signature_manager.log_signature_action(
                    user_id=user.id,
                    action='APPROVAL',
                    signature_type=signature_type if 'signature_type' in locals() else 'new',
                    attendance_id=attendance_id,
                    additional_data={
                        'approver_role': current_role,
                        'approver_name': user.name,
                        'approval_status': attendance.status
                    }
                )
        
        else:  # reject
            old_status = attendance.status
            attendance.status = 'rejected'
            attendance.note = f"Bị từ chối bởi {current_role}: {reason}"
            message = 'Từ chối thành công'
            
            log_audit_action(
                user_id=user.id,
                action='REJECT_ATTENDANCE',
                table_name='attendances',
                record_id=attendance_id,
                old_values={'status': old_status},
                new_values={'status': attendance.status, 'reason': reason}
            )
        
        # Commit cho TEAM_LEADER và MANAGER (ADMIN đã commit trước đó)
        if current_role in ['TEAM_LEADER', 'MANAGER']:
            try:
                db.session.commit()
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                try:
                    print(f"✅ [{current_role}_COMMIT] {timestamp} - Database committed", flush=True, file=sys.stderr)
                except Exception:
                    pass
            except Exception as e:
                db.session.rollback()
                timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                try:
                    print(f"❌ [{current_role}_COMMIT_ERROR] {timestamp} - Error: {e}", flush=True, file=sys.stderr)
                except Exception:
                    pass
                return jsonify({'error': 'Lỗi lưu database'}), 500
        
        timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        try:
            print(f"\n{'='*80}", flush=True, file=sys.stderr)
            print(f"✅ [APPROVE_SUCCESS] {timestamp} - Phê duyệt thành công!", flush=True, file=sys.stderr)
            print(f"   User: {user.name} ({current_role})", flush=True, file=sys.stderr)
            print(f"   Attendance ID: {attendance_id}", flush=True, file=sys.stderr)
            print(f"   New Status: {attendance.status}", flush=True, file=sys.stderr)
            print(f"{'='*80}\n", flush=True, file=sys.stderr)
        except Exception:
            pass
        
        return jsonify({'message': message})
        
    except ValidationError as ve:
        db.session.rollback()
        timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        try:
            print(f"❌ [VALIDATION_ERROR] {timestamp} - {ve.message}", flush=True, file=sys.stderr)
        except Exception:
            pass
        return jsonify({'error': ve.message}), 400
        
    except SQLAlchemyError as se:
        db.session.rollback()
        timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        try:
            print(f"❌ [DB_ERROR] {timestamp} - {str(se)}", flush=True, file=sys.stderr)
        except Exception:
            pass
        return jsonify({'error': 'Lỗi cơ sở dữ liệu'}), 500
        
    except Exception as e:
        db.session.rollback()
        timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        try:
            import traceback
            print(f"\n{'='*80}", flush=True, file=sys.stderr)
            print(f"❌ [APPROVE_ERROR] {timestamp} - Lỗi không mong muốn", flush=True, file=sys.stderr)
            print(f"   Error: {str(e)}", flush=True, file=sys.stderr)
            print(f"   Type: {type(e).__name__}", flush=True, file=sys.stderr)
            print(f"   Traceback:", flush=True, file=sys.stderr)
            print(traceback.format_exc(), flush=True, file=sys.stderr)
            print(f"{'='*80}\n", flush=True, file=sys.stderr)
        except Exception:
            pass
        return jsonify({'error': f'Lỗi hệ thống: {str(e)}'}), 500

@app.route('/test-google-api')
@require_admin
def test_google_api():
    """Test endpoint để kiểm tra Google API"""
    try:
        google_api = GoogleDriveAPI()
        
        result = {
            'sheets_service_available': bool(google_api.sheets_service),
            'drive_service_available': bool(google_api.drive_service),
            'creds_exists': bool(google_api.creds),
            'creds_valid': google_api.creds.valid if google_api.creds else False,
            'creds_expired': google_api.creds.expired if google_api.creds else None,
            'has_refresh_token': bool(google_api.creds.refresh_token) if google_api.creds else False
        }
        
        # Test list files
        try:
            if google_api.drive_service:
                results = google_api.drive_service.files().list(pageSize=1).execute()
                result['can_list_files'] = True
                result['test_list_success'] = True
            else:
                result['can_list_files'] = False
                result['test_list_success'] = False
        except Exception as e:
            result['can_list_files'] = False
            result['test_list_error'] = str(e)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'error_type': type(e).__name__
        }), 500


def format_hours_minutes(hours):
    try:
        if hours is None:
            return "0:00"
        # Nếu là chuỗi số, chuyển sang float
        if isinstance(hours, str):
            hours = float(hours)
        if hours != hours or hours < 0:  # kiểm tra NaN hoặc âm
            return "0:00"
        h = int(hours)
        m = int(round((hours - h) * 60))
        if m == 60:
            h += 1
            m = 0
        return f"{h}:{m:02d}"
    except Exception:
        return "0:00"

def translate_holiday_type(holiday_type_en):
    """Translates holiday type from English to Vietnamese."""
    if not holiday_type_en:
        return '-'
    translations = {
        'normal': 'Ngày thường',
        'weekend': 'Cuối tuần',
        'vietnamese_holiday': 'Lễ Việt Nam',
        'japanese_holiday': 'Lễ Nhật Bản'
    }
    return translations.get(holiday_type_en, holiday_type_en)

@app.route('/api/attendance/pending')
def get_pending_attendance():
    if 'user_id' not in session:
        return jsonify({'total': 0, 'page': 1, 'per_page': 10, 'data': []})
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'total': 0, 'page': 1, 'per_page': 10, 'data': []})
    current_role = session.get('current_role', user.roles.split(',')[0])
    page = validate_int(request.args.get('page', 1), min_val=1)
    per_page = validate_int(request.args.get('per_page', 10), min_val=1, max_val=100)
    search = validate_input_sanitize(request.args.get('search', '').strip())
    department = validate_input_sanitize(request.args.get('department', '').strip())
    date_from = validate_date(request.args.get('date_from', '').strip()) if request.args.get('date_from') else None
    date_to = validate_date(request.args.get('date_to', '').strip()) if request.args.get('date_to') else None
    force_refresh = request.args.get('force_refresh', '0') == '1'
    requested_role = request.args.get('role', '')
    
    if page is None or per_page is None:
        return jsonify({'error': 'Tham số phân trang không hợp lệ'}), 400
    
    # Add real-time data freshness check
    from utils.realtime_updates import check_data_freshness, invalidate_role_cache
    freshness = check_data_freshness(user.id, current_role)
    
    # Force refresh if requested or role mismatch
    if force_refresh or (requested_role and requested_role != current_role):
        freshness['needs_refresh'] = True
        freshness['force_refresh'] = True
    
    # Use optimized query for fastest performance
    from utils.query_optimizer import optimize_pending_attendance_query
    
    # Debug logging để kiểm tra vấn đề
    print(f"🔍 [PENDING_API] User: {user.name}, ID: {user.id}", flush=True)
    print(f"🔍 [PENDING_API] Role: {current_role}, Department: [{user.department}]", flush=True)
    
    records, total = optimize_pending_attendance_query(
        current_role=current_role, user=user, search=search, department=department,
        date_from=date_from, date_to=date_to, page=page, per_page=per_page
    )
    
    print(f"🔍 [PENDING_API] Found {total} records for role {current_role}", flush=True)
    
    result = []
    for att in records:
        result.append({
            'id': att.id,
            'date': att.date.strftime('%d/%m/%Y'),
            'check_in': att.check_in.strftime('%H:%M') if att.check_in else None,
            'check_out': att.check_out.strftime('%H:%M') if att.check_out else None,
            'break_time': att._format_hours_minutes(att.break_time),
            'comp_time_regular': att._format_minutes_to_hhmm(att.comp_time_regular_minutes),
            'comp_time_overtime': att._format_minutes_to_hhmm(att.comp_time_overtime_minutes),
            'comp_time_ot_before_22': att._format_minutes_to_hhmm(att.comp_time_ot_before_22_minutes),
            'comp_time_ot_after_22': att._format_minutes_to_hhmm(att.comp_time_ot_after_22_minutes),
            'overtime_comp_time': att._format_minutes_to_hhmm(att.overtime_comp_time_minutes),
            'total_work_hours': att._format_hours_minutes(att.total_work_hours) if att.total_work_hours is not None else "0:00",
            'work_hours_display': att._format_hours_minutes(att.calculate_regular_work_hours()),
            'overtime_before_22': att.overtime_before_22,
            'overtime_after_22': att.overtime_after_22,
            'holiday_type': translate_holiday_type(att.holiday_type),
            'user_name': att.user.name if att.user else '',
            'department': att.user.department if att.user else '',
            'note': att.note,
            'status': att.status,
            'approved': att.approved,
            'signature': att.signature,
            'team_leader_signature': att.team_leader_signature,
            'manager_signature': att.manager_signature
        })
    resp = jsonify({
        'total': total,
        'page': page,
        'per_page': per_page,
        'data': result,
        'freshness': freshness  # Include real-time freshness data
    })
    # Disable caching to ensure fresh data after role switch
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/api/attendance/debug/status')
def debug_attendance_status():
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    current_role = session.get('current_role', user.roles.split(',')[0])
    if current_role != 'ADMIN':
        return jsonify({'error': 'Chỉ ADMIN mới có thể truy cập endpoint này'}), 403
    if not has_role(session['user_id'], 'ADMIN'):
        return jsonify({'error': 'Bạn không có quyền truy cập debug endpoint'}), 403
    all_statuses = db.session.query(Attendance.status).distinct().all()
    status_counts = {}
    for status in ['pending', 'pending_manager', 'pending_admin', 'approved', 'rejected']:
        count = Attendance.query.filter_by(status=status).count()
        status_counts[status] = count
    sample_records = {}
    for status in ['pending', 'pending_manager', 'pending_admin']:
        records = Attendance.query.options(joinedload(Attendance.user)).filter_by(status=status).limit(5).all()
        sample_records[status] = [
            {
                'id': r.id,
                'user_id': r.user_id,
                'date': r.date.strftime('%d/%m/%Y'),
                'status': r.status,
                'approved': r.approved,
                'user_name': r.user.name if r.user else 'Unknown'
            }
            for r in records
        ]
    return jsonify({
        'all_statuses': [s[0] for s in all_statuses],
        'status_counts': status_counts,
        'sample_records': sample_records
    })


@app.route('/api/attendance/debug/team-leader')
def debug_team_leader_attendance():
    """Debug endpoint for TEAM_LEADER to check department and pending attendance"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    update_session_activity()
    
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    current_role = session.get('current_role', user.roles.split(',')[0])
    
    # Thông tin user hiện tại
    user_info = {
        'id': user.id,
        'name': user.name,
        'employee_id': user.employee_id,
        'department': user.department,
        'department_upper': (user.department or '').strip().upper(),
        'role': current_role
    }
    
    # Lấy tất cả phòng ban trong database
    all_departments = db.session.query(User.department, func.count(User.id)).filter(
        User.is_deleted == False
    ).group_by(User.department).all()
    
    dept_list = [{'name': d[0], 'upper': (d[0] or '').strip().upper(), 'count': d[1]} for d in all_departments]
    
    # Tìm nhân viên cùng phòng ban (exact match)
    user_dept = user.department
    exact_match_employees = User.query.filter(
        User.department == user_dept,
        User.is_deleted == False
    ).all()
    
    # Tìm nhân viên cùng phòng ban (case-insensitive)
    user_dept_upper = (user.department or '').strip().upper()
    case_insensitive_employees = User.query.filter(
        func.upper(func.trim(User.department)) == user_dept_upper,
        User.is_deleted == False
    ).all()
    
    # Lấy tất cả attendance có status='pending'
    all_pending = Attendance.query.filter(Attendance.status == 'pending').all()
    
    pending_list = []
    for att in all_pending:
        emp = db.session.get(User, att.user_id)
        pending_list.append({
            'id': att.id,
            'user_id': att.user_id,
            'user_name': emp.name if emp else 'Unknown',
            'user_dept': emp.department if emp else 'Unknown',
            'user_dept_upper': (emp.department or '').strip().upper() if emp else 'Unknown',
            'date': att.date.strftime('%Y-%m-%d'),
            'status': att.status,
            'matches_exact': (emp.department == user_dept) if emp else False,
            'matches_case_insensitive': ((emp.department or '').strip().upper() == user_dept_upper) if emp else False
        })
    
    # Lấy pending attendance của nhân viên cùng phòng ban (case-insensitive)
    team_user_ids = [u.id for u in case_insensitive_employees]
    team_pending = Attendance.query.filter(
        Attendance.status == 'pending',
        Attendance.user_id.in_(team_user_ids)
    ).all() if team_user_ids else []
    
    return jsonify({
        'current_user': user_info,
        'all_departments': dept_list,
        'exact_match_employees': [
            {'id': u.id, 'name': u.name, 'department': u.department}
            for u in exact_match_employees
        ],
        'case_insensitive_employees': [
            {'id': u.id, 'name': u.name, 'department': u.department}
            for u in case_insensitive_employees
        ],
        'all_pending_records': pending_list,
        'team_pending_count': len(team_pending),
        'team_pending_records': [
            {
                'id': att.id,
                'user_id': att.user_id,
                'date': att.date.strftime('%Y-%m-%d'),
                'status': att.status
            }
            for att in team_pending
        ]
    })












# Exempt certain API endpoints from CSRF protection if needed
# GET endpoints don't need CSRF protection
try:
    csrf.exempt(app.view_functions['get_attendance'])
    csrf.exempt(app.view_functions['get_attendance_history'])
    csrf.exempt(app.view_functions['get_pending_attendance'])
    csrf.exempt(app.view_functions['debug_attendance_status'])
    # Temporarily exempt signature APIs for testing
    csrf.exempt(app.view_functions['check_signature_status'])
    csrf.exempt(app.view_functions['save_signature_to_session'])
    csrf.exempt(app.view_functions['clear_session_signature'])
except KeyError:
    pass  # Routes might not exist yet


@app.route('/admin/users/<int:user_id>/reset_password', methods=['POST'])
@require_admin
def admin_reset_user_password(user_id):
    """Admin reset password cho user"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    
    update_session_activity()
    
    current_user = db.session.get(User, session['user_id'])
    if not current_user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    if 'ADMIN' not in current_user.roles.split(','):
        return jsonify({'error': 'Chỉ quản trị viên mới có quyền thực hiện hành động này'}), 403
    
    try:
        data = request.get_json()
        if not data or 'password' not in data:
            return jsonify({'error': 'Thiếu mật khẩu mới'}), 400
        
        new_password = data.get('password', '').strip()
        if len(new_password) < 6:
            return jsonify({'error': 'Mật khẩu phải có ít nhất 6 ký tự'}), 400
        
        target_user = db.session.get(User, user_id)
        if not target_user:
            return jsonify({'error': 'Không tìm thấy người dùng'}), 404
        
        # Lưu giá trị cũ để log
        old_password_hash = target_user.password_hash
        
        # Đặt mật khẩu mới
        target_user.set_password(new_password)
        
        db.session.commit()
        
        # Log audit
        log_audit_action(
            user_id=current_user.id,
            action='ADMIN_RESET_PASSWORD',
            table_name='users',
            record_id=user_id,
            old_values={'password': '***'},
            new_values={'password': '***', 'reset_by': current_user.id}
        )
        
        return jsonify({
            'success': True,
            'message': f'Đã đặt lại mật khẩu cho {target_user.name} (Mã NV: {target_user.employee_id})'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi đặt lại mật khẩu: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Đã xảy ra lỗi: {str(e)}'}), 500

@app.route('/admin/users/<int:user_id>/soft_delete', methods=['POST'])
@require_admin
def soft_delete_user(user_id):
    """Soft delete user - set is_deleted to True"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    # Không cho phép xóa chính mình
    if user.id == session['user_id']:
        return jsonify({'error': 'Không thể xóa tài khoản của chính mình'}), 400
    
    try:
        # Soft delete user
        user.soft_delete()
        db.session.commit()
        
        # Log the action
        log_audit_action(
            user_id=session['user_id'],
            action='SOFT_DELETE_USER',
            table_name='users',
            record_id=user_id,
            old_values={'is_deleted': False, 'is_active': True},
            new_values={'is_deleted': True, 'is_active': False}
        )
        
        return jsonify({
            'success': True,
            'message': f'Đã xóa người dùng {user.name} thành công'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error soft deleting user: {str(e)}")
        return jsonify({'error': 'Đã xảy ra lỗi khi xóa người dùng'}), 500

@app.route('/admin/users/<int:user_id>/restore', methods=['POST'])
@require_admin
def restore_user(user_id):
    """Restore soft deleted user"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    if not user.is_soft_deleted():
        return jsonify({'error': 'Người dùng này chưa bị xóa'}), 400
    
    try:
        # Restore user
        user.restore()
        db.session.commit()
        
        # Log the action
        log_audit_action(
            user_id=session['user_id'],
            action='RESTORE_USER',
            table_name='users',
            record_id=user_id,
            old_values={'is_deleted': True, 'is_active': False},
            new_values={'is_deleted': False, 'is_active': True}
        )
        
        return jsonify({
            'success': True,
            'message': f'Đã khôi phục người dùng {user.name} thành công'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error restoring user: {str(e)}")
        return jsonify({'error': 'Đã xảy ra lỗi khi khôi phục người dùng'}), 500

@app.route('/admin/yearly-reset/status', methods=['GET'])
@require_admin
def admin_yearly_reset_status():
    """Kiểm tra trạng thái reset lịch hàng năm và dữ liệu tháng 12"""
    try:
        if not session.get('user_id'):
            return jsonify({'error': 'Chưa đăng nhập'}), 401
        
        user = User.query.get(session['user_id'])
        if not user or not user.has_role('ADMIN'):
            return jsonify({'error': 'Không có quyền'}), 403
        
        current_year = datetime.now().year
        next_year = current_year + 1
        next_jan_1 = datetime(next_year, 1, 1)
        days_until_reset = (next_jan_1 - datetime.now()).days
        
        # Kiểm tra dữ liệu tháng 12
        check_result = check_december_data_complete(current_year)
        
        # Kiểm tra năm đã reset gần nhất
        global _last_reset_year
        last_reset_info = f"Năm {_last_reset_year}" if _last_reset_year else "Chưa có"
        
        return jsonify({
            'current_year': current_year,
            'next_reset_date': f"1/1/{next_year}",
            'days_until_reset': days_until_reset,
            'last_reset_year': _last_reset_year,
            'december_data': check_result,
            'status': 'ready' if days_until_reset > 0 else 'reset_pending'
        }), 200
    except Exception as e:
        return jsonify({'error': f'Lỗi: {str(e)}'}), 500

@app.route('/admin/yearly-reset/manual', methods=['POST'])
@require_admin
def admin_manual_yearly_reset():
    """Reset lịch hàng năm thủ công (chỉ dùng trong trường hợp đặc biệt)"""
    try:
        if not session.get('user_id'):
            return jsonify({'error': 'Chưa đăng nhập'}), 401
        
        user = User.query.get(session['user_id'])
        if not user or not user.has_role('ADMIN'):
            return jsonify({'error': 'Không có quyền'}), 403
        
        # Kiểm tra xác nhận
        confirm = request.json.get('confirm', False)
        if not confirm:
            return jsonify({'error': 'Cần xác nhận để thực hiện reset'}), 400
        
        # Thực hiện reset
        success, deleted_count = reset_yearly_schedule()
        
        if success:
            global _last_reset_year
            _last_reset_year = datetime.now().year
            
            log_audit_action(
                user_id=user.id,
                action='MANUAL_YEARLY_RESET',
                table_name='attendances',
                record_id=None,
                old_values={},
                new_values={'deleted_count': deleted_count, 'reset_by': user.name}
            )
            
            return jsonify({
                'message': f'Đã reset lịch thành công. Đã xóa {deleted_count} bản ghi.',
                'deleted_count': deleted_count
            }), 200
        else:
            return jsonify({'error': 'Lỗi khi reset lịch'}), 500
    except Exception as e:
        return jsonify({'error': f'Lỗi: {str(e)}'}), 500

@app.route('/admin/users/deleted/delete-all', methods=['POST'])
@require_admin
def delete_all_deleted_users():
    """Xóa vĩnh viễn tất cả người dùng đã bị soft delete - yêu cầu xác minh mật khẩu"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    try:
        # Lấy dữ liệu từ request
        data = request.get_json()
        password = data.get('password', '').strip() if data else ''
        
        if not password:
            return jsonify({'error': 'Vui lòng nhập mật khẩu để xác nhận'}), 400
        
        # Lấy current user và xác minh mật khẩu
        current_user = db.session.get(User, session['user_id'])
        if not current_user:
            return jsonify({'error': 'Không tìm thấy người dùng hiện tại'}), 404
        
        if not current_user.check_password(password):
            return jsonify({'error': 'Mật khẩu không đúng. Vui lòng thử lại.'}), 401
        
        # Lấy tất cả users đã bị soft delete
        deleted_users = User.query.filter_by(is_deleted=True).all()
        total_count = len(deleted_users)
        
        if total_count == 0:
            return jsonify({'error': 'Không có người dùng nào để xóa'}), 400
        
        # Lưu thông tin để audit log
        deleted_user_ids = [user.id for user in deleted_users]
        deleted_user_info = [
            {
                'employee_id': user.employee_id,
                'name': user.name,
                'email': user.email,
                'department': user.department
            }
            for user in deleted_users
        ]
        
        # Xóa vĩnh viễn tất cả users (hard delete)
        for user in deleted_users:
            db.session.delete(user)
        
        db.session.commit()
        
        # Log the action
        log_audit_action(
            user_id=session['user_id'],
            action='DELETE_ALL_DELETED_USERS',
            table_name='users',
            record_id=None,
            old_values={'deleted_users': deleted_user_info},
            new_values={'deleted_count': total_count}
        )
        
        return jsonify({
            'success': True,
            'message': f'Đã xóa vĩnh viễn {total_count} người dùng đã xóa'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi xóa tất cả người dùng đã xóa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi khi xóa tất cả người dùng: {str(e)}'}), 500

@app.route('/admin/system/clear-all-data', methods=['POST'])
@require_admin
def clear_all_data():
    """Xóa toàn bộ dữ liệu hệ thống, chỉ giữ lại admin Nguyễn Công Đạt"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    try:
        # Lấy dữ liệu từ request
        data = request.get_json()
        password = data.get('password', '').strip() if data else ''
        
        if not password:
            return jsonify({'error': 'Vui lòng nhập mật khẩu để xác nhận'}), 400
        
        # Lấy current user và xác minh mật khẩu
        current_user = db.session.get(User, session['user_id'])
        if not current_user:
            return jsonify({'error': 'Không tìm thấy người dùng hiện tại'}), 404
        
        if not current_user.check_password(password):
            return jsonify({'error': 'Mật khẩu không đúng. Vui lòng thử lại.'}), 401
        
        # Tìm admin Nguyễn Công Đạt để giữ lại (case-insensitive)
        all_users = User.query.all()
        admin_user = None
        for user in all_users:
            name_lower = user.name.lower()
            if 'nguyễn công đạt' in name_lower or 'nguyen cong dat' in name_lower:
                admin_user = user
                break
        
        if not admin_user:
            return jsonify({'error': 'Không tìm thấy admin Nguyễn Công Đạt. Không thể xóa dữ liệu.'}), 400
        
        # Đảm bảo admin này có quyền ADMIN
        if 'ADMIN' not in admin_user.get_roles_list():
            admin_user.roles = 'ADMIN'
            admin_user.is_active = True
            admin_user.is_deleted = False
        
        # Đếm số lượng dữ liệu trước khi xóa
        attendance_count = Attendance.query.count()
        leave_request_count = LeaveRequest.query.count()
        request_count = Request.query.count()
        audit_log_count = AuditLog.query.count()
        password_token_count = PasswordResetToken.query.count()
        user_count = User.query.count()
        
        # Xóa tất cả Attendance records
        Attendance.query.delete()
        
        # Xóa tất cả LeaveRequest records
        LeaveRequest.query.delete()
        
        # Xóa tất cả Request records
        Request.query.delete()
        
        # Xóa tất cả AuditLog records
        AuditLog.query.delete()
        
        # Xóa tất cả PasswordResetToken records
        PasswordResetToken.query.delete()
        
        # Xóa tất cả Users trừ admin Nguyễn Công Đạt
        User.query.filter(User.id != admin_user.id).delete()
        
        # Commit tất cả thay đổi
        db.session.commit()
        
        # Log action (sau khi commit để tránh lỗi)
        try:
            from utils.session import log_audit_action
            log_audit_action(
                user_id=current_user.id,
                action='CLEAR_ALL_SYSTEM_DATA',
                table_name='system',
                record_id=None,
                old_values={
                    'attendance_count': attendance_count,
                    'leave_request_count': leave_request_count,
                    'request_count': request_count,
                    'audit_log_count': audit_log_count,
                    'password_token_count': password_token_count,
                    'user_count': user_count
                },
                new_values={
                    'kept_admin': {
                        'id': admin_user.id,
                        'name': admin_user.name,
                        'employee_id': admin_user.employee_id,
                        'email': admin_user.email
                    },
                    'final_user_count': 1
                }
            )
        except Exception as log_err:
            print(f"Warning: Không thể log audit action: {log_err}")
        
        return jsonify({
            'success': True,
            'message': f'Đã xóa toàn bộ dữ liệu hệ thống thành công. Đã giữ lại admin: {admin_user.name} (Mã NV: {admin_user.employee_id})',
            'deleted_counts': {
                'attendances': attendance_count,
                'leave_requests': leave_request_count,
                'requests': request_count,
                'audit_logs': audit_log_count,
                'password_tokens': password_token_count,
                'users': user_count - 1  # Trừ admin được giữ lại
            },
            'kept_admin': {
                'name': admin_user.name,
                'employee_id': admin_user.employee_id,
                'email': admin_user.email
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi xóa toàn bộ dữ liệu: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi khi xóa dữ liệu: {str(e)}'}), 500

@app.route('/admin/system/clear-attendances', methods=['POST'])
@require_admin
def clear_all_attendances():
    """Xóa toàn bộ bản ghi chấm công, giữ nguyên thông tin nhân viên và nghỉ phép"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    try:
        # Lấy dữ liệu từ request
        data = request.get_json()
        password = data.get('password', '').strip() if data else ''
        
        if not password:
            return jsonify({'error': 'Vui lòng nhập mật khẩu để xác nhận'}), 400
        
        # Lấy current user và xác minh mật khẩu
        current_user = db.session.get(User, session['user_id'])
        if not current_user:
            return jsonify({'error': 'Không tìm thấy người dùng hiện tại'}), 404
        
        if not current_user.check_password(password):
            return jsonify({'error': 'Mật khẩu không đúng. Vui lòng thử lại.'}), 401
        
        # Đếm số lượng dữ liệu trước khi xóa
        attendance_count = Attendance.query.count()
        
        # Xóa tất cả Attendance records
        Attendance.query.delete()
        
        # Commit thay đổi
        db.session.commit()
        
        # Log action
        try:
            audit_logger.audit_action(
                action='CLEAR_ALL_ATTENDANCES',
                table_name='attendances',
                record_id=None,
                old_values={'attendance_count': attendance_count},
                new_values={'cleared_by': current_user.name}
            )
        except Exception as log_err:
            print(f"Warning: Không thể log audit action: {log_err}")
        
        return jsonify({
            'success': True,
            'message': f'Đã xóa thành công {attendance_count} bản ghi chấm công. Thông tin nhân viên và nghỉ phép vẫn được giữ nguyên.',
            'deleted_count': attendance_count
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi xóa chấm công: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi khi xóa chấm công: {str(e)}'}), 500

@app.route('/admin/system/clear-leave-requests', methods=['POST'])
@require_admin
def clear_all_leave_requests():
    """Xóa toàn bộ đơn nghỉ phép, giữ nguyên thông tin nhân viên và chấm công"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    try:
        # Lấy dữ liệu từ request
        data = request.get_json()
        password = data.get('password', '').strip() if data else ''
        
        if not password:
            return jsonify({'error': 'Vui lòng nhập mật khẩu để xác nhận'}), 400
        
        # Lấy current user và xác minh mật khẩu
        current_user = db.session.get(User, session['user_id'])
        if not current_user:
            return jsonify({'error': 'Không tìm thấy người dùng hiện tại'}), 404
        
        if not current_user.check_password(password):
            return jsonify({'error': 'Mật khẩu không đúng. Vui lòng thử lại.'}), 401
        
        # Đếm số lượng dữ liệu trước khi xóa
        leave_request_count = LeaveRequest.query.count()
        request_count = Request.query.count()
        
        # Xóa tất cả LeaveRequest records
        LeaveRequest.query.delete()
        
        # Xóa tất cả Request records (legacy)
        Request.query.delete()
        
        # Commit thay đổi
        db.session.commit()
        
        # Log action
        try:
            audit_logger.audit_action(
                action='CLEAR_ALL_LEAVE_REQUESTS',
                table_name='leave_requests',
                record_id=None,
                old_values={'leave_request_count': leave_request_count, 'request_count': request_count},
                new_values={'cleared_by': current_user.name}
            )
        except Exception as log_err:
            print(f"Warning: Không thể log audit action: {log_err}")
        
        return jsonify({
            'success': True,
            'message': f'Đã xóa thành công {leave_request_count} đơn nghỉ phép và {request_count} request khác. Thông tin nhân viên và chấm công vẫn được giữ nguyên.',
            'deleted_counts': {
                'leave_requests': leave_request_count,
                'requests': request_count
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi xóa nghỉ phép: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi khi xóa nghỉ phép: {str(e)}'}), 500

@app.route('/admin/system/clear-records-only', methods=['POST'])
@require_admin
def clear_records_only():
    """Xóa toàn bộ chấm công và nghỉ phép, giữ nguyên thông tin nhân viên"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    try:
        # Lấy dữ liệu từ request
        data = request.get_json()
        password = data.get('password', '').strip() if data else ''
        
        if not password:
            return jsonify({'error': 'Vui lòng nhập mật khẩu để xác nhận'}), 400
        
        # Lấy current user và xác minh mật khẩu
        current_user = db.session.get(User, session['user_id'])
        if not current_user:
            return jsonify({'error': 'Không tìm thấy người dùng hiện tại'}), 404
        
        if not current_user.check_password(password):
            return jsonify({'error': 'Mật khẩu không đúng. Vui lòng thử lại.'}), 401
        
        # Đếm số lượng dữ liệu trước khi xóa
        attendance_count = Attendance.query.count()
        leave_request_count = LeaveRequest.query.count()
        request_count = Request.query.count()
        
        # Xóa tất cả Attendance records
        Attendance.query.delete()
        
        # Xóa tất cả LeaveRequest records
        LeaveRequest.query.delete()
        
        # Xóa tất cả Request records (legacy)
        Request.query.delete()
        
        # Commit thay đổi
        db.session.commit()
        
        # Log action
        try:
            audit_logger.audit_action(
                action='CLEAR_RECORDS_ONLY',
                table_name='attendances+leave_requests',
                record_id=None,
                old_values={
                    'attendance_count': attendance_count,
                    'leave_request_count': leave_request_count,
                    'request_count': request_count
                },
                new_values={'cleared_by': current_user.name}
            )
        except Exception as log_err:
            print(f"Warning: Không thể log audit action: {log_err}")
        
        return jsonify({
            'success': True,
            'message': f'Đã xóa thành công {attendance_count} bản ghi chấm công và {leave_request_count} đơn nghỉ phép. Thông tin nhân viên vẫn được giữ nguyên.',
            'deleted_counts': {
                'attendances': attendance_count,
                'leave_requests': leave_request_count,
                'requests': request_count
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi xóa bản ghi: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi khi xóa bản ghi: {str(e)}'}), 500

@app.route('/admin/users/deleted')
@require_admin
def admin_deleted_users():
    """Show soft deleted users"""
    page = request.args.get('page', 1, type=int)
    per_page = 10
    search = request.args.get('search', '', type=str).strip()
    department_filter = request.args.get('department', '', type=str).strip()

    query = User.query.filter_by(is_deleted=True)  # Chỉ hiển thị users đã bị soft delete
    if search:
        search_lower = search.lower().strip()
        # Tách từ khóa tìm kiếm thành các từ riêng lẻ
        search_words = search_lower.split()
        
        # Tạo điều kiện tìm kiếm cho từng từ
        name_conditions = []
        for word in search_words:
            name_conditions.append(func.lower(User.name).contains(word))
        
        # Tạo điều kiện tìm kiếm đơn giản - tìm theo từng từ riêng lẻ
        name_conditions = []
        for word in search_words:
            name_conditions.append(func.lower(User.name).contains(word))
        
        # Thêm điều kiện tìm kiếm theo mã nhân viên
        name_conditions.append(func.lower(func.cast(User.employee_id, db.String)).contains(search_lower))
        
        # Kết hợp tất cả điều kiện với OR
        query = query.filter(db.or_(*name_conditions))
    if department_filter:
        query = query.filter(User.department == department_filter)
    query = query.order_by(User.name.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    users = pagination.items

    # Lấy danh sách phòng ban từ bảng Department trong database
    from database.models import Department
    db_departments = Department.query.filter_by(is_active=True).order_by(Department.name).all()
    departments = [d.name for d in db_departments]

    # Calculate statistics
    deleted_count = len(users)
    
    # Tính toán phân trang
    start_page = max(1, pagination.page - 2)
    end_page = min(pagination.pages, pagination.page + 2)
    if end_page - start_page < 4:
        end_page = min(pagination.pages, start_page + 4)
        start_page = max(1, end_page - 4)
    page_range = range(start_page, end_page + 1)

    return render_template(
        'admin/deleted_users.html',
        users=users,
        deleted_count=deleted_count,
        pagination=pagination,
        search=search,
        departments=departments,
        department_filter=department_filter,
        per_page=per_page,
        page_range=page_range
    )

@app.route('/admin/users/upload', methods=['POST'])
@require_admin
def upload_users():
    """Upload nhân viên từ file TXT hoặc XLSX
    
    Format file TXT:
    - Mỗi dòng là 1 nhân viên
    - Format: Mã NV|Họ và Tên|Phòng Ban|Vai Trò|Email|Mật khẩu
    - Nhiều vai trò cách nhau bằng dấu phẩy: EMPLOYEE,TEAM_LEADER
    
    Format file XLSX:
    - Hàng đầu tiên có thể là header (sẽ bỏ qua)
    - Mỗi hàng là 1 nhân viên
    - Cột A: Mã NV
    - Cột B: Họ và Tên
    - Cột C: Phòng Ban
    - Cột D: Vai Trò
    - Cột E: Email (tùy chọn)
    - Cột F: Mật khẩu (tùy chọn)
    
    Ví dụ TXT:
    1395|Nguyễn Văn A|OFFICE|EMPLOYEE|email@dmi.com|123456
    1396|Trần Thị B|PRODUCTION|EMPLOYEE,TEAM_LEADER||
    """
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    
    update_session_activity()
    
    # Lấy user hiện tại
    current_user = db.session.get(User, session['user_id'])
    if not current_user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    # Kiểm tra quyền admin
    if 'ADMIN' not in current_user.roles.split(','):
        return jsonify({'error': 'Chỉ quản trị viên mới có quyền thực hiện hành động này'}), 403
    
    # Kiểm tra file
    if 'file' not in request.files:
        return jsonify({'error': 'Không có file được tải lên'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Chưa chọn file'}), 400
    
    # Kiểm tra định dạng file
    filename_lower = file.filename.lower()
    is_txt = filename_lower.endswith('.txt')
    is_xlsx = filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls')
    
    if not (is_txt or is_xlsx):
        return jsonify({'error': 'File phải có định dạng .txt hoặc .xlsx/.xls'}), 400
    
    try:
        # Đọc dữ liệu từ file
        rows_data = []
        
        if is_txt:
            # Đọc file TXT
            content = file.read()
            try:
                # Thử decode UTF-8
                text = content.decode('utf-8')
            except UnicodeDecodeError:
                try:
                    # Thử decode UTF-8 với BOM
                    text = content.decode('utf-8-sig')
                except UnicodeDecodeError:
                    # Fallback về Windows-1252 (thường dùng cho tiếng Việt)
                    text = content.decode('windows-1252')
            
            lines = text.strip().split('\n')
            for line in lines:
                line = line.strip()
                # Bỏ qua dòng trống và dòng comment
                if not line or line.startswith('#'):
                    continue
                # Parse dòng: Mã NV|Họ và Tên|Phòng Ban|Vai Trò|Email|Mật khẩu
                parts = line.split('|')
                rows_data.append(parts)
        else:
            # Đọc file Excel
            file.seek(0)  # Reset file pointer
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            
            # Tự động phát hiện thứ tự cột từ header (nếu có)
            column_mapping = None  # [idx_ma_nv, idx_ten, idx_phong_ban, idx_vai_tro, idx_email, idx_mat_khau]
            
            # Đọc hàng đầu tiên để phát hiện header
            first_row = None
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                first_row = [str(cell).strip().upper() if cell is not None else '' for cell in row]
                break
            
            # Phát hiện thứ tự cột từ header
            if first_row:
                idx_ma_nv = -1
                idx_ten = -1
                idx_phong_ban = -1
                idx_vai_tro = -1
                idx_email = -1
                idx_mat_khau = -1
                
                for i, cell in enumerate(first_row):
                    cell_upper = str(cell).upper()
                    # Tìm cột Mã NV
                    if idx_ma_nv == -1 and any(kw in cell_upper for kw in ['MÃ', 'MÃ NV', 'MÃ NHÂN VIÊN', 'EMPLOYEE ID', 'ID', 'MÃ NV']):
                        idx_ma_nv = i
                    # Tìm cột Họ và Tên
                    elif idx_ten == -1 and any(kw in cell_upper for kw in ['HỌ', 'TÊN', 'HỌ VÀ TÊN', 'NAME', 'TÊN NHÂN VIÊN']):
                        idx_ten = i
                    # Tìm cột Phòng Ban
                    elif idx_phong_ban == -1 and any(kw in cell_upper for kw in ['PHÒNG', 'BAN', 'PHÒNG BAN', 'DEPARTMENT', 'TEAM']):
                        idx_phong_ban = i
                    # Tìm cột Vai Trò
                    elif idx_vai_tro == -1 and any(kw in cell_upper for kw in ['VAI', 'TRÒ', 'VAI TRÒ', 'ROLE', 'ROLES']):
                        idx_vai_tro = i
                    # Tìm cột Email
                    elif idx_email == -1 and 'EMAIL' in cell_upper:
                        idx_email = i
                    # Tìm cột Mật khẩu
                    elif idx_mat_khau == -1 and any(kw in cell_upper for kw in ['MẬT', 'KHẨU', 'MẬT KHẨU', 'PASSWORD', 'PASS']):
                        idx_mat_khau = i
                
                # Nếu tìm thấy đủ các cột bắt buộc, tạo mapping
                if idx_ma_nv != -1 and idx_ten != -1 and idx_phong_ban != -1 and idx_vai_tro != -1:
                    column_mapping = [idx_ma_nv, idx_ten, idx_phong_ban, idx_vai_tro, idx_email, idx_mat_khau]
            
            # Đọc tất cả các hàng dữ liệu
            start_row = 2 if column_mapping else 1  # Bỏ qua header nếu có mapping
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                # Bỏ qua hàng trống
                if not any(cell for cell in row):
                    continue
                
                # Chuyển đổi giá trị None thành chuỗi rỗng và strip
                row_data = [str(cell).strip() if cell is not None else '' for cell in row]
                
                # Nếu có mapping, sắp xếp lại theo thứ tự chuẩn
                if column_mapping:
                    mapped_data = []
                    for idx in column_mapping:
                        if idx != -1 and idx < len(row_data):
                            mapped_data.append(row_data[idx])
                        else:
                            mapped_data.append('')
                    row_data = mapped_data
                else:
                    # Không có mapping, giả định thứ tự chuẩn: Mã NV | Họ và Tên | Phòng Ban | Vai Trò | Email | Mật khẩu
                    # Hoặc thử phát hiện: nếu cột đầu là số thì là Mã NV, nếu là chữ thì là Tên
                    if len(row_data) >= 2:
                        first_cell = row_data[0]
                        second_cell = row_data[1] if len(row_data) > 1 else ''
                        # Nếu cột đầu là số (Mã NV) và cột 2 là chữ (Tên) -> đúng thứ tự
                        # Nếu cột đầu là chữ (Tên) và cột 2 là số (Mã NV) -> đảo ngược
                        try:
                            int(first_cell)
                            # Cột đầu là số -> đúng thứ tự
                            row_data = row_data[:6]
                        except (ValueError, TypeError):
                            # Cột đầu không phải số -> đảo ngược cột 0 và 1
                            if len(row_data) >= 2:
                                row_data = [row_data[1], row_data[0]] + row_data[2:6]
                            else:
                                row_data = row_data[:6]
                    else:
                        row_data = row_data[:6]
                
                rows_data.append(row_data)
            
            wb.close()
        
        lines = rows_data
        
        # Parse từng dòng
        results = {
            'success': [],
            'errors': [],
            'skipped': [],
            'conflicts': []  # Thêm conflicts để hỏi user
        }
        
        default_password = request.form.get('default_password', '123456')  # Mật khẩu mặc định
        
        # Tối ưu: Load tất cả existing users và emails vào memory trước để tránh query trong vòng lặp
        all_users = User.query.filter_by(is_deleted=False).all()
        existing_users = {user.employee_id: user for user in all_users}
        # Load emails với normalize (lowercase và strip) để so sánh chính xác
        # Đảm bảo normalize giống hệt như khi kiểm tra trong vòng lặp
        existing_emails = {}
        print(f"[DEBUG] Loading {len(all_users)} existing users from database...")
        for user in all_users:
            if user.email:
                # Normalize email: strip whitespace, lowercase, và loại bỏ tất cả whitespace ẩn
                # Cách normalize này phải giống hệt với cách normalize khi kiểm tra trong vòng lặp
                email_normalized = str(user.email).strip().lower()
                # Loại bỏ tất cả các ký tự whitespace ẩn (space, tab, newline, etc.)
                email_normalized = ''.join(email_normalized.split())
                if email_normalized:
                    existing_emails[email_normalized] = user
                    # Debug: log để kiểm tra
                    print(f"[DEBUG] Loaded email: '{user.email}' (raw) -> '{email_normalized}' (normalized) for user {user.employee_id}")
        print(f"[DEBUG] Total normalized emails in cache: {len(existing_emails)}")
        
        # Track emails và employee_ids đã thêm trong batch này để tránh duplicate trong cùng file
        batch_emails = set()
        batch_employee_ids = set()
        
        for line_num, parts in enumerate(lines, 1):
            # parts đã là list từ việc parse TXT hoặc Excel
            
            # Lọc bỏ các phần tử rỗng ở cuối
            parts = [p for p in parts if p] if parts else []
            
            if len(parts) < 4:
                row_content = '|'.join(parts) if parts else '(dòng trống)'
                results['errors'].append({
                    'line': line_num,
                    'content': row_content,
                    'error': 'Định dạng không đúng. Cần: Mã NV|Họ và Tên|Phòng Ban|Vai Trò (hoặc thêm Email|Mật khẩu)'
                })
                continue
            
            employee_id_str = str(parts[0]).strip() if parts[0] else ''
            name = str(parts[1]).strip() if len(parts) > 1 and parts[1] else ''
            department = str(parts[2]).strip() if len(parts) > 2 and parts[2] else ''
            roles_str = str(parts[3]).strip() if len(parts) > 3 and parts[3] else ''
            
            # Email và mật khẩu (tùy chọn - nếu có 6 trường)
            # Ưu tiên: Mật khẩu từ file Excel > Mật khẩu mặc định từ form
            email = str(parts[4]).strip() if len(parts) > 4 and parts[4] else None
            password_from_file = str(parts[5]).strip() if len(parts) > 5 and parts[5] else None
            
            # Làm sạch email (loại bỏ None và chuỗi rỗng)
            email = email if email and email.lower() != 'none' and email.strip() else None
            
            # Xử lý mật khẩu: Ưu tiên từ file, nếu không có hoặc rỗng thì dùng mặc định
            password = None
            if password_from_file and password_from_file.lower() != 'none' and password_from_file.strip():
                password = password_from_file.strip()
            # Nếu không có mật khẩu trong file, sẽ dùng default_password sau khi validate
            
            # Validate dữ liệu
            try:
                employee_id = validate_employee_id(employee_id_str)
                if not employee_id:
                    row_content = '|'.join([str(p) for p in parts[:6]])
                    results['errors'].append({
                        'line': line_num,
                        'content': row_content,
                        'error': f'Mã nhân viên không hợp lệ: {employee_id_str}'
                    })
                    continue
            except Exception as e:
                row_content = '|'.join([str(p) for p in parts[:6]])
                results['errors'].append({
                    'line': line_num,
                    'content': row_content,
                    'error': f'Mã nhân viên không hợp lệ: {str(e)}'
                })
                continue
            
            # Validate name
            name = validate_input_sanitize(name)
            if not name:
                row_content = '|'.join([str(p) for p in parts[:6]])
                results['errors'].append({
                    'line': line_num,
                    'content': row_content,
                    'error': 'Tên người dùng không hợp lệ'
                })
                continue
            
            # Validate và chuẩn hóa department
            department = validate_input_sanitize(department)
            if not department:
                row_content = '|'.join([str(p) for p in parts[:6]])
                results['errors'].append({
                    'line': line_num,
                    'content': row_content,
                    'error': 'Phòng ban không hợp lệ'
                })
                continue
            
            # Chuẩn hóa tên phòng ban: trim, loại bỏ khoảng trắng thừa, chuyển về chữ hoa
            department = ' '.join(department.split()).strip().upper()
            
            # Parse roles (cách nhau bằng dấu phẩy) và chuẩn hóa về chữ hoa
            roles_list = [r.strip().upper() for r in roles_str.split(',') if r.strip()]
            
            # Chuẩn hóa tên vai trò (TEAMLEADER -> TEAM_LEADER, etc.)
            normalized_roles = []
            role_mapping = {
                'TEAMLEADER': 'TEAM_LEADER',
                'TEAM LEADER': 'TEAM_LEADER',
                'TEAM-LEADER': 'TEAM_LEADER',
            }
            for role in roles_list:
                # Loại bỏ khoảng trắng thừa
                role = role.replace(' ', '_').replace('-', '_')
                # Áp dụng mapping nếu có
                normalized_role = role_mapping.get(role, role)
                normalized_roles.append(normalized_role)
            roles_list = normalized_roles
            
            # Validate roles (đã chuẩn hóa về chữ hoa)
            valid_roles = ['EMPLOYEE', 'TEAM_LEADER', 'MANAGER', 'ADMIN']
            invalid_roles = [r for r in roles_list if r not in valid_roles]
            if invalid_roles:
                row_content = '|'.join([str(p) for p in parts[:6]])
                results['errors'].append({
                    'line': line_num,
                    'content': row_content,
                    'error': f'Vai trò không hợp lệ: {", ".join(invalid_roles)}. Vai trò hợp lệ: {", ".join(valid_roles)}'
                })
                continue
            
            if not roles_list:
                row_content = '|'.join([str(p) for p in parts[:6]])
                results['errors'].append({
                    'line': line_num,
                    'content': row_content,
                    'error': 'Phải có ít nhất một vai trò'
                })
                continue
            
            # Validate email (nếu có)
            if email:
                email = email.strip()
                # Kiểm tra format email cơ bản
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_pattern, email):
                    row_content = '|'.join([str(p) for p in parts[:6]])
                    results['errors'].append({
                    'line': line_num,
                        'content': row_content,
                        'error': f'Email không hợp lệ: {email}'
                })
                continue
            
                # Bỏ qua kiểm tra trùng lặp email - người dùng sẽ kiểm tra trước khi upload
            
            # Validate và xử lý mật khẩu
            # Ưu tiên: Mật khẩu từ file Excel > Mật khẩu mặc định từ form
            if password:
                # Có mật khẩu từ file, validate độ dài
                if len(password) < 6:
                    row_content = '|'.join([str(p) for p in parts[:6]])
                    results['errors'].append({
                        'line': line_num,
                        'content': row_content,
                        'error': 'Mật khẩu từ file phải có ít nhất 6 ký tự'
                    })
                    continue
            else:
                # Không có mật khẩu trong file hoặc mật khẩu rỗng, dùng mật khẩu mặc định từ form
                password = default_password
            
            # Bỏ qua kiểm tra trùng lặp employee_id và email - người dùng sẽ kiểm tra trước khi upload
            
            # Tạo user mới - bỏ qua kiểm tra trùng lặp, chỉ cần try-except để bỏ qua nếu lỗi
            try:
                new_user = User(
                    employee_id=employee_id,
                    name=name,
                    department=department,
                    roles=','.join(roles_list),
                    email=email if email else None,
                    is_active=True
                )
                new_user.set_password(password)
                
                db.session.add(new_user)
                db.session.flush()
                
                # Cập nhật cache để tránh duplicate trong cùng batch
                existing_users[employee_id] = new_user
                batch_employee_ids.add(employee_id)
                if email:
                    # Normalize email giống như khi kiểm tra
                    email_normalized = email.strip().lower()
                    email_normalized = ''.join(email_normalized.split())  # Loại bỏ whitespace ẩn
                    existing_emails[email_normalized] = new_user
                    batch_emails.add(email_normalized)
                
                # Lưu thông tin để log sau (batch log để tránh commit từng cái)
                results['success'].append({
                    'line': line_num,
                        'employee_id': employee_id,
                        'name': name,
                        'department': department,
                        'roles': ','.join(roles_list),
                    'email': email if email else None,
                    'user_id': new_user.id  # Lưu ID để log sau
                })
                
            except Exception as e:
                db.session.rollback()
                error_str = str(e)
                
                # Nếu là lỗi UNIQUE constraint (email hoặc employee_id), tìm user đã tồn tại và update
                if 'UNIQUE constraint' in error_str:
                    existing_user = None
                    
                    # Sau khi rollback, query lại từ database
                    # Tìm user đã tồn tại theo employee_id trước (ưu tiên)
                    # Thử tìm cả user đã bị soft delete (có thể cần restore)
                    print(f"[DEBUG] Line {line_num}: UNIQUE constraint error, searching for employee_id={employee_id}")
                    existing_user = User.query.filter_by(employee_id=employee_id).first()
                    
                    if existing_user:
                        print(f"[DEBUG] Line {line_num}: Found user by employee_id: {existing_user.id}, is_deleted={existing_user.is_deleted}")
                    else:
                        print(f"[DEBUG] Line {line_num}: Not found by employee_id, trying email...")
                    
                    # Nếu không tìm thấy theo employee_id, thử tìm theo email
                    if not existing_user and email:
                        email_normalized = email.strip().lower()
                        email_normalized = ''.join(email_normalized.split())
                        print(f"[DEBUG] Line {line_num}: Searching by email (normalized): {email_normalized}")
                        all_db_users = User.query.all()  # Tìm cả user đã bị soft delete
                        for db_user in all_db_users:
                            if db_user.email:
                                db_email_normalized = str(db_user.email).strip().lower()
                                db_email_normalized = ''.join(db_email_normalized.split())
                                if db_email_normalized == email_normalized:
                                    existing_user = db_user
                                    print(f"[DEBUG] Line {line_num}: Found user by email: {existing_user.id}, employee_id={existing_user.employee_id}, is_deleted={existing_user.is_deleted}")
                                    break
                    
                    if not existing_user:
                        print(f"[DEBUG] Line {line_num}: User not found! employee_id={employee_id}, email={email}")
                    
                    if existing_user:
                        # Update thông tin user đã tồn tại
                        # Nếu user bị soft delete, restore lại
                        if existing_user.is_deleted:
                            existing_user.is_deleted = False
                            existing_user.is_active = True
                        
                        existing_user.name = name
                        existing_user.department = department
                        existing_user.roles = ','.join(roles_list)
                        if email:
                            existing_user.email = email
                        # Chỉ update mật khẩu nếu có mật khẩu từ file (không phải mật khẩu mặc định)
                        if password and password != default_password:
                            existing_user.set_password(password)
                        
                        db.session.add(existing_user)
                        try:
                            db.session.flush()
                        except Exception as update_error:
                            db.session.rollback()
                            row_content = '|'.join([str(p) for p in parts[:6]])
                            results['errors'].append({
                                'line': line_num,
                                'content': row_content,
                                'error': f'Lỗi khi cập nhật user: {str(update_error)}'
                            })
                            continue

                        # Thêm vào success với action là 'updated'
                        results['success'].append({
                            'line': line_num,
                            'employee_id': employee_id,
                            'name': name,
                            'department': department,
                            'roles': ','.join(roles_list),
                            'email': email if email else None,
                            'user_id': existing_user.id,
                            'action': 'updated'  # Đánh dấu là đã update
                        })
                        # Cập nhật cache
                        existing_users[employee_id] = existing_user
                        if email:
                            email_normalized = email.strip().lower()
                            email_normalized = ''.join(email_normalized.split())
                            existing_emails[email_normalized] = existing_user
                        continue
                    else:
                        # Không tìm thấy user, bỏ qua
                        row_content = '|'.join([str(p) for p in parts[:6]])
                        results['skipped'].append({
                            'line': line_num,
                            'content': row_content,
                            'employee_id': employee_id_str,
                            'name': name,
                            'reason': f'Đã tồn tại trong hệ thống nhưng không tìm thấy để cập nhật'
                        })
                        continue
                
                # Các lỗi khác
                row_content = '|'.join([str(p) for p in parts[:6]])
                results['errors'].append({
                    'line': line_num,
                    'content': row_content,
                    'error': f'Lỗi khi tạo user: {str(e)}'
                })
        
        # Commit tất cả users thành công (batch commit)
        if results['success']:
            try:
                db.session.commit()
                
                # Batch log audit actions (tạo audit logs nhưng không commit từng cái)
                audit_logs = []
                for success_item in results['success']:
                    user_id = success_item.get('user_id')
                    if user_id:
                        try:
                            from database.models import AuditLog
                            # Xác định action: CREATE hoặc UPDATE
                            action = 'UPDATE_USER_UPLOAD' if success_item.get('action') == 'updated' else 'CREATE_USER_UPLOAD'
                            audit_log = AuditLog(
                                user_id=current_user.id,
                                action=action,
                                table_name='users',
                                record_id=user_id,
                                new_values={
                                    'employee_id': success_item['employee_id'],
                                    'name': success_item['name'],
                                    'department': success_item['department'],
                                    'roles': success_item['roles'],
                                    'email': success_item.get('email'),
                                    'source': 'file_upload'
                                },
                                ip_address=request.remote_addr,
                                user_agent=request.headers.get('User-Agent')
                            )
                            db.session.add(audit_log)
                            audit_logs.append(audit_log)
                        except Exception as log_err:
                            # Nếu log lỗi, bỏ qua nhưng không làm fail upload
                            print(f"Warning: Không thể log audit cho user {user_id}: {log_err}")
                
                # Commit tất cả audit logs một lần
                if audit_logs:
                    try:
                        db.session.commit()
                    except Exception as log_commit_err:
                        # Nếu commit log lỗi, rollback nhưng không ảnh hưởng users đã tạo
                        db.session.rollback()
                        print(f"Warning: Không thể commit audit logs: {log_commit_err}")
                
                # Xóa user_id khỏi dict để tránh serialize lỗi khi trả về JSON
                for success_item in results['success']:
                    success_item.pop('user_id', None)
            except Exception as e:
                db.session.rollback()
                return jsonify({
                    'error': f'Lỗi khi lưu dữ liệu: {str(e)}',
                    'partial_results': results
                }), 500
        
        # Tổng kết
        total_lines = len(lines)
        success_count = len(results['success'])
        # Phân loại created vs updated
        created_count = sum(1 for item in results['success'] if item.get('action') != 'updated')
        updated_count = sum(1 for item in results['success'] if item.get('action') == 'updated')
        error_count = len(results['errors'])
        skipped_count = len(results['skipped'])
        conflict_count = len(results['conflicts'])
        
        message_parts = []
        if created_count > 0:
            message_parts.append(f'Đã tạo thành công {created_count} nhân viên')
        if updated_count > 0:
            message_parts.append(f'Đã cập nhật {updated_count} nhân viên')
        if not message_parts:
            message_parts.append('Không có nhân viên nào được tạo hoặc cập nhật')
        if skipped_count > 0:
            message_parts.append(f'bỏ qua {skipped_count} nhân viên')
        if conflict_count > 0:
            message_parts.append(f'{conflict_count} nhân viên cần xác nhận (trùng lặp)')
        if error_count > 0:
            message_parts.append(f'{error_count} dòng lỗi')
        message = ', '.join(message_parts)
        
        # Nếu có conflicts, trả về để hỏi user
        has_conflicts = conflict_count > 0
        
        return jsonify({
            'success': True,
            'message': message,
            'results': results,
            'summary': {
                'total_lines': total_lines,
                'success_count': success_count,
                'error_count': error_count,
                'skipped_count': skipped_count,
                'conflict_count': conflict_count
            },
            'has_conflicts': has_conflicts
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi upload users: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Đã xảy ra lỗi khi xử lý file: {str(e)}'}), 500

@app.route('/admin/users/download', methods=['GET'])
@require_admin
def download_users():
    """Tải danh sách nhân viên ra file XLSX."""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'Không có quyền truy cập'}), 401

        if check_session_timeout():
            return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401

        update_session_activity()

        current_user = db.session.get(User, session['user_id'])
        if not current_user or 'ADMIN' not in current_user.roles.split(','):
            return jsonify({'error': 'Chỉ quản trị viên mới có quyền thực hiện hành động này'}), 403

        users = User.query.order_by(User.employee_id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        headers = ["Mã NV", "Họ và Tên", "Phòng Ban", "Vai Trò", "Email", "Mật khẩu"]
        ws.append(headers)

        for u in users:
            ws.append([
                u.employee_id or "",
                u.name or "",
                u.department or "",
                u.roles or "",
                u.email or "",
                ""  # Không xuất mật khẩu (để trống)
            ])

        # Căn giữa toàn bộ nội dung
        from openpyxl.styles import Alignment
        center_align = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_align

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 40))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='thongtin_nhanvien.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"❌ Lỗi khi download users: {e}")
        return jsonify({'error': f'Lỗi khi tải danh sách: {str(e)}'}), 500

@app.route('/admin/users/upload/resolve-conflicts', methods=['POST'])
@require_admin
def resolve_upload_conflicts():
    """Xử lý conflicts khi upload nhân viên - user chọn giữ nguyên hoặc cập nhật"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    
    update_session_activity()
    
    current_user = db.session.get(User, session['user_id'])
    if not current_user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    if 'ADMIN' not in current_user.roles.split(','):
        return jsonify({'error': 'Chỉ quản trị viên mới có quyền thực hiện hành động này'}), 403
    
    try:
        data = request.get_json()
        if not data or 'decisions' not in data:
            return jsonify({'error': 'Thiếu dữ liệu quyết định'}), 400
        
        decisions = data['decisions']  # List of {conflict_id, action: 'keep' or 'update', fields_to_update: []}
        
        results = {
            'updated': [],
            'skipped': [],
            'errors': []
        }
        
        for decision in decisions:
            conflict_id = decision.get('conflict_id')
            action = decision.get('action')  # 'keep' or 'update'
            conflict_data = decision.get('conflict_data')
            fields_to_update = decision.get('fields_to_update', [])  # Danh sách các trường muốn cập nhật: ['name', 'department', 'roles', 'email', 'password']
            
            if not conflict_data:
                results['errors'].append({
                    'conflict_id': conflict_id,
                    'error': 'Thiếu dữ liệu conflict'
                })
                continue
            
            existing_user_id = conflict_data.get('existing_user_id')
            existing_user = db.session.get(User, existing_user_id)
            
            if not existing_user:
                results['errors'].append({
                    'conflict_id': conflict_id,
                    'error': f'Không tìm thấy user với ID {existing_user_id}'
                })
                continue
            
            if action == 'keep':
                # Giữ nguyên, bỏ qua
                results['skipped'].append({
                    'conflict_id': conflict_id,
                    'employee_id': conflict_data.get('employee_id'),
                    'name': conflict_data.get('name'),
                    'reason': 'Người dùng chọn giữ nguyên dữ liệu cũ'
                })
            elif action == 'update':
                # Cập nhật dữ liệu mới - chỉ cập nhật các trường được chọn
                try:
                    old_values = {
                        'name': existing_user.name,
                        'department': existing_user.department,
                        'roles': existing_user.roles,
                        'email': existing_user.email
                    }
                    
                    updated_fields = []
                    
                    # Cập nhật từng trường nếu được chọn
                    if 'name' in fields_to_update:
                        existing_user.name = conflict_data.get('name', existing_user.name)
                        updated_fields.append('name')
                    
                    if 'department' in fields_to_update:
                        existing_user.department = conflict_data.get('department', existing_user.department)
                        updated_fields.append('department')
                    
                    if 'roles' in fields_to_update:
                        existing_user.roles = conflict_data.get('roles', existing_user.roles)
                        updated_fields.append('roles')
                    
                    # Xử lý email - chỉ cập nhật nếu được chọn và không trùng với user khác
                    if 'email' in fields_to_update:
                        new_email = conflict_data.get('email')
                        if new_email and new_email.strip():
                            new_email_lower = new_email.strip().lower()
                            # Kiểm tra email có trùng với user khác không
                            other_user = User.query.filter(
                                User.email.ilike(new_email_lower),
                                User.id != existing_user.id,
                                User.is_deleted == False
                            ).first()
                            if other_user:
                                results['errors'].append({
                                    'conflict_id': conflict_id,
                                    'employee_id': conflict_data.get('employee_id'),
                                    'error': f'Email {new_email} đã được sử dụng bởi nhân viên khác (Mã NV: {other_user.employee_id})'
                                })
                                continue
                            existing_user.email = new_email.strip()
                            updated_fields.append('email')
                    
                    # Cập nhật mật khẩu nếu được chọn
                    if 'password' in fields_to_update:
                        new_password = conflict_data.get('password')
                        if new_password and new_password.strip():
                            existing_user.set_password(new_password.strip())
                            updated_fields.append('password')
                    
                    if not updated_fields:
                        # Không có trường nào được chọn để cập nhật
                        results['skipped'].append({
                            'conflict_id': conflict_id,
                            'employee_id': conflict_data.get('employee_id'),
                            'name': conflict_data.get('name'),
                            'reason': 'Không có trường nào được chọn để cập nhật'
                        })
                        continue
                    
                    db.session.flush()
                    
                    # Log audit
                    new_values = {}
                    for field in updated_fields:
                        if field == 'password':
                            new_values['password'] = '***'  # Không lưu mật khẩu vào log
                        else:
                            new_values[field] = getattr(existing_user, field)
                    
                    log_audit_action(
                        user_id=current_user.id,
                        action='UPDATE_USER_UPLOAD_CONFLICT',
                        table_name='users',
                        record_id=existing_user.id,
                        old_values=old_values,
                        new_values=new_values
                    )
                    
                    results['updated'].append({
                        'conflict_id': conflict_id,
                        'employee_id': conflict_data.get('employee_id'),
                        'name': conflict_data.get('name'),
                        'existing_user_id': existing_user.id,
                        'updated_fields': updated_fields
                    })
                    
                except Exception as e:
                    db.session.rollback()
                    results['errors'].append({
                        'conflict_id': conflict_id,
                        'employee_id': conflict_data.get('employee_id'),
                        'error': f'Lỗi khi cập nhật: {str(e)}'
                    })
        
        # Commit tất cả updates
        if results['updated']:
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                return jsonify({
                    'error': f'Lỗi khi lưu dữ liệu: {str(e)}',
                    'partial_results': results
                }), 500
        
        return jsonify({
            'success': True,
            'message': f'Đã xử lý {len(decisions)} conflicts: {len(results["updated"])} cập nhật, {len(results["skipped"])} giữ nguyên, {len(results["errors"])} lỗi',
            'results': results
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi xử lý conflicts: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Đã xảy ra lỗi: {str(e)}'}), 500

@app.route('/admin/users/delete-all', methods=['POST'])
@require_admin
def delete_all_users():
    """Xóa tất cả nhân viên (trừ user hiện tại) - yêu cầu xác nhận mật khẩu"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    
    update_session_activity()
    
    # Lấy user hiện tại
    current_user = db.session.get(User, session['user_id'])
    if not current_user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    # Kiểm tra quyền admin
    if 'ADMIN' not in current_user.roles.split(','):
        return jsonify({'error': 'Chỉ quản trị viên mới có quyền thực hiện hành động này'}), 403
    
    # Lấy mật khẩu từ request
    data = request.get_json()
    password = data.get('password', '')
    
    if not password:
        return jsonify({'error': 'Vui lòng nhập mật khẩu để xác nhận'}), 400
    
    # Xác thực mật khẩu - sử dụng method check_password của User model
    if not current_user.check_password(password):
        return jsonify({'error': 'Mật khẩu không đúng. Vui lòng thử lại.'}), 400
    
    try:
        # Đếm số lượng users sẽ bị xóa
        users_to_delete = User.query.filter(
            User.id != current_user.id,
            User.is_deleted == False
        ).all()
        
        total_count = len(users_to_delete)
        
        if total_count == 0:
            return jsonify({'error': 'Không có nhân viên nào để xóa'}), 400
        
        # Đếm số admin còn lại (không tính user hiện tại)
        other_admins = [u for u in users_to_delete if 'ADMIN' in u.roles.split(',')]
        other_admins_count = len(other_admins)
        
        # Cho phép xóa tất cả nhân viên (không phải admin)
        # User hiện tại (admin) sẽ được giữ lại tự động vì không có trong users_to_delete
        
        # Soft delete tất cả users (trừ user hiện tại và các admin khác)
        # Chỉ xóa các user không phải admin - sử dụng bulk update để nhanh hơn
        # Lấy danh sách ID của các admin để loại trừ (giữ lại tất cả admin)
        admin_ids = [u.id for u in users_to_delete if 'ADMIN' in u.roles.split(',')]
        
        # Lấy danh sách ID của các user sẽ bị xóa (không phải admin, không phải user hiện tại)
        user_ids_to_delete = [u.id for u in users_to_delete if 'ADMIN' not in u.roles.split(',')]
        
        if not user_ids_to_delete:
            return jsonify({
                'error': 'Không có nhân viên nào để xóa. Tất cả người dùng còn lại đều là quản trị viên và cần được giữ lại để đảm bảo hệ thống hoạt động.'
            }), 400
        
        # Bulk update: xóa tất cả users trong danh sách (chỉ nhân viên thường, không phải admin)
        deleted_count = User.query.filter(
            User.id.in_(user_ids_to_delete)
        ).update({
            'is_deleted': True,
            'is_active': False
        }, synchronize_session=False)
        
        # Ghi log hành động
        remaining_admins = other_admins_count + 1  # +1 cho user hiện tại
        log_audit_action(
            user_id=current_user.id,
            action='DELETE_ALL_USERS',
            table_name='users',
            record_id=None,
            old_values={'total_users': total_count, 'total_admins': other_admins_count + 1},
            new_values={
                'deleted_count': deleted_count, 
                'remaining_users': total_count - deleted_count + 1,  # +1 cho user hiện tại
                'remaining_admins': remaining_admins
            }
        )
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Đã xóa thành công {deleted_count} nhân viên. Hệ thống vẫn còn {remaining_admins} quản trị viên.',
            'deleted_count': deleted_count,
            'remaining_admins': remaining_admins
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Lỗi khi xóa tất cả nhân viên: {e}")
        import traceback
        traceback.print_exc()
        # Trả về JSON error response thay vì HTML error page
        return jsonify({
            'success': False,
            'error': f'Đã xảy ra lỗi khi xóa nhân viên: {str(e)}'
        }), 500

@app.route('/admin/users/<int:user_id>/toggle_active', methods=['POST'])
@require_admin
def toggle_user_active(user_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    if int(user_id) == int(session['user_id']):
        return jsonify({'error': 'Không thể tự khoá tài khoản của mình!'}), 400
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()
    return jsonify({
        'success': True,
        'is_active': user.is_active,
        'user_name': user.name,
        'status_label': 'Hoạt Động' if user.is_active else 'Đã Khoá',
        'status_class': 'bg-success' if user.is_active else 'bg-secondary'
    })

@app.route('/admin/attendance/<int:attendance_id>/export-overtime-pdf')
@require_admin
def export_overtime_pdf(attendance_id):
    try:
        # Load attendance với tất cả các relationship cần thiết
        attendance = Attendance.query.options(
            joinedload(Attendance.user).load_only(User.name, User.employee_id, User.department),
            joinedload(Attendance.team_leader_signer).load_only(User.name),
            joinedload(Attendance.manager_signer).load_only(User.name)
        ).get_or_404(attendance_id)
        
        buffer = io.BytesIO()
        
        # Sử dụng hàm create_overtime_pdf đã tách riêng
        create_overtime_pdf(attendance, buffer)
        
        # Tạo tên file
        safe_name = remove_vietnamese_accents(attendance.user.name)
        safe_empid = str(attendance.user.employee_id)
        safe_date = attendance.date.strftime('%d%m%Y')
        filename = f"tangca_{safe_name}_{safe_empid}_{safe_date}.pdf"
        
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
        
    except Exception as e:
        print('PDF export error:', e)
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': 'Lỗi khi sinh file PDF', 'detail': str(e)})

@app.route('/admin/attendance/<int:attendance_id>/test-signature-pdf')
@require_admin
def test_signature_on_overtime_pdf(attendance_id):
    """Test hiển thị chữ ký trên form tăng ca thực tế"""
    try:
        # Lấy bản ghi attendance với tất cả các relationship cần thiết
        attendance = Attendance.query.options(
            joinedload(Attendance.user).load_only(User.name, User.employee_id, User.department),
            joinedload(Attendance.team_leader_signer).load_only(User.name),
            joinedload(Attendance.manager_signer).load_only(User.name)
        ).get_or_404(attendance_id)
        
        # Tạo chữ ký mẫu cho test
        sample_signature = create_sample_signature_base64()
        
        # Tạo bản copy của attendance với chữ ký mẫu
        test_attendance = type('TestAttendance', (), {
            'id': attendance.id,
            'date': attendance.date,
            'check_in': attendance.check_in,
            'check_out': attendance.check_out,
            'break_time': attendance.break_time,
            'total_work_hours': attendance.total_work_hours,
            'overtime_before_22': attendance.overtime_before_22,
            'overtime_after_22': attendance.overtime_after_22,
            'note': attendance.note,
            'user': attendance.user,
            'signature': sample_signature,
            'team_leader_signature': sample_signature,
            'manager_signature': sample_signature,
            'team_leader_signer_id': attendance.team_leader_signer_id,
            'manager_signer_id': attendance.manager_signer_id,
            'team_leader_signer': attendance.team_leader_signer,
            'manager_signer': attendance.manager_signer,
            'approved': True,
            'approved_at': datetime.now()
        })()
        
        buffer = io.BytesIO()
        
        # Tạo PDF với chữ ký mẫu
        create_overtime_pdf(test_attendance, buffer)
        
        # Tạo tên file test
        safe_name = remove_vietnamese_accents(attendance.user.name)
        safe_empid = str(attendance.user.employee_id)
        safe_date = attendance.date.strftime('%d%m%Y')
        filename = f"test_chu_ky_tangca_{safe_name}_{safe_empid}_{safe_date}.pdf"
        
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
        
    except Exception as e:
        print('Test signature PDF error:', e)
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': 'Lỗi khi tạo PDF test chữ ký', 'detail': str(e)})

def create_sample_signature_base64():
    """Tạo chữ ký mẫu dạng base64"""
    try:
        # Tạo canvas để vẽ chữ ký mẫu
        from PIL import Image, ImageDraw
        
        # Tạo ảnh trắng
        img = Image.new('RGB', (200, 100), color='white')
        draw = ImageDraw.Draw(img)
        
        # Vẽ chữ ký mẫu với màu xanh như bút bi
        draw.line([(20, 50), (40, 30), (60, 70), (80, 40), (100, 60), (120, 35), (140, 65), (160, 45), (180, 55)], fill='blue', width=2)
        
        # Chuyển thành base64
        import io
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        import base64
        return base64.b64encode(buffer.getvalue()).decode('utf-8')
        
    except Exception as e:
        print(f"Error creating sample signature: {e}")
        # Trả về chữ ký mẫu đơn giản nếu có lỗi
        return "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="

# Hàm wrap_text cho phần ghi chú (đặt phía trên đoạn sử dụng)
def wrap_text(text, font_name, font_size, max_width, canvas_obj):
    from reportlab.pdfbase.pdfmetrics import stringWidth
    words = text.split(' ')
    lines = []
    current_line = ''
    for word in words:
        test_line = current_line + (' ' if current_line else '') + word
        if stringWidth(test_line, font_name, font_size) <= max_width:
            current_line = test_line
        else:
            if current_line:
                lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)
    return lines

@app.route('/admin/attendance/export-overtime-bulk')
@require_admin
def export_overtime_bulk():
    try:
        month = request.args.get('month')  # Có thể None nếu xuất theo năm
        year = int(request.args.get('year', 0))
        
        if not (2000 <= year <= 2100):
            return abort(400, 'Tham số năm không hợp lệ')
        
        # Xây dựng query filter
        query_filter = [
            db.extract('year', Attendance.date) == year,
            Attendance.approved == True
        ]
        
        # Thêm filter tháng nếu có
        if month:
            month = int(month)
            if not (1 <= month <= 12):
                return abort(400, 'Tham số tháng không hợp lệ')
            query_filter.append(db.extract('month', Attendance.date) == month)
        
        # Lấy tất cả bản ghi Attendance đã được phê duyệt
        # Tối ưu: chỉ lấy các trường cần thiết
        attendances = Attendance.query.filter(*query_filter).options(
            joinedload(Attendance.user).load_only(User.name, User.employee_id, User.department)
        ).all()
        
        if not attendances:
            if month:
                return abort(404, 'Không có bản ghi nào trong tháng này')
            else:
                return abort(404, 'Không có bản ghi nào trong năm này')
        
        print(f'Creating ZIP for {len(attendances)} records...')
        
        # Tạo file ZIP trong bộ nhớ với compression level cao hơn
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, compresslevel=9) as zipf:
            for i, att in enumerate(attendances, 1):
                try:
                    # Tạo PDF cho từng bản ghi
                    pdf_buffer = io.BytesIO()
                    
                    # Gọi hàm tạo PDF (tái sử dụng logic từ export_overtime_pdf)
                    create_overtime_pdf(att, pdf_buffer)
                    
                    # Đặt tên file cho từng PDF (loại bỏ dấu tiếng Việt)
                    
                    safe_name = remove_vietnamese_accents(att.user.name) if att.user and att.user.name else str(att.id)
                    safe_empid = str(att.user.employee_id) if att.user and att.user.employee_id else str(att.id)
                    safe_date = att.date.strftime('%d%m%Y')
                    filename = f"tangca_{safe_name}_{safe_empid}_{safe_date}.pdf"
                    
                    # Đảm bảo buffer ở đầu file
                    pdf_buffer.seek(0)
                    zipf.writestr(filename, pdf_buffer.read())
                    
                    # Log progress mỗi 10 records
                    if i % 10 == 0:
                        print(f'Processed {i}/{len(attendances)} records...')
                    
                except Exception as e:
                    print(f'Error creating PDF for attendance {att.id}: {e}')
                    continue
        
        zip_buffer.seek(0)
        
        # Tạo tên file ZIP
        if month:
            zip_filename = f"tangca_{month:02d}_{year}.zip"
        else:
            zip_filename = f"tangca_{year}.zip"
            
        print(f'ZIP creation completed: {zip_filename}')
        return send_file(zip_buffer, as_attachment=True, download_name=zip_filename, mimetype='application/zip')
        
    except Exception as e:
        print('Bulk export error:', e)
        return jsonify({'error': 'Lỗi khi xuất file ZIP', 'detail': str(e)})

# Cache fonts để tránh đăng ký lại mỗi lần
_fonts_registered = False

def register_pdf_fonts():
    """Đăng ký fonts cho PDF một lần duy nhất"""
    global _fonts_registered
    if _fonts_registered:
        return
    
    try:
        # Thử đăng ký DejaVuSans cho tiếng Việt
        registerFont(TTFont('DejaVuSans', 'static/fonts/DejaVuSans.ttf'))
        registerFont(TTFont('DejaVuSans-Bold', 'static/fonts/DejaVuSans.ttf'))  # Sử dụng cùng font cho bold
        
        # Đăng ký NotoSansJP cho tiếng Nhật
        registerFont(TTFont('NotoSansJP', 'static/fonts/NotoSansJP-Regular.ttf'))
        registerFont(TTFont('NotoSansJP-Bold', 'static/fonts/NotoSansJP-Bold.ttf'))
        registerFont(TTFont('NotoSansJP-Medium', 'static/fonts/NotoSansJP-Medium.ttf'))
        registerFont(TTFont('NotoSansJP-Light', 'static/fonts/NotoSansJP-Light.ttf'))
        registerFont(TTFont('NotoSansJP-Black', 'static/fonts/NotoSansJP-Black.ttf'))
        registerFont(TTFont('NotoSansJP-ExtraBold', 'static/fonts/NotoSansJP-ExtraBold.ttf'))
        registerFont(TTFont('NotoSansJP-ExtraLight', 'static/fonts/NotoSansJP-ExtraLight.ttf'))
        registerFont(TTFont('NotoSansJP-SemiBold', 'static/fonts/NotoSansJP-SemiBold.ttf'))
        registerFont(TTFont('NotoSansJP-Thin', 'static/fonts/NotoSansJP-Thin.ttf'))
        
        _fonts_registered = True
        print('PDF fonts registered successfully')
    except Exception as e:
        print('PDF font register error:', e)
        # Fallback: sử dụng font mặc định
        _fonts_registered = True

def fix_base64_padding(base64_string):
    """
    Sửa lỗi base64 padding để đảm bảo độ dài là bội số của 4
    """
    if not base64_string:
        return base64_string
    
    # Loại bỏ khoảng trắng và ký tự xuống dòng
    base64_string = base64_string.strip()
    
    # Tính số ký tự cần thêm để đạt bội số của 4
    padding_length = len(base64_string) % 4
    if padding_length > 0:
        # Thêm dấu = để đạt bội số của 4
        base64_string += '=' * (4 - padding_length)
        # print(f"DEBUG: Fixed base64 padding, added {4 - padding_length} padding characters")
    
    return base64_string

def looks_like_fernet_token(token):
    """
    Nhận diện nhanh chuỗi có khả năng là Fernet token để tránh thử giải mã sai dữ liệu
    """
    try:
        if not isinstance(token, str) or len(token) < 50:
            return False
        normalized = token.strip().replace(' ', '+')
        padding_length = len(normalized) % 4
        if padding_length > 0:
            normalized += '=' * (4 - padding_length)
        raw = base64.urlsafe_b64decode(normalized.encode('utf-8'))
        return len(raw) > 9 and raw[0] == 0x80
    except Exception:
        return False

def process_signature_for_pdf(signature_data):
    """
    Xử lý chữ ký để hiển thị trong PDF - IMPROVED VERSION với xử lý lỗi base64 an toàn
    """
    if not signature_data:
        # print("DEBUG: No signature data provided")
        return None
    
    try:
        if isinstance(signature_data, str):
            # print(f"DEBUG: Processing signature string, length: {len(signature_data)}")
            
            # Nếu là base64 từ frontend (data:image/png;base64,...)
            if signature_data.startswith('data:image'):
                # print("DEBUG: Found data:image format, extracting base64")
                try:
                    # Kiểm tra định dạng data:image
                    if not signature_data.startswith('data:image/png;base64,'):
                        # print("DEBUG: Not PNG format, trying to convert")
                        # Thử chuyển đổi từ các định dạng khác
                        if signature_data.startswith('data:image/jpeg;base64,'):
                            signature_data = signature_data.replace('data:image/jpeg;base64,', 'data:image/png;base64,')
                    
                    base64_data = signature_data.split(',')[1]
                    
                    # Sửa lỗi base64 padding
                    base64_data = fix_base64_padding(base64_data)
                    
                    # Kiểm tra base64 có hợp lệ không
                    try:
                        decoded = base64.b64decode(base64_data)
                        # print(f"DEBUG: Base64 decode successful, decoded length: {len(decoded)}")
                        
                        # Kiểm tra có phải là ảnh PNG không
                        if len(decoded) >= 8 and decoded.startswith(b'\x89PNG\r\n\x1a\n'):
                            # print("DEBUG: Valid PNG image confirmed")
                            return base64_data
                        else:
                            # print("DEBUG: Not a valid PNG image")
                            return None
                    except Exception as decode_error:
                        # print(f"DEBUG: Base64 decode failed after padding fix: {decode_error}")
                        return None
                        
                except Exception as e:
                    # print(f"DEBUG: Base64 decode failed: {e}")
                    return None
                    
            # Nếu là base64 thuần túy
            elif len(signature_data) > 100:
                try:
                    # Sửa lỗi base64 padding trước khi decode
                    fixed_signature = fix_base64_padding(signature_data)
                    
                    # Thử decode để kiểm tra
                    decoded = base64.b64decode(fixed_signature)
                    # print(f"DEBUG: Valid base64 signature found, decoded length: {len(decoded)}")
                    
                    # Kiểm tra có phải là ảnh PNG không
                    if len(decoded) >= 8 and decoded.startswith(b'\x89PNG\r\n\x1a\n'):
                        # print("DEBUG: Valid PNG image confirmed")
                        return fixed_signature
                    else:
                        # print("DEBUG: Not a valid PNG image")
                        return None
                        
                except Exception as base64_error:
                    print(f"DEBUG: Base64 decode failed: {base64_error}")
                    # Chỉ thử giải mã nếu thật sự trông giống Fernet token
                    if looks_like_fernet_token(signature_data):
                        try:
                            decrypted = signature_manager.decrypt_signature(signature_data)
                            # print(f"DEBUG: Decrypted signature, length: {len(decrypted) if decrypted else 0}")
                            
                            if decrypted:
                                # Nếu giải mã thành công và có data:image
                                if decrypted.startswith('data:image'):
                                    base64_data = decrypted.split(',')[1]
                                    try:
                                        # Sửa lỗi base64 padding sau khi giải mã
                                        base64_data = fix_base64_padding(base64_data)
                                        
                                        # Kiểm tra base64 sau khi giải mã
                                        decoded = base64.b64decode(base64_data)
                                        # print(f"DEBUG: Decrypted base64 decode successful, decoded length: {len(decoded)}")
                                        
                                        # Kiểm tra có phải là ảnh PNG không
                                        if len(decoded) >= 8 and decoded.startswith(b'\x89PNG\r\n\x1a\n'):
                                            # print("DEBUG: Valid PNG image confirmed after decryption")
                                            return base64_data
                                        else:
                                            # print("DEBUG: Not a valid PNG image after decryption")
                                            return None
                                            
                                    except Exception as e:
                                        # print(f"DEBUG: Decrypted base64 decode failed: {e}")
                                        return None
                                # Nếu giải mã thành công và là base64 thuần túy
                                elif len(decrypted) > 100:
                                    try:
                                        # Sửa lỗi base64 padding sau khi giải mã
                                        fixed_decrypted = fix_base64_padding(decrypted)
                                        
                                        decoded = base64.b64decode(fixed_decrypted)
                                        # print(f"DEBUG: Decrypted base64 decode successful, decoded length: {len(decoded)}")
                                        
                                        # Kiểm tra có phải là ảnh PNG không
                                        if len(decoded) >= 8 and decoded.startswith(b'\x89PNG\r\n\x1a\n'):
                                            # print("DEBUG: Valid PNG image confirmed after decryption")
                                            return fixed_decrypted
                                        else:
                                            # print("DEBUG: Not a valid PNG image after decryption")
                                            return None
                                            
                                    except Exception as e:
                                        # print(f"DEBUG: Decrypted base64 decode failed: {e}")
                                        return None
                        except Exception:
                            # Silent on decryption failure to avoid noisy logs; logic unchanged
                            return None
                    else:
                        print("DEBUG: Not valid base64 and not a Fernet token, skip decryption")
                        return None
            else:
                print(f"DEBUG: Short signature string: {signature_data}")
                return signature_data
        else:
            print(f"DEBUG: Non-string signature data type: {type(signature_data)}")
            return None
    except Exception as e:
        print(f"Error processing signature: {e}")
        import traceback
        traceback.print_exc()
        return None

def draw_signature_with_proper_scaling(canvas, signature_data, x, y, box_width, box_height):
    """
    Vẽ chữ ký với tỷ lệ đúng và màu xanh như bút bi - SỬ DỤNG SIGNATURE FIT ADAPTER
    """
    if not signature_data:
        print("DEBUG: No signature data provided to draw")
        return False
    
    try:
        # Sử dụng signature fit adapter để điều chỉnh chữ ký vừa khít với ô
        from utils.signature_manager import signature_manager
        
        # Xác định loại ô dựa trên kích thước
        box_type = 'default'
        if box_width >= 140 and box_height >= 70:
            box_type = 'manager'  # Ô quản lý
        elif box_width >= 120 and box_height >= 60:
            box_type = 'supervisor'  # Ô cấp trên
        elif box_width >= 100 and box_height >= 50:
            box_type = 'applicant'  # Ô người xin phép
        
        print(f"DEBUG: Using signature fit adapter for box type: {box_type}")
        
        # Điều chỉnh chữ ký vừa khít với ô
        fitted_signature = signature_manager.fit_signature_to_form_box(
            signature_data, 
            box_type=box_type
        )
        
        if not fitted_signature:
            print("DEBUG: Failed to fit signature to box")
            return False
                
        print(f"DEBUG: Fitted signature length: {len(fitted_signature)}")
        
        # Decode base64
        try:
            if fitted_signature.startswith('data:image'):
                fitted_signature = fitted_signature.split(',')[1]
            
            decoded_data = base64.b64decode(fitted_signature)
            print(f"DEBUG: Successfully decoded fitted signature, length: {len(decoded_data)}")
            
        except Exception as decode_error:
            print(f"DEBUG: Failed to decode fitted signature: {decode_error}")
            return False
        
        # Mở và chuẩn hóa ảnh, đồng thời chuẩn bị để nội suy theo kích thước vẽ thực tế
        try:
            from PIL import Image
            import io
            
            pil_image = Image.open(io.BytesIO(decoded_data))
            if pil_image.mode != 'RGBA':
                pil_image = pil_image.convert('RGBA')
        except Exception as img_open_err:
            print(f"DEBUG: Failed to open image for processing: {img_open_err}")
            return False
        
        # Tính tỷ lệ để giữ nguyên tỷ lệ khung hình và vừa khít với ô
        img_width, img_height = pil_image.size
        print(f"DEBUG: Fitted image size (PIL): {img_width}x{img_height}")
        aspect_ratio = img_width / img_height
        box_aspect_ratio = box_width / box_height
        
        # Tính kích thước thực tế để vẽ - sử dụng 99% kích thước ô để đảm bảo gần như kín mà không tràn
        if aspect_ratio > box_aspect_ratio:
            # Ảnh rộng hơn, căn theo chiều rộng
            draw_width = box_width * 0.99
            draw_height = draw_width / aspect_ratio
        else:
            # Ảnh cao hơn, căn theo chiều cao
            draw_height = box_height * 0.99
            draw_width = draw_height * aspect_ratio
        
        # Kiểm tra kích thước vẽ hợp lệ
        if draw_width <= 0 or draw_height <= 0:
            print(f"DEBUG: Invalid draw dimensions: {draw_width}x{draw_height}")
            return False
        
        # Nội suy ảnh tới độ phân giải mục tiêu dựa trên kích thước vẽ để luôn sắc nét
        try:
            target_dpi = 220  # DPI mục tiêu cho ảnh nhúng vào PDF (MaxFill - chất lượng cao)
            target_px_w = max(1, int(draw_width * target_dpi / 72.0))
            target_px_h = max(1, int(draw_height * target_dpi / 72.0))
            
            if pil_image.size != (target_px_w, target_px_h):
                pil_image = pil_image.resize((target_px_w, target_px_h), Image.Resampling.LANCZOS)
            
            # Chuyển màu chữ ký sang xanh bút bi sau khi đã resize để giữ cạnh mịn
            data = pil_image.getdata()
            blue_pen_color = (0, 0, 255, 255)
            new_data = []
            for item in data:
                if item[0] < 50 and item[1] < 50 and item[2] < 50 and item[3] > 100:
                    new_data.append(blue_pen_color)
                else:
                    new_data.append(item)
            new_image = Image.new('RGBA', pil_image.size)
            new_image.putdata(new_data)
            
            new_image_buffer = io.BytesIO()
            new_image.save(new_image_buffer, format='PNG')
            new_image_buffer.seek(0)
            img = ImageReader(new_image_buffer)
            print("DEBUG: Image prepared and ImageReader created at target DPI")
        except Exception as prep_err:
            print(f"DEBUG: Failed to prepare high-DPI image: {prep_err}")
            try:
                img = ImageReader(io.BytesIO(decoded_data))
            except Exception:
                return False
        
        # Tính vị trí căn giữa
        x_offset = (box_width - draw_width) / 2
        y_offset = (box_height - draw_height) / 2
        
        # Vẽ nền trắng cho ô chữ ký để tránh bị đen
        canvas.setFillColor(colors.white)
        canvas.rect(x, y, box_width, box_height, fill=1, stroke=0)
        canvas.setFillColor(colors.black)  # Reset về màu đen cho text
        
        # Vẽ chữ ký với kích thước đã tính toán
        try:
            final_x = x + x_offset
            final_y = y + y_offset
            
            # Kiểm tra vị trí hợp lệ
            if final_x < 0 or final_y < 0:
                print(f"DEBUG: Invalid position: ({final_x}, {final_y})")
                return False
                
            # Kiểm tra vị trí có vượt quá trang không
            if final_x + draw_width > canvas._pagesize[0] or final_y + draw_height > canvas._pagesize[1]:
                print(f"DEBUG: Position out of page bounds")
                return False
            
            canvas.drawImage(img, final_x, final_y, width=draw_width, height=draw_height)
            print(f"DEBUG: Blue signature drawn successfully with signature fit adapter")
            print(f"DEBUG: Fitted size: {img_width}x{img_height}, Draw size: {draw_width:.1f}x{draw_height:.1f}")
            print(f"DEBUG: Position: ({final_x:.1f}, {final_y:.1f})")
            return True
        except Exception as draw_error:
            print(f"DEBUG: Failed to draw image: {draw_error}")
            import traceback
            traceback.print_exc()
            return False
        
    except Exception as e:
        print(f"DEBUG: Error drawing signature with signature fit adapter: {e}")
        import traceback
        traceback.print_exc()
        return False

def create_signature_placeholder(canvas, x, y, box_width, box_height, text="Chữ ký"):
    """Tạo placeholder cho chữ ký khi không thể hiển thị"""
    try:
        # Vẽ nền trắng
        canvas.setFillColor(colors.white)
        canvas.rect(x, y, box_width, box_height, fill=1, stroke=0)
        
        # Vẽ border
        canvas.setStrokeColor(colors.grey)
        canvas.setLineWidth(0.5)
        canvas.rect(x, y, box_width, box_height, stroke=1, fill=0)
        
        # Vẽ text placeholder
        canvas.setFillColor(colors.grey)
        canvas.setFont("DejaVuSans", 8)
        
        # Căn giữa text
        text_width = canvas.stringWidth(text, "DejaVuSans", 8)
        text_x = x + (box_width - text_width) / 2
        text_y = y + box_height / 2 + 3  # +3 để căn giữa theo chiều dọc
        
        canvas.drawString(text_x, text_y, text)
        
        # Reset màu
        canvas.setFillColor(colors.black)
        canvas.setStrokeColor(colors.black)
        
        return True
    except Exception as e:
        print(f"DEBUG: Error creating signature placeholder: {e}")
        return False

def create_overtime_pdf(attendance, buffer):
    """Tạo PDF giấy tăng ca cho một bản ghi attendance"""
    # Đăng ký fonts một lần duy nhất
    register_pdf_fonts()
    
    user = attendance.user
    employee_signature = attendance.signature if attendance.signature else None
    team_leader_signature = attendance.team_leader_signature if attendance.team_leader_signature else None
    manager_signature = attendance.manager_signature if attendance.manager_signature else None
    
    # Lấy thông tin người ký từ database
    from database.models import User
    
    # Thông tin người ký employee (người tạo đơn)
    employee_signer_name = user.name if user else "Không xác định"
    
    # Thông tin người ký team leader và manager - load relationship và xử lý an toàn
    team_leader_signer_name = "Chưa ký"
    manager_signer_name = "Chưa ký"
    
    # Kiểm tra và lấy tên người ký team leader
    if hasattr(attendance, 'team_leader_signer') and attendance.team_leader_signer:
        team_leader_signer_name = attendance.team_leader_signer.name
    elif hasattr(attendance, 'team_leader_signer_id') and attendance.team_leader_signer_id:
        # Nếu có ID nhưng relationship chưa load, query trực tiếp
        team_leader = db.session.get(User, attendance.team_leader_signer_id)
        if team_leader:
            team_leader_signer_name = team_leader.name
    
    # Kiểm tra và lấy tên người ký manager
    if hasattr(attendance, 'manager_signer') and attendance.manager_signer:
        manager_signer_name = attendance.manager_signer.name
    elif hasattr(attendance, 'manager_signer_id') and attendance.manager_signer_id:
        # Nếu có ID nhưng relationship chưa load, query trực tiếp
        manager = db.session.get(User, attendance.manager_signer_id)
        if manager:
            manager_signer_name = manager.name
    


    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 30
    y = height - margin

    # Header: Bảng 6 cột như trong hình
    header_data = [
        [
            Paragraph('<b>DMI HUẾ</b>', ParagraphStyle('h', fontName='DejaVuSans', fontSize=9, alignment=1)),
            Paragraph('<b>総務<br/>TỔNG VỤ</b>', ParagraphStyle('h', fontName='NotoSansJP', fontSize=8, alignment=1)),
            Paragraph('<b>分類番号：<br/>Số hiệu phân loại：</b>', ParagraphStyle('h', fontName='NotoSansJP', fontSize=7, alignment=1)),
            Paragraph('', ParagraphStyle('h', fontName='DejaVuSans', fontSize=8, alignment=1)),  # Ô trắng sau ô 3
            Paragraph('<b>記入 FORM<br/>NHẬP FORM</b>', ParagraphStyle('h', fontName='NotoSansJP', fontSize=8, alignment=1)),
            Paragraph('<b>Form作成：<br/>Tác thành：</b>', ParagraphStyle('h', fontName='NotoSansJP', fontSize=7, alignment=1)),
            Paragraph('', ParagraphStyle('h', fontName='DejaVuSans', fontSize=8, alignment=1)),  # Ô trắng sau ô tác thành
            Paragraph('', ParagraphStyle('h', fontName='DejaVuSans', fontSize=8, alignment=1)),  # Ô trắng thứ 2 sau ô tác thành
        ]
    ]
    
    col_widths = [60, 80, 100, 50, 80, 80, 50, 50]  # Tổng = 570, gần bằng width A4
    header_table_width = sum(col_widths)
    x_header = (width - header_table_width) / 2
    header_table = Table(header_data, colWidths=col_widths, rowHeights=25)
    header_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
    ]))
    header_table.wrapOn(c, width-2*margin, 30)
    header_table.drawOn(c, x_header, y-25)
    y -= 40

    # Thông tin công ty
    c.setFont("DejaVuSans", 10)
    c.drawString(margin, y, "Công ty TNHH DMI HUẾ")
    y -= 12
    c.setFont("DejaVuSans", 8)
    c.drawString(margin, y, "174 Bà Triệu- tòa nhà 4 tầng Phong Phú Plaza, phường Phú Hội, Thành phố Huế, Tỉnh Thừa Thiên Huế,Việt Nam.")
    y -= 25

    # Tiêu đề chính
    c.setFont("DejaVuSans", 14)
    c.drawCentredString(width/2, y, "GIẤY ĐỀ NGHỊ TĂNG CA/ĐI LÀM NGÀY NGHỈ")
    y -= 16
    c.setFont("NotoSansJP", 11)
    c.drawCentredString(width/2, y, "(残業/休日出勤申請書)")
    y -= 20
    c.setFont("DejaVuSans", 9)
    c.drawCentredString(width/2, y, "Nộp tại bộ phận tổng vụ")
    c.setFont("NotoSansJP-Light", 9)
    c.drawCentredString(width/2, y-10, "(総務部署で提出)")
    y -= 30

    # Phần checkbox và thông tin cá nhân
    c.setFont("DejaVuSans", 10)
    
    # Dòng checkbox
    checkbox_y = y
    c.rect(margin, checkbox_y-3, 8, 8)  # Checkbox tăng ca
    c.drawString(margin+15, checkbox_y, "Tăng ca /")
    c.setFont("NotoSansJP", 10)
    c.drawString(margin+70, checkbox_y, "残業")
    
    c.rect(margin+200, checkbox_y-3, 8, 8)  # Checkbox đi làm ngày nghỉ
    c.setFont("DejaVuSans", 10)
    c.drawString(margin+215, checkbox_y, "Đi làm ngày nghỉ /")
    c.setFont("NotoSansJP", 10)
    c.drawString(margin+320, checkbox_y, "休日出勤")
    y -= 20

    # Thông tin nhân viên
    c.setFont("NotoSansJP-Light", 10)
    c.drawString(margin, y, f"Họ tên (氏名)：{user.name}")
    c.drawString(margin+200, y, f"Nhóm (チーム)：{user.department}")
    c.drawString(margin+350, y, f"Mã NV (社員コード): {user.employee_id}")
    y -= 15
    
    c.drawString(margin, y, f"Lý do tăng ca (理由): {attendance.note}")
    y -= 15
    
    c.drawString(margin, y, "Đề nghị công ty chấp thuận cho tôi được tăng ca/đi làm vào ngày nghỉ.")
    y -= 10
    c.setFont("NotoSansJP-Light", 9)
    c.drawString(margin, y, "残業/休日出勤を許可お願いします。")
    y -= 25
    
    # Thêm khoảng cách trước khi vẽ bảng thời gian
    y -= 15

    # Bảng chấm công chi tiết
    table_y = y
    table_width = width - 2*margin
    
    # Định nghĩa style cho tiêu đề
    header_style_vn = ParagraphStyle('header_vn', fontName='DejaVuSans', fontSize=8, alignment=1)
    header_style_jp = ParagraphStyle('header_jp', fontName='NotoSansJP', fontSize=8, alignment=1)
    
    # Tạo chuỗi thời gian làm việc
    time_str = f"{attendance.check_in.strftime('%H:%M') if attendance.check_in else '-'} - {attendance.check_out.strftime('%H:%M') if attendance.check_out else '-'}"
    
    # Xác định hình thức (1 hoặc 2)
    holiday_type = getattr(attendance, 'holiday_type', None)
    special_day_types = {'weekend', 'vietnamese_holiday', 'japanese_holiday'}
    form_type = "2" if holiday_type in special_day_types else "1"
    
    # Hàng 1: Tiếng Việt
    header_row1 = [
        Paragraph('No.', header_style_vn),
        Paragraph('NGÀY THÁNG NĂM', header_style_vn),
        Paragraph('HÌNH THỨC', header_style_vn),
        Paragraph('CA LÀM VIỆC', header_style_vn),
        Paragraph('GIỜ VÀO - GIỜ RA', header_style_vn),
        Paragraph('Thời gian nghỉ đối ứng công việc', header_style_vn),
        Paragraph('XÁC NHẬN', header_style_vn)
    ]
    # Hàng 2: Tiếng Nhật/Hán
    header_row2 = [
        Paragraph('', header_style_jp),
        Paragraph('日付', header_style_jp),
        Paragraph('種類', header_style_jp),
        Paragraph('シフト', header_style_jp),
        Paragraph('出勤時間-退勤時間', header_style_jp),
        Paragraph('業務対応時間', header_style_jp),
        Paragraph('ラボマネ承認', header_style_jp)
    ]
    # Hàng dữ liệu
    # Tách riêng thời gian và đối ứng để dễ đọc
    time_info = f"{attendance.check_in.strftime('%H:%M') if attendance.check_in else '-'} - {attendance.check_out.strftime('%H:%M') if attendance.check_out else '-'}"
    
    # Tính tổng thời gian đối ứng - chỉ hiển thị 1 giá trị duy nhất
    total_comp_time = 0.0
    
    # Cộng tất cả các loại đối ứng - SỬ DỤNG CỘT MINUTES MỚI
    total_comp_minutes = 0
    total_comp_minutes += attendance.comp_time_regular_minutes or 0
    total_comp_minutes += attendance.comp_time_overtime_minutes or 0
    total_comp_minutes += attendance.comp_time_ot_before_22_minutes or 0
    total_comp_minutes += attendance.comp_time_ot_after_22_minutes or 0
    total_comp_minutes += attendance.overtime_comp_time_minutes or 0
    
    total_comp_time = total_comp_minutes / 60.0
    
    # Định dạng tổng thời gian đối ứng
    if total_comp_time > 0:
        comp_time_display = attendance._format_hours_minutes(total_comp_time)
    else:
        comp_time_display = "0:00"
    
    # Tạo dữ liệu hàng với thông tin rõ ràng
    row_data = [
        '1',
        attendance.date.strftime('%d/%m/%Y'),
        form_type,
        attendance.shift_code or '-',
        time_info,
        comp_time_display,  # Chỉ hiển thị 1 giá trị tổng thời gian đối ứng
        ''
    ]
    
    table_data = [header_row1, header_row2, row_data]
    col_widths = [30, 80, 50, 65, 80, 110, 70]  # Tổng nhỏ hơn width, luôn còn margin hai bên
    row_heights = [40, 14, 18]  # Hàng dữ liệu bình thường vì chỉ hiển thị 1 giá trị
    
    detail_table_width = sum(col_widths)
    x_detail = (width - detail_table_width) / 2
    table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
    table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'DejaVuSans'),
        ('FONTNAME', (0,1), (-1,1), 'NotoSansJP'),
        ('FONTSIZE', (0,0), (-1,1), 8),
        ('FONTSIZE', (0,2), (-1,2), 9),
        # Xóa dòng kẻ ngang giữa hàng 0 và 1
        ('LINEBELOW', (0,0), (-1,0), 0, colors.white),
    ]))
    table.wrapOn(c, width-2*margin, 50)
    table.drawOn(c, x_detail, table_y - 46)
    y = table_y - 46 - 36  # cập nhật y cho phần tiếp theo
    
    # Ghi chú dưới bảng
    note_sections = [
        ("DejaVuSans", 8, "* Ghi chú: Tại cột Hình thức: Tăng ca ngày bình thường ghi số 1 Đi làm ngày nghỉ, tăng ca ghi số 2"),
        ("NotoSansJP-Light", 8, "備考：平日の残業の場合：1番を記入してください。 休日出勤の場合：2番を記入してください。"),
        ("DejaVuSans", 8, "*Về việc nghỉ giải lao (60 phút) ngày thường trong tuần, trường hợp nếu nghỉ dài hơn vì đối ứng công việc ：Hãy nộp đơn cho bộ phận văn phòng."),
        ("NotoSansJP-Light", 8, "通常（1の場合）の昼休憩（60分）に、休憩途中で業務対応する場合、申請をして下さい。"),
        ("DejaVuSans", 8, "*Trong trường hợp không xin phép trước, thì tăng ca và đi làm ngày nghỉ không được chấp nhận."),
        ("DejaVuSans", 8, "Phải ghi giấy tăng ca sau khi tăng ca (chậm nhất là ngày mai) ,sang ngày mốt ghi tăng ca thì không được chấp nhận."),
        ("NotoSansJP-Light", 8, "※1分単位で申請して下さい。申請をしない限り、残業と休日出勤は反映されません。"),
        ("NotoSansJP-Light", 8, "必ず、残業をした日に申請すること。（次の日までの申請は認めますが、それ以外の申請は認めません）")
    ]
    max_note_width = width - 2*margin - 10
    for i, (font_name, font_size, text) in enumerate(note_sections):
        lines = wrap_text(text, font_name, font_size, max_note_width, c)
        for line in lines:
            c.setFont(font_name, font_size)
            c.drawString(margin, y, line)
            y -= font_size + 1
        # Thêm dòng trắng sau mỗi đoạn bắt đầu bằng * (trừ đoạn cuối)
        if text.startswith('*') and i < len(note_sections)-1:
            y -= font_size + 1
    
    # Thêm khoảng cách giữa phần ghi chú và dòng ngày tháng
    y -= 25
    # Ngày tháng - Đặt ở vị trí cao hơn để không bị đè
    date_y = y + 20  # Đặt dòng ngày tháng cao hơn
    c.setFont("DejaVuSans", 10)
    c.drawRightString(width-margin, date_y, f"Huế, ngày {attendance.date.day} tháng {attendance.date.month} năm {attendance.date.year}")
    y -= 10  # Đẩy dòng ngày tháng xuống thấp hơn
    y -= 95  # Tăng thêm khoảng cách để không bị đè lên phần ghi chú và dòng ngày tháng
    
    # --- Căn chỉnh lại phần chữ ký và tiêu đề phía trên ---
    # Số ô và kích thước - GIẢM KÍCH THƯỚC Ô ĐỂ VỪA TRANG VÀ CÓ BORDER
    num_boxes = 3
    box_width = 140  # Giảm từ 180 xuống 140 để vừa trang
    box_height = 70  # Giảm từ 80 xuống 70 để cân đối
    box_spacing = 30  # Giảm khoảng cách từ 40 xuống 30 để vừa trang
    total_width = num_boxes * box_width + (num_boxes - 1) * box_spacing
    start_x = (width - total_width) / 2
    box_y = y  # y là vị trí đáy các ô
    label_font_size = 10
    sublabel_font_size = 8
    # Tiêu đề các ô
    box_titles = [
        ("Quản lí", "ラボマネジャー"),
        ("Cấp trên trực tiếp", "□室長　□リーダー　□他"),
        ("Người xin phép", "申請者")
    ]
    # Vẽ tiêu đề và sublabel căn giữa trên mỗi ô
    for i, (title, sublabel) in enumerate(box_titles):
        x = start_x + i * (box_width + box_spacing)
        # Căn giữa tiêu đề
        c.setFont("DejaVuSans", label_font_size)
        c.drawCentredString(x + box_width/2, box_y + box_height + 22, title)
        c.setFont("NotoSansJP-Light", sublabel_font_size)
        c.drawCentredString(x + box_width/2, box_y + box_height + 10, sublabel)
    # Vẽ các ô chữ ký với border - SẼ ĐƯỢC VẼ LẠI SAU KHI VẼ CHỮ KÝ
    signature_boxes = []
    for i in range(num_boxes):
        x = start_x + i * (box_width + box_spacing)
        signature_boxes.append((x, box_y, box_width, box_height))
    # Hiển thị chữ ký hoặc (chưa ký) căn giữa trong từng ô
    # Quản lý
    x0 = start_x
    signature_area_height = box_height - 18  # Giảm vùng chữ ký (để lại 18px cho tên)
    signature_y = box_y + 18  # Chữ ký ở phần trên (cách đáy 18px)
    signature_center_y = signature_y + signature_area_height/2 - 8/2  # Căn giữa chữ ký
    name_y = box_y + 8  # Tên người ký ở phần dưới (cách đáy 8px)
    
    if manager_signature:
        print(f"DEBUG: Processing manager signature for PDF")
        debug_signature_data(manager_signature, "manager")
        success = draw_signature_with_proper_scaling(c, manager_signature, x0, signature_y, box_width, signature_area_height)
        if not success:
            print(f"DEBUG: Failed to draw manager signature, creating placeholder")
            create_signature_placeholder(c, x0, signature_y, box_width, signature_area_height, "Lỗi hiển thị")
    else:
        c.setFont("DejaVuSans", 8)
        c.drawCentredString(x0 + box_width/2, signature_center_y, "(chưa ký)")
    
    # Thêm tên người ký quản lý bên trong ô chữ ký (phía dưới chữ ký)
    c.setFont("DejaVuSans", 8)
    c.drawCentredString(x0 + box_width/2, name_y, manager_signer_name)
    
    # Trưởng nhóm
    x1 = start_x + 1 * (box_width + box_spacing)
    
    if team_leader_signature:
        print(f"DEBUG: Processing team leader signature for PDF")
        debug_signature_data(team_leader_signature, "team_leader")
        success = draw_signature_with_proper_scaling(c, team_leader_signature, x1, signature_y, box_width, signature_area_height)
        if not success:
            print(f"DEBUG: Failed to draw team leader signature, creating placeholder")
            create_signature_placeholder(c, x1, signature_y, box_width, signature_area_height, "Lỗi hiển thị")
    else:
        c.setFont("DejaVuSans", 8)
        c.drawCentredString(x1 + box_width/2, signature_center_y, "(chưa ký)")
    
    # Thêm tên người ký trưởng nhóm bên trong ô chữ ký (phía dưới chữ ký)
    c.setFont("DejaVuSans", 8)
    c.drawCentredString(x1 + box_width/2, name_y, team_leader_signer_name)
    
    # Nhân viên
    x2 = start_x + 2 * (box_width + box_spacing)
    
    if employee_signature:
        print(f"DEBUG: Processing employee signature for PDF")
        debug_signature_data(employee_signature, "employee")
        success = draw_signature_with_proper_scaling(c, employee_signature, x2, signature_y, box_width, signature_area_height)
        if not success:
            print(f"DEBUG: Failed to draw employee signature, creating placeholder")
            create_signature_placeholder(c, x2, signature_y, box_width, signature_area_height, "Lỗi hiển thị")
    else:
        c.setFont("DejaVuSans", 8)
        c.drawCentredString(x2 + box_width/2, signature_center_y, "(chưa ký)")
    
    # Thêm tên người ký nhân viên bên trong ô chữ ký (phía dưới chữ ký)
    c.setFont("DejaVuSans", 8)
    c.drawCentredString(x2 + box_width/2, name_y, employee_signer_name)
    
    # Vẽ lại border cho tất cả các ô chữ ký sau khi đã vẽ chữ ký
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.5)
    for x, y, w, h in signature_boxes:
        c.rect(x, y, w, h, stroke=1, fill=0)
    
    c.save()

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        if not email:
            flash('Vui lòng nhập email đã đăng ký!', 'error')
            return render_template('forgot_password.html')
        user = User.query.filter_by(email=email).first()
        if not user:
            flash('Không tìm thấy tài khoản với email này!', 'error')
            return render_template('forgot_password.html')
        # Tạo token
        import secrets
        token = secrets.token_urlsafe(48)
        expires_at = datetime.utcnow() + timedelta(hours=1)
        reset_token = PasswordResetToken(user_id=user.id, token=token, expires_at=expires_at)
        db.session.add(reset_token)
        db.session.commit()
        # Gửi email
        reset_link = url_for('reset_password', token=token, _external=True)
        email_sent = send_reset_email(user.email, user.name, reset_link)
        if email_sent:
            flash('Đã gửi email hướng dẫn đặt lại mật khẩu. Vui lòng kiểm tra hộp thư!', 'success')
        else:
            flash('Không thể gửi email. Vui lòng liên hệ quản trị viên để được hỗ trợ.', 'error')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    reset_token = PasswordResetToken.query.filter_by(token=token, used=False).first()
    if not reset_token or reset_token.is_expired():
        flash('Link đặt lại mật khẩu không hợp lệ hoặc đã hết hạn!', 'error')
        return redirect(url_for('login'))
    user = db.session.get(User, reset_token.user_id)
    if request.method == 'POST':
        password = request.form.get('password')
        confirm = request.form.get('confirm')
        if not password or not confirm or password != confirm:
            flash('Mật khẩu không khớp hoặc không hợp lệ!', 'error')
            return render_template('reset_password.html', token=token)
        user.set_password(password)
        db.session.commit()
        reset_token.used = True
        db.session.commit()
        flash('Đặt lại mật khẩu thành công! Bạn có thể đăng nhập lại.', 'success')
        return redirect(url_for('login'))
    return render_template('reset_password.html', token=token)

def send_reset_email(to_email, to_name, reset_link):
    # Cấu hình SMTP từ config
    smtp_server = app.config['SMTP_SERVER']
    smtp_port = app.config['SMTP_PORT']
    smtp_user = app.config['SMTP_USER']
    smtp_password = app.config['SMTP_PASSWORD']
    from_email = app.config['MAIL_FROM']
    
    # Kiểm tra và đặt giá trị mặc định cho from_email nếu không có
    if not from_email:
        from_email = smtp_user if smtp_user else 'noreply@dmi.com'
    
    subject = 'Đặt lại mật khẩu hệ thống chấm công DMI'
    
    # Plain text version
    text_body = f"""Xin chào {to_name},

Bạn vừa yêu cầu đặt lại mật khẩu cho tài khoản hệ thống chấm công DMI.

Vui lòng copy link dưới đây vào trình duyệt để đặt lại mật khẩu (có hiệu lực trong 1 giờ):

{reset_link}

Lưu ý: Nếu bạn không yêu cầu đặt lại mật khẩu, vui lòng bỏ qua email này.

Trân trọng,
Hệ thống chấm công DMI"""
    
    # HTML version
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Đặt lại mật khẩu</title>
    </head>
    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background-color: #f9f9f9; padding: 30px; border-radius: 10px; border: 1px solid #ddd;">
            <h2 style="color: #1976d2; margin-bottom: 20px;">Đặt lại mật khẩu hệ thống chấm công DMI</h2>
            
            <p>Xin chào <strong>{to_name}</strong>,</p>
            
            <p>Bạn vừa yêu cầu đặt lại mật khẩu cho tài khoản hệ thống chấm công DMI.</p>
            
            <p>Vui lòng nhấn vào link dưới đây để đặt lại mật khẩu (có hiệu lực trong 1 giờ):</p>
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_link}" style="background-color: #1976d2; color: white; padding: 12px 24px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">Đặt lại mật khẩu</a>
            </div>
            
            <p style="font-size: 14px; color: #666;">Hoặc copy link này vào trình duyệt:</p>
            <p style="word-break: break-all; background-color: #f5f5f5; padding: 10px; border-radius: 5px; font-size: 12px;">{reset_link}</p>
            
            <p style="color: #d32f2f; font-size: 14px;"><strong>Lưu ý:</strong> Nếu bạn không yêu cầu đặt lại mật khẩu, vui lòng bỏ qua email này.</p>
            
            <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
            
            <p style="color: #666; font-size: 14px;">Trân trọng,<br>
            <strong>Hệ thống chấm công DMI</strong></p>
        </div>
    </body>
    </html>
    """
    
    # Kiểm tra cấu hình SMTP trước khi gửi email
    if not all([smtp_server, smtp_port, smtp_user, smtp_password]):
        print('SMTP configuration incomplete. Cannot send email.')
        return False
    
    # Create multipart message
    from email.mime.multipart import MIMEMultipart
    msg = MIMEMultipart('alternative')
    msg['Subject'] = Header(subject, 'utf-8')
    msg['From'] = formataddr((str(Header('DMI Attendance', 'utf-8')), from_email))
    msg['To'] = to_email
    
    # Attach both plain text and HTML versions
    text_part = MIMEText(text_body, 'plain', 'utf-8')
    html_part = MIMEText(html_body, 'html', 'utf-8')
    
    msg.attach(text_part)
    msg.attach(html_part)
    
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.sendmail(from_email, [to_email], msg.as_string())
        server.quit()
        return True
    except Exception as e:
        print('Email send error:', e)
        # Không raise để không lộ thông tin cho user
        return False

# Đổi mật khẩu khi đã đăng nhập
@app.route('/change-password-legacy', methods=['GET', 'POST'])
@login_required
def change_password_legacy():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = db.session.get(User, session['user_id'])
    if request.method == 'POST':
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        confirm = request.form.get('confirm')
        if not user.check_password(old_password):
            flash('Mật khẩu cũ không đúng!', 'error')
            return render_template('change_password.html')
        if not new_password or new_password != confirm:
            flash('Mật khẩu mới không hợp lệ hoặc không khớp!', 'error')
            return render_template('change_password.html')
        user.set_password(new_password)
        db.session.commit()
        flash('Đổi mật khẩu thành công!', 'success')
        return redirect(url_for('dashboard'))
    return render_template('change_password.html')

@app.route('/api/signature/debug/<int:attendance_id>')
@require_admin
def debug_signature(attendance_id):
    """Debug endpoint để kiểm tra chữ ký trong database"""
    try:
        attendance = db.session.get(Attendance, attendance_id)
        if not attendance:
            return jsonify({'error': 'Attendance not found'}), 404
        
        debug_info = {
            'attendance_id': attendance_id,
            'employee_signature': {
                'exists': bool(attendance.signature),
                'length': len(attendance.signature) if attendance.signature else 0,
                'type': type(attendance.signature).__name__ if attendance.signature else None,
                'starts_with_data_image': attendance.signature.startswith('data:image') if attendance.signature else False,
                'processed': process_signature_for_pdf(attendance.signature) is not None if attendance.signature else False
            },
            'team_leader_signature': {
                'exists': bool(attendance.team_leader_signature),
                'length': len(attendance.team_leader_signature) if attendance.team_leader_signature else 0,
                'type': type(attendance.team_leader_signature).__name__ if attendance.team_leader_signature else None,
                'starts_with_data_image': attendance.team_leader_signature.startswith('data:image') if attendance.team_leader_signature else False,
                'processed': process_signature_for_pdf(attendance.team_leader_signature) is not None if attendance.team_leader_signature else False
            },
            'manager_signature': {
                'exists': bool(attendance.manager_signature),
                'length': len(attendance.manager_signature) if attendance.manager_signature else 0,
                'type': type(attendance.manager_signature).__name__ if attendance.manager_signature else None,
                'starts_with_data_image': attendance.manager_signature.startswith('data:image') if attendance.manager_signature else False,
                'processed': process_signature_for_pdf(attendance.manager_signature) is not None if attendance.manager_signature else False
            }
        }
        
        # Thêm debug chi tiết cho từng chữ ký
        if attendance.signature:
            debug_signature_data(attendance.signature, "employee_debug")
        if attendance.team_leader_signature:
            debug_signature_data(attendance.team_leader_signature, "team_leader_debug")
        if attendance.manager_signature:
            debug_signature_data(attendance.manager_signature, "manager_debug")
        
        return jsonify(debug_info)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Thêm hàm debug chữ ký
def debug_signature_data(signature_data, source="unknown"):
    """Debug chi tiết dữ liệu chữ ký"""
    print(f"=== DEBUG SIGNATURE DATA ({source}) ===")
    if not signature_data:
        print("Signature data is None or empty")
        return
    
    print(f"Type: {type(signature_data)}")
    print(f"Length: {len(signature_data)}")
    
    if isinstance(signature_data, str):
        print(f"Starts with 'data:image': {signature_data.startswith('data:image')}")
        print(f"First 100 chars: {signature_data[:100]}")
        print(f"Last 100 chars: {signature_data[-100:]}")
        
        # Kiểm tra có phải base64 không
        try:
            decoded = base64.b64decode(signature_data)
            print(f"Valid base64: Yes, decoded length: {len(decoded)}")
        except Exception:
            print("Valid base64: No")
            
            # Thử giải mã nếu có thể
            try:
                decrypted = signature_manager.decrypt_signature(signature_data)
                if decrypted:
                    print(f"Decrypted successfully, length: {len(decrypted)}")
                    print(f"Decrypted starts with 'data:image': {decrypted.startswith('data:image')}")
                else:
                    print("Decryption failed or returned empty")
            except Exception as e:
                print(f"Decryption error: {e}")
    
    print("=== END DEBUG ===")

@app.route('/personal-signature', methods=['GET', 'POST'])
@login_required
def personal_signature():
    """Trang quản lý chữ ký cá nhân"""
    if request.method == 'POST':
        signature = request.form.get('signature')
        if signature:
            # 使用签名处理器优化签名质量
            processed_signature = signature_manager.process_signature_for_display(signature)
            
            user = db.session.get(User, session['user_id'])
            user.personal_signature = processed_signature
            db.session.commit()
            
            # 记录签名操作
            signature_manager.log_signature_action(
                user_id=user.id,
                action='UPDATE_PERSONAL',
                signature_type='personal_signature'
            )
            
            flash('Đã cập nhật chữ ký cá nhân thành công! Hệ thống đã tự động tối ưu hóa chất lượng chữ ký.', 'success')
            return redirect(url_for('personal_signature'))
    
    user = db.session.get(User, session['user_id'])
    return render_template('personal_signature.html', user=user)

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    """Trang cài đặt thông tin cá nhân và chữ ký"""
    print("DEBUG: Settings route accessed")
    print("DEBUG: Session user_id:", session.get('user_id'))
    print("DEBUG: Session keys:", list(session.keys()))
    
    if 'user_id' not in session:
        print("DEBUG: No user_id in session, redirecting to login")
        return redirect(url_for('login'))
    
    user = db.session.get(User, session['user_id'])
    if not user:
        print("DEBUG: User not found, redirecting to login")
        session.clear()
        flash('Phiên đăng nhập không hợp lệ!', 'error')
        return redirect(url_for('login'))
    
    # Kiểm tra user có active không
    if not user.is_active:
        session.clear()
        flash('Tài khoản đã bị khóa!', 'error')
        return redirect(url_for('login'))
    
    # Kiểm tra session timeout
    if check_session_timeout():
        flash('Phiên đăng nhập đã hết hạn!', 'error')
        return redirect(url_for('login'))
    
    # Cập nhật thời gian hoạt động cuối
    update_session_activity()
    
    print("DEBUG: User found:", user.name)
    
    if request.method == 'POST':
        # Phân nhánh action: lưu chữ ký hoặc đổi mật khẩu
        action = request.form.get('action')
        if action == 'change_password':
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')

            if not current_password or not new_password or not confirm_password:
                flash('Vui lòng nhập đầy đủ mật khẩu hiện tại và mật khẩu mới', 'error')
                return redirect(url_for('settings'))
            if new_password != confirm_password:
                flash('Xác nhận mật khẩu mới không khớp', 'error')
                return redirect(url_for('settings'))

            # Verify current password
            try:
                from werkzeug.security import check_password_hash, generate_password_hash
                if not user.password or not check_password_hash(user.password, current_password):
                    flash('Mật khẩu hiện tại không đúng', 'error')
                    return redirect(url_for('settings'))
                # Update password
                user.password = generate_password_hash(new_password)
                db.session.commit()
                flash('Đổi mật khẩu thành công', 'success')
                return redirect(url_for('settings'))
            except Exception as e:
                db.session.rollback()
                flash('Có lỗi khi đổi mật khẩu', 'error')
                return redirect(url_for('settings'))
        else:
            # Lưu chữ ký cá nhân
            signature = request.form.get('signature')
            if signature:
                user.personal_signature = signature
                try:
                    db.session.commit()
                    flash('Lưu chữ ký thành công!', 'success')
                    return redirect(url_for('settings'))
                except Exception as e:
                    db.session.rollback()
                    flash('Đã xảy ra lỗi khi lưu chữ ký', 'error')
            else:
                flash('Chưa có chữ ký để lưu', 'error')
    
    return render_template('settings.html', user=user)

@app.route('/settings/check-password', methods=['POST'])
def check_current_password():
    """AJAX: Kiểm tra mật khẩu hiện tại có đúng không"""
    if 'user_id' not in session:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401
    try:
        user = db.session.get(User, session['user_id'])
        if not user:
            return jsonify({'ok': False, 'error': 'User not found'}), 404
        data = request.get_json(silent=True) or {}
        current_password = (data.get('current_password') or '').strip()
        from werkzeug.security import check_password_hash
        if user.password and check_password_hash(user.password, current_password):
            return jsonify({'ok': True})
        return jsonify({'ok': False}), 200
    except Exception as e:
        return jsonify({'ok': False, 'error': 'Internal error'}), 500

@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    """Trang đổi mật khẩu riêng biệt"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = db.session.get(User, session['user_id'])
    if not user:
        session.clear()
        flash('Phiên đăng nhập không hợp lệ!', 'error')
        return redirect(url_for('login'))
    if check_session_timeout():
        flash('Phiên đăng nhập đã hết hạn!', 'error')
        return redirect(url_for('login'))
    update_session_activity()

    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        if not current_password or not new_password or not confirm_password:
            flash('Vui lòng nhập đầy đủ mật khẩu', 'error')
            return redirect(url_for('change_password'))
        if new_password != confirm_password:
            flash('Xác nhận mật khẩu mới không khớp', 'error')
            return redirect(url_for('change_password'))
        from werkzeug.security import check_password_hash, generate_password_hash
        # Hệ thống dùng trường password_hash và method check_password/set_password trong model User
        # Ưu tiên dùng method nếu có, fallback dùng trực tiếp password_hash
        try:
            valid_current = user.check_password(current_password)
        except Exception:
            valid_current = bool(getattr(user, 'password_hash', None) and check_password_hash(user.password_hash, current_password))
        if not valid_current:
            flash('Mật khẩu hiện tại không đúng', 'error')
            return redirect(url_for('change_password'))
        try:
            try:
                user.set_password(new_password)
            except Exception:
                user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            flash('Đổi mật khẩu thành công', 'success')
            return redirect(url_for('settings'))
        except Exception:
            db.session.rollback()
            flash('Có lỗi khi đổi mật khẩu', 'error')
            return redirect(url_for('change_password'))

    return render_template('change_password.html', user=user)

@app.route('/signature-test', methods=['GET', 'POST'])
def signature_test():
    """Trang test hiển thị chữ ký cho cả 3 vai trò"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = db.session.get(User, session['user_id'])
    if not user:
        session.clear()
        flash('Phiên đăng nhập không hợp lệ!', 'error')
        return redirect(url_for('login'))
    
    # Kiểm tra user có active không
    if not user.is_active:
        session.clear()
        flash('Tài khoản đã bị khóa!', 'error')
        return redirect(url_for('login'))
    
    # Kiểm tra session timeout
    if check_session_timeout():
        flash('Phiên đăng nhập đã hết hạn!', 'error')
        return redirect(url_for('login'))
    
    # Cập nhật thời gian hoạt động cuối
    update_session_activity()
    
    if request.method == 'POST':
        # Xử lý lưu test chữ ký
        employee_signature = request.form.get('employee_signature')
        team_leader_signature = request.form.get('team_leader_signature')
        manager_signature = request.form.get('manager_signature')
        test_date = request.form.get('test_date')
        test_note = request.form.get('test_note', 'Test hiển thị chữ ký')
        
        # Lưu vào session để sử dụng cho PDF
        session['test_signatures'] = {
            'employee': employee_signature,
            'team_leader': team_leader_signature,
            'manager': manager_signature,
            'date': test_date,
            'note': test_note
        }
        
        flash('Đã lưu test chữ ký thành công!', 'success')
        return redirect(url_for('signature_test'))
    
    return render_template('signature_test.html', user=user, today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/signature-test/download-pdf', methods=['POST'])
def download_signature_test_pdf():
    """Tải PDF test chữ ký"""
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    # Lấy dữ liệu chữ ký từ form
    employee_signature = request.form.get('employee_signature')
    team_leader_signature = request.form.get('team_leader_signature')
    manager_signature = request.form.get('manager_signature')
    test_date = request.form.get('test_date')
    test_note = request.form.get('test_note', 'Test hiển thị chữ ký')
    
    # Tạo buffer cho PDF
    buffer = io.BytesIO()
    
    # Đăng ký font
    register_pdf_fonts()
    
    # Tạo PDF
    canvas_obj = canvas.Canvas(buffer, pagesize=A4)
    canvas_obj.setTitle('Test Chữ ký - DMI Attendance')
    
    # Header
    canvas_obj.setFont('NotoSansJP-Bold', 18)
    canvas_obj.drawString(50, 800, 'TEST HIỂN THỊ CHỮ KÝ')
    canvas_obj.setFont('NotoSansJP-Regular', 12)
    canvas_obj.drawString(50, 780, f'Ngày test: {test_date}')
    canvas_obj.drawString(50, 760, f'Ghi chú: {test_note}')
    canvas_obj.drawString(50, 740, f'Người tạo: {user.name}')
    
    # Vẽ đường kẻ
    canvas_obj.line(50, 720, 550, 720)
    
    # Chữ ký Nhân viên
    y_position = 680
    canvas_obj.setFont('NotoSansJP-Bold', 14)
    canvas_obj.drawString(50, y_position, '1. Chữ ký Nhân viên:')
    
    if employee_signature:
        try:
            draw_signature_with_proper_scaling(canvas_obj, employee_signature, 50, y_position - 80, 200, 60)
        except Exception as e:
            print(f"Error drawing employee signature: {e}")
            create_signature_placeholder(canvas_obj, 50, y_position - 80, 200, 60, "Lỗi hiển thị")
    else:
        create_signature_placeholder(canvas_obj, 50, y_position - 80, 200, 60, "Chưa có chữ ký")
    
    # Chữ ký Trưởng nhóm
    y_position = 540
    canvas_obj.setFont('NotoSansJP-Bold', 14)
    canvas_obj.drawString(50, y_position, '2. Chữ ký Trưởng nhóm:')
    
    if team_leader_signature:
        try:
            draw_signature_with_proper_scaling(canvas_obj, team_leader_signature, 50, y_position - 80, 200, 60)
        except Exception as e:
            print(f"Error drawing team leader signature: {e}")
            create_signature_placeholder(canvas_obj, 50, y_position - 80, 200, 60, "Lỗi hiển thị")
    else:
        create_signature_placeholder(canvas_obj, 50, y_position - 80, 200, 60, "Chưa có chữ ký")
    
    # Chữ ký Quản lý
    y_position = 400
    canvas_obj.setFont('NotoSansJP-Bold', 14)
    canvas_obj.drawString(50, y_position, '3. Chữ ký Quản lý:')
    
    if manager_signature:
        try:
            draw_signature_with_proper_scaling(canvas_obj, manager_signature, 50, y_position - 80, 200, 60)
        except Exception as e:
            print(f"Error drawing manager signature: {e}")
            create_signature_placeholder(canvas_obj, 50, y_position - 80, 200, 60, "Lỗi hiển thị")
    else:
        create_signature_placeholder(canvas_obj, 50, y_position - 80, 200, 60, "Chưa có chữ ký")
    
    # Footer
    y_position = 200
    canvas_obj.line(50, y_position, 550, y_position)
    canvas_obj.setFont('NotoSansJP-Regular', 10)
    canvas_obj.drawString(50, y_position - 20, f'Được tạo bởi: {user.name} - {user.employee_id}')
    canvas_obj.drawString(50, y_position - 40, f'Thời gian: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
    
    canvas_obj.save()
    buffer.seek(0)
    
    # Tạo response
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=test_chu_ky_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@app.route('/settings/test-signature-pdf', methods=['POST'])
def test_signature_pdf():
    """Tạo PDF test chữ ký cá nhân trên mẫu phiếu tăng ca thực tế"""
    print("DEBUG: test_signature_pdf route accessed")
    print("DEBUG: Session user_id:", session.get('user_id'))
    print("DEBUG: Form data:", request.form)
    
    if 'user_id' not in session:
        # print("DEBUG: No user_id in session")
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    user = db.session.get(User, session['user_id'])
    if not user:
        print("DEBUG: User not found")
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    signature = request.form.get('signature')
    if not signature:
        print("DEBUG: No signature provided")
        return jsonify({'error': 'Chưa có chữ ký'}), 400
    
    print("DEBUG: Signature length:", len(signature) if signature else 0)
    try:
        # Tạo buffer cho PDF
        buffer = io.BytesIO()
        register_pdf_fonts()
        canvas_obj = canvas.Canvas(buffer, pagesize=A4)
        canvas_obj.setTitle('Test Chữ ký trên Phiếu Tăng Ca - DMI Attendance')
        print("DEBUG: PDF canvas created successfully")
        
        # Tạo mẫu phiếu tăng ca với chữ ký test
        create_overtime_test_pdf(canvas_obj, user, signature)
        
        canvas_obj.save()
        buffer.seek(0)
        print("DEBUG: PDF created successfully, size:", len(buffer.getvalue()))
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=test_phieu_tang_ca_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return response
    except Exception as e:
        print(f"DEBUG: Error creating PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi tạo PDF: {str(e)}'}), 500

def create_overtime_test_pdf(canvas_obj, user, signature):
    """Tạo PDF test với mẫu phiếu tăng ca thực tế"""
    width, height = A4
    margin = 30
    y = height - margin

    # Header: Bảng 6 cột như trong mẫu thực tế - sử dụng font an toàn
    header_data = [
        [
            Paragraph('<b>DMI HUẾ</b>', ParagraphStyle('h', fontName='DejaVuSans', fontSize=9, alignment=1)),
            Paragraph('<b>総務<br/>TONG VU</b>', ParagraphStyle('h', fontName='NotoSansJP', fontSize=8, alignment=1)),
            Paragraph('<b>分類番号：<br/>So hieu phan loai：</b>', ParagraphStyle('h', fontName='NotoSansJP', fontSize=7, alignment=1)),
            Paragraph('', ParagraphStyle('h', fontName='DejaVuSans', fontSize=8, alignment=1)),
            Paragraph('<b>記入 FORM<br/>NHAP FORM</b>', ParagraphStyle('h', fontName='NotoSansJP', fontSize=8, alignment=1)),
            Paragraph('<b>Form作成：<br/>Tac thanh：</b>', ParagraphStyle('h', fontName='NotoSansJP', fontSize=7, alignment=1)),
            Paragraph('', ParagraphStyle('h', fontName='DejaVuSans', fontSize=8, alignment=1)),
            Paragraph('', ParagraphStyle('h', fontName='DejaVuSans', fontSize=8, alignment=1)),
        ]
    ]
    
    col_widths = [60, 80, 100, 50, 80, 80, 50, 50]
    header_table_width = sum(col_widths)
    x_header = (width - header_table_width) / 2
    header_table = Table(header_data, colWidths=col_widths, rowHeights=25)
    header_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
    ]))
    header_table.wrapOn(canvas_obj, width-2*margin, 30)
    header_table.drawOn(canvas_obj, x_header, y-25)
    y -= 40

    # Thông tin công ty
    canvas_obj.setFont("DejaVuSans", 10)
    canvas_obj.drawString(margin, y, "Công ty TNHH DMI HUẾ")
    y -= 12
    canvas_obj.setFont("DejaVuSans", 8)
    canvas_obj.drawString(margin, y, "174 Bà Triệu- tòa nhà 4 tầng Phong Phú Plaza, phường Phú Hội, Thành phố Huế, Tỉnh Thừa Thiên Huế,Việt Nam.")
    y -= 25

    # Tiêu đề chính
    canvas_obj.setFont("DejaVuSans", 14)
    canvas_obj.drawCentredString(width/2, y, "GIẤY ĐỀ NGHỊ TĂNG CA/ĐI LÀM NGÀY NGHỈ")
    y -= 16
    canvas_obj.setFont("NotoSansJP", 11)
    canvas_obj.drawCentredString(width/2, y, "(残業/休日出勤申請書)")
    y -= 20
    canvas_obj.setFont("DejaVuSans", 9)
    canvas_obj.drawCentredString(width/2, y, "Nộp tại bộ phận tổng vụ")
    canvas_obj.setFont("NotoSansJP-Light", 9)
    canvas_obj.drawCentredString(width/2, y-10, "(総務部署で提出)")
    y -= 30

    # Phần checkbox và thông tin cá nhân
    canvas_obj.setFont("DejaVuSans", 10)
    
    # Dòng checkbox
    checkbox_y = y
    canvas_obj.rect(margin, checkbox_y-3, 8, 8)  # Checkbox tăng ca
    canvas_obj.drawString(margin+15, checkbox_y, "Tăng ca /")
    canvas_obj.setFont("NotoSansJP", 10)
    canvas_obj.drawString(margin+70, checkbox_y, "残業")
    
    canvas_obj.rect(margin+200, checkbox_y-3, 8, 8)  # Checkbox đi làm ngày nghỉ
    canvas_obj.setFont("DejaVuSans", 10)
    canvas_obj.drawString(margin+215, checkbox_y, "Đi làm ngày nghỉ /")
    canvas_obj.setFont("NotoSansJP", 10)
    canvas_obj.drawString(margin+320, checkbox_y, "休日出勤")
    y -= 20

    # Thông tin nhân viên
    canvas_obj.setFont("DejaVuSans", 10)
    canvas_obj.drawString(margin, y, f"Họ tên: {user.name}")
    canvas_obj.drawString(margin+200, y, f"Nhóm: {user.department}")
    canvas_obj.drawString(margin+350, y, f"Mã NV: {user.employee_id}")
    y -= 15
    canvas_obj.drawString(margin, y, f"Lý do tăng ca: Test hiển thị chữ ký trên phiếu tăng ca")
    y -= 15
    canvas_obj.drawString(margin, y, "Đề nghị công ty chấp thuận cho tôi được tăng ca/đi làm vào ngày nghỉ.")
    y -= 10
    canvas_obj.setFont("NotoSansJP-Light", 9)
    canvas_obj.drawString(margin, y, "残業/休日出勤を許可お願いします。")
    y -= 25
    
    # Thêm khoảng cách trước khi vẽ bảng thời gian
    y -= 15

    # Bảng chấm công chi tiết
    table_y = y
    table_width = width - 2*margin
    
    # Định nghĩa style cho tiêu đề
    header_style_vn = ParagraphStyle('header_vn', fontName='DejaVuSans', fontSize=8, alignment=1)
    header_style_jp = ParagraphStyle('header_jp', fontName='NotoSansJP', fontSize=8, alignment=1)
    
    # Tạo chuỗi thời gian làm việc mẫu
    time_str = "18:00 - 22:00"
    
    # Hàng 1: Tiếng Việt
    header_row1 = [
        Paragraph('No.', header_style_vn),
        Paragraph('NGÀY THÁNG NĂM', header_style_vn),
        Paragraph('HÌNH THỨC', header_style_vn),
        Paragraph('CA LÀM VIỆC', header_style_vn),
        Paragraph('GIỜ VÀO - GIỜ RA', header_style_vn),
        Paragraph('Thời gian nghỉ đối ứng công việc', header_style_vn),
        Paragraph('XÁC NHẬN', header_style_vn)
    ]
    # Hàng 2: Tiếng Nhật/Hán
    header_row2 = [
        Paragraph('', header_style_jp),
        Paragraph('日付', header_style_jp),
        Paragraph('種類', header_style_jp),
        Paragraph('シフト', header_style_jp),
        Paragraph('出勤時間-退勤時間', header_style_jp),
        Paragraph('業務対応時間', header_style_jp),
        Paragraph('ラボマネ承認', header_style_jp)
    ]
    # Hàng dữ liệu mẫu - chỉ hiển thị giá trị thời gian
    row_data = [
        '1',
        '15/07/2025',
        '1',
        'Tăng ca',
        time_str,
        '3:30',  # Chỉ hiển thị 1 giá trị tổng thời gian đối ứng (0:30 + 2:00 + 1:00 = 3:30)
        ''
    ]
    
    table_data = [header_row1, header_row2, row_data]
    col_widths = [30, 80, 50, 65, 80, 110, 70]
    row_heights = [40, 14, 18]  # Hàng dữ liệu bình thường vì chỉ hiển thị 1 giá trị
    
    detail_table_width = sum(col_widths)
    x_detail = (width - detail_table_width) / 2
    table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
    table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'DejaVuSans'),
        ('FONTNAME', (0,1), (-1,1), 'NotoSansJP'),
        ('FONTSIZE', (0,0), (-1,1), 8),
        ('FONTSIZE', (0,2), (-1,2), 9),
        ('LINEBELOW', (0,0), (-1,0), 0, colors.white),
    ]))
    table.wrapOn(canvas_obj, width-2*margin, 50)
    table.drawOn(canvas_obj, x_detail, table_y - 46)
    y = table_y - 46 - 36
    
    # Ghi chú dưới bảng
    note_sections = [
        ("DejaVuSans", 8, "* Ghi chú: Tại cột Hình thức: Tăng ca ngày bình thường ghi số 1 Đi làm ngày nghỉ, tăng ca ghi số 2"),
        ("NotoSansJP-Light", 8, "備考：平日の残業の場合：1番を記入してください。 休日出勤の場合：2番を記入してください。"),
        ("DejaVuSans", 8, "*Về việc nghỉ giải lao (60 phút) ngày thường trong tuần, trường hợp nếu nghỉ dài hơn vì đối ứng công việc ：Hãy nộp đơn cho bộ phận văn phòng."),
        ("NotoSansJP-Light", 8, "通常（1の場合）の昼休憩（60分）に、休憩途中で業務対応する場合、申請をして下さい。"),
        ("DejaVuSans", 8, "*Trong trường hợp không xin phép trước, thì tăng ca và đi làm ngày nghỉ không được chấp nhận."),
        ("DejaVuSans", 8, "Phải ghi giấy tăng ca sau khi tăng ca (chậm nhất là ngày mai) ,sang ngày mốt ghi tăng ca thì không được chấp nhận."),
        ("NotoSansJP-Light", 8, "※1分単位で申請して下さい。申請をしない限り、残業と休日出勤は反映されません。"),
        ("NotoSansJP-Light", 8, "必ず、残業をした日に申請すること。（次の日までの申請は認めますが、それ以外の申請は認めません）")
    ]
    max_note_width = width - 2*margin - 10
    for i, (font_name, font_size, text) in enumerate(note_sections):
        lines = wrap_text(text, font_name, font_size, max_note_width, canvas_obj)
        for line in lines:
            canvas_obj.setFont(font_name, font_size)
            canvas_obj.drawString(margin, y, line)
            y -= font_size + 1
        if text.startswith('*') and i < len(note_sections)-1:
            y -= font_size + 1
    
    # Thêm khoảng cách giữa phần ghi chú và dòng ngày tháng
    y -= 25
    # Ngày tháng - Đặt ở vị trí cao hơn để không bị đè
    date_y = y + 20  # Đặt dòng ngày tháng cao hơn
    canvas_obj.setFont("DejaVuSans", 10)
    canvas_obj.drawRightString(width-margin, date_y, f"Huế, ngày 15 tháng 07 năm 2025")
    y -= 10
    y -= 95  # Tăng thêm khoảng cách để không bị đè lên phần ghi chú và dòng ngày tháng
    
    # Số ô và kích thước chữ ký - GIẢM KÍCH THƯỚC Ô ĐỂ VỪA TRANG VÀ CÓ BORDER
    num_boxes = 3
    box_width = 140  # Giảm từ 180 xuống 140 để vừa trang
    box_height = 70  # Giảm từ 80 xuống 70 để cân đối
    box_spacing = 30  # Giảm khoảng cách từ 40 xuống 30 để vừa trang
    total_width = num_boxes * box_width + (num_boxes - 1) * box_spacing
    start_x = (width - total_width) / 2
    box_y = y
    label_font_size = 10
    sublabel_font_size = 8
    
    # Tiêu đề các ô
    box_titles = [
        ("Quản lí", "ラボマネジャー"),
        ("Cấp trên trực tiếp", "□室長　□リーダー　□他"),
        ("Người xin phép", "申請者")
    ]
    
    # Vẽ tiêu đề và sublabel căn giữa trên mỗi ô
    for i, (title, sublabel) in enumerate(box_titles):
        x = start_x + i * (box_width + box_spacing)
        canvas_obj.setFont("DejaVuSans", label_font_size)
        canvas_obj.drawCentredString(x + box_width/2, box_y + box_height + 22, title)
        canvas_obj.setFont("NotoSansJP-Light", sublabel_font_size)
        canvas_obj.drawCentredString(x + box_width/2, box_y + box_height + 10, sublabel)
    
    # Vẽ các ô chữ ký với border
    signature_boxes = []
    for i in range(num_boxes):
        x = start_x + i * (box_width + box_spacing)
        signature_boxes.append((x, box_y, box_width, box_height))
    
    # Hiển thị chữ ký trong từng ô
    signature_area_height = box_height - 18  # Giảm vùng chữ ký (để lại 18px cho tên)
    signature_y = box_y + 18  # Chữ ký ở phần trên (cách đáy 18px)
    signature_center_y = signature_y + signature_area_height/2 - 8/2  # Căn giữa chữ ký
    name_y = box_y + 8  # Tên người ký ở phần dưới (cách đáy 8px)
    
    # Quản lý
    x0 = start_x
    
    print("DEBUG: Processing manager signature for PDF")
    try:
        success = draw_signature_with_proper_scaling(canvas_obj, signature, x0, signature_y, box_width, signature_area_height)
        if not success:
            print("DEBUG: Failed to draw manager signature, creating placeholder")
            create_signature_placeholder(canvas_obj, x0, signature_y, box_width, signature_area_height, "Lỗi hiển thị")
    except Exception as e:
        print(f"Error drawing manager signature: {e}")
        create_signature_placeholder(canvas_obj, x0, signature_y, box_width, signature_area_height, "Lỗi")
    
    # Thêm tên người ký quản lý bên trong ô chữ ký (phía dưới chữ ký)
    canvas_obj.setFont("DejaVuSans", 8)
    canvas_obj.drawCentredString(x0 + box_width/2, name_y, "Test Quản lý")
    
    # Trưởng nhóm
    x1 = start_x + 1 * (box_width + box_spacing)
    
    print("DEBUG: Processing team leader signature for PDF")
    try:
        success = draw_signature_with_proper_scaling(canvas_obj, signature, x1, signature_y, box_width, signature_area_height)
        if not success:
            print("DEBUG: Failed to draw team leader signature, creating placeholder")
            create_signature_placeholder(canvas_obj, x1, signature_y, box_width, signature_area_height, "Lỗi hiển thị")
    except Exception as e:
        print(f"Error drawing team leader signature: {e}")
        create_signature_placeholder(canvas_obj, x1, signature_y, box_width, signature_area_height, "Lỗi")
    
    # Thêm tên người ký trưởng nhóm bên trong ô chữ ký (phía dưới chữ ký)
    canvas_obj.setFont("DejaVuSans", 8)
    canvas_obj.drawCentredString(x1 + box_width/2, name_y, "Test Trưởng nhóm")
    
    # Nhân viên
    x2 = start_x + 2 * (box_width + box_spacing)
    
    print("DEBUG: Processing employee signature for PDF")
    try:
        success = draw_signature_with_proper_scaling(canvas_obj, signature, x2, signature_y, box_width, signature_area_height)
        if not success:
            print("DEBUG: Failed to draw employee signature, creating placeholder")
            create_signature_placeholder(canvas_obj, x2, signature_y, box_width, signature_area_height, "Lỗi hiển thị")
    except Exception as e:
        print(f"Error drawing employee signature: {e}")
        create_signature_placeholder(canvas_obj, x2, signature_y, box_width, signature_area_height, "Lỗi")
    
    # Thêm tên người ký nhân viên bên trong ô chữ ký (phía dưới chữ ký)
    canvas_obj.setFont("DejaVuSans", 8)
    canvas_obj.drawCentredString(x2 + box_width/2, name_y, user.name)
    
    # Vẽ lại border cho tất cả các ô chữ ký sau khi đã vẽ chữ ký
    canvas_obj.setStrokeColor(colors.black)
    canvas_obj.setLineWidth(0.5)
    for x, y, w, h in signature_boxes:
        canvas_obj.rect(x, y, w, h, stroke=1, fill=0)
    
    # Thêm ghi chú test ở cuối
    canvas_obj.setFont("DejaVuSans", 8)
    canvas_obj.drawString(margin, 50, "*** Đây là PDF test để kiểm tra hiển thị chữ ký cá nhân trên mẫu phiếu tăng ca thực tế ***")
    canvas_obj.drawString(margin, 35, f"Được tạo bởi: {user.name} - {user.employee_id} - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    canvas_obj.drawString(margin, 20, "Phiếu này chỉ dùng để test hiển thị chữ ký, không có giá trị pháp lý.")

def remove_vietnamese_accents(text):
    """Loại bỏ dấu tiếng Việt và chuyển thành chữ thường, loại bỏ khoảng trắng"""
    if not text:
        return ""
    
    # Mapping dấu tiếng Việt
    vietnamese_map = {
        'à': 'a', 'á': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
        'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
        'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
        'è': 'e', 'é': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
        'ê': 'e', 'ề': 'e', 'ế': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
        'ì': 'i', 'í': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
        'ò': 'o', 'ó': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
        'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
        'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
        'ù': 'u', 'ú': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
        'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
        'ỳ': 'y', 'ý': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
        'đ': 'd',
        'À': 'A', 'Á': 'A', 'Ả': 'A', 'Ã': 'A', 'Ạ': 'A',
        'Ă': 'A', 'Ằ': 'A', 'Ắ': 'A', 'Ẳ': 'A', 'Ẵ': 'A', 'Ặ': 'A',
        'Â': 'A', 'Ầ': 'A', 'Ấ': 'A', 'Ẩ': 'A', 'Ẫ': 'A', 'Ậ': 'A',
        'È': 'E', 'É': 'E', 'Ẻ': 'E', 'Ẽ': 'E', 'Ẹ': 'E',
        'Ê': 'E', 'Ề': 'E', 'Ế': 'E', 'Ể': 'E', 'Ễ': 'E', 'Ệ': 'E',
        'Ì': 'I', 'Í': 'I', 'Ỉ': 'I', 'Ĩ': 'I', 'Ị': 'I',
        'Ò': 'O', 'Ó': 'O', 'Ỏ': 'O', 'Õ': 'O', 'Ọ': 'O',
        'Ô': 'O', 'Ồ': 'O', 'Ố': 'O', 'Ổ': 'O', 'Ỗ': 'O', 'Ộ': 'O',
        'Ơ': 'O', 'Ờ': 'O', 'Ớ': 'O', 'Ở': 'O', 'Ỡ': 'O', 'Ợ': 'O',
        'Ù': 'U', 'Ú': 'U', 'Ủ': 'U', 'Ũ': 'U', 'Ụ': 'U',
        'Ư': 'U', 'Ừ': 'U', 'Ứ': 'U', 'Ử': 'U', 'Ữ': 'U', 'Ự': 'U',
        'Ỳ': 'Y', 'Ý': 'Y', 'Ỷ': 'Y', 'Ỹ': 'Y', 'Ỵ': 'Y',
        'Đ': 'D'
    }
    
    result = ""
    for char in text:
        result += vietnamese_map.get(char, char)
    
    # Chuyển thành chữ thường và loại bỏ khoảng trắng
    result = result.lower().replace(' ', '')
    
    # Loại bỏ các ký tự đặc biệt khác, chỉ giữ lại chữ cái và số
    import re
    result = re.sub(r'[^a-z0-9]', '', result)
    
    return result

# API endpoint để phê duyệt tất cả attendance records
@app.route('/api/attendance/approve-all', methods=['POST'])
@rate_limit(max_requests=10, window_seconds=60)  # Giới hạn 10 lần gọi API trong 1 phút
def approve_all_attendances():
    if 'user_id' not in session:
        return jsonify({'error': 'Không có quyền truy cập'}), 401
    
    if check_session_timeout():
        return jsonify({'error': 'Phiên đăng nhập đã hết hạn'}), 401
    
    update_session_activity()
    
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'error': 'Không tìm thấy người dùng'}), 404
    
    current_role = session.get('current_role', user.roles.split(',')[0])
    if current_role not in ['TEAM_LEADER', 'MANAGER', 'ADMIN']:
        return jsonify({'error': 'Bạn không có quyền phê duyệt hàng loạt'}), 403
    
    data = request.get_json()
    action = data.get('action')  # 'approve' hoặc 'reject'
    reason = validate_reason(data.get('reason', '')) if data.get('action') == 'reject' else ''
    
    if action not in ['approve', 'reject']:
        return jsonify({'error': 'Hành động không hợp lệ'}), 400
    
    if action == 'reject' and not reason:
        return jsonify({'error': 'Lý do từ chối không hợp lệ'}), 400
    
    try:
        # Xác định phạm vi attendance records cần phê duyệt
        if current_role == 'ADMIN':
            # Admin có thể phê duyệt tất cả
            attendances_query = Attendance.query.filter(
                Attendance.approved == False
            )
        elif current_role == 'MANAGER':
            # Manager có thể phê duyệt tất cả nhân viên (không phân biệt phòng ban)
            # Bao gồm cả nhân viên từ các phòng ban khác
            attendances_query = Attendance.query.filter(
                Attendance.approved == False
            )
        else:  # TEAM_LEADER
            # Team leader chỉ có thể phê duyệt nhân viên cùng phòng ban
            attendances_query = Attendance.query.join(User, Attendance.user_id == User.id).filter(
                Attendance.approved == False,
                User.department == user.department
            )
        
        # Lọc theo trạng thái hiện tại
        if current_role == 'TEAM_LEADER':
            attendances_query = attendances_query.filter(Attendance.status == 'pending')
        elif current_role == 'MANAGER':
            # Manager chỉ có thể phê duyệt pending và pending_manager
            attendances_query = attendances_query.filter(
                Attendance.status.in_(['pending', 'pending_manager'])
            )
        elif current_role == 'ADMIN':
            # Admin có thể phê duyệt tất cả trạng thái chờ duyệt
            attendances_query = attendances_query.filter(
                Attendance.status.in_(['pending', 'pending_manager', 'pending_admin'])
            )
        
        attendances = attendances_query.all()
        
        if not attendances:
            return jsonify({'message': 'Không có bản ghi nào cần phê duyệt', 'count': 0}), 200
        
        approved_count = 0
        rejected_count = 0
        
        for attendance in attendances:
            # Kiểm tra quyền phê duyệt từng record
            has_permission, error_message = check_approval_permission(user.id, attendance.id, current_role)
            if not has_permission:
                continue
            
            if action == 'approve':
                # Xử lý phê duyệt
                if current_role == 'TEAM_LEADER':
                    attendance.status = 'pending_manager'
                    attendance.approved_by = user.id
                    attendance.approved_at = datetime.now()
                    # Lưu chữ ký và ID người ký nếu có
                    if user.has_personal_signature():
                        attendance.team_leader_signature = user.personal_signature
                    attendance.team_leader_signer_id = user.id  # Cập nhật ID người ký trưởng nhóm
                elif current_role == 'MANAGER':
                    # Manager chuyển lên QUẢN TRỊ VIÊN để kiểm tra cuối cùng
                    # Nếu trạng thái là pending, cần lưu chữ ký trưởng nhóm (nếu có)
                    if attendance.status == 'pending' and user.has_personal_signature():
                        attendance.team_leader_signature = user.personal_signature
                        # Cập nhật ID người ký trưởng nhóm nếu chưa có
                        if not attendance.team_leader_signer_id:
                            attendance.team_leader_signer_id = user.id
                    
                    attendance.status = 'pending_admin'
                    attendance.approved_by = user.id
                    attendance.approved_at = datetime.now()
                    # Lưu chữ ký quản lý nếu có
                    if user.has_personal_signature():
                        attendance.manager_signature = user.personal_signature
                    attendance.manager_signer_id = user.id  # Cập nhật ID người ký quản lý
                elif current_role == 'ADMIN':
                    # Check Google API token trước khi ADMIN approve
                    token_status = check_google_token_status()
                    if not token_status.get('can_approve', False):
                        # Publish notification to all admins
                        publish_token_status('expired', token_status.get('message', 'Token hết hạn'), needs_reauth=True)
                        return jsonify({
                            'error': f"⚠️ Token Google API hết hạn. {token_status.get('message', 'Vui lòng refresh token trước khi phê duyệt.')}",
                            'error_code': 'token_expired',
                            'needs_reauth': True
                        }), 503
                    
                    # Admin có thể phê duyệt trực tiếp lên trạng thái cuối cùng
                    # Lưu chữ ký trưởng nhóm nếu trạng thái là pending và có chữ ký
                    if attendance.status == 'pending' and user.has_personal_signature():
                        attendance.team_leader_signature = user.personal_signature
                        attendance.team_leader_signer_id = user.id  # Cập nhật ID người ký trưởng nhóm
                    
                    # Lưu chữ ký quản lý nếu trạng thái là pending_manager và có chữ ký
                    if attendance.status == 'pending_manager' and user.has_personal_signature():
                        attendance.manager_signature = user.personal_signature
                        attendance.manager_signer_id = user.id  # Cập nhật ID người ký quản lý
                    
                    # Lưu chữ ký quản lý nếu trạng thái là pending_admin và có chữ ký
                    if attendance.status == 'pending_admin' and user.has_personal_signature():
                        attendance.manager_signature = user.personal_signature
                        attendance.manager_signer_id = user.id  # Cập nhật ID người ký quản lý
                    
                    attendance.status = 'approved'
                    attendance.approved = True
                    attendance.approved_by = user.id
                    attendance.approved_at = datetime.now()
                
                approved_count += 1
                
                # Log audit action
                log_audit_action(
                    user_id=user.id,
                    action='BULK_APPROVE_ATTENDANCE',
                    table_name='attendances',
                    record_id=attendance.id,
                    old_values={'status': attendance.status},
                    new_values={
                        'status': attendance.status, 
                        'approved_by': user.id,
                        'team_leader_signer_id': getattr(attendance, 'team_leader_signer_id', None),
                        'manager_signer_id': getattr(attendance, 'manager_signer_id', None)
                    }
                )
                
            else:  # reject
                attendance.status = 'rejected'
                attendance.reject_reason = reason
                attendance.approved_by = user.id
                attendance.approved_at = datetime.now()
                rejected_count += 1
                
                # Log audit action
                log_audit_action(
                    user_id=user.id,
                    action='BULK_REJECT_ATTENDANCE',
                    table_name='attendances',
                    record_id=attendance.id,
                    old_values={'status': attendance.status},
                    new_values={'status': 'rejected', 'reject_reason': reason, 'approved_by': user.id}
                )
        
        db.session.commit()
        
        total_processed = approved_count + rejected_count
        message = f'Đã xử lý {total_processed} bản ghi: {approved_count} phê duyệt, {rejected_count} từ chối'
        
        return jsonify({
            'success': True,
            'message': message,
            'total_processed': total_processed,
            'approved_count': approved_count,
            'rejected_count': rejected_count
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in bulk approval: {e}")
        return jsonify({'error': f'Lỗi khi xử lý hàng loạt: {str(e)}'}), 500

# ============================================================================
# LEAVE REQUEST ROUTES
# ============================================================================

@app.route('/test-auth')
def test_auth():
    """Test route để kiểm tra authentication"""
    if 'user_id' not in session:
        return jsonify({'authenticated': False, 'error': 'Not logged in'}), 401
    
    user = db.session.get(User, session['user_id'])
    if not user:
        return jsonify({'authenticated': False, 'error': 'Invalid user'}), 401
    
    return jsonify({
        'authenticated': True,
        'user_id': user.id,
        'user_name': user.name,
        'roles': user.roles
    })

@app.route('/leave-request', methods=['GET'])
def leave_request_form():
    """Hiển thị form xin nghỉ phép"""
    try:
        # Kiểm tra user đã đăng nhập
        if 'user_id' not in session:
            flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
            return redirect(url_for('login'))
        
        user = db.session.get(User, session['user_id'])
        if not user:
            session.clear()
            flash('Phiên đăng nhập không hợp lệ!', 'error')
            return redirect(url_for('login'))
        
        # Lấy vai trò hiện tại từ session
        current_role = session.get('current_role', user.roles.split(',')[0])
        # Current role setup for leave request form
        work_shift = '08:00 - 17:00'
        return render_template('leave_request_form.html', user=user, current_role=current_role, work_shift=work_shift)
    except Exception as e:
        print(f"Error in leave_request_form: {e}")
        flash('Có lỗi xảy ra, vui lòng thử lại', 'error')
        return redirect(url_for('dashboard'))


@app.route('/activate', methods=['GET', 'POST'])
def activate():
    """Trang nhập key kích hoạt ứng dụng."""
    activation = get_activation_record()

    # Nếu đã kích hoạt rồi thì chuyển về trang đăng nhập / dashboard
    if activation.is_activated:
        if 'user_id' in session:
            return redirect(url_for('dashboard'))
        return redirect(url_for('login'))

    if request.method == 'POST':
        input_key = (request.form.get('license_key') or '').strip()
        if not input_key:
            flash('Vui lòng nhập key kích hoạt!', 'error')
            return render_template('activate.html')

        # So sánh với key chuẩn trên server
        if input_key == APP_LICENSE_KEY:
            activation.is_activated = True
            activation.license_key = input_key
            activation.activated_at = datetime.utcnow()
            try:
                db.session.commit()
                flash('Kích hoạt thành công! Bạn có thể đăng nhập và sử dụng hệ thống.', 'success')
                return redirect(url_for('login'))
            except Exception as e:
                db.session.rollback()
                print(f"[ERROR] Lỗi lưu kích hoạt: {e}")
                flash('Có lỗi khi lưu thông tin kích hoạt. Vui lòng thử lại.', 'error')
        else:
            flash('Key kích hoạt không hợp lệ. Vui lòng kiểm tra lại.', 'error')

    return render_template('activate.html')
@app.route('/leave-request', methods=['POST'])
def submit_leave_request():
    """Xử lý đơn xin nghỉ phép"""
    try:
        print("[Leave][Create] submit_leave_request called")
        # Kiểm tra user đã đăng nhập
        if 'user_id' not in session:
            flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
            return redirect(url_for('login'))
        
        user = db.session.get(User, session['user_id'])
        if not user:
            session.clear()
            flash('Phiên đăng nhập không hợp lệ!', 'error')
            return redirect(url_for('login'))
        
        # Lấy dữ liệu từ form
        data = request.form
        
        # Xử lý file upload
        attachments_info = []
        if 'attachments' in request.files:
            files = request.files.getlist('attachments')
            for file in files:
                if file and file.filename:
                    # Tạo tên file unique
                    filename = f"{uuid.uuid4()}_{file.filename}"
                    
                    # Tạo thư mục uploads nếu chưa có
                    upload_dir = os.path.join(app.root_path, 'uploads', 'leave_requests')
                    os.makedirs(upload_dir, exist_ok=True)
                    
                    # Lưu file
                    file_path = os.path.join(upload_dir, filename)
                    file.save(file_path)
                    
                    attachments_info.append({
                        'original_name': file.filename,
                        'saved_name': filename,
                        'size': file.content_length or 0
                    })
        
        # Parse trước một số trường ngày để ràng buộc hợp lệ
        from_date_str = data.get('leave_from_date', '2024-01-01')
        to_date_str = data.get('leave_to_date', '2024-01-01')
        try:
            from_date_dt = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date_dt = datetime.strptime(to_date_str, '%Y-%m-%d')
        except Exception:
            flash('Định dạng ngày không hợp lệ', 'error')
            return redirect(url_for('leave_request_form'))

        requested_annual = float(data.get('annual_leave_days', 0) or 0)
        requested_unpaid = float(data.get('unpaid_leave_days', 0) or 0)
        requested_special = float(data.get('special_leave_days', 0) or 0)
        total_requested_days = requested_annual + requested_unpaid + requested_special
        # Sử dụng kết quả tính toán từ frontend (đã tính đúng theo ca làm việc)
        # Frontend đã tính toán chính xác theo ca làm việc và giờ nghỉ trưa
        from_time_str = data.get('leave_from_time') or '00:00'
        to_time_str = data.get('leave_to_time') or '00:00'

        # Validate: Thời gian bắt đầu phải nhỏ hơn thời gian kết thúc
        try:
            clean_from_time = clean_time_format(from_time_str)
            clean_to_time = clean_time_format(to_time_str)
            
            start_dt = datetime.combine(from_date_dt.date(), datetime.strptime(clean_from_time, '%H:%M').time())
            end_dt = datetime.combine(to_date_dt.date(), datetime.strptime(clean_to_time, '%H:%M').time())
        except Exception:
            flash('Định dạng giờ không hợp lệ', 'error')
            return redirect(url_for('leave_request_form'))
        if start_dt >= end_dt:
            flash('Thời gian từ ngày giờ phải nhỏ hơn đến ngày giờ', 'error')
            return redirect(url_for('leave_request_form'))
        
        # Validate: Giờ kết thúc không được vượt quá giờ ra khỏi ca
        shift_code = data.get('leave_shift_code', '1')
        shift_ranges = {
            '1': {'start': '07:30', 'end': '16:30'},
            '2': {'start': '09:00', 'end': '18:00'},
            '3': {'start': '11:00', 'end': '20:00'},
            '4': {'start': '08:00', 'end': '17:00'}
        }
        
        if shift_code in shift_ranges:
            shift_end_time = shift_ranges[shift_code]['end']
            shift_end_dt = datetime.combine(to_date_dt.date(), datetime.strptime(shift_end_time, '%H:%M').time())
            
            # Chỉ kiểm tra nếu cùng ngày (không kiểm tra khi nghỉ qua nhiều ngày)
            if from_date_dt.date() == to_date_dt.date() and end_dt > shift_end_dt:
                flash(f'Giờ kết thúc nghỉ ({clean_to_time}) không được vượt quá giờ ra khỏi ca ({shift_end_time})', 'error')
                return redirect(url_for('leave_request_form'))
        
        # Lấy kết quả tính toán từ frontend nếu có
        frontend_calculated_days = data.get('calculated_leave_days')
        if frontend_calculated_days is not None:
            available_units = float(frontend_calculated_days)
        else:
            # Fallback: tính theo logic cũ (không chính xác cho tất cả ca)
            available_units = _compute_leave_units_generic(from_date_dt, from_time_str, to_date_dt, to_time_str)
        
        if total_requested_days > available_units + 1e-9:
            flash('Tổng số ngày xin nghỉ vượt quá số ngày có thể xin trong khoảng thời gian đã chọn (theo ca làm việc).', 'error')
            return redirect(url_for('leave_request_form'))

        # Tạo đơn xin nghỉ phép mới
        leave_request = LeaveRequest(
            user_id=user.id,
            employee_name=data.get('employee_name'),
            team=data.get('team'),
            employee_code=data.get('employee_code'),
            request_type=data.get('request_type', 'leave'),
            late_early_type=data.get('late_early_type'),
            leave_reason=data.get('leave_reason'),
            attachments=json.dumps(attachments_info) if attachments_info else None,
            reason_sick_child=bool(data.get('reason_sick_child')),
            reason_sick=bool(data.get('reason_sick')),
            reason_death_anniversary=bool(data.get('reason_death_anniversary')),
            reason_other=bool(data.get('reason_other')),
            reason_other_detail=data.get('reason_other_detail'),
            hospital_confirmation=bool(data.get('hospital_confirmation')),
            wedding_invitation=bool(data.get('wedding_invitation')),
            death_birth_certificate=bool(data.get('death_birth_certificate')),
            leave_from_hour=int(clean_from_time.split(':')[0]),
            leave_from_minute=int(clean_from_time.split(':')[1]),
            leave_from_day=int(data.get('leave_from_date', '2024-01-01').split('-')[2]),
            leave_from_month=int(data.get('leave_from_date', '2024-01-01').split('-')[1]),
            leave_from_year=int(data.get('leave_from_date', '2024-01-01').split('-')[0]),
            leave_to_hour=int(clean_to_time.split(':')[0]),
            leave_to_minute=int(clean_to_time.split(':')[1]),
            leave_to_day=int(data.get('leave_to_date', '2024-01-01').split('-')[2]),
            leave_to_month=int(data.get('leave_to_date', '2024-01-01').split('-')[1]),
            leave_to_year=int(data.get('leave_to_date', '2024-01-01').split('-')[0]),
            annual_leave_days=float(data.get('annual_leave_days', 0) or 0),
            unpaid_leave_days=float(data.get('unpaid_leave_days', 0) or 0),
            special_leave_days=float(data.get('special_leave_days', 0) or 0),
            special_leave_type=data.get('special_leave_type'),
            substitute_name=data.get('substitute_name'),
            substitute_employee_id=data.get('substitute_employee_id'),
            notes=data.get('notes'),
            # Lưu ca áp dụng khi xin nghỉ (tùy chọn)
            # Tương thích: nếu không có, giữ None
            shift_code=data.get('leave_shift_code') if data.get('leave_shift_code') in ['1','2','3','4'] else None,
            status='pending'
        )
        
        # Lưu use_lunch_break vào notes dưới dạng JSON
        use_lunch_break_value = data.get('use_lunch_break')
        if use_lunch_break_value in ['true', 'false']:
            import json
            notes_data = {}
            original_notes_text = None
            
            if leave_request.notes:
                try:
                    # Thử parse JSON
                    notes_data = json.loads(leave_request.notes)
                    if not isinstance(notes_data, dict):
                        # Nếu không phải dict, giữ lại text gốc
                        original_notes_text = leave_request.notes
                        notes_data = {}
                except:
                    # Nếu không phải JSON, giữ lại text gốc
                    original_notes_text = leave_request.notes
                    notes_data = {}
            
            # Lưu use_lunch_break
            notes_data['use_lunch_break'] = use_lunch_break_value == 'true'
            
            # Nếu có notes text gốc, thêm vào notes_data
            if original_notes_text:
                notes_data['_original_notes'] = original_notes_text
            
            leave_request.notes = json.dumps(notes_data, ensure_ascii=False)
            
            try:
                _safe_print(f"[Leave][Create] Đã lưu use_lunch_break={use_lunch_break_value == 'true'} vào notes cho đơn #{leave_request.id}")
            except Exception:
                pass
        
        # Ràng buộc: các số ngày phải là bội số 0.5
        def ensure_half_step(x):
            return (int(round(x * 2)) / 2.0)
        leave_request.annual_leave_days = ensure_half_step(leave_request.annual_leave_days or 0.0)
        leave_request.unpaid_leave_days = ensure_half_step(leave_request.unpaid_leave_days or 0.0)
        leave_request.special_leave_days = ensure_half_step(leave_request.special_leave_days or 0.0)

        # Kiểm tra lần nữa sau chuẩn hóa: tổng không vượt quá đơn vị nghỉ tính được
        if (leave_request.annual_leave_days + leave_request.unpaid_leave_days + leave_request.special_leave_days) > available_units + 1e-9:
            flash('Tổng số ngày xin nghỉ vượt quá số ngày có thể xin trong khoảng thời gian đã chọn (theo ca làm việc).', 'error')
            return redirect(url_for('leave_request_form'))

        # Lưu vào cơ sở dữ liệu
        db.session.add(leave_request)
        db.session.commit()
        
        # Kiểm tra xem người dùng có muốn gửi email hay không
        email_consent = data.get('email_consent', 'no').lower()
        send_email = email_consent == 'yes'
        print(f"[DEBUG] Email consent received: '{email_consent}', send_email: {send_email}")
        
        if send_email:
            # Gửi email thông báo đến HR (bất đồng bộ)
            try:
                print(f"[Mail] Attempting to send create email for leave_request #{leave_request.id} by user #{user.id} ({user.name})")
                send_leave_request_email_async(leave_request, user, action='create')
                # Process any pending DB updates from async threads
                from utils.email_utils import process_db_updates
                process_db_updates()
                # Persist 'sending' immediately
                upsert_email_status(leave_request.id, 'sending', 'Đang gửi email thông báo...')
                # Lưu trạng thái email vào session cho tất cả vai trò
                session['email_status'] = {
                    'request_id': leave_request.id,
                    'status': 'sending',
                    'message': 'Đang gửi email thông báo...'
                }
                # Chỉ thông báo về đơn; tiến trình email sẽ do toast hiển thị
                # flash('Đơn xin nghỉ phép đã được gửi thành công! Email thông báo đã được gửi đến phòng nhân sự.', 'success')
            except Exception as e:
                print(f"[Mail] Error scheduling leave create email: {e}")
                # Lưu trạng thái email vào session cho tất cả vai trò
                session['email_status'] = {
                    'request_id': leave_request.id,
                    'status': 'error',
                    'message': f'Lỗi khi gửi email: {str(e)}'
                }
                flash('Đơn đã gửi thành công, nhưng có lỗi khi gửi email thông báo.', 'warning')
        else:
            # Không gửi email
            print(f"[Mail] User chose not to send email for leave_request #{leave_request.id}")
            session['email_status'] = {
                'request_id': leave_request.id,
                'status': 'skipped',
                'message': 'Người dùng đã chọn không gửi email thông báo'
            }
            # flash('Đơn xin nghỉ phép đã được gửi thành công! (Không gửi email thông báo)', 'success')
        print(f"[DEBUG] Redirecting to leave_requests_list with request_id={leave_request.id}")
        print(f"[DEBUG] Session email_status before redirect: {session.get('email_status')}")
        return redirect(url_for('leave_requests_list', request_id=leave_request.id))
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in submit_leave_request: {e}")
        flash(f'Lỗi khi gửi đơn xin nghỉ phép: {str(e)}', 'error')
        return redirect(url_for('leave_request_form'))

@app.route('/leave-request/<int:request_id>/attachment/<filename>')
def download_leave_attachment(request_id, filename):
    """Download attachment file for leave request"""
    try:
        # Kiểm tra user đã đăng nhập
        if 'user_id' not in session:
            flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
            return redirect(url_for('login'))
        
        user = db.session.get(User, session['user_id'])
        if not user:
            session.clear()
            flash('Phiên đăng nhập không hợp lệ!', 'error')
            return redirect(url_for('login'))
        
        # Lấy leave request
        leave_request = LeaveRequest.query.get_or_404(request_id)
        
        # Kiểm tra quyền truy cập (chỉ user tạo đơn hoặc admin/manager mới xem được)
        current_role = session.get('current_role', user.roles.split(',')[0])
        if not (user.id == leave_request.user_id or 
                current_role in ['ADMIN', 'MANAGER', 'TEAM_LEADER']):
            flash('Bạn không có quyền truy cập file này', 'error')
            return redirect(url_for('leave_requests_list'))
        
        # Kiểm tra file có trong attachments không
        if not leave_request.attachments:
            flash('Không tìm thấy file đính kèm', 'error')
            return redirect(url_for('view_leave_request', request_id=request_id))
        
        attachments = json.loads(leave_request.attachments)
        file_info = None
        for att in attachments:
            if att['saved_name'] == filename:
                file_info = att
                break
        
        if not file_info:
            flash('File không tồn tại', 'error')
            return redirect(url_for('view_leave_request', request_id=request_id))
        
        # Đường dẫn file
        file_path = os.path.join(app.root_path, 'uploads', 'leave_requests', filename)
        
        if not os.path.exists(file_path):
            flash('File không tồn tại trên server', 'error')
            return redirect(url_for('view_leave_request', request_id=request_id))
        
        return send_file(file_path, as_attachment=True, download_name=file_info['original_name'])
        
    except Exception as e:
        print(f"Error in download_leave_attachment: {e}")
        flash('Có lỗi xảy ra khi tải file', 'error')
        return redirect(url_for('leave_requests_list'))

@app.route('/leave-request/<int:request_id>/download-all')
def download_all_leave_attachments(request_id):
    """Download all attachments as ZIP file for leave request"""
    try:
        # Kiểm tra user đã đăng nhập
        if 'user_id' not in session:
            flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
            return redirect(url_for('login'))
        
        user = db.session.get(User, session['user_id'])
        if not user:
            session.clear()
            flash('Phiên đăng nhập không hợp lệ!', 'error')
            return redirect(url_for('login'))
        
        # Lấy leave request
        leave_request = LeaveRequest.query.get_or_404(request_id)
        
        # Kiểm tra quyền truy cập
        current_role = session.get('current_role', user.roles.split(',')[0])
        if not (user.id == leave_request.user_id or 
                current_role in ['ADMIN', 'MANAGER', 'TEAM_LEADER']):
            flash('Bạn không có quyền truy cập file này', 'error')
            return redirect(url_for('leave_requests_list'))
        
        # Kiểm tra có attachments không
        if not leave_request.attachments:
            flash('Không có chứng từ để tải xuống', 'error')
            return redirect(url_for('view_leave_request', request_id=request_id))
        
        attachments = json.loads(leave_request.attachments)
        if not attachments:
            flash('Không có chứng từ để tải xuống', 'error')
            return redirect(url_for('view_leave_request', request_id=request_id))
        
        # Tạo ZIP file
        import zipfile
        import io
        
        zip_buffer = io.BytesIO()
        upload_dir = os.path.join(app.root_path, 'uploads', 'leave_requests')
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for attachment in attachments:
                file_path = os.path.join(upload_dir, attachment['saved_name'])
                if os.path.exists(file_path):
                    # Sử dụng tên file gốc trong ZIP
                    zip_file.write(file_path, attachment['original_name'])
                else:
                    print(f"Warning: File not found: {file_path}")
        
        zip_buffer.seek(0)
        
        # Tên file ZIP
        zip_filename = f"Chứng_từ_nghỉ_phép_{request_id}_{leave_request.employee_name.replace(' ', '_')}.zip"
        
        return send_file(zip_buffer, as_attachment=True, download_name=zip_filename, mimetype='application/zip')
        
    except Exception as e:
        print(f"Error in download_all_leave_attachments: {e}")
        flash('Có lỗi xảy ra khi tạo file ZIP', 'error')
        return redirect(url_for('view_leave_request', request_id=request_id))

@app.route('/leave-requests')
def leave_requests_list():
    """Hiển thị danh sách đơn xin nghỉ phép"""
    try:
        # Kiểm tra user đã đăng nhập
        if 'user_id' not in session:
            flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
            return redirect(url_for('login'))
        
        user = db.session.get(User, session['user_id'])
        if not user:
            session.clear()
            flash('Phiên đăng nhập không hợp lệ!', 'error')
            return redirect(url_for('login'))
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        request_id = request.args.get('request_id', type=int)
        
        # Xây dựng query cơ bản
        query = LeaveRequest.query
        
        # Lọc theo trạng thái
        status = request.args.get('status')
        if status:
            query = query.filter(LeaveRequest.status == status)
        
        # Lọc theo nhân viên
        employee = request.args.get('employee')
        if employee:
            query = query.filter(
                db.or_(
                    LeaveRequest.employee_name.contains(employee),
                    LeaveRequest.employee_code.contains(employee)
                )
            )
        
        # Lọc theo phòng ban (chỉ cho ADMIN và MANAGER)
        department = request.args.get('department')
        if department and current_role in ['ADMIN', 'MANAGER']:
            query = query.join(User, User.id == LeaveRequest.user_id).filter(User.department == department)
        
        # Lọc theo loại đơn
        request_type = request.args.get('request_type')
        if request_type:
            query = query.filter(LeaveRequest.request_type == request_type)
        
        # Lọc theo ngày xin nghỉ
        date_from = request.args.get('date_from')
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d')
                # Lọc các đơn có ngày kết thúc nghỉ >= ngày bắt đầu lọc
                query = query.filter(
                    db.or_(
                        # Ngày kết thúc nghỉ >= ngày lọc
                        db.and_(
                            LeaveRequest.leave_to_year > from_date.year
                        ),
                        db.and_(
                            LeaveRequest.leave_to_year == from_date.year,
                            LeaveRequest.leave_to_month > from_date.month
                        ),
                        db.and_(
                            LeaveRequest.leave_to_year == from_date.year,
                            LeaveRequest.leave_to_month == from_date.month,
                            LeaveRequest.leave_to_day >= from_date.day
                        )
                    )
                )
            except ValueError:
                pass
        
        date_to = request.args.get('date_to')
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d')
                # Lọc các đơn có ngày bắt đầu nghỉ <= ngày kết thúc lọc
                query = query.filter(
                    db.or_(
                        # Ngày bắt đầu nghỉ <= ngày lọc
                        db.and_(
                            LeaveRequest.leave_from_year < to_date.year
                        ),
                        db.and_(
                            LeaveRequest.leave_from_year == to_date.year,
                            LeaveRequest.leave_from_month < to_date.month
                        ),
                        db.and_(
                            LeaveRequest.leave_from_year == to_date.year,
                            LeaveRequest.leave_from_month == to_date.month,
                            LeaveRequest.leave_from_day <= to_date.day
                        )
                    )
                )
            except ValueError:
                pass
        
        # Lọc theo vai trò và trạng thái
        user_roles = user.get_roles_list()
        current_role = session.get('current_role', user_roles[0] if user_roles else 'EMPLOYEE')
        
        if current_role == 'TEAM_LEADER':
            # TEAM_LEADER chỉ thấy đơn pending (chưa được phê duyệt) của cùng phòng ban
            query = query.filter(
                LeaveRequest.status == 'pending',
                LeaveRequest.user.has(User.department == user.department)
            )
        elif current_role == 'MANAGER':
            # MANAGER chỉ thấy đơn pending_manager (đã được TEAM_LEADER phê duyệt)
            query = query.filter(LeaveRequest.status == 'pending_manager')
        elif current_role == 'ADMIN':
            # ADMIN chỉ thấy đơn pending_admin (đã được MANAGER phê duyệt)
            query = query.filter(LeaveRequest.status == 'pending_admin')
        else:
            # EMPLOYEE chỉ thấy đơn của mình
            query = query.filter(LeaveRequest.user_id == user.id)
        
        # Sắp xếp
        sort_by = request.args.get('sort_by', 'created_at')
        sort_dir = request.args.get('sort_dir', 'desc')
        sort_field = LeaveRequest.created_at
        if sort_by == 'status':
            sort_field = LeaveRequest.status
        elif sort_by == 'created_at':
            sort_field = LeaveRequest.created_at
        if sort_dir == 'asc':
            query = query.order_by(sort_field.asc())
        else:
            query = query.order_by(sort_field.desc())
        
        # Phân trang
            pagination = query.paginate(
                page=page, per_page=per_page, error_out=False
            )
        
        # Lấy vai trò hiện tại từ session
        current_role = session.get('current_role', user.roles.split(',')[0])
        
        # Parse attachments cho mỗi leave request
        for leave_request in pagination.items:
            leave_request.attachments_list = []
            if leave_request.attachments:
                try:
                    leave_request.attachments_list = json.loads(leave_request.attachments)
                except (json.JSONDecodeError, TypeError, AttributeError):
                    leave_request.attachments_list = []
        
        # Lấy danh sách phòng ban cho bộ lọc
        try:
            # Thử lấy từ bảng Department trước
            dept_objects = Department.query.filter(Department.is_active == True).order_by(Department.name.asc()).all()
            if dept_objects:
                departments = [d.name for d in dept_objects]
            else:
                # Fallback: distinct từ User nếu Department trống
                departments = sorted({u.department for u in User.query.filter(User.department.isnot(None), User.department != '').all()})
        except Exception:
            # Fallback: distinct từ User
            departments = sorted({u.department for u in User.query.filter(User.department.isnot(None), User.department != '').all()})
        
        return render_template('leave_requests_list.html', 
                             leave_requests=pagination.items,
                             pagination=pagination,
                             user=user,
                             current_role=current_role,
                             request_id=request_id,
                             departments=departments)
    except Exception:
        flash('Có lỗi xảy ra khi tải danh sách đơn nghỉ phép', 'error')
        return redirect(url_for('dashboard'))

@app.route('/leave-request/<int:request_id>')
def view_leave_request(request_id):
    """Xem chi tiết đơn xin nghỉ phép"""
    if 'user_id' not in session:
        flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
        return redirect(url_for('login'))
    
    user = db.session.get(User, session['user_id'])
    if not user:
        session.clear()
        flash('Phiên đăng nhập không hợp lệ!', 'error')
        return redirect(url_for('login'))
    
    leave_request = LeaveRequest.query.get_or_404(request_id)
    
    # Kiểm tra quyền truy cập
    user_roles = user.get_roles_list()
    if not (any(role in ['ADMIN', 'MANAGER', 'TEAM_LEADER'] for role in user_roles) or 
            user.id == leave_request.user_id):
        abort(403)
    
    # Lấy vai trò hiện tại từ session
    current_role = session.get('current_role', user.roles.split(',')[0])
    
    # Parse existing attachments for template
    attachments_list = []
    if leave_request.attachments:
        try:
            attachments_list = json.loads(leave_request.attachments)
        except (json.JSONDecodeError, TypeError, AttributeError):
            attachments_list = []
    
    return render_template('view_leave_request.html', leave_request=leave_request, user=user, current_role=current_role, request_id=request_id, attachments_list=attachments_list)

@app.route('/api/email-status/<int:request_id>')
def get_email_status(request_id):
    """API để kiểm tra trạng thái gửi email"""
    print(f"[API] Email status request for request_id: {request_id}")
    
    if 'user_id' not in session:
        print("[API] Unauthorized - no user_id in session")
        return jsonify({'error': 'Unauthorized'}), 401
    
    user = db.session.get(User, session['user_id'])
    if not user:
        print("[API] User not found")
        return jsonify({'error': 'User not found'}), 401
    
    print(f"[API] User: {user.name}, Role: {user.roles}")
    
    # Trả về email status cho tất cả vai trò
    # (Đã loại bỏ giới hạn chỉ cho nhân viên)
    
    # Ưu tiên trạng thái trong DB trước
    db_rec = get_email_status_record(request_id)
    if db_rec:
        resp = {'request_id': request_id, 'status': db_rec.status, 'message': db_rec.message}
        print(f"[API] DB status: {resp}")
        # Nếu là kết quả cuối cùng, dọn trạng thái để không lặp lại
        if db_rec.status in ['success', 'error']:
            try:
                db.session.delete(db_rec)
                db.session.commit()
            except Exception as _:
                db.session.rollback()
            session.pop('email_status', None)
        return jsonify(resp)
    
    # Nếu không có global status, kiểm tra session
    if 'email_status' in session and session['email_status'].get('request_id') == request_id:
        session_status = session['email_status']
        print(f"[API] Found session status: {session_status}")
        # Nếu là kết quả cuối cùng, dọn session
        if session_status.get('status') in ['success', 'error']:
            session.pop('email_status', None)
        return jsonify(session_status)
    
    # Fallback về unknown
    status = {'status': 'unknown', 'message': 'Không có thông tin'}
    print(f"[API] No status found, returning unknown: {status}")
    return jsonify(status)

@app.route('/api/email-status/latest')
def get_latest_email_status():
    """API: lấy trạng thái email gần nhất từ session, không cần request_id trên URL.
    Nếu session đang là 'sending' và có request_id, sẽ đối chiếu với global email_status
    để trả về kết quả cuối cùng khi có (success/error).
    """
    # Không dùng print ở đây để tránh lỗi I/O trên stdout khi server chạy nền
    if 'user_id' not in session:
        print("[API] Latest: Unauthorized - no user_id in session")
        return jsonify({'error': 'Unauthorized'}), 401

    # Lấy từ session nếu có
    sess = session.get('email_status')
    if not sess:
        return jsonify({'status': 'unknown', 'message': 'Không có thông tin'})

    request_id = sess.get('request_id')
    if request_id:
        # Kiểm tra DB trước
        db_rec = get_email_status_record(request_id)
        print(f"[API] Latest: DB status = {db_rec.status if db_rec else None}")
        if db_rec and db_rec.status in ['success', 'error']:
            response_payload = {
                'request_id': request_id,
                'status': db_rec.status,
                'message': db_rec.message
            }
            # Dọn DB và session để không lặp lại
            try:
                db.session.delete(db_rec)
                db.session.commit()
            except Exception:
                db.session.rollback()
            session.pop('email_status', None)
            return jsonify(response_payload)

    # Ngược lại trả về session hiện tại
    # Nếu sess đã là kết quả cuối cùng thì dọn luôn và trả một lần
    if sess and sess.get('status') in ['success', 'error']:
        payload = sess
        session.pop('email_status', None)
        return jsonify(payload)
    return jsonify(sess)

# ===================== SSE: Email Status Push =====================
# In-memory subscribers per user_id
_email_sse_subscribers = defaultdict(list)

def _sse_subscribe(user_id: int) -> Queue:
    q = Queue()
    _email_sse_subscribers[user_id].append(q)
    return q

def _sse_unsubscribe(user_id: int, q: Queue) -> None:
    try:
        if q in _email_sse_subscribers.get(user_id, []):
            _email_sse_subscribers[user_id].remove(q)
    except Exception:
        pass

def publish_email_status(user_id: int, request_id: int, status: str, message: str) -> None:
    """Publish an email status event to all live SSE subscribers of the user."""
    payload = {
        'request_id': request_id,
        'status': status,
        'message': message,
    }
    for q in list(_email_sse_subscribers.get(user_id, [])):
        try:
            q.put_nowait(payload)
        except Exception:
            # if queue full/broken, ignore
            pass

@app.route('/sse/email-status')
def sse_email_status():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    user_id = session['user_id']
    q = _sse_subscribe(user_id)

    def stream():
        # advise client reconnection delay
        yield 'retry: 3000\n\n'
        try:
            # send a heartbeat every 15s if idle
            last = time_module.time()
            while True:
                try:
                    item = q.get(timeout=1.0)
                    import json as _json
                    data = _json.dumps(item, ensure_ascii=False)
                    yield f"event: email_status\ndata: {data}\n\n"
                    last = time_module.time()
                except Exception:
                    now = time_module.time()
                    if now - last > 15:
                        # comment heartbeat to keep connection alive
                        yield ": keep-alive\n\n"
                        last = now
        finally:
            _sse_unsubscribe(user_id, q)

    from flask import Response
    return Response(stream(), mimetype='text/event-stream', headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

@app.route('/api/test-email-status')
def test_email_status():
    """Test endpoint để kiểm tra email status"""
    return jsonify({
        'global_status': dict(email_status),
        'session_status': session.get('email_status'),
        'message': 'Test endpoint for debugging'
    })

@app.route('/api/set-test-email-status/<int:request_id>')
def set_test_email_status(request_id):
    """Test endpoint để set email status manually"""
    # Set global status
    email_status[request_id] = {
        'status': 'success',
        'message': 'Test email sent successfully!',
        'timestamp': time_module.time()
    }
    
    # Set session status
    session['email_status'] = {
        'request_id': request_id,
        'status': 'success',
        'message': 'Test email sent successfully!'
    }
    
    return jsonify({
        'message': f'Set test email status for request {request_id}',
        'status': 'success'
    })

# ===================== SSE: Token Status Push =====================
# In-memory subscribers for token status (admin only)
_token_sse_subscribers = defaultdict(list)
# Global token status flag
_token_status = {
    'valid': True,
    'needs_reauth': False,
    'message': '',
    'last_check': None
}
_token_status_lock = threading.Lock()

# Cache trạng thái cảnh báo LICENSE để UI có thể hiển thị ngay khi load trang
_license_warning_state = {
    'active': False,
    'payload': None,  # payload giống với object gửi qua SSE (status, message, needs_reauth, timestamp)
    'updated_at': None,
}
_license_warning_lock = threading.Lock()

def _sse_token_subscribe(user_id: int) -> Queue:
    q = Queue()
    _token_sse_subscribers[user_id].append(q)
    return q

def _sse_token_unsubscribe(user_id: int, q: Queue) -> None:
    try:
        if q in _token_sse_subscribers.get(user_id, []):
            _token_sse_subscribers[user_id].remove(q)
    except Exception:
        pass

def publish_token_status(status: str, message: str, needs_reauth: bool = False) -> None:
    """Publish token status event to all admin SSE subscribers."""
    global _token_status
    with _token_status_lock:
        _token_status = {
            'valid': status == 'valid',
            'needs_reauth': needs_reauth,
            'message': message,
            'last_check': datetime.now().isoformat()
        }

    # Nếu đây là cảnh báo LICENSE (sử dụng chung cơ chế token_status), lưu lại vào cache riêng
    if 'ỨNG DỤNG CHẤM CÔNG' in (message or '') or 'LICENSE' in (message or ''):
        global _license_warning_state
        with _license_warning_lock:
            _license_warning_state = {
                'active': True,
                'payload': {
                    'status': status,
                    'message': message,
                    'needs_reauth': needs_reauth,
                    'timestamp': time_module.time()
                },
                'updated_at': datetime.now().isoformat(),
            }
    
    payload = {
        'status': status,
        'message': message,
        'needs_reauth': needs_reauth,
        'timestamp': time_module.time()
    }
    
    # Broadcast to all admin subscribers
    for user_id, queues in list(_token_sse_subscribers.items()):
        for q in list(queues):
            try:
                q.put_nowait(payload)
            except Exception:
                pass
    
    print(f"🔔 [Token Status] Published: {status} - {message}")

# Cache token status để tránh kiểm tra quá nhiều lần
_token_status_cache = None
_token_status_cache_time = 0
_token_status_cache_ttl = 5  # Cache trong 5 giây

def check_google_token_status(use_cache=True) -> dict:
    """Check current Google API token status without auto-authenticating.
    
    Args:
        use_cache: Nếu True, sử dụng cache nếu còn hiệu lực (mặc định True)
    """
    global _token_status_cache, _token_status_cache_time
    
    # Kiểm tra cache
    if use_cache and _token_status_cache and (time_module.time() - _token_status_cache_time) < _token_status_cache_ttl:
        return _token_status_cache
    
    try:
        # Chỉ load token từ file, không tạo instance để tránh auto-authenticate
        creds = None
        if os.path.exists('token.pickle'):
            try:
                with open('token.pickle', 'rb') as token:
                    creds = pickle.load(token)
            except Exception as e:
                result = {
                    'valid': False,
                    'needs_reauth': True,
                    'message': f'Lỗi đọc token: {str(e)}',
                    'can_approve': False
                }
                _token_status_cache = result
                _token_status_cache_time = time_module.time()
                return result
        
        if not creds:
            result = {
                'valid': False,
                'needs_reauth': True,
                'message': 'Không có credentials. Cần xác thực lại với Google.',
                'can_approve': False
            }
            _token_status_cache = result
            _token_status_cache_time = time_module.time()
            return result
        
        # Kiểm tra token có hết hạn không (chỉ kiểm tra expired, không test API để nhanh hơn)
        if creds.expired:
            # Try to refresh nếu có refresh_token (chỉ khi thực sự cần)
            if creds.refresh_token:
                try:
                    # Thử refresh token
                    creds.refresh(GoogleRequest())
                    # Lưu token mới
                    with open('token.pickle', 'wb') as token:
                        pickle.dump(creds, token)
                    result = {
                        'valid': True,
                        'needs_reauth': False,
                        'message': 'Token đã được refresh thành công.',
                        'can_approve': True
                    }
                    _token_status_cache = result
                    _token_status_cache_time = time_module.time()
                    return result
                except Exception as e:
                    error_str = str(e).lower()
                    if 'invalid_grant' in error_str:
                        result = {
                            'valid': False,
                            'needs_reauth': True,
                            'message': 'Token không hợp lệ (invalid_grant). Vui lòng bấm nút Refresh Token để ủy quyền lại.',
                            'can_approve': False
                        }
                    else:
                        result = {
                            'valid': False,
                            'needs_reauth': True,
                            'message': 'Token hết hạn và không thể refresh tự động. Vui lòng bấm nút Refresh Token để ủy quyền lại.',
                            'can_approve': False
                        }
                    _token_status_cache = result
                    _token_status_cache_time = time_module.time()
                    return result
            else:
                result = {
                    'valid': False,
                    'needs_reauth': True,
                    'message': 'Token hết hạn và không có refresh_token. Vui lòng bấm nút Refresh Token để ủy quyền lại.',
                    'can_approve': False
                }
                _token_status_cache = result
                _token_status_cache_time = time_module.time()
                return result
        
        # Token còn hiệu lực (không expired), không cần test API để nhanh hơn
        # Chỉ kiểm tra expired là đủ vì Google token sẽ tự động đánh dấu expired khi hết hạn
        result = {
            'valid': True,
            'needs_reauth': False,
            'message': 'Token hợp lệ.',
            'can_approve': True
        }
        _token_status_cache = result
        _token_status_cache_time = time_module.time()
        return result
        
    except Exception as e:
        result = {
            'valid': False,
            'needs_reauth': True,
            'message': f'Lỗi kiểm tra token: {str(e)}',
            'can_approve': False
        }
        _token_status_cache = result
        _token_status_cache_time = time_module.time()
        return result

@app.route('/sse/token-status')
def sse_token_status():
    """SSE endpoint for realtime token status (admin only)."""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    user_id = session['user_id']
    q = _sse_token_subscribe(user_id)
    
    def stream():
        yield 'retry: 5000\n\n'
        try:
            last = time_module.time()
            # Send initial status
            initial_status = check_google_token_status()
            import json as _json
            yield f"event: token_status\ndata: {_json.dumps(initial_status, ensure_ascii=False)}\n\n"
            
            while True:
                try:
                    item = q.get(timeout=1.0)
                    data = _json.dumps(item, ensure_ascii=False)
                    yield f"event: token_status\ndata: {data}\n\n"
                    last = time_module.time()
                except Exception:
                    now = time_module.time()
                    if now - last > 30:
                        yield ": keep-alive\n\n"
                        last = now
        finally:
            _sse_token_unsubscribe(user_id, q)
    
    from flask import Response
    return Response(stream(), mimetype='text/event-stream', headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/api/license/warning-status')
def api_license_warning_status():
    """
    Trả về trạng thái cảnh báo LICENSE gần nhất để UI có thể hiển thị ngay khi load trang,
    không phải đợi worker kiểm tra lại hoặc SSE push lần tiếp theo.
    """
    try:
        with _license_warning_lock:
            state = dict(_license_warning_state)
        return jsonify(state)
    except Exception as e:
        return jsonify({'active': False, 'error': str(e)}), 500

@app.route('/api/token/status')
def api_token_status():
    """API endpoint to check Google API token status."""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    user = db.session.get(User, session['user_id'])
    if not user or 'ADMIN' not in user.roles:
        return jsonify({'error': 'Admin only'}), 403

    force = request.args.get('force', '').lower() in ['1', 'true', 'yes']
    status = check_google_token_status(use_cache=not force)
    return jsonify(status)

@app.route('/api/token/authorize', methods=['POST'])
def api_token_authorize():
    """API endpoint to open Chrome for OAuth authorization."""
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        
        user = db.session.get(User, session['user_id'])
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401
        
        if 'ADMIN' not in user.roles:
            return jsonify({'success': False, 'error': 'Admin only'}), 403
        
        if not GOOGLE_API_AVAILABLE:
            return jsonify({
                'success': False,
                'message': 'Google API libraries không có sẵn. Vui lòng cài đặt các thư viện cần thiết.'
            }), 500
        
        if not os.path.exists('credentials.json'):
            return jsonify({
                'success': False,
                'message': 'Không tìm thấy file credentials.json. Vui lòng kiểm tra cấu hình.'
            }), 500
        
        # Đọc credentials để lấy redirect_uri đã đăng ký
        import json as json_module
        with open('credentials.json', 'r') as f:
            creds_data = json_module.load(f)
        
        # Sử dụng InstalledAppFlow, callback về Flask (không run_local_server để tránh mismatching_state)
        from google_auth_oauthlib.flow import InstalledAppFlow
        from urllib.parse import urlparse

        # Dùng loopback fixed để tránh lỗi device_id/device_name
        redirect_uri = 'http://127.0.0.1:5000/api/token/callback'

        flow = InstalledAppFlow.from_client_secrets_file(
            'credentials.json',
            ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets']
        )
        flow.redirect_uri = redirect_uri

        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent'
        )

        print(f"[OAuth] Redirect URI set to: {redirect_uri}")
        print(f"[OAuth] Authorization URL created (first 300 chars): {authorization_url[:300]}")
        print(f"[OAuth] State: {state}")

        # Lưu flow theo state để callback dùng lại
        if not hasattr(app, '_oauth_flow_store'):
            app._oauth_flow_store = {}
        app._oauth_flow_store[state] = flow

        # Lưu state vào session (không bắt buộc nhưng giúp debug)
        session['oauth_state'] = state
        session['oauth_redirect_uri'] = redirect_uri

        return jsonify({
            'success': True,
            'message': 'Đã tạo URL ủy quyền. Vui lòng mở Chrome để hoàn tất quá trình ủy quyền.',
            'auth_url': authorization_url
        })
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        error_trace = traceback.format_exc()
        print(f"Error in api_token_authorize: {error_msg}")
        print(f"Traceback: {error_trace}")
        
        return jsonify({
            'success': False,
            'message': f'Lỗi khi tạo URL ủy quyền: {error_msg}'
        }), 500

@app.route('/api/token/callback')
def api_token_callback():
    """Callback endpoint để nhận authorization code từ Google."""
    try:
        # Lấy authorization code từ query string
        code = request.args.get('code')
        state = request.args.get('state')
        error = request.args.get('error')
        
        if error:
            return f'''
            <html>
            <head><title>Authorization Error</title></head>
            <body>
                <h1>Lỗi ủy quyền</h1>
                <p>Google đã trả về lỗi: {error}</p>
                <p>Vui lòng thử lại.</p>
                <script>
                    setTimeout(function() {{
                        window.close();
                    }}, 3000);
                </script>
            </body>
            </html>
            ''', 400
        
        if not code:
            return '''
            <html>
            <head><title>Authorization Error</title></head>
            <body>
                <h1>Lỗi ủy quyền</h1>
                <p>Không nhận được authorization code từ Google.</p>
                <p>Vui lòng thử lại.</p>
                <script>
                    setTimeout(function() {
                        window.close();
                    }, 3000);
                </script>
            </body>
            </html>
            ''', 400
        
        # Lấy flow và redirect_uri theo state đã lưu
        flow = None
        redirect_uri = None

        # Ưu tiên lấy từ store (an toàn hơn, tránh mismatching_state)
        if hasattr(app, '_oauth_flow_store') and state in app._oauth_flow_store:
            flow = app._oauth_flow_store.pop(state)
            redirect_uri = getattr(flow, 'redirect_uri', None)
        else:
            # Fallback: kiểm tra session (có thể mất nếu reload)
            if 'oauth_state' not in session or session['oauth_state'] != state:
                return '''
                <html>
                <head><title>Authorization Error</title></head>
                <body>
                    <h1>Lỗi bảo mật</h1>
                    <p>State không khớp. Vui lòng thử lại.</p>
                    <script>
                        setTimeout(function() {
                            window.close();
                        }, 3000);
                    </script>
                </body>
                </html>
                ''', 400
            redirect_uri = session.get('oauth_redirect_uri')

        if not flow:
            # Nếu không có flow, tạo mới với redirect_uri đã biết
            if not redirect_uri:
                redirect_uri = 'http://127.0.0.1:5000/api/token/callback'
            from google_auth_oauthlib.flow import InstalledAppFlow
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json',
                ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets']
            )
            flow.redirect_uri = redirect_uri

        # Exchange authorization code để lấy credentials
        flow.fetch_token(code=code)
        creds = flow.credentials
        
        # Lưu credentials vào file
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
        
        # Xóa session data
        session.pop('oauth_state', None)
        session.pop('oauth_redirect_uri', None)
        # Dọn store
        if hasattr(app, '_oauth_flow_store'):
            app._oauth_flow_store.pop(state, None)
        
        # Publish token status update
        try:
            publish_token_status('valid', 'Token đã được ủy quyền thành công!')
        except Exception:
            pass
        
        return '''
        <html>
        <head><title>Authorization Success</title></head>
        <body>
            <h1>✅ Ủy quyền thành công!</h1>
            <p>Token đã được lưu thành công. Bạn có thể đóng cửa sổ này.</p>
            <script>
                // Thông báo cho parent window (nếu có)
                if (window.opener) {
                    window.opener.postMessage({type: 'token_authorized', success: true}, '*');
                }
                setTimeout(function() {
                    window.close();
                }, 2000);
            </script>
        </body>
        </html>
        '''
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        error_trace = traceback.format_exc()
        print(f"Error in api_token_callback: {error_msg}")
        print(f"Traceback: {error_trace}")
        
        return f'''
        <html>
        <head><title>Authorization Error</title></head>
        <body>
            <h1>Lỗi xử lý ủy quyền</h1>
            <p>Đã xảy ra lỗi: {error_msg}</p>
            <p>Vui lòng thử lại.</p>
            <script>
                setTimeout(function() {{
                    window.close();
                }}, 5000);
            </script>
        </body>
        </html>
        ''', 500

@app.route('/api/token/refresh', methods=['POST'])
def api_token_refresh():
    """API endpoint to try refreshing token."""
    try:
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        
        user = db.session.get(User, session['user_id'])
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401
        
        if 'ADMIN' not in user.roles:
            return jsonify({'success': False, 'error': 'Admin only'}), 403
        
        google_api = GoogleDriveAPI()
        if google_api.refresh_token():
            google_api.save_last_refresh_time()
            publish_token_status('valid', 'Token đã được refresh thành công!')
            return jsonify({
                'success': True,
                'message': 'Token đã được refresh thành công!'
            })
        else:
            publish_token_status('expired', 'Không thể refresh token. Cần xác thực lại.', needs_reauth=True)
            return jsonify({
                'success': False,
                'message': 'Không thể refresh token. Cần chạy refresh_token.py để xác thực lại.'
            }), 400
    except Exception as e:
        import traceback
        error_msg = str(e)
        error_trace = traceback.format_exc()
        print(f"Error in api_token_refresh: {error_msg}")
        print(f"Traceback: {error_trace}")
        
        if 'invalid_grant' in error_msg.lower():
            publish_token_status('expired', 'Token không hợp lệ (invalid_grant). Cần xác thực lại.', needs_reauth=True)
            return jsonify({
                'success': False,
                'message': 'Token không hợp lệ (invalid_grant). Cần chạy refresh_token.py để xác thực lại.'
            }), 400
        
        # Đảm bảo luôn trả về JSON
        return jsonify({
            'success': False,
            'message': f'Lỗi khi refresh token: {error_msg}'
        }), 500

@app.route('/leave-request/<int:request_id>/edit', methods=['GET', 'POST'])
def edit_leave_request(request_id):
    """Sửa đơn xin nghỉ phép"""
    if 'user_id' not in session:
        flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
        return redirect(url_for('login'))
    
    user = db.session.get(User, session['user_id'])
    if not user:
        session.clear()
        flash('Phiên đăng nhập không hợp lệ!', 'error')
        return redirect(url_for('login'))
    
    leave_request = LeaveRequest.query.get_or_404(request_id)
    
    # Chỉ người tạo đơn mới có thể sửa
    if user.id != leave_request.user_id:
        abort(403)
    
    # Chỉ có thể sửa khi đơn đang ở trạng thái chờ phê duyệt hoặc bị từ chối
    if leave_request.status not in ['pending', 'rejected']:
        flash('Chỉ có thể sửa đơn khi đang chờ phê duyệt hoặc bị từ chối', 'error')
        return redirect(url_for('view_leave_request', request_id=request_id))
    
    if request.method == 'POST':
        try:
            data = request.form
            
            # Cập nhật thông tin nhân viên và lý do
            leave_request.employee_name = data.get('employee_name')
            leave_request.team = data.get('team')
            leave_request.employee_code = data.get('employee_code')
            leave_request.request_type = data.get('request_type', 'leave')
            leave_request.late_early_type = data.get('late_early_type')
            leave_request.leave_reason = data.get('leave_reason')

            # Cập nhật thời gian từ các trường date/time hiện có trên form
            from_time_str = data.get('leave_from_time') or '00:00'
            to_time_str = data.get('leave_to_time') or '00:00'
            
            clean_from_time = clean_time_format(from_time_str)
            clean_to_time = clean_time_format(to_time_str)
            
            from_time = clean_from_time.split(':')
            to_time = clean_to_time.split(':')
            from_date_parts = (data.get('leave_from_date') or '2024-01-01').split('-')
            to_date_parts = (data.get('leave_to_date') or '2024-01-01').split('-')

            leave_request.leave_from_hour = int(from_time[0])
            leave_request.leave_from_minute = int(from_time[1])
            leave_request.leave_from_year = int(from_date_parts[0])
            leave_request.leave_from_month = int(from_date_parts[1])
            leave_request.leave_from_day = int(from_date_parts[2])

            leave_request.leave_to_hour = int(to_time[0])
            leave_request.leave_to_minute = int(to_time[1])
            leave_request.leave_to_year = int(to_date_parts[0])
            leave_request.leave_to_month = int(to_date_parts[1])
            leave_request.leave_to_day = int(to_date_parts[2])

            # Cập nhật hình thức nghỉ
            leave_request.annual_leave_days = float(data.get('annual_leave_days', 0) or 0)
            leave_request.unpaid_leave_days = float(data.get('unpaid_leave_days', 0) or 0)
            leave_request.special_leave_days = float(data.get('special_leave_days', 0) or 0)
            leave_request.special_leave_type = data.get('special_leave_type')
            # Cập nhật ca làm việc áp dụng khi xin nghỉ (nếu có chọn)
            sel_shift = data.get('leave_shift_code')
            if sel_shift in ['1','2','3','4']:
                leave_request.shift_code = sel_shift

            # Chuẩn hóa bội số 0.5 cho số ngày
            def ensure_half_step(x):
                return (int(round((x or 0.0) * 2)) / 2.0)
            leave_request.annual_leave_days = ensure_half_step(leave_request.annual_leave_days)
            leave_request.unpaid_leave_days = ensure_half_step(leave_request.unpaid_leave_days)
            leave_request.special_leave_days = ensure_half_step(leave_request.special_leave_days)

            # Nếu đơn từng bị từ chối, khi người dùng sửa và gửi lại => reset về trạng thái chờ trưởng nhóm duyệt
            if leave_request.status == 'rejected':
                leave_request.status = 'pending'
                leave_request.step = 'leader'
                leave_request.current_approver_id = None
                leave_request.reject_reason = None
                # Xóa toàn bộ chữ ký/phê duyệt cũ để quy trình duyệt lại từ đầu
                leave_request.team_leader_signature = None
                leave_request.team_leader_signer_id = None
                leave_request.team_leader_approved_at = None
                leave_request.manager_signature = None
                leave_request.manager_signer_id = None
                leave_request.manager_approved_at = None
                leave_request.admin_signature = None
                leave_request.admin_signer_id = None
                leave_request.admin_approved_at = None

            # Ràng buộc: tổng ngày xin nghỉ không vượt quá số ngày trong khoảng từ ngày-đến ngày
            try:
                from_date_dt = datetime.strptime(data.get('leave_from_date', '2024-01-01'), '%Y-%m-%d')
                to_date_dt = datetime.strptime(data.get('leave_to_date', '2024-01-01'), '%Y-%m-%d')
                from_time_str = data.get('leave_from_time') or '00:00'
                to_time_str = data.get('leave_to_time') or '00:00'
                
                clean_from_time = clean_time_format(from_time_str)
                clean_to_time = clean_time_format(to_time_str)
                
                # Validate: Thời gian bắt đầu phải nhỏ hơn thời gian kết thúc
                start_dt = datetime.combine(from_date_dt.date(), datetime.strptime(clean_from_time, '%H:%M').time())
                end_dt = datetime.combine(to_date_dt.date(), datetime.strptime(clean_to_time, '%H:%M').time())
                
                if start_dt >= end_dt:
                    flash('Thời gian từ ngày giờ phải nhỏ hơn đến ngày giờ', 'error')
                    return redirect(url_for('edit_leave_request', request_id=request_id))
                
                # Validate: Giờ kết thúc không được vượt quá giờ ra khỏi ca
                shift_code = data.get('leave_shift_code', '1')
                shift_ranges = {
                    '1': {'start': '07:30', 'end': '16:30'},
                    '2': {'start': '09:00', 'end': '18:00'},
                    '3': {'start': '11:00', 'end': '20:00'},
                    '4': {'start': '08:00', 'end': '17:00'}
                }
                
                if shift_code in shift_ranges:
                    shift_end_time = shift_ranges[shift_code]['end']
                    shift_end_dt = datetime.combine(to_date_dt.date(), datetime.strptime(shift_end_time, '%H:%M').time())
                    
                    # Chỉ kiểm tra nếu cùng ngày (không kiểm tra khi nghỉ qua nhiều ngày)
                    if from_date_dt.date() == to_date_dt.date() and end_dt > shift_end_dt:
                        flash(f'Giờ kết thúc nghỉ ({clean_to_time}) không được vượt quá giờ ra khỏi ca ({shift_end_time})', 'error')
                        return redirect(url_for('edit_leave_request', request_id=request_id))
                
                available_units = _compute_leave_units_generic(from_date_dt, clean_from_time, to_date_dt, clean_to_time)
            except Exception:
                available_units = None
            if available_units is not None:
                if (leave_request.annual_leave_days + leave_request.unpaid_leave_days + leave_request.special_leave_days) > available_units + 1e-9:
                    flash('Tổng số ngày xin nghỉ vượt quá số ngày có thể xin trong khoảng thời gian đã chọn (theo ca làm việc).', 'error')
                    return redirect(url_for('edit_leave_request', request_id=request_id))

            leave_request.substitute_name = data.get('substitute_name')
            leave_request.substitute_employee_id = data.get('substitute_employee_id')
            
            # Xử lý notes và use_lunch_break khi edit
            import json
            notes_from_form = data.get('notes', '').strip()
            use_lunch_break_value = data.get('use_lunch_break')
            
            # Lấy use_lunch_break từ notes cũ (nếu có)
            existing_use_lunch_break = None
            if leave_request.notes:
                try:
                    existing_notes_data = json.loads(leave_request.notes)
                    if isinstance(existing_notes_data, dict) and 'use_lunch_break' in existing_notes_data:
                        existing_use_lunch_break = existing_notes_data['use_lunch_break']
                except (json.JSONDecodeError, TypeError, ValueError):
                    pass
            
            # Nếu form có gửi use_lunch_break, dùng giá trị từ form
            # Nếu không, giữ lại giá trị cũ (nếu có)
            final_use_lunch_break = None
            if use_lunch_break_value in ['true', 'false']:
                final_use_lunch_break = use_lunch_break_value == 'true'
            elif existing_use_lunch_break is not None:
                final_use_lunch_break = existing_use_lunch_break
            
            # Lưu notes dưới dạng JSON nếu có use_lunch_break hoặc có notes text
            if final_use_lunch_break is not None or notes_from_form:
                notes_data = {}
                if notes_from_form:
                    notes_data['_original_notes'] = notes_from_form
                if final_use_lunch_break is not None:
                    notes_data['use_lunch_break'] = final_use_lunch_break
                leave_request.notes = json.dumps(notes_data, ensure_ascii=False) if notes_data else None
            else:
                leave_request.notes = notes_from_form if notes_from_form else None
            
            # Xử lý file upload mới
            if 'attachments' in request.files:
                files = request.files.getlist('attachments')
                new_attachments = []
                
                # Lấy danh sách attachments hiện có
                existing_attachments = []
                if leave_request.attachments:
                    try:
                        existing_attachments = json.loads(leave_request.attachments)
                    except (json.JSONDecodeError, TypeError, AttributeError):
                        existing_attachments = []
                
                # Xử lý các file mới được upload
                for file in files:
                    if file and file.filename:
                        # Tạo tên file unique
                        filename = f"{uuid.uuid4()}_{file.filename}"
                        
                        # Tạo thư mục uploads nếu chưa có
                        upload_dir = os.path.join(app.root_path, 'uploads', 'leave_requests')
                        os.makedirs(upload_dir, exist_ok=True)
                        
                        # Lưu file
                        file_path = os.path.join(upload_dir, filename)
                        file.save(file_path)
                        
                        new_attachments.append({
                            'original_name': file.filename,
                            'saved_name': filename,
                            'size': file.content_length or 0
                        })
                
                # Kết hợp attachments cũ và mới
                all_attachments = existing_attachments + new_attachments
                leave_request.attachments = json.dumps(all_attachments) if all_attachments else None
            
            db.session.commit()
            
            # Kiểm tra xem người dùng có muốn gửi email hay không
            email_consent = data.get('email_consent', 'no').lower()
            send_email = email_consent == 'yes'
            print(f"[DEBUG] Edit - Email consent received: '{email_consent}', send_email: {send_email}")
            
            if send_email:
                try:
                    print(f"[Mail] Attempting to send update email for leave_request #{leave_request.id} by user #{user.id} ({user.name})")
                    send_leave_request_email_async(leave_request, user, action='update')
                    # Process any pending DB updates from async threads
                    from utils.email_utils import process_db_updates
                    process_db_updates()
                    # Lưu trạng thái email vào session cho tất cả vai trò
                    session['email_status'] = {
                        'request_id': leave_request.id,
                        'status': 'sending',
                        'message': 'Đang gửi email thông báo...'
                    }
                    # Chỉ thông báo về cập nhật; tiến trình email sẽ do toast hiển thị
                    # flash('Đã cập nhật đơn thành công! Email thông báo đã được gửi đến phòng nhân sự.', 'success')
                except Exception as e:
                    print(f"[Mail] Error scheduling leave update email: {e}")
                    # Lưu trạng thái email vào session cho tất cả vai trò
                    session['email_status'] = {
                        'request_id': leave_request.id,
                        'status': 'error',
                        'message': f'Lỗi khi gửi email: {str(e)}'
                    }
                    flash('Đơn đã cập nhật thành công, nhưng có lỗi khi gửi email thông báo.', 'warning')
            else:
                # Không gửi email
                print(f"[Mail] User chose not to send email for leave_request update #{leave_request.id}")
                session['email_status'] = {
                    'request_id': leave_request.id,
                    'status': 'skipped',
                    'message': 'Người dùng đã chọn không gửi email thông báo'
                }
                # flash('Đã cập nhật đơn thành công! (Không gửi email thông báo)', 'success')
            return redirect(url_for('leave_requests_list', status='pending'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi khi cập nhật đơn xin nghỉ phép: {str(e)}', 'error')
    
    # Xử lý GET request - hiển thị form sửa
    work_shift = '08:00 - 17:00'
    current_role = session.get('current_role', user.roles.split(',')[0])
    
    # Parse existing attachments for template
    existing_attachments_list = []
    if leave_request.attachments:
        try:
            existing_attachments_list = json.loads(leave_request.attachments)
        except (json.JSONDecodeError, TypeError, AttributeError):
            existing_attachments_list = []
    
    return render_template('leave_request_form.html', 
                         leave_request=leave_request, 
                         is_edit=True, 
                         user=user, 
                         current_role=current_role, 
                         work_shift=work_shift,
                         existing_attachments_list=existing_attachments_list)


def _format_leave_days_summary_for_sheet(leave_request):
    """Tạo chuỗi mô tả số ngày nghỉ để ghi vào Google Sheet."""
    def _fmt(value):
        if value is None:
            return None
        try:
            value = float(value)
        except (TypeError, ValueError):
            return None
        if value <= 0:
            return None
        if value.is_integer():
            return str(int(value))
        return f"{value:.1f}".rstrip('0').rstrip('.')
    
    parts = []
    
    annual_text = _fmt(leave_request.annual_leave_days)
    if annual_text:
        parts.append(f"Phép năm: {annual_text} ngày")
    
    unpaid_text = _fmt(leave_request.unpaid_leave_days)
    if unpaid_text:
        parts.append(f"Nghỉ không lương: {unpaid_text} ngày")
    
    special_text = _fmt(leave_request.special_leave_days)
    if special_text:
        special_label = "Nghỉ đặc biệt"
        if leave_request.special_leave_type:
            special_label += f" ({leave_request.special_leave_type})"
        parts.append(f"{special_label}: {special_text} ngày")
    
    if not parts:
        return "0 ngày"
    
    return " | ".join(parts)


def _generate_leave_date_range(leave_request):
    """Trả về danh sách ngày (date) nằm trong khoảng nghỉ phép."""
    try:
        start_date = leave_request.get_leave_from_datetime().date()
        end_date = leave_request.get_leave_to_datetime().date()
    except Exception:
        return []
    
    dates = []
    current = start_date
    while current <= end_date:
        dates.append(current)
        current += timedelta(days=1)
    return dates


def schedule_leave_sheet_updates(leave_request, approver=None):
    """Đưa thông tin số ngày nghỉ lên Google Sheet (cột P) sau khi Admin phê duyệt."""
    import sys
    from datetime import datetime as dt
    
    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    # Dùng print trực tiếp vào cả stdout và stderr để đảm bảo log được hiển thị
    try:
        print(f"\n{'='*80}", flush=True, file=sys.stderr)
        print(f"🔄 [SCHEDULE_LEAVE_SHEET] {timestamp} - Bắt đầu lên lịch cập nhật Google Sheet", flush=True, file=sys.stderr)
        print(f"   Leave Request ID: {leave_request.id if leave_request else 'None'}", flush=True, file=sys.stderr)
        print(f"   Approver: {approver.name if approver else 'None'}", flush=True, file=sys.stderr)
        print(f"{'='*80}\n", flush=True, file=sys.stderr)
        sys.stderr.flush()
    except Exception:
        pass
    
    try:
        print(f"\n{'='*80}", flush=True)
        print(f"🔄 [SCHEDULE_LEAVE_SHEET] {timestamp} - Bắt đầu lên lịch cập nhật Google Sheet", flush=True)
        print(f"   Leave Request ID: {leave_request.id if leave_request else 'None'}", flush=True)
        print(f"   Approver: {approver.name if approver else 'None'}", flush=True)
        print(f"{'='*80}\n", flush=True)
    except Exception:
        pass
    
    try:
        _safe_print(f"\n{'='*80}")
        _safe_print(f"🔄 [SCHEDULE_LEAVE_SHEET] {timestamp} - Bắt đầu lên lịch cập nhật Google Sheet")
        _safe_print(f"   Leave Request ID: {leave_request.id if leave_request else 'None'}")
        _safe_print(f"   Approver: {approver.name if approver else 'None'}")
        _safe_print(f"{'='*80}\n")
        sys.stdout.flush()
    except Exception:
        pass
    
    try:
        # Dùng print trực tiếp vào cả stdout và stderr để đảm bảo log được hiển thị
        try:
            print(f"\n{'='*80}", flush=True, file=sys.stderr)
            print(f"🔍 [SCHEDULE_LEAVE_SHEET] Bắt đầu kiểm tra điều kiện", flush=True, file=sys.stderr)
            print(f"   Leave Request: {'✅ Có' if leave_request else '❌ None'}", flush=True, file=sys.stderr)
            if leave_request:
                print(f"   Request Type: {leave_request.request_type}", flush=True, file=sys.stderr)
                print(f"   Request ID: {leave_request.id}", flush=True, file=sys.stderr)
            print(f"   Điều kiện (request_type == 'leave'): {leave_request.request_type == 'leave' if leave_request else False}", flush=True, file=sys.stderr)
            print(f"{'='*80}\n", flush=True, file=sys.stderr)
            sys.stderr.flush()
        except Exception:
            pass
        
        try:
            print(f"\n{'='*80}", flush=True)
            print(f"🔍 [SCHEDULE_LEAVE_SHEET] Bắt đầu kiểm tra điều kiện", flush=True)
            print(f"   Leave Request: {'✅ Có' if leave_request else '❌ None'}", flush=True)
            if leave_request:
                print(f"   Request Type: {leave_request.request_type}", flush=True)
                print(f"   Request ID: {leave_request.id}", flush=True)
            print(f"   Điều kiện (request_type == 'leave'): {leave_request.request_type == 'leave' if leave_request else False}", flush=True)
            print(f"{'='*80}\n", flush=True)
        except Exception:
            pass
        
        try:
            _safe_print(f"\n{'='*80}")
            _safe_print(f"🔍 [SCHEDULE_LEAVE_SHEET] Bắt đầu kiểm tra điều kiện")
            _safe_print(f"   Leave Request: {'✅ Có' if leave_request else '❌ None'}")
            if leave_request:
                _safe_print(f"   Request Type: {leave_request.request_type}")
                _safe_print(f"   Request ID: {leave_request.id}")
            _safe_print(f"   Điều kiện (request_type == 'leave'): {leave_request.request_type == 'leave' if leave_request else False}")
            _safe_print(f"{'='*80}\n")
            sys.stdout.flush()
        except Exception:
            pass
        
        # Không có đơn nghỉ thì dừng
        if not leave_request:
            try:
                _safe_print(f"\n{'='*80}")
                _safe_print(f"⚠️ [SCHEDULE_LEAVE_SHEET] ❌ DỪNG - Không có đơn nghỉ phép")
                _safe_print(f"{'='*80}\n")
                sys.stdout.flush()
            except Exception:
                pass
            return

        # Xử lý riêng cho đơn nghỉ 30 phút: chỉ đẩy memo vào cột P
        if leave_request.request_type == '30min_break':
            try:
                from_hour = getattr(leave_request, 'leave_from_hour', 0) or 0
                from_minute = getattr(leave_request, 'leave_from_minute', 0) or 0
                to_hour = getattr(leave_request, 'leave_to_hour', 0) or 0
                to_minute = getattr(leave_request, 'leave_to_minute', 0) or 0
                from_time_str = f"{from_hour:02d}:{from_minute:02d}"
                to_time_str = f"{to_hour:02d}:{to_minute:02d}"
            except Exception:
                from_time_str = ""
                to_time_str = ""

            try:
                # Lấy nhân viên & team
                employee = leave_request.user
                employee_team = employee.department or leave_request.team or "Unknown"
                employee_id = employee.employee_id
            except Exception:
                employee = None
                employee_team = leave_request.team or "Unknown"
                employee_id = getattr(leave_request, 'employee_code', 'Unknown')

            from_datetime = leave_request.get_leave_from_datetime()
            leave_date = from_datetime.date() if from_datetime else datetime.utcnow().date()

            approved_by = approver.name if approver else "Admin"
            approved_at = (leave_request.admin_approved_at or datetime.utcnow()).strftime('%Y-%m-%d %H:%M:%S')

            summary_text = f"Nghỉ 30 phút: {from_time_str} - {to_time_str}".strip()

            attendance_data = {
                'date': leave_date.isoformat(),
                'user_name': leave_request.employee_name,
                'approved_by': approved_by,
                'approved_at': approved_at,
                'leave_summary': summary_text,
                'memo_only': True
            }

            try:
                _safe_print(f"📌 [SCHEDULE_LEAVE_SHEET] Đơn nghỉ 30 phút -> chỉ cập nhật cột P")
                _safe_print(f"   Ngày: {leave_date.isoformat()}")
                _safe_print(f"   Memo: {summary_text}")
            except Exception:
                pass

            try:
                update_google_sheet_background_safe(
                    attendance_id=f"break30-{leave_request.id}",
                    employee_team=employee_team,
                    employee_id=employee_id,
                    attendance_data=attendance_data
                )
            except Exception as memo_err:
                try:
                    _safe_print(f"❌ [SCHEDULE_LEAVE_SHEET] Lỗi khi cập nhật memo nghỉ 30 phút: {memo_err}")
                except Exception:
                    pass

        # Xử lý riêng cho đơn đi trễ/về sớm: cập nhật cột P và trừ giờ từ cột G hoặc K
        if leave_request.request_type == 'late_early':
            try:
                from_hour = getattr(leave_request, 'leave_from_hour', 0) or 0
                from_minute = getattr(leave_request, 'leave_from_minute', 0) or 0
                to_hour = getattr(leave_request, 'leave_to_hour', 0) or 0
                to_minute = getattr(leave_request, 'leave_to_minute', 0) or 0
                from_time_str = f"{from_hour:02d}:{from_minute:02d}"
                to_time_str = f"{to_hour:02d}:{to_minute:02d}"
                
                # Tính số phút đi trễ/về sớm
                from_total_minutes = from_hour * 60 + from_minute
                to_total_minutes = to_hour * 60 + to_minute
                late_early_minutes = abs(to_total_minutes - from_total_minutes)
            except Exception:
                from_time_str = ""
                to_time_str = ""
                late_early_minutes = 0

            try:
                # Lấy nhân viên & team
                employee = leave_request.user
                employee_team = employee.department or leave_request.team or "Unknown"
                employee_id = employee.employee_id
            except Exception:
                employee = None
                employee_team = leave_request.team or "Unknown"
                employee_id = getattr(leave_request, 'employee_code', 'Unknown')

            from_datetime = leave_request.get_leave_from_datetime()
            leave_date = from_datetime.date() if from_datetime else datetime.utcnow().date()
            leave_date_str = leave_date.strftime('%d/%m/%Y')

            approved_by = approver.name if approver else "Admin"
            approved_at = (leave_request.admin_approved_at or datetime.utcnow()).strftime('%Y-%m-%d %H:%M:%S')

            # Xác định loại đi trễ/về sớm
            late_early_type = getattr(leave_request, 'late_early_type', None) or 'late'
            
            # Tạo memo text cho cột P (ngắn gọn)
            if late_early_type == 'late':
                summary_text = f"Đi trễ từ {from_time_str}-{to_time_str}"
            else:  # early
                summary_text = f"Về sớm {from_time_str}-{to_time_str}"

            attendance_data = {
                'date': leave_date.isoformat(),
                'user_name': leave_request.employee_name,
                'approved_by': approved_by,
                'approved_at': approved_at,
                'leave_summary': summary_text,
                'late_early_type': late_early_type,
                'late_early_minutes': late_early_minutes,
                'memo_only': True  # memo_only = True để không động vào các cột N, O, E, M. Logic trừ giờ G/K chạy riêng.
            }

            try:
                _safe_print(f"📌 [SCHEDULE_LEAVE_SHEET] Đơn đi trễ/về sớm -> cập nhật cột P và trừ giờ")
                _safe_print(f"   Ngày: {leave_date.isoformat()}")
                _safe_print(f"   Loại: {late_early_type}")
                _safe_print(f"   Memo: {summary_text}")
                _safe_print(f"   Số phút: {late_early_minutes}")
            except Exception:
                pass

            try:
                update_google_sheet_background_safe(
                    attendance_id=f"late_early-{leave_request.id}",
                    employee_team=employee_team,
                    employee_id=employee_id,
                    attendance_data=attendance_data
                )
            except Exception as late_early_err:
                try:
                    _safe_print(f"❌ [SCHEDULE_LEAVE_SHEET] Lỗi khi cập nhật đơn đi trễ/về sớm: {late_early_err}")
                except Exception:
                    pass

        # Các loại khác ngoài 'leave', '30min_break' và 'late_early' thì không xử lý
        if leave_request.request_type != 'leave':
            try:
                _safe_print(f"\n{'='*80}")
                _safe_print(f"⚠️ [SCHEDULE_LEAVE_SHEET] ❌ DỪNG - Không phải đơn nghỉ phép")
                _safe_print(f"   Leave Request: {'✅ Có' if leave_request else '❌ None'}")
                if leave_request:
                    _safe_print(f"   Request Type: {leave_request.request_type}")
                _safe_print(f"{'='*80}\n")
                sys.stdout.flush()
            except Exception:
                pass
            return

        # Dùng cùng logic phân bổ ngày như Excel export để mỗi ngày có đúng số ngày nghỉ
        try:
            print(f"📊 [SCHEDULE_LEAVE_SHEET] Đang phân bổ ngày nghỉ...", flush=True, file=sys.stderr)
            _safe_print(f"📊 [SCHEDULE_LEAVE_SHEET] Đang phân bổ ngày nghỉ...")
            
            # Log thông tin đơn nghỉ phép trước khi xử lý
            try:
                # Lấy thông tin giờ từ hour và minute
                from_hour = getattr(leave_request, 'leave_from_hour', 0) or 0
                from_minute = getattr(leave_request, 'leave_from_minute', 0) or 0
                to_hour = getattr(leave_request, 'leave_to_hour', 0) or 0
                to_minute = getattr(leave_request, 'leave_to_minute', 0) or 0
                from_time_str = f"{from_hour:02d}:{from_minute:02d}"
                to_time_str = f"{to_hour:02d}:{to_minute:02d}"
                
                print(f"🔍 [SCHEDULE_LEAVE_SHEET] Thông tin đơn nghỉ phép:", flush=True, file=sys.stderr)
                print(f"   ID: {leave_request.id}", flush=True, file=sys.stderr)
                print(f"   Từ ngày: {leave_request.leave_from_year}-{leave_request.leave_from_month}-{leave_request.leave_from_day}", flush=True, file=sys.stderr)
                print(f"   Đến ngày: {leave_request.leave_to_year}-{leave_request.leave_to_month}-{leave_request.leave_to_day}", flush=True, file=sys.stderr)
                print(f"   Từ giờ: {from_time_str}", flush=True, file=sys.stderr)
                print(f"   Đến giờ: {to_time_str}", flush=True, file=sys.stderr)
                print(f"   Loại nghỉ: {getattr(leave_request, 'leave_type', 'N/A')}", flush=True, file=sys.stderr)
                print(f"   Annual leave days: {getattr(leave_request, 'annual_leave_days', 'N/A')}", flush=True, file=sys.stderr)
                print(f"   Unpaid leave days: {getattr(leave_request, 'unpaid_leave_days', 'N/A')}", flush=True, file=sys.stderr)
                print(f"   Special leave days: {getattr(leave_request, 'special_leave_days', 'N/A')}", flush=True, file=sys.stderr)
                _safe_print(f"🔍 [SCHEDULE_LEAVE_SHEET] Thông tin đơn nghỉ phép:")
                _safe_print(f"   ID: {leave_request.id}")
                _safe_print(f"   Từ ngày: {leave_request.leave_from_year}-{leave_request.leave_from_month}-{leave_request.leave_from_day}")
                _safe_print(f"   Đến ngày: {leave_request.leave_to_year}-{leave_request.leave_to_month}-{leave_request.leave_to_day}")
                _safe_print(f"   Từ giờ: {from_time_str}")
                _safe_print(f"   Đến giờ: {to_time_str}")
                _safe_print(f"   Loại nghỉ: {getattr(leave_request, 'leave_type', 'N/A')}")
                _safe_print(f"   Annual leave days: {getattr(leave_request, 'annual_leave_days', 'N/A')}")
                _safe_print(f"   Unpaid leave days: {getattr(leave_request, 'unpaid_leave_days', 'N/A')}")
                _safe_print(f"   Special leave days: {getattr(leave_request, 'special_leave_days', 'N/A')}")
            except Exception as info_err:
                try:
                    print(f"⚠️ [SCHEDULE_LEAVE_SHEET] Lỗi khi log thông tin đơn: {info_err}", flush=True, file=sys.stderr)
                except Exception:
                    pass
            
            # Gọi process_leave_requests_for_excel ngay cả khi log lỗi
            try:
                sys.stdout.flush()
            except Exception:
                pass
            
            from utils.excel_leave_processor import process_leave_requests_for_excel
            daily_leaves = process_leave_requests_for_excel([leave_request])
            
            print(f"✅ [SCHEDULE_LEAVE_SHEET] Phân bổ thành công: {len(daily_leaves)} ngày", flush=True, file=sys.stderr)
            if len(daily_leaves) > 0:
                print(f"📋 [SCHEDULE_LEAVE_SHEET] Chi tiết các ngày:", flush=True, file=sys.stderr)
                for idx, day in enumerate(daily_leaves[:5], 1):  # Chỉ log 5 ngày đầu
                    print(f"   Ngày {idx}: {day.get('date')} - {day.get('fractional_days', 'N/A')} ngày", flush=True, file=sys.stderr)
            else:
                print(f"⚠️ [SCHEDULE_LEAVE_SHEET] Không có ngày nào được phân bổ!", flush=True, file=sys.stderr)
            
            try:
                _safe_print(f"✅ [SCHEDULE_LEAVE_SHEET] Phân bổ thành công: {len(daily_leaves)} ngày")
                if len(daily_leaves) > 0:
                    _safe_print(f"📋 [SCHEDULE_LEAVE_SHEET] Chi tiết các ngày:")
                    for idx, day in enumerate(daily_leaves[:5], 1):
                        _safe_print(f"   Ngày {idx}: {day.get('date')} - {day.get('fractional_days', 'N/A')} ngày")
                else:
                    _safe_print(f"⚠️ [SCHEDULE_LEAVE_SHEET] Không có ngày nào được phân bổ!")
                try:
                    sys.stdout.flush()
                except Exception:
                    pass
            except Exception:
                pass
        except Exception as alloc_err:
            try:
                print(f"\n{'='*80}", flush=True, file=sys.stderr)
                print(f"❌ [SCHEDULE_LEAVE_SHEET] Lỗi khi phân bổ ngày nghỉ cho đơn #{leave_request.id}", flush=True, file=sys.stderr)
                print(f"   Error: {str(alloc_err)}", flush=True, file=sys.stderr)
                print(f"   Type: {type(alloc_err).__name__}", flush=True, file=sys.stderr)
                import traceback
                print(f"   Traceback:", flush=True, file=sys.stderr)
                print(traceback.format_exc(), flush=True, file=sys.stderr)
                print(f"{'='*80}\n", flush=True, file=sys.stderr)
            except Exception:
                pass
            
            try:
                _safe_print(f"\n{'='*80}")
                _safe_print(f"❌ [SCHEDULE_LEAVE_SHEET] Lỗi khi phân bổ ngày nghỉ cho đơn #{leave_request.id}")
                _safe_print(f"   Error: {str(alloc_err)}")
                _safe_print(f"   Type: {type(alloc_err).__name__}")
                import traceback
                _safe_print(f"   Traceback:")
                _safe_print(traceback.format_exc())
                _safe_print(f"{'='*80}\n")
            except Exception:
                pass
            
            daily_leaves = []
        
        try:
            print(f"🔍 [SCHEDULE_LEAVE_SHEET] Kiểm tra daily_leaves: {len(daily_leaves) if daily_leaves else 0} ngày", flush=True, file=sys.stderr)
            _safe_print(f"🔍 [SCHEDULE_LEAVE_SHEET] Kiểm tra daily_leaves: {len(daily_leaves) if daily_leaves else 0} ngày")
            sys.stdout.flush()
        except Exception:
            pass
        
        if not daily_leaves:
            try:
                print(f"\n{'='*80}", flush=True, file=sys.stderr)
                print(f"⚠️ [SCHEDULE_LEAVE_SHEET] ❌ DỪNG - Không có dữ liệu daily_leaves cho đơn #{leave_request.id}", flush=True, file=sys.stderr)
                print(f"   Số ngày: {len(daily_leaves)}", flush=True, file=sys.stderr)
                print(f"{'='*80}\n", flush=True, file=sys.stderr)
                _safe_print(f"\n{'='*80}")
                _safe_print(f"⚠️ [SCHEDULE_LEAVE_SHEET] ❌ DỪNG - Không có dữ liệu daily_leaves cho đơn #{leave_request.id}")
                _safe_print(f"   Số ngày: {len(daily_leaves)}")
                _safe_print(f"{'='*80}\n")
                sys.stdout.flush()
            except Exception:
                pass
            return
        
        try:
            _safe_print(f"👤 [SCHEDULE_LEAVE_SHEET] Đang lấy thông tin nhân viên...")
            sys.stdout.flush()
        except Exception:
            pass
        
        employee = leave_request.user or db.session.get(User, leave_request.user_id)
        
        try:
            _safe_print(f"🔍 [SCHEDULE_LEAVE_SHEET] Kiểm tra thông tin nhân viên:")
            _safe_print(f"   Employee: {'✅ Có' if employee else '❌ None'}")
            if employee:
                _safe_print(f"   Employee ID: {employee.employee_id if hasattr(employee, 'employee_id') else 'None'}")
            _safe_print(f"   Điều kiện (employee and employee.employee_id): {bool(employee and employee.employee_id)}")
            sys.stdout.flush()
        except Exception:
            pass
        
        if not employee or not employee.employee_id:
            try:
                _safe_print(f"\n{'='*80}")
                _safe_print(f"⚠️ [SCHEDULE_LEAVE_SHEET] ❌ DỪNG - Thiếu thông tin nhân viên cho đơn #{leave_request.id}")
                _safe_print(f"   Employee: {'✅ Có' if employee else '❌ None'}")
                if employee:
                    _safe_print(f"   Employee ID: {employee.employee_id if hasattr(employee, 'employee_id') else 'None'}")
                _safe_print(f"{'='*80}\n")
                sys.stdout.flush()
            except Exception:
                pass
            return
        
        employee_team = employee.department or leave_request.team or "Unknown"
        approved_by = approver.name if approver else "Admin"
        approved_at = (leave_request.admin_approved_at or datetime.utcnow()).strftime('%Y-%m-%d %H:%M:%S')
        attendance_prefix = f"leave-{leave_request.id}"
        employee_id = employee.employee_id
        
        def _worker():
            import sys
            try:
                print(f"\n{'='*80}", flush=True, file=sys.stderr)
                print(f"🚀 [LEAVE_SHEET_UPDATE] Bắt đầu xử lý {len(daily_leaves)} ngày nghỉ cho đơn #{leave_request.id}", flush=True, file=sys.stderr)
                print(f"   👤 Nhân viên: {leave_request.employee_name} (Employee ID: {employee_id})", flush=True, file=sys.stderr)
                print(f"   🏢 Phòng ban: {employee_team}", flush=True, file=sys.stderr)
                print(f"   ✅ Phê duyệt bởi: {approved_by}", flush=True, file=sys.stderr)
                print(f"   ⏰ Thời gian phê duyệt: {approved_at}", flush=True, file=sys.stderr)
                print(f"{'='*80}\n", flush=True, file=sys.stderr)
                _safe_print(f"\n{'='*80}")
                _safe_print(f"🚀 [LEAVE_SHEET_UPDATE] Bắt đầu xử lý {len(daily_leaves)} ngày nghỉ cho đơn #{leave_request.id}")
                _safe_print(f"   👤 Nhân viên: {leave_request.employee_name} (Employee ID: {employee_id})")
                _safe_print(f"   🏢 Phòng ban: {employee_team}")
                _safe_print(f"   ✅ Phê duyệt bởi: {approved_by}")
                _safe_print(f"   ⏰ Thời gian phê duyệt: {approved_at}")
                _safe_print(f"{'='*80}\n")
                try:
                    sys.stdout.flush()
                except Exception:
                    pass
            except Exception as e:
                try:
                    print(f"❌ [LEAVE_SHEET_UPDATE] Lỗi khi log đầu worker: {e}", flush=True, file=sys.stderr)
                except Exception:
                    pass
            
            try:
                _safe_print(f"📋 [LEAVE_SHEET_UPDATE] Bắt đầu xử lý SONG SONG {len(daily_leaves)} ngày nghỉ")
                sys.stdout.flush()
            except Exception:
                pass
            
            # Tạo thread riêng cho mỗi ngày để chạy song song (parallel)
            threads = []
            for idx, day_leave in enumerate(daily_leaves, start=1):
                def _process_day(day_idx, day_data):
                    """Xử lý một ngày trong thread riêng"""
                    import sys
                    try:
                        _safe_print(f"\n{'='*80}")
                        _safe_print(f"📅 [LEAVE_SHEET_UPDATE] Xử lý ngày {day_idx}/{len(daily_leaves)}")
                        try:
                            sys.stdout.flush()
                        except Exception:
                            pass
                    except Exception:
                        pass
                    
                    leave_date = day_data['date']
                    leave_type = day_data.get('leave_type', {}) or {}

                    # Lấy số ngày cho từng ngày (có thể lẻ 0.5, 1.5, ...)
                    raw_days = day_data.get('fractional_days', leave_type.get('days', 1.0))
                    try:
                        day_value = float(raw_days)
                    except (TypeError, ValueError):
                        day_value = 0.0

                    # Format số ngày: 1, 1.5, 2.5...
                    if day_value <= 0:
                        day_text = "0"
                    elif abs(day_value - round(day_value)) < 1e-9:
                        day_text = str(int(round(day_value)))
                    else:
                        day_text = f"{day_value:.1f}".rstrip('0').rstrip('.')

                    # Nhãn loại nghỉ cho ngày này
                    type_name = str(leave_type.get('name') or '').strip() or "Nghỉ"
                    special_type = leave_type.get('special_type')
                    if special_type:
                        type_name += f" ({special_type})"

                    summary_text = f"{type_name}: {day_text} ngày"
                    full_leave_day = abs(day_value - 1.0) < 1e-9
                    
                    # Lấy use_lunch_break từ notes (lưu dưới dạng JSON)
                    use_lunch_break = None
                    if leave_request.notes:
                        try:
                            import json
                            notes_data = json.loads(leave_request.notes)
                            if isinstance(notes_data, dict) and 'use_lunch_break' in notes_data:
                                use_lunch_break = bool(notes_data['use_lunch_break'])
                        except Exception:
                            pass

                    attendance_data = {
                        'date': leave_date.isoformat(),
                        'user_name': leave_request.employee_name,
                        'approved_by': approved_by,
                        'approved_at': approved_at,
                        'leave_summary': summary_text,
                        'full_leave_day': full_leave_day,
                        'use_lunch_break': use_lunch_break,
                        'leave_fraction_days': day_value
                    }
                    
                    try:
                        print(f"🚀 [LEAVE_SHEET_UPDATE] Đang gọi update_google_sheet_background_safe cho ngày {day_idx}/{len(daily_leaves)}", flush=True, file=sys.stderr)
                    except Exception:
                        pass
                    
                    # Gọi hàm cập nhật Google Sheet
                    try:
                        update_google_sheet_background_safe(
                            attendance_id=f"{attendance_prefix}-{day_idx}",
                            employee_team=employee_team,
                            employee_id=employee_id,
                            attendance_data=attendance_data
                        )
                        try:
                            print(f"✅ [LEAVE_SHEET_UPDATE] Đã HOÀN THÀNH ngày {day_idx}/{len(daily_leaves)}", flush=True, file=sys.stderr)
                        except Exception:
                            pass
                    except Exception as func_err:
                        try:
                            print(f"❌ [LEAVE_SHEET_UPDATE] Lỗi ngày {day_idx}: {func_err}", flush=True, file=sys.stderr)
                            import traceback
                            print(f"   Traceback: {traceback.format_exc()}", flush=True, file=sys.stderr)
                        except Exception:
                            pass
                
                # Tạo thread cho ngày này và chạy ngay lập tức
                thread = threading.Thread(
                    target=_process_day,
                    args=(idx, day_leave),
                    daemon=True
                )
                thread.start()
                threads.append(thread)
            
            # Không chờ các thread - để chúng chạy song song trong background
            # Lưu threads vào list để tránh garbage collection
            try:
                print(f"✅ [LEAVE_SHEET_UPDATE] Đã khởi động {len(threads)} thread song song để cập nhật {len(daily_leaves)} ngày", flush=True, file=sys.stderr)
            except Exception:
                pass
        
        try:
            print(f"🚀 [SCHEDULE_LEAVE_SHEET] Đang khởi động background thread để cập nhật Google Sheet...", flush=True, file=sys.stderr)
            _safe_print(f"🚀 [SCHEDULE_LEAVE_SHEET] Đang khởi động background thread để cập nhật Google Sheet...")
            sys.stdout.flush()
        except Exception:
            pass
        
        threading.Thread(target=_worker, daemon=True).start()
        
        try:
            print(f"✅ [SCHEDULE_LEAVE_SHEET] Đã khởi động background thread để cập nhật Google Sheet", flush=True, file=sys.stderr)
            _safe_print(f"✅ [SCHEDULE_LEAVE_SHEET] Đã khởi động background thread để cập nhật Google Sheet")
            sys.stdout.flush()
        except Exception:
            pass
    except Exception as sheet_error:
        import sys
        from datetime import datetime as dt
        timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        try:
            _safe_print(f"\n{'='*80}")
            _safe_print(f"❌ [SCHEDULE_LEAVE_SHEET] {timestamp} - Lỗi khi chuẩn bị cập nhật Google Sheet")
            _safe_print(f"   Leave Request ID: {leave_request.id if leave_request else 'None'}")
            _safe_print(f"   Error: {str(sheet_error)}")
            _safe_print(f"   Type: {type(sheet_error).__name__}")
            import traceback
            _safe_print(f"   Traceback:")
            _safe_print(traceback.format_exc())
            _safe_print(f"{'='*80}\n")
            sys.stdout.flush()
        except Exception:
            pass


def trigger_schedule_leave_sheet_updates_async(leave_request_id, approver_id=None):
    """Chạy schedule_leave_sheet_updates trong background để tránh block request."""
    def _runner():
        try:
            _safe_print(f"🧵 [LEAVE_SHEET_ASYNC] Thread bắt đầu cho đơn #{leave_request_id}")
            with app.app_context():
                lr = db.session.get(LeaveRequest, leave_request_id)
                approver = db.session.get(User, approver_id) if approver_id else None
                if not lr:
                    _safe_print(f"⚠️ [LEAVE_SHEET_ASYNC] Không tìm thấy đơn #{leave_request_id}")
                    return
                schedule_leave_sheet_updates(lr, approver)
        except Exception as async_err:
            try:
                import traceback
                _safe_print(f"❌ [LEAVE_SHEET_ASYNC] Lỗi khi chạy background cho đơn #{leave_request_id}: {async_err}")
                _safe_print(traceback.format_exc())
            except Exception:
                pass
        finally:
            try:
                _safe_print(f"🧵 [LEAVE_SHEET_ASYNC] Thread kết thúc cho đơn #{leave_request_id}")
            except Exception:
                pass

    thread = threading.Thread(target=_runner, name=f"leave-sheet-{leave_request_id}", daemon=True)
    thread.start()


@app.route('/leave-request/<int:request_id>/approve', methods=['POST'])
def approve_leave_request(request_id):
    """Phê duyệt hoặc từ chối đơn xin nghỉ phép - Logic đa cấp đồng bộ với chấm công"""
    import sys
    from datetime import datetime as dt
    
    # Ghi log vào cả stdout và stderr để đảm bảo hiển thị
    timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    try:
        print(f"\n{'='*80}", flush=True, file=sys.stderr)
        print(f"🚀 [LEAVE_APPROVE_START] {timestamp} - Bắt đầu xử lý phê duyệt đơn nghỉ phép", flush=True, file=sys.stderr)
        print(f"   Request ID: {request_id}", flush=True, file=sys.stderr)
        print(f"{'='*80}\n", flush=True, file=sys.stderr)
        sys.stderr.flush()
    except Exception as e:
        pass
    
    # Dùng print trực tiếp để đảm bảo log được hiển thị
    try:
        print(f"\n{'='*80}", flush=True)
        print(f"🚀 [LEAVE_APPROVE_START] {timestamp} - Bắt đầu xử lý phê duyệt đơn nghỉ phép", flush=True)
        print(f"   Request ID: {request_id}", flush=True)
        print(f"{'='*80}\n", flush=True)
    except Exception as e:
        pass
    
    try:
        _safe_print(f"\n{'='*80}")
        _safe_print(f"🚀 [LEAVE_APPROVE_START] {timestamp} - Bắt đầu xử lý phê duyệt đơn nghỉ phép")
        _safe_print(f"   Request ID: {request_id}")
        _safe_print(f"{'='*80}\n")
        sys.stdout.flush()
    except Exception:
        pass
    
    if 'user_id' not in session:
        return jsonify({'error': 'Vui lòng đăng nhập để sử dụng chức năng này'}), 401
    
    user = db.session.get(User, session['user_id'])
    if not user:
        session.clear()
        return jsonify({'error': 'Phiên đăng nhập không hợp lệ!'}), 401
    
    user_roles = user.get_roles_list()
    current_role = session.get('current_role', user_roles[0] if user_roles else 'EMPLOYEE')
    
    # Log chi tiết về role để debug
    try:
        print(f"🔍 [LEAVE_APPROVE_DEBUG] User: {user.name}", flush=True, file=sys.stderr)
        print(f"   User Roles (tất cả): {user_roles}", flush=True, file=sys.stderr)
        print(f"   Session current_role: {session.get('current_role', 'None')}", flush=True, file=sys.stderr)
        print(f"   Final current_role: {current_role}", flush=True, file=sys.stderr)
        print(f"   Is ADMIN in roles: {'ADMIN' in user_roles}", flush=True, file=sys.stderr)
        print(f"   Is current_role ADMIN: {current_role == 'ADMIN'}", flush=True, file=sys.stderr)
        sys.stderr.flush()
    except Exception:
        pass
    
    try:
        _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] User: {user.name}, Current Role: {current_role}, User Roles: {user_roles}")
    except Exception:
        pass
    
    if current_role not in ['ADMIN', 'MANAGER', 'TEAM_LEADER']:
        try:
            _safe_print(f"❌ [LEAVE_APPROVE] User không có quyền phê duyệt (role: {current_role})")
        except Exception:
            pass
        abort(403)
    
    # Kiểm tra quyền phê duyệt
    has_permission, error_message = check_leave_approval_permission(user.id, request_id, current_role)
    if not has_permission:
        try:
            _safe_print(f"❌ [LEAVE_APPROVE] Không có quyền phê duyệt: {error_message}")
        except Exception:
            pass
        return jsonify({'error': error_message}), 403
    
    leave_request = LeaveRequest.query.get(request_id)
    if not leave_request:
        try:
            _safe_print(f"❌ [LEAVE_APPROVE] Leave request {request_id} not found")
        except Exception:
            pass
        return jsonify({'error': 'Không tìm thấy đơn nghỉ phép'}), 404
    
    action = request.form.get('action')
    reason = request.form.get('rejection_reason', '')
    csrf_token = request.form.get('csrf_token')
    
    # Dùng print trực tiếp để đảm bảo log được hiển thị
    print(f"🔍 [LEAVE_APPROVE_DEBUG] Action: {action}, Reason: {reason}", flush=True)
    print(f"🔍 [LEAVE_APPROVE_DEBUG] Current Role: {current_role}", flush=True)
    print(f"🔍 [LEAVE_APPROVE_DEBUG] Form data: {dict(request.form)}", flush=True)
    
    try:
        _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] Action: {action}, Reason: {reason}")
        _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] CSRF token received: {csrf_token}")
        _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] Form data: {dict(request.form)}")
    except Exception:
        pass
    
    if not action:
        try:
            _safe_print(f"❌ [LEAVE_APPROVE] No action provided for request {request_id}")
        except Exception:
            pass
        return jsonify({'error': 'Không có hành động được chỉ định'}), 400
    
    try:
        try:
            _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] Approving leave request {request_id}, action: {action}, reason: {reason}")
            _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] User {user.name} ({current_role}) approving request {request_id}")
            _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] Leave request status: {leave_request.status}")
            _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] Leave request user: {leave_request.user.name if leave_request.user else 'None'}")
            _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] Leave request user department: {leave_request.user.department if leave_request.user else 'None'}")
            _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] Current user department: {user.department}")
            _safe_print(f"🔍 [LEAVE_APPROVE_DEBUG] Has permission: {has_permission}")
        except Exception:
            pass
        
        if action == 'approve':
            # Logic phê duyệt đa cấp đồng bộ với chấm công
            if current_role == 'TEAM_LEADER':
                if leave_request.status != 'pending':
                    return jsonify({'error': 'Đơn nghỉ phép không ở trạng thái chờ duyệt'}), 400
                
                # Chuyển lên Manager - tìm user có role MANAGER (quản lý tất cả phòng ban)
                manager = User.query.filter(
                    User.roles.like('%MANAGER%'),
                    User.is_deleted == False
                ).first()
                
                if manager:
                    # Có MANAGER - chuyển lên MANAGER
                    leave_request.status = 'pending_manager'
                    leave_request.step = 'manager'
                    leave_request.current_approver_id = manager.id
                    leave_request.team_leader_signer_id = user.id
                    leave_request.team_leader_approved_at = datetime.now()
                    message = 'Đã chuyển lên Quản lý phê duyệt'
                else:
                    # Không có MANAGER - chuyển lên ADMIN
                    admin = User.query.filter(
                        User.roles.like('%ADMIN%'),
                        User.is_deleted == False
                    ).first()
                    if admin:
                        leave_request.status = 'pending_admin'
                        leave_request.step = 'admin'
                        leave_request.current_approver_id = admin.id
                        leave_request.team_leader_signer_id = user.id
                        leave_request.team_leader_approved_at = datetime.now()
                        message = 'Đã chuyển lên Admin phê duyệt (không có Manager)'
                    else:
                        # Không có cả MANAGER và ADMIN - báo lỗi, không cho phê duyệt trực tiếp
                        return jsonify({'error': 'Không tìm thấy Quản lý hoặc Quản trị viên để phê duyệt. Vui lòng liên hệ quản trị hệ thống.'}), 400
                
            elif current_role == 'MANAGER':
                if leave_request.status != 'pending_manager':
                    return jsonify({'error': 'Đơn nghỉ phép chưa được Trưởng nhóm phê duyệt'}), 400
                
                # Chuyển lên Admin - tìm user có role ADMIN
                admin = User.query.filter(
                    User.roles.like('%ADMIN%'),
                    User.is_deleted == False
                ).first()
                if not admin:
                    return jsonify({'error': 'Không tìm thấy quản trị viên'}), 400
                
                leave_request.status = 'pending_admin'
                leave_request.step = 'admin'
                leave_request.current_approver_id = admin.id
                leave_request.manager_signer_id = user.id
                leave_request.manager_approved_at = datetime.now()
                message = 'Đã phê duyệt thành công'
                
            elif current_role == 'ADMIN':
                if leave_request.status != 'pending_admin':
                    return jsonify({'error': 'Đơn nghỉ phép chưa được Quản lý phê duyệt'}), 400
                
                # Check Google API token trước khi ADMIN approve
                token_status = check_google_token_status()
                if not token_status.get('can_approve', False):
                    # Publish notification to all admins
                    publish_token_status('expired', token_status.get('message', 'Token hết hạn'), needs_reauth=True)
                    return jsonify({
                        'error': f"⚠️ Token Google API hết hạn. {token_status.get('message', 'Vui lòng refresh token trước khi phê duyệt.')}",
                        'error_code': 'token_expired',
                        'needs_reauth': True
                    }), 503
                
                # Phê duyệt cuối cùng
                leave_request.status = 'approved'
                leave_request.step = 'done'
                leave_request.current_approver_id = None
                leave_request.admin_signer_id = user.id
                leave_request.admin_approved_at = datetime.now()
                message = 'Đơn xin nghỉ phép đã được phê duyệt hoàn tất!'
                
        elif action == 'reject':
            # Từ chối - chuyển về nhân viên chỉnh sửa
            leave_request.status = 'rejected'
            leave_request.step = 'employee_edit'
            leave_request.current_approver_id = leave_request.user_id
            leave_request.reject_reason = reason
            
            # Thêm vai trò người từ chối vào lý do từ chối
            if reason:
                leave_request.notes = f"Lý do từ chối: {reason} ( {current_role} )"
            else:
                leave_request.notes = f"Lý do từ chối: ( {current_role} )"
            message = 'Đơn xin nghỉ phép đã bị từ chối!'
        else:
            return jsonify({'error': 'Hành động không hợp lệ!'}), 400
        
        db.session.commit()
        
        # Debug log để kiểm tra điều kiện - Dùng print trực tiếp vào cả stdout và stderr
        import sys
        from datetime import datetime as dt
        timestamp = dt.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        try:
            print(f"\n{'='*80}", flush=True, file=sys.stderr)
            print(f"🔍 [LEAVE_APPROVAL_DEBUG] {timestamp} - Sau khi commit database", flush=True, file=sys.stderr)
            print(f"   Action: {action}", flush=True, file=sys.stderr)
            print(f"   Current Role: {current_role}", flush=True, file=sys.stderr)
            print(f"   Leave Request ID: {leave_request.id}", flush=True, file=sys.stderr)
            print(f"   Leave Request Status: {leave_request.status}", flush=True, file=sys.stderr)
            print(f"   Leave Request Type: {leave_request.request_type}", flush=True, file=sys.stderr)
            print(f"   Condition check (action=='approve'): {action == 'approve'}", flush=True, file=sys.stderr)
            print(f"   Condition check (current_role=='ADMIN'): {current_role == 'ADMIN'}", flush=True, file=sys.stderr)
            print(f"   Condition check (action=='approve' and current_role=='ADMIN'): {action == 'approve' and current_role == 'ADMIN'}", flush=True, file=sys.stderr)
            print(f"{'='*80}\n", flush=True, file=sys.stderr)
            sys.stderr.flush()
        except Exception:
            pass
        
        try:
            print(f"\n{'='*80}", flush=True)
            print(f"🔍 [LEAVE_APPROVAL_DEBUG] {timestamp} - Sau khi commit database", flush=True)
            print(f"   Action: {action}", flush=True)
            print(f"   Current Role: {current_role}", flush=True)
            print(f"   Leave Request ID: {leave_request.id}", flush=True)
            print(f"   Leave Request Status: {leave_request.status}", flush=True)
            print(f"   Leave Request Type: {leave_request.request_type}", flush=True)
            print(f"   Condition check (action=='approve'): {action == 'approve'}", flush=True)
            print(f"   Condition check (current_role=='ADMIN'): {current_role == 'ADMIN'}", flush=True)
            print(f"   Condition check (action=='approve' and current_role=='ADMIN'): {action == 'approve' and current_role == 'ADMIN'}", flush=True)
            print(f"{'='*80}\n", flush=True)
        except Exception:
            pass
        
        try:
            _safe_print(f"\n{'='*80}")
            _safe_print(f"🔍 [LEAVE_APPROVAL_DEBUG] {timestamp} - Sau khi commit database")
            _safe_print(f"   Action: {action}")
            _safe_print(f"   Current Role: {current_role}")
            _safe_print(f"   Leave Request ID: {leave_request.id}")
            _safe_print(f"   Leave Request Status: {leave_request.status}")
            _safe_print(f"   Leave Request Type: {leave_request.request_type}")
            _safe_print(f"   Condition check (action=='approve'): {action == 'approve'}")
            _safe_print(f"   Condition check (current_role=='ADMIN'): {current_role == 'ADMIN'}")
            _safe_print(f"   Condition check (action=='approve' and current_role=='ADMIN'): {action == 'approve' and current_role == 'ADMIN'}")
            _safe_print(f"{'='*80}\n")
            sys.stdout.flush()
        except Exception:
            pass
        
        # Log trước khi vào điều kiện - Dùng print trực tiếp vào cả stdout và stderr
        # Kiểm tra điều kiện mới: action == 'approve' và leave_request.status == 'approved' VÀ có admin_approved_at hoặc admin_signer_id
        has_admin_approval_check = (leave_request.admin_approved_at is not None or leave_request.admin_signer_id is not None)
        should_update_sheet_check = (action == 'approve' and leave_request.status == 'approved' and has_admin_approval_check)
        
        try:
            result_text = '✅ VÀO KHỐI CẬP NHẬT GOOGLE SHEET' if should_update_sheet_check else '❌ KHÔNG VÀO KHỐI'
            print(f"\n{'='*80}", flush=True, file=sys.stderr)
            print(f"🔍 [LEAVE_APPROVAL] Kiểm tra điều kiện để cập nhật Google Sheet", flush=True, file=sys.stderr)
            print(f"   action == 'approve': {action == 'approve'}", flush=True, file=sys.stderr)
            print(f"   current_role == 'ADMIN': {current_role == 'ADMIN'}", flush=True, file=sys.stderr)
            print(f"   leave_request.status == 'approved': {leave_request.status == 'approved'}", flush=True, file=sys.stderr)
            print(f"   admin_approved_at: {leave_request.admin_approved_at}", flush=True, file=sys.stderr)
            print(f"   admin_signer_id: {leave_request.admin_signer_id}", flush=True, file=sys.stderr)
            print(f"   has_admin_approval: {has_admin_approval_check}", flush=True, file=sys.stderr)
            print(f"   Điều kiện (action=='approve' AND status=='approved' AND has_admin_approval): {should_update_sheet_check}", flush=True, file=sys.stderr)
            print(f"   Kết quả: {result_text}", flush=True, file=sys.stderr)
            print(f"{'='*80}\n", flush=True, file=sys.stderr)
            sys.stderr.flush()
        except Exception:
            pass
        
        try:
            result_text = '✅ VÀO KHỐI CẬP NHẬT GOOGLE SHEET' if should_update_sheet_check else '❌ KHÔNG VÀO KHỐI'
            print(f"\n{'='*80}", flush=True)
            print(f"🔍 [LEAVE_APPROVAL] Kiểm tra điều kiện để cập nhật Google Sheet", flush=True)
            print(f"   action == 'approve': {action == 'approve'}", flush=True)
            print(f"   current_role == 'ADMIN': {current_role == 'ADMIN'}", flush=True)
            print(f"   leave_request.status == 'approved': {leave_request.status == 'approved'}", flush=True)
            print(f"   admin_approved_at: {leave_request.admin_approved_at}", flush=True)
            print(f"   admin_signer_id: {leave_request.admin_signer_id}", flush=True)
            print(f"   has_admin_approval: {has_admin_approval_check}", flush=True)
            print(f"   Điều kiện (action=='approve' AND status=='approved' AND has_admin_approval): {should_update_sheet_check}", flush=True)
            print(f"   Kết quả: {result_text}", flush=True)
            print(f"{'='*80}\n", flush=True)
        except Exception:
            pass
        
        try:
            _safe_print(f"\n{'='*80}")
            _safe_print(f"🔍 [LEAVE_APPROVAL] Kiểm tra điều kiện để cập nhật Google Sheet")
            _safe_print(f"   action == 'approve': {action == 'approve'}")
            _safe_print(f"   current_role == 'ADMIN': {current_role == 'ADMIN'}")
            _safe_print(f"   Kết quả: {'✅ VÀO KHỐI CẬP NHẬT GOOGLE SHEET' if (action == 'approve' and current_role == 'ADMIN') else '❌ KHÔNG VÀO KHỐI'}")
            _safe_print(f"{'='*80}\n")
            sys.stdout.flush()
        except Exception:
            pass
        
        # Kiểm tra điều kiện: action == 'approve' và leave_request.status == 'approved' 
        # VÀ có admin_approved_at hoặc admin_signer_id (đảm bảo chỉ khi ADMIN phê duyệt)
        has_admin_approval = (leave_request.admin_approved_at is not None or leave_request.admin_signer_id is not None)
        should_update_sheet = (action == 'approve' and leave_request.status == 'approved' and has_admin_approval)
        
        # Log điều kiện mới
        try:
            print(f"🔍 [LEAVE_APPROVAL] Điều kiện cập nhật Google Sheet:", flush=True, file=sys.stderr)
            print(f"   action == 'approve': {action == 'approve'}", flush=True, file=sys.stderr)
            print(f"   leave_request.status == 'approved': {leave_request.status == 'approved'}", flush=True, file=sys.stderr)
            print(f"   admin_approved_at: {leave_request.admin_approved_at}", flush=True, file=sys.stderr)
            print(f"   admin_signer_id: {leave_request.admin_signer_id}", flush=True, file=sys.stderr)
            print(f"   has_admin_approval: {has_admin_approval}", flush=True, file=sys.stderr)
            print(f"   should_update_sheet (action=='approve' AND status=='approved' AND has_admin_approval): {should_update_sheet}", flush=True, file=sys.stderr)
            sys.stderr.flush()
        except Exception:
            pass
        
        if should_update_sheet:
            # Dùng print trực tiếp vào cả stdout và stderr để đảm bảo log được hiển thị
            try:
                print(f"\n{'='*80}", flush=True, file=sys.stderr)
                print(f"✅ [LEAVE_APPROVAL] ĐÃ VÀO KHỐI CẬP NHẬT GOOGLE SHEET", flush=True, file=sys.stderr)
                print(f"   Leave Request ID: {leave_request.id}", flush=True, file=sys.stderr)
                print(f"   User: {user.name if user else 'None'}", flush=True, file=sys.stderr)
                print(f"   Current Role: {current_role}", flush=True, file=sys.stderr)
                print(f"{'='*80}\n", flush=True, file=sys.stderr)
                sys.stderr.flush()
            except Exception:
                pass
            
            try:
                print(f"\n{'='*80}", flush=True)
                print(f"✅ [LEAVE_APPROVAL] ĐÃ VÀO KHỐI CẬP NHẬT GOOGLE SHEET", flush=True)
                print(f"   Leave Request ID: {leave_request.id}", flush=True)
                print(f"   User: {user.name if user else 'None'}", flush=True)
                print(f"   Current Role: {current_role}", flush=True)
                print(f"{'='*80}\n", flush=True)
            except Exception:
                pass
            
            try:
                _safe_print(f"\n{'='*80}")
                _safe_print(f"✅ [LEAVE_APPROVAL] ĐÃ VÀO KHỐI CẬP NHẬT GOOGLE SHEET")
                _safe_print(f"   Leave Request ID: {leave_request.id}")
                _safe_print(f"   User: {user.name if user else 'None'}")
                _safe_print(f"   Current Role: {current_role}")
                _safe_print(f"{'='*80}\n")
                sys.stdout.flush()
            except Exception:
                pass
            try:
                _safe_print(f"\n{'='*80}")
                _safe_print(f"✅ [LEAVE_APPROVAL] ĐÃ VÀO KHỐI CẬP NHẬT GOOGLE SHEET")
                _safe_print(f"   Leave Request ID: {leave_request.id}")
                _safe_print(f"   User: {user.name if user else 'None'}")
                _safe_print(f"   Current Role: {current_role}")
                _safe_print(f"{'='*80}\n")
                sys.stdout.flush()
            except Exception:
                pass
            try:
                # Refresh leave_request từ database sau khi commit để đảm bảo có dữ liệu mới nhất
                db.session.refresh(leave_request)
                
                try:
                    _safe_print(f"\n{'='*80}")
                    _safe_print(f"📋 [LEAVE_APPROVAL] Admin {user.name} đã phê duyệt đơn nghỉ phép #{leave_request.id}")
                    _safe_print(f"   👤 Nhân viên: {leave_request.employee_name} (ID: {leave_request.user_id})")
                    _safe_print(f"   📅 Từ ngày: {leave_request.leave_from_day}/{leave_request.leave_from_month}/{leave_request.leave_from_year}")
                    _safe_print(f"   📅 Đến ngày: {leave_request.leave_to_day}/{leave_request.leave_to_month}/{leave_request.leave_to_year}")
                    _safe_print(f"   📊 Phép năm: {leave_request.annual_leave_days} ngày")
                    _safe_print(f"   📊 Nghỉ không lương: {leave_request.unpaid_leave_days} ngày")
                    _safe_print(f"   📊 Nghỉ đặc biệt: {leave_request.special_leave_days} ngày")
                    _safe_print(f"   🏢 Phòng ban: {leave_request.team}")
                    _safe_print(f"   📝 Lý do: {leave_request.leave_reason}")
                    _safe_print(f"   📝 Notes: {leave_request.notes}")
                    _safe_print(f"{'='*80}\n")
                    sys.stdout.flush()
                except Exception:
                    pass
                
                # Dùng print trực tiếp vào cả stdout và stderr để đảm bảo log được hiển thị
                try:
                    print(f"🚀 [LEAVE_APPROVAL] Đang khởi động background cập nhật Google Sheet cho đơn #{leave_request.id}", flush=True, file=sys.stderr)
                    sys.stderr.flush()
                except Exception:
                    pass
                
                try:
                    print(f"🚀 [LEAVE_APPROVAL] Đang khởi động background cập nhật Google Sheet cho đơn #{leave_request.id}", flush=True)
                except Exception:
                    pass
                
                try:
                    _safe_print(f"🚀 [LEAVE_APPROVAL] Đang khởi động background cập nhật Google Sheet cho đơn #{leave_request.id}")
                    sys.stdout.flush()
                except Exception:
                    pass
                
                trigger_schedule_leave_sheet_updates_async(leave_request.id, user.id if user else None)
                
                try:
                    print(f"✅ [LEAVE_APPROVAL] Đã khởi động background cập nhật Google Sheet cho đơn #{leave_request.id}", flush=True, file=sys.stderr)
                    sys.stderr.flush()
                except Exception:
                    pass
                
                try:
                    print(f"✅ [LEAVE_APPROVAL] Đã khởi động background cập nhật Google Sheet cho đơn #{leave_request.id}", flush=True)
                except Exception:
                    pass
                
                try:
                    _safe_print(f"✅ [LEAVE_APPROVAL] Đã khởi động background cập nhật Google Sheet cho đơn #{leave_request.id}")
                    sys.stdout.flush()
                except Exception:
                    pass
            except Exception as sheet_err:
                try:
                    _safe_print(f"\n{'='*80}")
                    _safe_print(f"❌ [LEAVE_APPROVAL] Không thể đẩy dữ liệu lên Google Sheet cho đơn #{leave_request.id}")
                    _safe_print(f"   Error: {str(sheet_err)}")
                    _safe_print(f"   Type: {type(sheet_err).__name__}")
                    import traceback
                    _safe_print(f"   Traceback:")
                    _safe_print(traceback.format_exc())
                    _safe_print(f"{'='*80}\n")
                except Exception:
                    pass
        
        return jsonify({'message': message})
        
    except Exception as e:
        db.session.rollback()
        # Ghi log vào cả stdout và stderr để đảm bảo hiển thị
        try:
            import traceback
            error_msg = f"[ERROR] Error in approve_leave_request: {e}"
            print(error_msg, flush=True, file=sys.stderr)
            print(f"[ERROR] Request ID: {request_id}", flush=True, file=sys.stderr)
            print(f"[ERROR] Action: {action if 'action' in locals() else 'Unknown'}", flush=True, file=sys.stderr)
            print(f"[ERROR] User: {user.name if 'user' in locals() and user else 'None'}", flush=True, file=sys.stderr)
            print(f"[ERROR] Current role: {current_role if 'current_role' in locals() else 'Unknown'}", flush=True, file=sys.stderr)
            print("Traceback:", flush=True, file=sys.stderr)
            print(traceback.format_exc(), flush=True, file=sys.stderr)
            sys.stderr.flush()
        except Exception:
            pass
        
        try:
            _safe_print(f"[ERROR] Error in approve_leave_request: {e}")
            _safe_print(f"[ERROR] Request ID: {request_id}")
            _safe_print(f"[ERROR] Action: {action if 'action' in locals() else 'Unknown'}")
            _safe_print(f"[ERROR] User: {user.name if 'user' in locals() and user else 'None'}")
            _safe_print(f"[ERROR] Current role: {current_role if 'current_role' in locals() else 'Unknown'}")
            import traceback
            _safe_print(traceback.format_exc())
        except Exception:
            pass
        return jsonify({'error': f'Lỗi khi xử lý đơn xin nghỉ phép: {str(e)}'}), 500

@app.route('/leave-request/<int:request_id>/delete')
def delete_leave_request(request_id):
    """Xóa đơn xin nghỉ phép"""
    if 'user_id' not in session:
        flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
        return redirect(url_for('login'))
    
    user = db.session.get(User, session['user_id'])
    if not user:
        session.clear()
        flash('Phiên đăng nhập không hợp lệ!', 'error')
        return redirect(url_for('login'))
    
    leave_request = LeaveRequest.query.get_or_404(request_id)
    
    # Chỉ người tạo đơn mới có thể xóa
    if user.id != leave_request.user_id:
        abort(403)
    
    # Chỉ có thể xóa khi đơn đang ở trạng thái chờ phê duyệt hoặc bị từ chối
    if leave_request.status not in ['pending', 'rejected']:
        flash('Chỉ có thể xóa đơn khi đang chờ phê duyệt hoặc bị từ chối', 'error')
        return redirect(url_for('view_leave_request', request_id=request_id))
    
    try:
        try:
            # Gửi email thông báo tới HR về việc người dùng hủy/xóa đơn
            send_leave_request_email_async(leave_request, user, action='delete')
            from utils.email_utils import process_db_updates
            process_db_updates()
            upsert_email_status(leave_request.id, 'sending', 'Đang gửi email thông báo hủy/xóa đơn...')
            session['email_status'] = {
                'request_id': leave_request.id,
                'status': 'sending',
                'message': 'Đang gửi email thông báo hủy/xóa đơn...'
            }
        except Exception as mail_err:
            session['email_status'] = {
                'request_id': leave_request.id,
                'status': 'error',
                'message': f'Lỗi khi gửi email thông báo hủy/xóa đơn: {str(mail_err)}'
            }
        db.session.delete(leave_request)
        db.session.commit()
        # flash('Đơn xin nghỉ phép đã được xóa thành công!', 'success')
        return redirect(url_for('leave_requests_list'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi khi xóa đơn xin nghỉ phép: {str(e)}', 'error')
        return redirect(url_for('view_leave_request', request_id=request_id))

@app.route('/leave-history')
def leave_history():
    """Hiển thị lịch sử nghỉ phép: tất cả đơn nghỉ phép của chính người dùng"""
    try:
        if 'user_id' not in session:
            flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
            return redirect(url_for('login'))

        user = db.session.get(User, session['user_id'])
        if not user:
            session.clear()
            flash('Phiên đăng nhập không hợp lệ!', 'error')
            return redirect(url_for('login'))

        current_role = session.get('current_role', user.roles.split(',')[0])

        page = request.args.get('page', 1, type=int)
        per_page = 10

        try:
            # Lịch sử nghỉ phép
            # - ADMIN: xem toàn bộ lịch sử đã được phê duyệt của tất cả nhân viên
            # - Người dùng khác: xem lịch sử đã được phê duyệt của chính mình
            if current_role == 'ADMIN':
                query = LeaveRequest.query.filter(LeaveRequest.status == 'approved')
            else:
                query = LeaveRequest.query.filter(LeaveRequest.user_id == user.id, LeaveRequest.status == 'approved')

            # Bộ lọc (GET)
            keyword = (request.args.get('q') or '').strip()
            department = (request.args.get('department') or '').strip()
            from_date_str = (request.args.get('from_date') or '').strip()
            to_date_str = (request.args.get('to_date') or '').strip()
            status_filter = (request.args.get('status') or '').strip()
            request_type_filter = (request.args.get('request_type') or '').strip()

            if keyword:
                # Tìm theo tên hoặc mã nhân viên
                query = query.filter(
                    db.or_(
                        LeaveRequest.employee_name.ilike(f"%{keyword}%"),
                        LeaveRequest.employee_code.ilike(f"%{keyword}%")
                    )
                )

            if department and current_role == 'ADMIN':
                # Join sang User để lọc theo phòng ban
                query = query.join(User, User.id == LeaveRequest.user_id).filter(User.department == department)

            if status_filter:
                query = query.filter(LeaveRequest.status == status_filter)

            if request_type_filter:
                # Hỗ trợ: leave | late_early | 30min_break
                query = query.filter(LeaveRequest.request_type == request_type_filter)

            # Lọc theo ngày xin nghỉ thực tế
            try:
                if from_date_str:
                    from_dt = datetime.strptime(from_date_str, '%Y-%m-%d')
                    # Lọc các đơn có ngày kết thúc nghỉ >= ngày bắt đầu lọc
                    query = query.filter(
                        db.or_(
                            # Ngày kết thúc nghỉ >= ngày lọc
                            db.and_(
                                LeaveRequest.leave_to_year > from_dt.year
                            ),
                            db.and_(
                                LeaveRequest.leave_to_year == from_dt.year,
                                LeaveRequest.leave_to_month > from_dt.month
                            ),
                            db.and_(
                                LeaveRequest.leave_to_year == from_dt.year,
                                LeaveRequest.leave_to_month == from_dt.month,
                                LeaveRequest.leave_to_day >= from_dt.day
                            )
                        )
                    )
                if to_date_str:
                    to_dt = datetime.strptime(to_date_str, '%Y-%m-%d')
                    # Lọc các đơn có ngày bắt đầu nghỉ <= ngày kết thúc lọc
                    query = query.filter(
                        db.or_(
                            # Ngày bắt đầu nghỉ <= ngày lọc
                            db.and_(
                                LeaveRequest.leave_from_year < to_dt.year
                            ),
                            db.and_(
                                LeaveRequest.leave_from_year == to_dt.year,
                                LeaveRequest.leave_from_month < to_dt.month
                            ),
                            db.and_(
                                LeaveRequest.leave_from_year == to_dt.year,
                                LeaveRequest.leave_from_month == to_dt.month,
                                LeaveRequest.leave_from_day <= to_dt.day
                            )
                        )
                    )
            except Exception:
                pass

            query = query.order_by(LeaveRequest.created_at.desc())

            pagination = query.paginate(page=page, per_page=per_page, error_out=False)
            
            print(f"[DEBUG] leave_history: Found {len(pagination.items)} leave requests for user {user.id}")
            
        except Exception as query_error:
            print(f"[ERROR] Database query error in leave_history: {query_error}")
            import traceback
            traceback.print_exc()
            flash(f'Lỗi truy vấn dữ liệu: {str(query_error)}', 'error')
            return render_template('leave_history.html',
                                   leave_requests=[],
                                   pagination=None,
                                   user=user,
                                   current_role=current_role)

        # Danh sách phòng ban cho filter
        try:
            # Thử lấy từ bảng Department trước
            dept_objects = Department.query.filter(Department.is_active == True).order_by(Department.name.asc()).all()
            if dept_objects:
                departments = [d.name for d in dept_objects]
                print(f"[DEBUG] Got {len(departments)} departments from Department table: {departments}")
            else:
                # Fallback: distinct từ User nếu Department trống
                departments = sorted({u.department for u in User.query.filter(User.department.isnot(None), User.department != '').all()})
                print(f"[DEBUG] Got {len(departments)} departments from User table: {departments}")
        except Exception as e:
            print(f"[DEBUG] Error getting departments: {e}")
            # Fallback: distinct từ User
            departments = sorted({u.department for u in User.query.filter(User.department.isnot(None), User.department != '').all()})
            print(f"[DEBUG] Fallback: Got {len(departments)} departments from User table: {departments}")

        return render_template('leave_history.html',
                               leave_requests=pagination.items,
                               pagination=pagination,
                               user=user,
                               current_role=current_role,
                               departments=departments,
                               current_filters={
                                   'q': keyword,
                                   'department': department,
                                   'from_date': from_date_str,
                                   'to_date': to_date_str,
                                   'status': status_filter,
                                   'request_type': request_type_filter
                               })
    except Exception as e:
        print(f"[ERROR] Error in leave_history: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Có lỗi xảy ra: {str(e)}', 'error')
        return redirect(url_for('dashboard'))
@app.route('/leave-request/back-to-dashboard')
def back_to_dashboard():
    """Quay về dashboard với vai trò hiện tại"""
    if 'user_id' not in session:
        flash('Vui lòng đăng nhập để sử dụng chức năng này', 'error')
        return redirect(url_for('login'))
    
    # Lấy vai trò hiện tại từ session
    current_role = session.get('current_role', 'EMPLOYEE')
    # Redirect to dashboard with current role
    
    # Chuyển hướng về dashboard với vai trò hiện tại
    return redirect(url_for('dashboard', role=current_role))

@app.route('/api/pending-leave-count')
def api_pending_leave_count():
    """API để lấy số lượng đơn nghỉ phép cần phê duyệt"""
    try:
        # Kiểm tra user đã đăng nhập
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = db.session.get(User, session['user_id'])
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Chỉ trưởng nhóm, quản lý và quản trị viên mới có quyền xem
        user_roles = user.get_roles_list()
        if not any(role in ['TEAM_LEADER', 'MANAGER', 'ADMIN'] for role in user_roles):
            return jsonify({'error': 'Forbidden'}), 403
        
        # Đếm số đơn nghỉ phép đang chờ phê duyệt theo vai trò
        current_role = session.get('current_role', user_roles[0] if user_roles else 'EMPLOYEE')
        
        if current_role == 'TEAM_LEADER':
            # TEAM_LEADER chỉ đếm đơn pending của nhân viên cùng phòng ban
            pending_count = LeaveRequest.query.filter(
                LeaveRequest.status == 'pending',
                LeaveRequest.user.has(User.department == user.department)
            ).count()
        elif current_role == 'MANAGER':
            pending_count = LeaveRequest.query.filter_by(status='pending_manager').count()
        elif current_role == 'ADMIN':
            pending_count = LeaveRequest.query.filter_by(status='pending_admin').count()
        else:
            pending_count = 0
        
        return jsonify({'count': pending_count})
        
    except Exception as e:
        print(f"Error in api_pending_leave_count: {e}")
        return jsonify({'error': 'Internal server error'}), 500



@csrf.exempt  # Tạm thời bỏ qua CSRF để test
@app.route('/test-excel')
def test_excel():
    """Test endpoint để kiểm tra Excel export"""
    print(f"[DEBUG] Test Excel endpoint called")
    try:
        # Tạo file Excel đơn giản
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws.cell(row=1, column=1, value="Test")
        ws.cell(row=2, column=1, value="Hello World")
        
        # Lưu file vào memory
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Tạo response
        filename = f"test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"[DEBUG] Test response with filename: {filename}")
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"[ERROR] Error in test Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi test Excel: {str(e)}'}), 500

@csrf.exempt  # Tạm thời bỏ qua CSRF để test
@app.route('/export-leave-history-excel')
def export_leave_history_excel():
    """Xuất lịch sử nghỉ phép ra file Excel cho ADMIN - Tách từng ngày riêng biệt"""
    print(f"[DEBUG] Excel export endpoint called")
    try:
        # Import utility functions
        from utils.excel_leave_processor import process_leave_requests_for_excel
        
        # Lấy dữ liệu theo bộ lọc giống trang danh sách
        print(f"[DEBUG] Getting leave requests with filters for Excel")
        if 'user_id' not in session:
            return jsonify({'error': 'Chưa đăng nhập'}), 401

        user = db.session.get(User, session['user_id'])
        current_role = session.get('current_role', user.roles.split(',')[0]) if user else 'EMPLOYEE'

        if current_role == 'ADMIN':
            query = LeaveRequest.query.filter(LeaveRequest.status == 'approved')
        else:
            query = LeaveRequest.query.filter(LeaveRequest.user_id == user.id, LeaveRequest.status == 'approved')

        # Nhận tham số filter từ query string
        keyword = (request.args.get('q') or '').strip()
        department = (request.args.get('department') or '').strip()
        from_date_str = (request.args.get('from_date') or '').strip()
        to_date_str = (request.args.get('to_date') or '').strip()
        status_filter = (request.args.get('status') or '').strip()

        if keyword:
            query = query.filter(
                db.or_(
                    LeaveRequest.employee_name.ilike(f"%{keyword}%"),
                    LeaveRequest.employee_code.ilike(f"%{keyword}%")
                )
            )

        if department and current_role == 'ADMIN':
            query = query.join(User, User.id == LeaveRequest.user_id).filter(User.department == department)

        if status_filter:
            query = query.filter(LeaveRequest.status == status_filter)

        # Lọc theo khoảng ngày tương tự trang danh sách
        try:
            if from_date_str:
                from_dt = datetime.strptime(from_date_str, '%Y-%m-%d')
                query = query.filter(
                    db.or_(
                        db.and_(LeaveRequest.leave_to_year > from_dt.year),
                        db.and_(LeaveRequest.leave_to_year == from_dt.year, LeaveRequest.leave_to_month > from_dt.month),
                        db.and_(
                            LeaveRequest.leave_to_year == from_dt.year,
                            LeaveRequest.leave_to_month == from_dt.month,
                            LeaveRequest.leave_to_day >= from_dt.day
                        )
                    )
                )
            if to_date_str:
                to_dt = datetime.strptime(to_date_str, '%Y-%m-%d')
                query = query.filter(
                    db.or_(
                        db.and_(LeaveRequest.leave_from_year < to_dt.year),
                        db.and_(LeaveRequest.leave_from_year == to_dt.year, LeaveRequest.leave_from_month < to_dt.month),
                        db.and_(
                            LeaveRequest.leave_from_year == to_dt.year,
                            LeaveRequest.leave_from_month == to_dt.month,
                            LeaveRequest.leave_from_day <= to_dt.day
                        )
                    )
                )
        except Exception:
            pass

        leave_requests = query.order_by(LeaveRequest.created_at.desc()).all()
        print(f"[DEBUG] Found {len(leave_requests)} leave requests after filters")
        
        # Xử lý dữ liệu để tách từng ngày
        print(f"[DEBUG] Processing leave requests to split by days")
        daily_leaves = process_leave_requests_for_excel(leave_requests)
        print(f"[DEBUG] Generated {len(daily_leaves)} daily leave entries")
        
        # Tạo file Excel
        wb = Workbook()
        ws = wb.active
        # Đặt tiêu đề sheet tiếng Việt (<=31 ký tự, không chứa: : \\ / ? * [ ])
        ws.title = "Lịch sử nghỉ phép"
        
        # Định dạng header
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Tạo header mới với thông tin chi tiết hơn
        headers = [
            "Nhân viên", 
            "Ngày nghỉ", 
            "Thời gian nghỉ", 
            "Lý do", 
            "Loại nghỉ", 
            "Số ngày",
            "Ngày tạo"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Thêm dữ liệu đã được tách theo ngày
        print(f"[DEBUG] Adding {len(daily_leaves)} daily leave entries to Excel")
        for row, day_leave in enumerate(daily_leaves, 2):
            try:
                # Nhân viên - xử lý encoding an toàn
                employee_name = str(day_leave['employee_name']).replace('\x00', '').replace('\r', '').replace('\n', ' ')
                employee_code = str(day_leave['employee_code']).replace('\x00', '').replace('\r', '').replace('\n', ' ')
                employee_info = f"{employee_name} ({employee_code})"
                cell_a = ws.cell(row=row, column=1, value=employee_info)
                cell_a.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Ngày nghỉ
                date_str = day_leave['date'].strftime('%d/%m/%Y')
                cell_b = ws.cell(row=row, column=2, value=date_str)
                cell_b.alignment = Alignment(horizontal="center", vertical="center")
                
                # Thời gian nghỉ (giờ bắt đầu - giờ kết thúc)
                start_time_str = day_leave['start_time'].strftime('%H:%M')
                end_time_str = day_leave['end_time'].strftime('%H:%M')
                time_info = f"{start_time_str} - {end_time_str}"
                cell_c = ws.cell(row=row, column=3, value=time_info)
                cell_c.alignment = Alignment(horizontal="center", vertical="center")
                
                # Lý do - xử lý encoding an toàn
                reason_text = str(day_leave['reason']) if day_leave['reason'] else ""
                # Loại bỏ ký tự đặc biệt có thể gây lỗi
                reason_text = reason_text.replace('\x00', '').replace('\r', '').replace('\n', ' ')
                cell_d = ws.cell(row=row, column=4, value=reason_text)
                cell_d.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Loại nghỉ - xử lý encoding an toàn
                leave_type = day_leave['leave_type']
                leave_type_text = str(leave_type['name']) if leave_type['name'] else ""
                if leave_type.get('special_type'):
                    special_type = str(leave_type['special_type']).replace('\x00', '').replace('\r', '').replace('\n', ' ')
                    leave_type_text += f" ({special_type})"
                # Loại bỏ ký tự đặc biệt
                leave_type_text = leave_type_text.replace('\x00', '').replace('\r', '').replace('\n', ' ')
                cell_e = ws.cell(row=row, column=5, value=leave_type_text)
                cell_e.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Số ngày (có thể lẻ)
                days_value = day_leave.get('fractional_days', leave_type.get('days', 1.0))
                cell_f = ws.cell(row=row, column=6, value=days_value)
                cell_f.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ngày tạo
                created_date = _vn_datetime_format(day_leave['created_at'], '%d/%m/%Y %H:%M')
                cell_g = ws.cell(row=row, column=7, value=created_date)
                cell_g.alignment = Alignment(horizontal="center", vertical="center")
                
            except Exception as e:
                print(f"[ERROR] Error adding row {row}: {e}")
                import traceback
                traceback.print_exc()
                # Thêm dữ liệu cơ bản nếu có lỗi
                try:
                    employee_name = str(day_leave.get('employee_name', 'N/A')).replace('\x00', '').replace('\r', '').replace('\n', ' ')
                    employee_code = str(day_leave.get('employee_code', 'N/A')).replace('\x00', '').replace('\r', '').replace('\n', ' ')
                    ws.cell(row=row, column=1, value=f"{employee_name} ({employee_code})")
                except (KeyError, AttributeError, TypeError, ValueError):
                    ws.cell(row=row, column=1, value="Lỗi dữ liệu")
                
                ws.cell(row=row, column=2, value="Lỗi hiển thị ngày")
                ws.cell(row=row, column=3, value="Lỗi hiển thị thời gian")
                ws.cell(row=row, column=4, value="Lỗi hiển thị lý do")
                ws.cell(row=row, column=5, value="Lỗi hiển thị loại nghỉ")
                ws.cell(row=row, column=6, value="Lỗi")
                ws.cell(row=row, column=7, value="Lỗi")
        
        # Điều chỉnh độ rộng cột - tối ưu cho 7 cột
        column_widths = [30, 15, 18, 50, 30, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Điều chỉnh chiều cao hàng
        for row in range(2, len(daily_leaves) + 2):
            ws.row_dimensions[row].height = 30  # Chiều cao phù hợp cho dữ liệu đã tách
        
        # Thêm filter cho header
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(daily_leaves) + 1}"
        
        # Lưu file vào memory
        print(f"[DEBUG] Saving Excel file to memory")
        from io import BytesIO
        output = BytesIO()
        
        try:
            wb.save(output)
            output.seek(0)
            print(f"[DEBUG] Excel file saved successfully, size: {len(output.getvalue())} bytes")
        except Exception as save_error:
            print(f"[ERROR] Error saving Excel file: {save_error}")
            raise save_error
        
        # Tạo response
        # Tên file tiếng Việt + fallback ASCII theo RFC 5987
        vn_filename = f"Lịch_sử_nghỉ_phép_chi_tiết_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"[DEBUG] Creating response with filename: {vn_filename}")
        from urllib.parse import quote
        ascii_fallback = "lich_su_nghi_phep_chi_tiet.xlsx"
        content_disposition = (
            f"attachment; filename=\"{ascii_fallback}\"; "
            f"filename*=UTF-8''{quote(vn_filename)}"
        )
        
        # Tạo response với encoding đúng
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = content_disposition
        response.headers['Content-Length'] = len(output.getvalue())
        # Thêm header để tự động tải xuống Downloads
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        print(f"[ERROR] Error exporting Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi khi xuất file Excel: {str(e)}'}), 500


@csrf.exempt
@app.route('/export-leave-cases-excel')
def export_leave_cases_excel():
    """Tạo file Excel tổng hợp mọi trường hợp nghỉ để kiểm tra hiển thị.
    Không ghi DB, dùng dữ liệu giả lập bao phủ: ngày nguyên, ngày lẻ 0.5/1.5/2.5/3.5,
    kết hợp nhiều loại nghỉ, có/không special_type, khoảng ngày nhiều ngày.
    """
    try:
        from utils.excel_leave_processor import process_leave_requests_for_excel
        from dataclasses import dataclass
        from datetime import datetime, timedelta

        @dataclass
        class DummyLeave:
            employee_name: str
            employee_code: str
            team: str
            leave_reason: str
            leave_from_year: int
            leave_from_month: int
            leave_from_day: int
            leave_from_hour: int
            leave_from_minute: int
            leave_to_year: int
            leave_to_month: int
            leave_to_day: int
            leave_to_hour: int
            leave_to_minute: int
            annual_leave_days: float = 0.0
            unpaid_leave_days: float = 0.0
            special_leave_days: float = 0.0
            special_leave_type: str | None = None
            created_at: datetime = datetime.utcnow()
            substitute_name: str | None = None
            substitute_employee_id: str | None = None
            status: str = 'approved'

            # API tương thích models.LeaveRequest
            def get_leave_from_datetime(self):
                return datetime(self.leave_from_year, self.leave_from_month, self.leave_from_day,
                               self.leave_from_hour, self.leave_from_minute)

            def get_leave_to_datetime(self):
                return datetime(self.leave_to_year, self.leave_to_month, self.leave_to_day,
                               self.leave_to_hour, self.leave_to_minute)

            def get_reason_text(self):
                return self.leave_reason

        # Tạo danh sách test cases bao phủ các tình huống chính
        base_date = datetime.utcnow().replace(hour=7, minute=30, second=0, microsecond=0)
        cases: list[DummyLeave] = [
            # 1) Chỉ phép năm 1, 2 ngày
            DummyLeave('Nguyễn Văn A', 'A001', 'Kế toán', 'Nghỉ phép năm 1 ngày',
                       base_date.year, base_date.month, base_date.day, 7, 30,
                       base_date.year, base_date.month, base_date.day, 16, 30,
                       annual_leave_days=1.0),
            DummyLeave('Nguyễn Văn B', 'A002', 'Kế toán', 'Nghỉ phép năm 2 ngày',
                       base_date.year, base_date.month, base_date.day, 7, 30,
                       (base_date + timedelta(days=1)).year, (base_date + timedelta(days=1)).month, (base_date + timedelta(days=1)).day, 16, 30,
                       annual_leave_days=2.0),
            # 2) Chỉ không lương 1.5 ngày
            DummyLeave('Trần Thị C', 'A003', 'Hành chính', 'Nghỉ không lương 1.5 ngày',
                       base_date.year, base_date.month, base_date.day, 7, 30,
                       base_date.year, base_date.month, base_date.day, 12, 0,
                       unpaid_leave_days=1.5),
            # 3) Đặc biệt 1 ngày (kết hôn)
            DummyLeave('Lê Văn D', 'A004', 'Kho vận', 'Nghỉ đặc biệt (kết hôn)',
                       base_date.year, base_date.month, base_date.day, 7, 30,
                       base_date.year, base_date.month, base_date.day, 16, 30,
                       special_leave_days=1.0, special_leave_type='Kết hôn'),
            # 4) Kết hợp 3 loại mỗi loại 1 ngày (3 ngày liên tiếp)
            DummyLeave('Phạm Thị E', 'A005', 'Nhân sự', 'Kết hợp 3 loại nghỉ (mỗi loại 1 ngày)',
                       base_date.year, base_date.month, base_date.day, 7, 30,
                       (base_date + timedelta(days=2)).year, (base_date + timedelta(days=2)).month, (base_date + timedelta(days=2)).day, 16, 30,
                       annual_leave_days=1.0, unpaid_leave_days=1.0, special_leave_days=1.0, special_leave_type='Tang lễ'),
            # 5) Tổng 2.5 ngày: 1.0 phép năm + 1.5 không lương
            DummyLeave('Đỗ Văn G', 'A006', 'Sản xuất', 'Phép năm 1.0 + Không lương 1.5',
                       base_date.year, base_date.month, base_date.day, 7, 30,
                       (base_date + timedelta(days=2)).year, (base_date + timedelta(days=2)).month, (base_date + timedelta(days=2)).day, 12, 0,
                       annual_leave_days=1.0, unpaid_leave_days=1.5),
            # 6) Tổng 3.5 ngày: 3.0 không lương + 0.5 đặc biệt
            DummyLeave('Ngô Thị H', 'A007', 'Bán hàng', 'Không lương 3.0 + Đặc biệt 0.5',
                       base_date.year, base_date.month, base_date.day, 7, 30,
                       (base_date + timedelta(days=3)).year, (base_date + timedelta(days=3)).month, (base_date + timedelta(days=3)).day, 12, 0,
                       unpaid_leave_days=3.0, special_leave_days=0.5, special_leave_type='Khác'),
            # 7) Tổng 7.5 ngày: 3 annual + 3.5 unpaid + 1 special
            DummyLeave('Bùi Minh I', 'A008', 'Kỹ thuật', 'Tổng 7.5 ngày: 3 PN + 3.5 NKL + 1 ĐB',
                       base_date.year, base_date.month, base_date.day, 7, 30,
                       (base_date + timedelta(days=7)).year, (base_date + timedelta(days=7)).month, (base_date + timedelta(days=7)).day, 12, 0,
                       annual_leave_days=3.0, unpaid_leave_days=3.5, special_leave_days=1.0, special_leave_type='Gia đình'),
        ]

        # Xử lý thành daily rows
        daily_leaves = process_leave_requests_for_excel(cases)

        # Xuất Excel (dùng cùng format với export chính)
        wb = Workbook()
        ws = wb.active
        # Đặt tiêu đề sheet tiếng Việt (<=31 ký tự, không ký tự cấm)
        ws.title = "Lịch sử nghỉ phép"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = ["Nhân viên", "Ngày nghỉ", "Thời gian nghỉ", "Lý do", "Loại nghỉ", "Số ngày", "Ngày tạo"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row, day_leave in enumerate(daily_leaves, 2):
            employee_info = f"{day_leave['employee_name']} ({day_leave['employee_code']})"
            ws.cell(row=row, column=1, value=employee_info)
            ws.cell(row=row, column=2, value=day_leave['date'].strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=f"{day_leave['start_time'].strftime('%H:%M')} - {day_leave['end_time'].strftime('%H:%M')}")
            ws.cell(row=row, column=4, value=str(day_leave['reason'] or ''))
            lt = day_leave['leave_type']
            lt_text = lt['name'] + (f" ({lt['special_type']})" if lt.get('special_type') else '')
            ws.cell(row=row, column=5, value=lt_text)
            ws.cell(row=row, column=6, value=day_leave.get('fractional_days', lt.get('days', 1.0)))
            ws.cell(row=row, column=7, value=_vn_datetime_format(day_leave['created_at'], '%d/%m/%Y %H:%M'))

        column_widths = [30, 15, 18, 50, 30, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        for row in range(2, len(daily_leaves) + 2):
            ws.row_dimensions[row].height = 30

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        vn_filename = f"Bộ_test_các_trường_hợp_nghỉ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        from urllib.parse import quote
        ascii_fallback = "bo_test_cac_truong_hop_nghi.xlsx"
        content_disposition = (
            f"attachment; filename=\"{ascii_fallback}\"; "
            f"filename*=UTF-8''{quote(vn_filename)}"
        )
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = content_disposition
        response.headers['Content-Length'] = len(output.getvalue())
        # Thêm header để tự động tải xuống Downloads
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        print(f"[ERROR] Error exporting test cases Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi khi xuất test cases Excel: {str(e)}'}), 500
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        convert_overtime_to_hhmm()

    # --- Bước 1: Kiểm tra license NGAY KHI KHỞI ĐỘNG ---
    print("[LICENSE] Đang kiểm tra license trước khi khởi động server...", flush=True)
    is_valid, expired, status, msg = _check_license_once()

    if (not is_valid) or expired or (status not in ("active", "đang hoạt động", "")):
        # License không hợp lệ -> set flag để chặn truy cập nhưng vẫn khởi động server
        _license_is_valid = False
        try:
            # Lấy lại license_key giống trong _check_license_once (ưu tiên APP_LICENSE_KEY)
            license_key = None
            license_key = (APP_LICENSE_KEY or '').strip()
            if not license_key:
                activation = None
                try:
                    activation = get_activation_record()
                except Exception:
                    activation = None
                if activation is not None:
                    license_key = (getattr(activation, 'license_key', None) or '').strip()
            print(f"[LICENSE] License key đang dùng: {license_key}", flush=True)
        except Exception:
            pass

        final_msg = msg or "License không hợp lệ hoặc đã hết hạn"
        contact_msg = (
            f"{final_msg}\n\n"
            "Vui lòng liên hệ ADMIN để gia hạn:\n"
            "Nguyễn Công Đạt - 0375097105."
        )
        print(f"[LICENSE] License KHÔNG HỢP LỆ / HẾT HẠN - Server vẫn khởi động nhưng sẽ chặn tất cả truy cập.", flush=True)
        print(f"[LICENSE] Chi tiết: {contact_msg}", flush=True)
        print(f"[LICENSE] Hệ thống sẽ tự động kiểm tra lại license mỗi 60 giây.", flush=True)
    else:
        # License hợp lệ
        _license_is_valid = True
        print("[LICENSE] License hợp lệ, tiếp tục khởi động server...", flush=True)

    # --- Bước 2: Khởi động các dịch vụ nền (trong đó có license checker mỗi 60 giây) ---
    try:
        start_all_background_services()
        print("🚀 Tất cả dịch vụ nền đã được khởi động:")
        print("   🛡️ Backup: mỗi 60 phút, giữ 3 bản + Telegram")
        print("   🔑 Token Keep-Alive: mỗi 30 phút")
        print("   📅 Yearly reset: kiểm tra mỗi ngày, tự reset vào 1/1 hằng năm")
        print("   🔐 License check: verify online mỗi 60 giây")
    except Exception as e:
        print(f"⚠️ Lỗi khởi động dịch vụ nền: {e}")
        # Fallback: khởi động từng dịch vụ riêng lẻ
        try:
            ensure_backup_scheduler_started(interval_minutes=60, backup_dir="backups", retention=3, send_to_telegram=True)
            print("🛡️ Backup scheduler đã được khởi động riêng lẻ")
        except Exception as e2:
            print(f"⚠️ Không thể khởi động backup scheduler: {e2}")

        try:
            ensure_token_keepalive_started(interval_minutes=30)
            print("🔑 Token Keep-Alive đã được khởi động riêng lẻ")
        except Exception as e3:
            print(f"⚠️ Không thể khởi động Token Keep-Alive: {e3}")

        try:
            ensure_license_check_started(interval_seconds=60)
            print("🔐 License online checker đã được khởi động riêng lẻ")
        except Exception as e4:
            print(f"[LICENSE] Không thể khởi động license online checker: {e4}")

    # --- Bước 3: Chỉ chạy Flask server khi license hợp lệ ---
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    host = os.environ.get('FLASK_HOST', '0.0.0.0')
    port = int(os.environ.get('FLASK_PORT', 5000))

    app.run(debug=debug_mode, host=host, port=port)